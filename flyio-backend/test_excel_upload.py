"""엑셀 업로드 파싱 단위 검증 스크립트.
_parse_keyword_excel 로직을 독립적으로 검증 (네이버 API 호출 없음).
"""
import io
import re
import sys
from typing import Any, Dict, List

import openpyxl


def _parse_keyword_excel(file_bytes: bytes, default_bid: int, force_default_bid: bool = False) -> Dict[str, Any]:
    """routers/naver_ad.py의 _parse_keyword_excel 복사본 (단위 테스트용)"""
    items: List[Dict[str, Any]] = []
    errors: List[str] = []
    seen = set()

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("cp949", errors="replace")
        import csv
        rows = [tuple(r) for r in csv.reader(io.StringIO(text))]

    if not rows:
        return {"items": [], "errors": ["파일이 비어있습니다"], "total": 0}

    header = rows[0]
    header_cells = [str(c).strip() if c is not None else "" for c in header]
    header_lower = [h.lower() for h in header_cells]

    kw_idx = 0
    bid_idx = 1
    has_header = False
    kw_aliases = ["키워드", "keyword", "kw"]
    bid_aliases = ["입찰가", "bid", "bidamt", "입찰", "cpc"]

    for i, h in enumerate(header_lower):
        if h in kw_aliases:
            kw_idx = i
            has_header = True
        elif h in bid_aliases:
            bid_idx = i
            has_header = True

    data_rows = rows[1:] if has_header else rows
    kw_pattern = re.compile(r"^[\w가-힣\s\-\+]{1,40}$", re.UNICODE)

    for lineno, row in enumerate(data_rows, start=2 if has_header else 1):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        raw_kw = row[kw_idx] if kw_idx < len(row) else None
        raw_bid = row[bid_idx] if bid_idx < len(row) else None

        if raw_kw is None:
            continue
        keyword = str(raw_kw).strip()
        if not keyword:
            continue

        if len(keyword) > 40:
            errors.append(f"{lineno}행: 키워드 길이 초과 ({keyword[:20]}...)")
            continue
        if not kw_pattern.match(keyword):
            errors.append(f"{lineno}행: 허용되지 않는 문자 ({keyword})")
            continue
        if keyword in seen:
            continue
        seen.add(keyword)

        bid = default_bid
        if not force_default_bid and raw_bid is not None and str(raw_bid).strip() != "":
            try:
                bid_val = int(float(str(raw_bid).replace(",", "").strip()))
                if bid_val < 70:
                    errors.append(f"{lineno}행: 입찰가 최소 70원 ({bid_val}) → {default_bid}원 적용")
                    bid = default_bid
                elif bid_val > 100000:
                    errors.append(f"{lineno}행: 입찰가 최대 100000원 초과 ({bid_val}) → 100000원 적용")
                    bid = 100000
                else:
                    bid = bid_val
            except (ValueError, TypeError):
                errors.append(f"{lineno}행: 입찰가 파싱 실패 ({raw_bid}) → {default_bid}원 적용")
                bid = default_bid

        items.append({"keyword": keyword, "bid": bid, "row": lineno})

    return {"items": items, "errors": errors, "total": len(items)}


def make_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_keyword_only_force_default():
    data = make_excel([
        ("키워드",),
        ("병원 블로그 마케팅",),
        ("치과 광고",),
        ("성형외과 블로그",),
    ])
    result = _parse_keyword_excel(data, default_bid=150, force_default_bid=True)
    assert result["total"] == 3, f"expected 3, got {result['total']}"
    assert all(item["bid"] == 150 for item in result["items"]), "모든 bid가 150이어야 함"
    assert result["items"][0]["keyword"] == "병원 블로그 마케팅"
    print(f"[PASS] 키워드만 + 일괄 적용 → {result['total']}개, 모두 150원")


def test_bid_in_excel_but_forced():
    data = make_excel([
        ("키워드", "입찰가"),
        ("키워드1", 500),
        ("키워드2", 1000),
        ("키워드3", 2000),
    ])
    result = _parse_keyword_excel(data, default_bid=250, force_default_bid=True)
    assert result["total"] == 3
    assert all(item["bid"] == 250 for item in result["items"]), \
        f"강제 덮어쓰기 실패: {[item['bid'] for item in result['items']]}"
    print(f"[PASS] B열 입찰가 있지만 강제 덮어쓰기 → 모두 250원")


def test_bid_in_excel_respected():
    data = make_excel([
        ("키워드", "입찰가"),
        ("키워드1", 500),
        ("키워드2", 1000),
        ("키워드3", None),
    ])
    result = _parse_keyword_excel(data, default_bid=100, force_default_bid=False)
    assert result["total"] == 3
    bids = [item["bid"] for item in result["items"]]
    assert bids == [500, 1000, 100], f"예상 [500,1000,100], 실제 {bids}"
    print(f"[PASS] B열 입찰가 우선 → {bids}")


def test_invalid_bid_clamping():
    data = make_excel([
        ("키워드", "입찰가"),
        ("정상", 500),
        ("너무낮음", 50),
        ("너무높음", 200000),
        ("이상한값", "abc"),
    ])
    result = _parse_keyword_excel(data, default_bid=100, force_default_bid=False)
    bids = [item["bid"] for item in result["items"]]
    assert bids == [500, 100, 100000, 100], f"실제 {bids}"
    assert len(result["errors"]) == 3
    print(f"[PASS] 입찰가 범위 검증 → {bids} + 경고 {len(result['errors'])}건")


def test_dedup():
    data = make_excel([
        ("키워드",),
        ("중복키워드",),
        ("유니크",),
        ("중복키워드",),
        ("중복키워드",),
    ])
    result = _parse_keyword_excel(data, default_bid=100, force_default_bid=True)
    assert result["total"] == 2, f"중복 제거 실패: {result['total']}"
    print(f"[PASS] 중복 제거 → {result['total']}개")


def test_no_header():
    data = make_excel([
        ("바로키워드1", 300),
        ("바로키워드2", 400),
    ])
    result = _parse_keyword_excel(data, default_bid=100, force_default_bid=False)
    assert result["total"] == 2
    assert result["items"][0]["bid"] == 300
    print(f"[PASS] 헤더 없음 자동 감지 → {result['total']}개")


def test_invalid_keywords():
    data = make_excel([
        ("키워드",),
        ("정상키워드",),
        ("!@#$%^&*()",),
        ("가" * 50,),
    ])
    result = _parse_keyword_excel(data, default_bid=100, force_default_bid=True)
    assert result["total"] == 1, f"예상 1개, 실제 {result['total']}"
    assert len(result["errors"]) >= 2
    print(f"[PASS] 유효성 검증 → 유효 {result['total']}개, 경고 {len(result['errors'])}건")


def test_empty_excel():
    data = make_excel([])
    result = _parse_keyword_excel(data, default_bid=100, force_default_bid=True)
    assert result["total"] == 0
    print(f"[PASS] 빈 엑셀 처리")


def test_large_scale():
    rows = [("키워드",)] + [(f"테스트키워드{i}",) for i in range(1000)]
    data = make_excel(rows)
    result = _parse_keyword_excel(data, default_bid=250, force_default_bid=True)
    assert result["total"] == 1000, f"예상 1000, 실제 {result['total']}"
    assert all(item["bid"] == 250 for item in result["items"])
    print(f"[PASS] 대량 1000개 → 전부 250원")


def test_csv_fallback():
    csv_bytes = "키워드,입찰가\n키워드A,200\n키워드B,300\n".encode("utf-8-sig")
    result = _parse_keyword_excel(csv_bytes, default_bid=100, force_default_bid=False)
    assert result["total"] == 2
    bids = [item["bid"] for item in result["items"]]
    assert bids == [200, 300]
    print(f"[PASS] CSV 파일 처리 → {result['total']}개")


if __name__ == "__main__":
    tests = [
        test_keyword_only_force_default,
        test_bid_in_excel_but_forced,
        test_bid_in_excel_respected,
        test_invalid_bid_clamping,
        test_dedup,
        test_no_header,
        test_invalid_keywords,
        test_empty_excel,
        test_large_scale,
        test_csv_fallback,
    ]
    passed = 0
    failed = 0
    for t in tests:
        try:
            t()
            passed += 1
        except AssertionError as e:
            print(f"[FAIL] {t.__name__} - {e}")
            failed += 1
        except Exception as e:
            print(f"[ERROR] {t.__name__} - {type(e).__name__}: {e}")
            failed += 1

    print(f"\n{'='*50}")
    print(f"Total {passed + failed} / Passed {passed} / Failed {failed}")
    sys.exit(0 if failed == 0 else 1)
