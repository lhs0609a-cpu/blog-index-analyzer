"""
네이버 광고 자동 최적화 API 라우터
"""
from fastapi import APIRouter, HTTPException, Query, BackgroundTasks, Depends, UploadFile, File, Form
from pydantic import BaseModel, Field
from routers.auth_deps import get_user_id_with_fallback
from routers.admin import require_admin
from typing import Optional, List, Dict, Any, Tuple, Set, Union
from datetime import datetime, timedelta
import logging
import asyncio
import io
import json as _json_lib
import random
import re
import httpx

from services.naver_ad_service import (
    NaverAdOptimizer,
    BidStrategy,
    get_optimizer
)
from database.naver_ad_db import (
    init_naver_ad_tables,
    get_optimization_settings,
    save_optimization_settings,
    get_bid_history,
    get_bid_changes_summary,
    get_keyword_performance,
    get_performance_summary,
    get_excluded_keywords,
    restore_excluded_keyword,
    get_discovered_keywords,
    update_discovered_keyword_status,
    get_daily_reports,
    get_optimization_logs,
    get_dashboard_stats,
    save_bid_change,
    save_excluded_keyword,
    save_discovered_keywords,
    save_optimization_log,
    # 새로 추가된 함수들
    save_ad_account,
    get_ad_account,
    update_ad_account_status,
    delete_ad_account,
    save_efficiency_tracking,
    get_efficiency_summary,
    get_efficiency_history,
    save_trending_keywords,
    get_trending_keywords,
    update_trending_keyword_status,
    # 대량 등록
    create_bulk_upload_job,
    update_bulk_upload_job,
    get_bulk_upload_job,
    list_bulk_upload_jobs,
    get_bulk_upload_failures,
    # 검색량 필터
    create_volume_filter_job,
    update_volume_filter_job,
    get_volume_filter_job,
    list_volume_filter_jobs,
    get_volume_filter_results,
    count_volume_filter_results,
    set_volume_filter_control,
)

logger = logging.getLogger(__name__)
router = APIRouter()

# cleanup-by-score 의 BackgroundTask 동시 실행 제한 — 광고주별 1개만.
# 사용자가 긴급삭제 버튼을 연타하거나 두 탭에서 동시에 누르면 50k DELETE 작업이
# 여러 개 쌓여서 event loop CPU + Naver API rate limit 폭주. customer_id 단위로
# 진행 중 표식 두고 두 번째 요청은 즉시 409 반환.
_BULK_CLEANUP_RUNNING: set[int] = set()

# 테이블 초기화
try:
    init_naver_ad_tables()
except Exception as e:
    logger.error(f"Failed to initialize naver ad tables: {e}")


# ============ Pydantic 모델 ============

class OptimizationSettingsRequest(BaseModel):
    strategy: str = Field(default="balanced", description="입찰 전략 (balanced, target_roas, target_position, target_cpa, maximize_conversions)")
    target_roas: float = Field(default=300, description="목표 ROAS (%)")
    target_position: int = Field(default=3, description="목표 순위")
    target_cpa: int = Field(default=20000, description="목표 CPA (전환당 비용)")
    conversion_value: int = Field(default=59400, description="전환 가치 (LTV)")
    max_bid_change_ratio: float = Field(default=0.2, description="최대 입찰 변경폭")
    min_bid: int = Field(default=70, description="최소 입찰가")
    max_bid: int = Field(default=100000, description="최대 입찰가")
    min_ctr: float = Field(default=0.01, description="최소 CTR")
    max_cost_no_conv: int = Field(default=50000, description="전환없이 최대 비용")
    min_quality_score: int = Field(default=4, description="최소 품질지수")
    evaluation_days: int = Field(default=7, description="평가 기간 (일)")
    optimization_interval: int = Field(default=60, description="최적화 주기 (초)")
    is_auto_optimization: bool = Field(default=False, description="자동 최적화 활성화")
    blacklist_keywords: List[str] = Field(default=[], description="제외할 키워드 패턴")
    core_terms: List[str] = Field(default=[], description="핵심 키워드")
    # 전환 키워드 자동 발굴 설정
    conversion_keywords: List[str] = Field(default=["가격", "비용", "구독", "결제", "신청", "구매", "추천", "비교", "후기"], description="전환 의도 키워드")


class KeywordDiscoveryRequest(BaseModel):
    seed_keywords: List[str] = Field(..., description="시드 키워드 목록")
    ad_group_id: Optional[str] = Field(None, description="추가할 광고그룹 ID")
    max_keywords: int = Field(default=50, description="최대 키워드 수")
    min_search_volume: int = Field(default=100, description="최소 검색량")
    max_competition: float = Field(default=0.85, description="최대 경쟁도")
    auto_add: bool = Field(default=False, description="자동 추가 여부")


class ManualBidUpdateRequest(BaseModel):
    keyword_id: str = Field(..., description="키워드 ID")
    new_bid: int = Field(..., description="새 입찰가")
    reason: Optional[str] = Field(default="수동 변경", description="변경 사유")


class BulkKeywordAddRequest(BaseModel):
    ad_group_id: str = Field(..., description="광고그룹 ID")
    keywords: List[str] = Field(..., description="키워드 목록")
    default_bid: int = Field(default=100, description="기본 입찰가")


class KeywordWithBid(BaseModel):
    keyword: str = Field(..., description="키워드")
    bid: int = Field(..., description="입찰가 (원)")


class BulkKeywordWithBidRequest(BaseModel):
    ad_group_id: str = Field(..., description="광고그룹 ID")
    items: List[KeywordWithBid] = Field(..., description="키워드+입찰가 목록")
    default_bid: int = Field(default=100, description="개별 입찰가가 없을 때의 기본값")


# ============ 대시보드 ============

@router.get("/dashboard")
async def get_dashboard(user_id: int = Depends(get_user_id_with_fallback)):
    """대시보드 통계 조회"""
    try:
        stats = get_dashboard_stats(user_id)
        return {
            "success": True,
            "data": stats
        }
    except Exception as e:
        logger.error(f"Dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/dashboard/realtime")
async def get_realtime_status(user_id: int = Depends(get_user_id_with_fallback)):
    """실시간 최적화 상태"""
    try:
        optimizer = get_optimizer()
        status = await optimizer.get_optimization_status()

        # 최근 입찰 변경
        recent_changes = get_bid_history(user_id, limit=10)

        return {
            "success": True,
            "data": {
                **status,
                "recent_changes": recent_changes
            }
        }
    except Exception as e:
        logger.error(f"Realtime status error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 최적화 설정 ============

@router.get("/settings")
async def get_settings(user_id: int = Depends(get_user_id_with_fallback)):
    """최적화 설정 조회"""
    settings = get_optimization_settings(user_id)

    if not settings:
        # 기본 설정 생성
        settings = save_optimization_settings(user_id, {})

    return {
        "success": True,
        "data": settings
    }


@router.post("/settings")
async def update_settings(
    request: OptimizationSettingsRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """최적화 설정 저장"""
    try:
        # 전략 유효성 검사
        valid_strategies = [s.value for s in BidStrategy]
        if request.strategy not in valid_strategies:
            raise HTTPException(
                status_code=400,
                detail=f"유효하지 않은 전략입니다. 가능한 값: {valid_strategies}"
            )

        settings = save_optimization_settings(user_id, request.dict())

        # 옵티마이저 설정 업데이트
        optimizer = get_optimizer()
        optimizer.bid_optimizer.set_strategy(
            strategy=BidStrategy(request.strategy),
            target_roas=request.target_roas,
            target_position=request.target_position,
            target_cpa=request.target_cpa,
            conversion_value=request.conversion_value,
            max_bid_change_ratio=request.max_bid_change_ratio,
            min_bid=request.min_bid,
            max_bid=request.max_bid
        )

        optimizer.exclusion.set_thresholds(
            min_ctr=request.min_ctr,
            max_cost_no_conv=request.max_cost_no_conv,
            min_quality_score=request.min_quality_score,
            evaluation_days=request.evaluation_days
        )

        optimizer.discovery.set_filters(
            blacklist=request.blacklist_keywords,
            core_terms=request.core_terms
        )

        save_optimization_log(user_id, "settings_update", "최적화 설정이 변경되었습니다", request.dict())

        return {
            "success": True,
            "message": "설정이 저장되었습니다",
            "data": settings
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Settings update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 자동 최적화 제어 ============

@router.post("/optimization/start")
async def start_optimization(
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_user_id_with_fallback),
    ad_group_ids: Optional[List[str]] = Query(None, description="광고그룹 ID 목록")
):
    """자동 최적화 시작"""
    try:
        optimizer = get_optimizer()

        if optimizer.is_running:
            return {
                "success": False,
                "message": "이미 최적화가 실행 중입니다"
            }

        # 설정 업데이트
        save_optimization_settings(user_id, {"is_auto_optimization": True})

        # 백그라운드에서 최적화 실행
        background_tasks.add_task(optimizer.start_auto_optimization, ad_group_ids)

        save_optimization_log(user_id, "optimization_start", "자동 최적화가 시작되었습니다")

        return {
            "success": True,
            "message": "자동 최적화가 시작되었습니다",
            "interval_seconds": optimizer.optimization_interval
        }
    except Exception as e:
        logger.error(f"Start optimization error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/optimization/stop")
async def stop_optimization(user_id: int = Depends(get_user_id_with_fallback)):
    """자동 최적화 중지"""
    try:
        optimizer = get_optimizer()
        optimizer.stop_auto_optimization()

        save_optimization_settings(user_id, {"is_auto_optimization": False})
        save_optimization_log(user_id, "optimization_stop", "자동 최적화가 중지되었습니다")

        return {
            "success": True,
            "message": "자동 최적화가 중지되었습니다"
        }
    except Exception as e:
        logger.error(f"Stop optimization error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/optimization/run-once")
async def run_optimization_once(
    user_id: int = Depends(get_user_id_with_fallback),
    ad_group_ids: Optional[List[str]] = Query(None, description="광고그룹 ID 목록")
):
    """입찰 최적화 1회 실행"""
    try:
        optimizer = get_optimizer()

        # 설정 로드 및 적용
        settings = get_optimization_settings(user_id)
        if settings:
            optimizer.bid_optimizer.set_strategy(
                strategy=BidStrategy(settings.get("strategy", "balanced")),
                target_roas=settings.get("target_roas", 300),
                target_position=settings.get("target_position", 3)
            )

        # 최적화 실행
        changes = await optimizer.bid_optimizer.optimize_all_keywords(ad_group_ids)

        # 변경 내역 저장
        for change in changes:
            save_bid_change(
                user_id=user_id,
                keyword_id=change.keyword_id,
                keyword_text=change.keyword,
                old_bid=change.old_bid,
                new_bid=change.new_bid,
                reason=change.reason,
                strategy=settings.get("strategy", "balanced") if settings else "balanced"
            )

        save_optimization_log(
            user_id, "optimization_run",
            f"입찰 최적화 완료: {len(changes)}개 키워드 변경",
            {"changes_count": len(changes)}
        )

        return {
            "success": True,
            "message": f"{len(changes)}개 키워드의 입찰가가 최적화되었습니다",
            "changes": [
                {
                    "keyword_id": c.keyword_id,
                    "keyword": c.keyword,
                    "old_bid": c.old_bid,
                    "new_bid": c.new_bid,
                    "reason": c.reason
                }
                for c in changes
            ]
        }
    except Exception as e:
        logger.error(f"Run optimization error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 키워드 발굴 ============

@router.post("/keywords/discover")
async def discover_keywords(
    request: KeywordDiscoveryRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """연관 키워드 발굴"""
    try:
        optimizer = get_optimizer()

        # 필터 설정
        settings = get_optimization_settings(user_id)
        optimizer.discovery.set_filters(
            min_search_volume=request.min_search_volume,
            max_competition=request.max_competition,
            blacklist=settings.get("blacklist_keywords", []) if settings else [],
            core_terms=settings.get("core_terms", []) if settings else []
        )

        # 키워드 발굴
        suggestions = await optimizer.discovery.discover_keywords(
            request.seed_keywords,
            request.max_keywords
        )

        # 발굴 결과 저장
        save_discovered_keywords(
            user_id,
            [
                {
                    "keyword": s.keyword,
                    "monthly_search_count": s.monthly_search_count,
                    "monthly_pc_search_count": s.monthly_pc_search_count,
                    "monthly_mobile_search_count": s.monthly_mobile_search_count,
                    "competition_level": s.competition_level,
                    "competition_index": s.competition_index,
                    "suggested_bid": s.suggested_bid,
                    "relevance_score": s.relevance_score,
                    "potential_score": s.potential_score
                }
                for s in suggestions
            ],
            seed_keyword=", ".join(request.seed_keywords)
        )

        # 자동 추가
        added_count = 0
        if request.auto_add and request.ad_group_id:
            added = await optimizer.bulk_manager.bulk_add_keywords(
                request.ad_group_id,
                suggestions
            )
            added_count = len(added)

            # 상태 업데이트
            for s in suggestions:
                update_discovered_keyword_status(
                    user_id, s.keyword, "added", request.ad_group_id
                )

        save_optimization_log(
            user_id, "keyword_discovery",
            f"키워드 발굴 완료: {len(suggestions)}개 발굴, {added_count}개 추가",
            {"discovered": len(suggestions), "added": added_count}
        )

        return {
            "success": True,
            "discovered": len(suggestions),
            "added": added_count,
            "keywords": [
                {
                    "keyword": s.keyword,
                    "monthly_search_count": s.monthly_search_count,
                    "competition_level": s.competition_level,
                    "suggested_bid": s.suggested_bid,
                    "relevance_score": round(s.relevance_score, 2),
                    "potential_score": round(s.potential_score, 2)
                }
                for s in suggestions
            ]
        }
    except Exception as e:
        logger.error(f"Keyword discovery error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/keywords/discover-conversion")
async def discover_conversion_keywords(
    request: KeywordDiscoveryRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """전환 키워드만 집중 발굴 (구매의도 높은 키워드)"""
    try:
        optimizer = get_optimizer()

        # 필터 설정
        settings = get_optimization_settings(user_id)
        optimizer.discovery.set_filters(
            min_search_volume=request.min_search_volume,
            max_competition=request.max_competition,
            blacklist=settings.get("blacklist_keywords", []) if settings else [],
            core_terms=settings.get("core_terms", []) if settings else []
        )

        # 전환 키워드 발굴
        suggestions = await optimizer.discovery.discover_conversion_keywords(
            request.seed_keywords,
            request.max_keywords
        )

        # 발굴 결과 저장
        save_discovered_keywords(
            user_id,
            [
                {
                    "keyword": s.keyword,
                    "monthly_search_count": s.monthly_search_count,
                    "monthly_pc_search_count": s.monthly_pc_search_count,
                    "monthly_mobile_search_count": s.monthly_mobile_search_count,
                    "competition_level": s.competition_level,
                    "competition_index": s.competition_index,
                    "suggested_bid": s.suggested_bid,
                    "relevance_score": s.relevance_score,
                    "potential_score": s.potential_score,
                    "is_conversion_keyword": True
                }
                for s in suggestions
            ],
            seed_keyword=", ".join(request.seed_keywords)
        )

        # 자동 추가
        added_count = 0
        if request.auto_add and request.ad_group_id:
            added = await optimizer.bulk_manager.bulk_add_keywords(
                request.ad_group_id,
                suggestions
            )
            added_count = len(added)

        save_optimization_log(
            user_id, "conversion_keyword_discovery",
            f"전환 키워드 발굴 완료: {len(suggestions)}개 발굴, {added_count}개 추가",
            {"discovered": len(suggestions), "added": added_count}
        )

        return {
            "success": True,
            "message": "전환 키워드 발굴 완료",
            "discovered": len(suggestions),
            "added": added_count,
            "keywords": [
                {
                    "keyword": s.keyword,
                    "monthly_search_count": s.monthly_search_count,
                    "competition_level": s.competition_level,
                    "suggested_bid": s.suggested_bid,
                    "relevance_score": round(s.relevance_score, 2),
                    "potential_score": round(s.potential_score, 2),
                    "conversion_intent": "높음" if s.potential_score > 50 else "중간" if s.potential_score > 20 else "낮음"
                }
                for s in suggestions
            ]
        }
    except Exception as e:
        logger.error(f"Conversion keyword discovery error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/keywords/discovered")
async def get_discovered(
    user_id: int = Depends(get_user_id_with_fallback),
    status: Optional[str] = Query(None, description="상태 필터"),
    limit: int = Query(100, description="조회 개수")
):
    """발굴된 키워드 목록 조회"""
    keywords = get_discovered_keywords(user_id, status, limit)
    return {
        "success": True,
        "count": len(keywords),
        "keywords": keywords
    }


@router.post("/keywords/bulk-add")
async def bulk_add_keywords(
    request: BulkKeywordAddRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """키워드 대량 추가"""
    try:
        optimizer = get_optimizer()

        # KeywordSuggestion 객체로 변환
        from services.naver_ad_service import KeywordSuggestion
        suggestions = [
            KeywordSuggestion(keyword=kw, suggested_bid=request.default_bid)
            for kw in request.keywords
        ]

        # 대량 추가
        added = await optimizer.bulk_manager.bulk_add_keywords(
            request.ad_group_id,
            suggestions,
            request.default_bid
        )

        save_optimization_log(
            user_id, "bulk_add",
            f"키워드 대량 추가: {len(added)}개",
            {"count": len(added), "ad_group_id": request.ad_group_id}
        )

        return {
            "success": True,
            "message": f"{len(added)}개 키워드가 추가되었습니다",
            "added_count": len(added)
        }
    except Exception as e:
        logger.error(f"Bulk add error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/keywords/bulk-add-with-bids")
async def bulk_add_keywords_with_bids(
    request: BulkKeywordWithBidRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """키워드별 개별 입찰가로 대량 추가"""
    try:
        optimizer = get_optimizer()

        from services.naver_ad_service import KeywordSuggestion
        suggestions = [
            KeywordSuggestion(
                keyword=item.keyword,
                suggested_bid=item.bid if item.bid and item.bid >= 70 else request.default_bid
            )
            for item in request.items
        ]

        added = await optimizer.bulk_manager.bulk_add_keywords(
            request.ad_group_id,
            suggestions,
            request.default_bid
        )

        save_optimization_log(
            user_id, "bulk_add_with_bids",
            f"키워드 대량 추가(개별 입찰가): {len(added)}개",
            {"count": len(added), "ad_group_id": request.ad_group_id}
        )

        return {
            "success": True,
            "message": f"{len(added)}개 키워드가 개별 입찰가로 추가되었습니다",
            "added_count": len(added),
            "total_requested": len(request.items)
        }
    except Exception as e:
        logger.error(f"Bulk add with bids error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _parse_keyword_excel(file_bytes: bytes, default_bid: int, force_default_bid: bool = False) -> Dict[str, Any]:
    """엑셀/CSV 파일에서 키워드+입찰가 파싱.
    컬럼: '키워드'(필수), '입찰가'(선택). 헤더가 없으면 1열=키워드, 2열=입찰가로 해석.
    force_default_bid=True면 엑셀의 입찰가를 무시하고 모든 키워드에 default_bid 적용.
    """
    import openpyxl

    items: List[Dict[str, Any]] = []
    errors: List[str] = []
    seen = set()

    # 엑셀 시도
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        # CSV fallback
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("cp949", errors="replace")
        import csv
        rows = [tuple(r) for r in csv.reader(io.StringIO(text))]

    if not rows:
        return {"items": [], "errors": ["파일이 비어있습니다"], "total": 0}

    # 헤더 감지
    header = rows[0]
    header_cells = [str(c).strip() if c is not None else "" for c in header]
    header_lower = [h.lower() for h in header_cells]

    kw_idx = 0
    bid_idx = 1
    has_header = False

    kw_aliases = ["키워드", "keyword", "kw"]
    bid_aliases = ["입찰가", "bid", "bidamt", "입찰", "cpc"]

    for i, h in enumerate(header_lower):
        if h in kw_aliases:
            kw_idx = i
            has_header = True
        elif h in bid_aliases:
            bid_idx = i
            has_header = True

    data_rows = rows[1:] if has_header else rows
    kw_pattern = re.compile(r"^[\w가-힣\s\-\+]{1,40}$", re.UNICODE)

    for lineno, row in enumerate(data_rows, start=2 if has_header else 1):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        raw_kw = row[kw_idx] if kw_idx < len(row) else None
        raw_bid = row[bid_idx] if bid_idx < len(row) else None

        if raw_kw is None:
            continue
        keyword = str(raw_kw).strip()
        if not keyword:
            continue

        # 네이버 키워드 제약: 공백/특수문자 과다 필터
        if len(keyword) > 40:
            errors.append(f"{lineno}행: 키워드 길이 초과 ({keyword[:20]}...)")
            continue
        if not kw_pattern.match(keyword):
            errors.append(f"{lineno}행: 허용되지 않는 문자 ({keyword})")
            continue
        if keyword in seen:
            continue
        seen.add(keyword)

        bid = default_bid
        if not force_default_bid and raw_bid is not None and str(raw_bid).strip() != "":
            try:
                bid_val = int(float(str(raw_bid).replace(",", "").strip()))
                if bid_val < 70:
                    errors.append(f"{lineno}행: 입찰가 최소 70원 ({bid_val}) → {default_bid}원 적용")
                    bid = default_bid
                elif bid_val > 100000:
                    errors.append(f"{lineno}행: 입찰가 최대 100000원 초과 ({bid_val}) → 100000원 적용")
                    bid = 100000
                else:
                    bid = bid_val
            except (ValueError, TypeError):
                errors.append(f"{lineno}행: 입찰가 파싱 실패 ({raw_bid}) → {default_bid}원 적용")
                bid = default_bid

        items.append({"keyword": keyword, "bid": bid, "row": lineno})

    return {"items": items, "errors": errors, "total": len(items)}


# ============ 검색량 필터링 (50만 규모) ============

@router.post("/keywords/volume-filter")
async def start_volume_filter(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="엑셀/CSV - A열 키워드"),
    min_volume: int = Form(default=10, description="월 총 검색량 최소치"),
    test_size: int = Form(default=10000, description="캐너리 테스트 크기 (0=비활성)"),
    min_pass_rate_pct: float = Form(default=2.0, description="캐너리 최소 통과율(%)"),
    auto_continue_on_canary: bool = Form(default=True, description="캐너리 통과시 자동 계속"),
    user_id: int = Depends(get_user_id_with_fallback),
):
    """검색량 필터링 작업 시작
    - 캐너리: 첫 test_size개 처리 후 통과율 평가 → 임계치 이상이면 자동 계속, 미만이면 중단
    - auto_continue_on_canary=False면 캐너리 통과 여부와 무관하게 미달 시 대기
    - 취소/일시정지/재개 가능
    """
    try:
        if min_volume < 0 or min_volume > 100000:
            raise HTTPException(status_code=400, detail="min_volume 범위 오류 (0~100000)")
        if test_size < 0:
            raise HTTPException(status_code=400, detail="test_size 범위 오류")

        content = await file.read()
        if len(content) > 100 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="파일은 100MB 이하")

        parsed = _parse_keyword_excel(content, default_bid=100, force_default_bid=True)
        if parsed["total"] == 0:
            raise HTTPException(status_code=400, detail="유효 키워드 없음")

        keywords = [item["keyword"] for item in parsed["items"]]
        total = len(keywords)

        account = get_ad_account(user_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="네이버 광고 계정을 먼저 연동하세요")

        # Job 생성 (keywords_file 경로 미리 확보하려면 id 필요해서 후처리)
        job_id = create_volume_filter_job(
            user_id=user_id,
            filename=file.filename or "uploaded.xlsx",
            min_volume=min_volume,
            total_keywords=total,
            test_size=test_size,
            min_pass_rate_pct=min_pass_rate_pct,
            auto_continue_on_canary=auto_continue_on_canary,
        )

        # 키워드를 파일로 저장 (재개용)
        from services.volume_filter import VolumeFilterService
        from database.naver_ad_db import DATA_DIR
        kw_path = VolumeFilterService.save_keywords_file(job_id, keywords, DATA_DIR)
        update_volume_filter_job(job_id, keywords_file=kw_path)

        async def _run():
            from services.volume_filter import VolumeFilterService, VolumeFilterConfig
            from services.naver_ad_service import NaverAdApiClient
            try:
                client = NaverAdApiClient()
                client.customer_id = account.get("customer_id")
                client.api_key = account.get("api_key")
                client.secret_key = account.get("secret_key")

                svc = VolumeFilterService(client)
                cfg = VolumeFilterConfig(
                    job_id=job_id, user_id=user_id, min_volume=min_volume,
                    test_size=test_size,
                    min_pass_rate_pct=min_pass_rate_pct,
                    auto_continue_on_canary=auto_continue_on_canary,
                )
                await svc.run(cfg, keywords)
            except Exception as e:
                logger.exception(f"[Filter {job_id}] 실행 실패")
                update_volume_filter_job(
                    job_id, status="failed",
                    error_message=str(e)[:1000],
                    completed_at=datetime.now().isoformat(),
                )

        background_tasks.add_task(_run)

        estimated_seconds = int(total / 5 * 0.4)

        save_optimization_log(
            user_id, "volume_filter_start",
            f"검색량 필터 시작 (job #{job_id}): {total}개 (임계치 {min_volume}, 캐너리 {test_size})",
            {"job_id": job_id, "total": total, "min_volume": min_volume,
             "test_size": test_size}
        )

        return {
            "success": True,
            "job_id": job_id,
            "total_keywords": total,
            "min_volume": min_volume,
            "test_size": test_size,
            "min_pass_rate_pct": min_pass_rate_pct,
            "estimated_seconds": estimated_seconds,
            "estimated_minutes": round(estimated_seconds / 60, 1),
            "canary_estimated_seconds": int(test_size / 5 * 0.4) if test_size else 0,
            "message": (
                f"백그라운드 필터링 시작. 캐너리 {test_size}개 테스트 "
                f"(예상 {int(test_size / 5 * 0.4 / 60)}분) 후 자동 판단"
                if test_size else
                f"백그라운드 필터링 시작 (예상 {estimated_seconds // 60}분)"
            ),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("volume filter start error")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 관리자 전용: AI 씨앗/앵커 제안 ============

class AiSuggestSeedsRequest(BaseModel):
    topic: str = Field(..., description="주제 또는 카테고리 (예: 대출, 성형외과, 인테리어)")
    target_count: int = Field(default=10000, description="목표 수집 키워드 수")


@router.post("/keywords/ai-suggest-seeds")
async def ai_suggest_seeds(
    request: AiSuggestSeedsRequest,
    admin: dict = Depends(require_admin),
):
    """주제 + 목표 개수 → GPT가 씨앗/앵커/블랙리스트 + BFS 파라미터 자동 제안.
    관리자 전용. 응답을 프론트에서 보여주고 사용자가 수정 후 실제 확장 돌림.
    """
    topic = (request.topic or "").strip()
    if not topic:
        raise HTTPException(status_code=400, detail="topic이 비어있습니다")
    if len(topic) > 100:
        raise HTTPException(status_code=400, detail="topic이 너무 깁니다 (최대 100자)")
    if request.target_count <= 0 or request.target_count > 1000000:
        raise HTTPException(status_code=400, detail="target_count 범위 오류 (1~1000000)")

    from services.ai_seed_suggester import suggest_keyword_setup
    result = await suggest_keyword_setup(topic, request.target_count)
    if not result.get("success"):
        raise HTTPException(status_code=502, detail=result.get("message", "AI 제안 실패"))

    return result


# ============ 관리자 전용: 씨앗 AI 증폭 ============

class AiAmplifySeedsRequest(BaseModel):
    seeds: List[str] = Field(..., description="원본 씨앗 (1~100개)")
    target_count: int = Field(default=50, description="목표 씨앗 수 (입력의 N배, 최대 500)")


@router.post("/keywords/ai-amplify-seeds")
async def ai_amplify_seeds(
    request: AiAmplifySeedsRequest,
    admin: dict = Depends(require_admin),
):
    """씨앗 N개를 GPT가 패턴 분석해서 target_count개로 펼침.
    예: 10개 → 50개 (5배). 원본 씨앗은 결과에 반드시 포함.
    """
    seeds = [s.strip() for s in request.seeds if s and s.strip()]
    if not seeds:
        raise HTTPException(status_code=400, detail="씨앗이 비어있습니다")
    if len(seeds) > 100:
        raise HTTPException(status_code=400, detail="원본 씨앗 최대 100개")
    if request.target_count < len(seeds) or request.target_count > 500:
        raise HTTPException(status_code=400, detail=f"target_count 범위 오류 ({len(seeds)}~500)")

    from services.ai_seed_suggester import amplify_seeds
    result = await amplify_seeds(seeds, request.target_count)
    if not result.get("success"):
        raise HTTPException(status_code=502, detail=result.get("message", "AI 증폭 실패"))
    return result


# ============ 관리자 전용: AI 키워드 자동 확장 ============

class AiKeywordExpandRequest(BaseModel):
    seeds: List[str] = Field(..., description="씨앗 키워드 목록 (1~50개)")
    min_volume: int = Field(default=5, description="월 총 검색량 최소치")
    max_total_kept: int = Field(default=10000, description="최종 저장 최대 키워드 수")
    max_api_calls: int = Field(default=2000, description="네이버 API 총 호출 상한")
    max_depth: int = Field(default=3, description="BFS 확장 깊이")
    top_n_per_level: int = Field(default=50, description="각 레벨에서 다음 확장 대상 상위 개수")
    core_terms: List[str] = Field(default=[], description="반드시 포함돼야 할 앵커 단어 목록 (비우면 씨앗에서 자동 추출)")
    blacklist: List[str] = Field(default=[], description="포함되면 즉시 제외할 단어 목록")
    # 실시간 캠페인 등록 옵션
    stream_register: bool = Field(default=False, description="수집과 동시에 네이버 캠페인 실시간 등록")
    campaign_prefix: str = Field(default="", description="실시간 등록 시 캠페인 이름 prefix")
    bid: int = Field(default=100, description="키워드 공통 입찰가 (원)")
    daily_budget: int = Field(default=10000, description="캠페인 일 예산 (원)")
    campaign_tp: str = Field(default="WEB_SITE", description="캠페인 유형")
    keywords_per_ad_group: int = Field(default=1000, description="광고그룹당 키워드 수")
    stream_batch_size: int = Field(default=10, description="몇 개 찰 때마다 등록할지 (작을수록 실시간)")


@router.post("/keywords/ai-expand")
async def start_ai_keyword_expand(
    request: AiKeywordExpandRequest,
    background_tasks: BackgroundTasks,
    admin: dict = Depends(require_admin),
):
    """씨앗 키워드에서 출발해 네이버 연관검색어를 BFS 확장하며 검색량 필터링.
    관리자만 사용 가능. 기존 volume_filter_jobs 테이블을 재사용하므로
    진행률 조회/결과 다운로드/광고 등록 플로우를 그대로 쓸 수 있다.
    """
    try:
        seeds = [s.strip() for s in request.seeds if s and s.strip()]
        if not seeds:
            raise HTTPException(status_code=400, detail="씨앗 키워드가 비어있습니다")
        if len(seeds) > 500:
            raise HTTPException(status_code=400, detail="씨앗은 최대 500개까지 허용됩니다")
        if request.min_volume < 0 or request.min_volume > 100000:
            raise HTTPException(status_code=400, detail="min_volume 범위 오류")
        if request.max_total_kept <= 0 or request.max_total_kept > 1000000:
            raise HTTPException(status_code=400, detail="max_total_kept 범위 오류 (1~1000000)")
        if request.max_api_calls <= 0 or request.max_api_calls > 50000:
            raise HTTPException(status_code=400, detail="max_api_calls 범위 오류 (1~50000)")
        if request.max_depth < 0 or request.max_depth > 5:
            raise HTTPException(status_code=400, detail="max_depth 범위 오류 (0~5)")
        if request.stream_register:
            if not request.campaign_prefix or len(request.campaign_prefix) < 2:
                raise HTTPException(status_code=400, detail="실시간 등록 시 campaign_prefix 필수 (2자 이상)")
            if request.bid < 70 or request.bid > 100000:
                raise HTTPException(status_code=400, detail="입찰가 70~100000원")
            if request.stream_batch_size < 1 or request.stream_batch_size > 100:
                raise HTTPException(status_code=400, detail="stream_batch_size 1~100")
            if request.keywords_per_ad_group < 10 or request.keywords_per_ad_group > 1000:
                raise HTTPException(status_code=400, detail="keywords_per_ad_group 10~1000")

        admin_id = admin["id"]
        account = get_ad_account(admin_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="네이버 광고 계정을 먼저 연동하세요")

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"AI확장_{seeds[0][:20]}_{ts}"

        # 기존 필터 테이블을 재사용 (total_keywords = max_api_calls 기준으로 표기)
        job_id = create_volume_filter_job(
            user_id=admin_id,
            filename=filename,
            min_volume=request.min_volume,
            total_keywords=request.max_api_calls,
            test_size=0,
            min_pass_rate_pct=0.0,
            auto_continue_on_canary=True,
        )
        update_volume_filter_job(
            job_id,
            current_step=f"AI 확장 대기: 씨앗 {len(seeds)}개",
        )

        async def _run():
            from services.ai_keyword_expander import AiKeywordExpander, AiExpandConfig
            from services.naver_ad_service import NaverAdApiClient
            try:
                client = NaverAdApiClient()
                client.customer_id = account.get("customer_id")
                client.api_key = account.get("api_key")
                client.secret_key = account.get("secret_key")

                expander = AiKeywordExpander(client)
                cfg = AiExpandConfig(
                    job_id=job_id,
                    user_id=admin_id,
                    seeds=seeds,
                    min_volume=request.min_volume,
                    max_total_kept=request.max_total_kept,
                    max_api_calls=request.max_api_calls,
                    max_depth=request.max_depth,
                    top_n_per_level=request.top_n_per_level,
                    core_terms=request.core_terms or None,
                    blacklist=request.blacklist or None,
                    stream_register=request.stream_register,
                    campaign_prefix=request.campaign_prefix,
                    bid=request.bid,
                    daily_budget=request.daily_budget,
                    campaign_tp=request.campaign_tp,
                    keywords_per_ad_group=request.keywords_per_ad_group,
                    stream_batch_size=request.stream_batch_size,
                )
                await expander.run(cfg)
            except Exception as e:
                logger.exception(f"[AiExpand {job_id}] 실행 실패")
                update_volume_filter_job(
                    job_id, status="failed",
                    error_message=str(e)[:1000],
                    completed_at=datetime.now().isoformat(),
                )

        background_tasks.add_task(_run)

        save_optimization_log(
            admin_id, "ai_keyword_expand_start",
            f"AI 키워드 확장 시작 (job #{job_id}): 씨앗 {len(seeds)}개, "
            f"depth {request.max_depth}, max_api {request.max_api_calls}",
            {"job_id": job_id, "seeds": seeds[:10], "seed_count": len(seeds),
             "min_volume": request.min_volume, "max_api_calls": request.max_api_calls},
        )

        return {
            "success": True,
            "job_id": job_id,
            "seed_count": len(seeds),
            "min_volume": request.min_volume,
            "max_api_calls": request.max_api_calls,
            "estimated_minutes": round(request.max_api_calls * 0.35 / 60, 1),
            "message": f"AI 확장 시작. 씨앗 {len(seeds)}개 → 예상 {round(request.max_api_calls * 0.35 / 60, 1)}분",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("ai keyword expand start error")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/keywords/volume-filter/{job_id}/cancel")
async def cancel_volume_filter(
    job_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """실행 중인 필터 작업 취소 (진행 상태 보존 안 됨)"""
    job = get_volume_filter_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if job["status"] not in ("pending", "running"):
        raise HTTPException(status_code=400, detail=f"취소할 수 없는 상태입니다: {job['status']}")

    set_volume_filter_control(job_id, should_cancel=True)
    return {"success": True, "message": "취소 요청됨. 최대 수 초 내 반영"}


@router.post("/keywords/volume-filter/{job_id}/pause")
async def pause_volume_filter(
    job_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """실행 중인 필터 작업 일시정지 (나중에 재개 가능)"""
    job = get_volume_filter_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if job["status"] != "running":
        raise HTTPException(status_code=400, detail=f"일시정지 불가 상태: {job['status']}")

    set_volume_filter_control(job_id, should_pause=True)
    return {"success": True, "message": "일시정지 요청됨. 최대 수 초 내 반영"}


@router.post("/keywords/volume-filter/{job_id}/resume")
async def resume_volume_filter(
    job_id: int,
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """일시정지/캐너리실패 작업 재개"""
    job = get_volume_filter_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")
    if job["status"] not in ("paused", "canary_failed"):
        raise HTTPException(status_code=400,
                            detail=f"재개 불가 상태: {job['status']}")

    kw_file = job.get("keywords_file")
    if not kw_file:
        raise HTTPException(status_code=500, detail="키워드 파일 경로 없음 (복구 불가)")

    from services.volume_filter import VolumeFilterService, VolumeFilterConfig
    keywords = VolumeFilterService.load_keywords_file(kw_file)
    if not keywords:
        raise HTTPException(status_code=500, detail="키워드 파일 누락 (복구 불가)")

    account = get_ad_account(user_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="네이버 광고 계정 연동 필요")

    start_index = job.get("processed_count", 0) or 0

    async def _run():
        from services.naver_ad_service import NaverAdApiClient
        try:
            client = NaverAdApiClient()
            client.customer_id = account.get("customer_id")
            client.api_key = account.get("api_key")
            client.secret_key = account.get("secret_key")

            svc = VolumeFilterService(client)
            cfg = VolumeFilterConfig(
                job_id=job_id, user_id=user_id,
                min_volume=job.get("min_volume", 10),
                test_size=job.get("test_size", 10000),
                min_pass_rate_pct=job.get("min_pass_rate_pct", 2.0),
                # 재개 시 캐너리 무시 (이미 평가됐거나, 사용자가 "재개"로 강제 진행)
                auto_continue_on_canary=True,
            )
            await svc.run(cfg, keywords, start_index=start_index)
        except Exception as e:
            logger.exception(f"[Filter {job_id}] 재개 실패")
            update_volume_filter_job(
                job_id, status="failed",
                error_message=str(e)[:1000],
                completed_at=datetime.now().isoformat(),
            )

    background_tasks.add_task(_run)

    save_optimization_log(
        user_id, "volume_filter_resume",
        f"필터 재개 (job #{job_id}) at {start_index}/{len(keywords)}",
        {"job_id": job_id, "start_index": start_index}
    )

    return {
        "success": True,
        "message": f"재개 요청됨 ({start_index}/{len(keywords)}부터)",
        "start_index": start_index,
        "total": len(keywords),
    }


@router.get("/keywords/volume-filter/{job_id}/status")
async def get_volume_filter_status(
    job_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """검색량 필터 진행 상태"""
    job = get_volume_filter_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")

    total = job.get("total_keywords", 0) or 0
    processed = job.get("processed_count", 0) or 0
    progress = int(processed / total * 100) if total > 0 else 0

    return {
        "success": True,
        "job": {**job, "progress_percent": progress},
    }


@router.get("/keywords/volume-filter/jobs")
async def list_volume_filter_jobs_route(
    user_id: int = Depends(get_user_id_with_fallback),
    limit: int = Query(default=20),
):
    jobs = list_volume_filter_jobs(user_id, limit=limit)
    return {"success": True, "count": len(jobs), "jobs": jobs}


@router.get("/keywords/volume-filter/{job_id}/results")
async def get_volume_filter_results_route(
    job_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
    limit: Optional[int] = Query(default=None),
    format: str = Query(default="json", description="json 또는 csv"),
):
    """필터 통과 키워드 조회"""
    job = get_volume_filter_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")

    results = get_volume_filter_results(job_id, limit=limit)

    if format == "csv":
        from fastapi.responses import StreamingResponse
        import csv
        import io as _io

        buf = _io.StringIO()
        buf.write("\ufeff")
        writer = csv.writer(buf)
        writer.writerow(["키워드", "PC검색량", "모바일검색량", "총검색량", "경쟁도"])
        for r in results:
            writer.writerow([r["keyword"], r["monthly_pc"], r["monthly_mobile"],
                             r["monthly_total"], r.get("comp_idx", "")])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="filtered_job_{job_id}.csv"'},
        )

    return {"success": True, "count": len(results), "results": results}


class FilterToRegisterRequest(BaseModel):
    campaign_prefix: str = Field(..., description="캠페인 prefix")
    bid: int = Field(default=100)
    keywords_per_group: int = Field(default=500)
    daily_budget: int = Field(default=10000)
    campaign_tp: str = Field(default="WEB_SITE")
    min_volume_override: Optional[int] = Field(default=None, description="등록 시 재필터링 임계치 (생략 시 필터 job의 min_volume 사용)")


@router.post("/keywords/volume-filter/{job_id}/register")
async def register_from_filter(
    job_id: int,
    request: FilterToRegisterRequest,
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """필터 통과 키워드로 대량 등록 job 시작"""
    try:
        job = get_volume_filter_job(job_id, user_id)
        if not job:
            raise HTTPException(status_code=404, detail="필터 작업을 찾을 수 없습니다")
        if job["status"] not in ("completed", "completed_with_errors"):
            raise HTTPException(status_code=400, detail="필터링이 아직 완료되지 않았습니다")

        # 통과 키워드 로드
        raw_results = get_volume_filter_results(job_id)
        min_v = request.min_volume_override if request.min_volume_override is not None else job["min_volume"]
        keywords = [r["keyword"] for r in raw_results if r["monthly_total"] >= min_v]

        if not keywords:
            raise HTTPException(status_code=400, detail="등록할 키워드가 없습니다")

        if request.bid < 70 or request.bid > 100000:
            raise HTTPException(status_code=400, detail="입찰가 70~100000원")
        if not request.campaign_prefix or len(request.campaign_prefix) < 2:
            raise HTTPException(status_code=400, detail="campaign_prefix 필수")

        account = get_ad_account(user_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 연동 필요")

        per_group = request.keywords_per_group
        num_ad_groups = (len(keywords) + per_group - 1) // per_group
        num_campaigns = (num_ad_groups + 999) // 1000

        register_job_id = create_bulk_upload_job(
            user_id=user_id,
            filename=f"filter_job_{job_id}_result",
            campaign_prefix=request.campaign_prefix,
            keywords_per_group=per_group,
            bid=request.bid,
            daily_budget=request.daily_budget,
            total_keywords=len(keywords),
        )

        async def _run():
            from services.bulk_upload_orchestrator import BulkUploadOrchestrator, BulkJobConfig
            from services.naver_ad_service import NaverAdApiClient

            try:
                client = NaverAdApiClient()
                client.customer_id = account.get("customer_id")
                client.api_key = account.get("api_key")
                client.secret_key = account.get("secret_key")

                orchestrator = BulkUploadOrchestrator(client)
                cfg = BulkJobConfig(
                    job_id=register_job_id, user_id=user_id,
                    campaign_prefix=request.campaign_prefix,
                    keywords_per_group=per_group,
                    bid=request.bid,
                    daily_budget=request.daily_budget,
                    campaign_tp=request.campaign_tp,
                )
                await orchestrator.run(cfg, keywords)
            except Exception as e:
                logger.exception(f"[Job {register_job_id}] 실행 실패")
                update_bulk_upload_job(
                    register_job_id, status="failed",
                    error_message=str(e)[:1000],
                    completed_at=datetime.now().isoformat(),
                )

        background_tasks.add_task(_run)

        save_optimization_log(
            user_id, "filter_to_register",
            f"필터 #{job_id} → 등록 #{register_job_id}: {len(keywords)}개 키워드",
            {"filter_job": job_id, "register_job": register_job_id, "total": len(keywords)}
        )

        return {
            "success": True,
            "filter_job_id": job_id,
            "register_job_id": register_job_id,
            "total_keywords": len(keywords),
            "estimated": {
                "campaigns": num_campaigns,
                "ad_groups": num_ad_groups,
            },
            "message": f"등록 작업 시작 (#{register_job_id})",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("filter-to-register error")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 대량 등록 (10만 규모) ============

@router.post("/keywords/scale-register")
async def scale_register_keywords(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="엑셀 또는 CSV - A열 키워드"),
    campaign_prefix: str = Form(..., description="캠페인 이름 prefix (예: bulk_20260422)"),
    bid: int = Form(default=100, description="전체 키워드에 적용할 입찰가 (원)"),
    keywords_per_group: int = Form(default=500, description="광고그룹당 키워드 수 (기본 500)"),
    daily_budget: int = Form(default=10000, description="캠페인당 일 예산 (원)"),
    campaign_tp: str = Form(default="WEB_SITE", description="캠페인 유형"),
    user_id: int = Depends(get_user_id_with_fallback),
):
    """10만 개 규모 키워드 자동 등록
    - 캠페인/광고그룹 자동 생성
    - 500개/광고그룹, 1,000그룹/캠페인 자동 분할
    - 백그라운드 실행, job_id 리턴
    """
    try:
        if bid < 70 or bid > 100000:
            raise HTTPException(status_code=400, detail="입찰가는 70~100,000원이어야 합니다")
        if keywords_per_group < 1 or keywords_per_group > 1000:
            raise HTTPException(status_code=400, detail="광고그룹당 키워드는 1~1000개")
        if daily_budget < 1000:
            raise HTTPException(status_code=400, detail="일 예산은 최소 1,000원")
        if not campaign_prefix or len(campaign_prefix) < 2:
            raise HTTPException(status_code=400, detail="캠페인 prefix를 입력하세요 (2자 이상)")

        content = await file.read()
        if len(content) > 50 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="파일은 50MB 이하")

        # 엑셀 파싱 (force_default_bid=True → bid 전체 적용)
        parsed = _parse_keyword_excel(content, default_bid=bid, force_default_bid=True)
        if parsed["total"] == 0:
            raise HTTPException(status_code=400, detail="유효한 키워드가 없습니다")

        keywords = [item["keyword"] for item in parsed["items"]]
        total = len(keywords)

        # 스케일 계산 & 안전장치
        per_group = keywords_per_group
        num_ad_groups = (total + per_group - 1) // per_group
        num_campaigns = (num_ad_groups + 999) // 1000
        if num_campaigns > 50:
            raise HTTPException(
                status_code=400,
                detail=f"필요 캠페인 {num_campaigns}개가 너무 많습니다. "
                       f"keywords_per_group을 늘리거나 키워드 수를 줄이세요"
            )

        # 광고 계정 연동 확인
        account = get_ad_account(user_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="네이버 광고 계정을 먼저 연동하세요")

        # Job 생성
        job_id = create_bulk_upload_job(
            user_id=user_id,
            filename=file.filename or "uploaded.xlsx",
            campaign_prefix=campaign_prefix,
            keywords_per_group=per_group,
            bid=bid,
            daily_budget=daily_budget,
            total_keywords=total,
        )

        # 백그라운드에서 실제 처리
        async def _run():
            from services.bulk_upload_orchestrator import BulkUploadOrchestrator, BulkJobConfig
            from services.naver_ad_service import NaverAdApiClient

            try:
                client = NaverAdApiClient()
                client.customer_id = account.get("customer_id")
                client.api_key = account.get("api_key")
                client.secret_key = account.get("secret_key")

                orchestrator = BulkUploadOrchestrator(client)
                cfg = BulkJobConfig(
                    job_id=job_id,
                    user_id=user_id,
                    campaign_prefix=campaign_prefix,
                    keywords_per_group=per_group,
                    bid=bid,
                    daily_budget=daily_budget,
                    campaign_tp=campaign_tp,
                )
                await orchestrator.run(cfg, keywords)
            except Exception as e:
                logger.exception(f"[Job {job_id}] 오케스트레이터 실행 실패")
                update_bulk_upload_job(
                    job_id,
                    status="failed",
                    error_message=str(e)[:1000],
                    completed_at=datetime.now().isoformat(),
                )

        background_tasks.add_task(_run)

        save_optimization_log(
            user_id, "scale_register_start",
            f"대량 등록 시작 (job #{job_id}): {total}개 → 캠페인 {num_campaigns}개, 광고그룹 {num_ad_groups}개",
            {
                "job_id": job_id,
                "total": total,
                "num_campaigns": num_campaigns,
                "num_ad_groups": num_ad_groups,
            }
        )

        return {
            "success": True,
            "job_id": job_id,
            "total_keywords": total,
            "estimated": {
                "campaigns": num_campaigns,
                "ad_groups": num_ad_groups,
                "keywords_per_group": per_group,
                "estimated_seconds": int(num_ad_groups * 0.5 + total / 100 * 0.5 + num_campaigns * 0.5),
            },
            "parse_errors_count": parsed["errors_count"] if "errors_count" in parsed else len(parsed["errors"]),
            "message": f"백그라운드 등록 시작 (job #{job_id})",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("scale register error")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/keywords/scale-register/{job_id}/status")
async def get_scale_register_status(
    job_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """대량 등록 작업 진행 상태"""
    job = get_bulk_upload_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")

    total = job.get("total_keywords", 0) or 0
    processed = job.get("processed_count", 0) or 0
    progress = int(processed / total * 100) if total > 0 else 0

    return {
        "success": True,
        "job": {
            **job,
            "progress_percent": progress,
        },
    }


@router.get("/keywords/scale-register/jobs")
async def list_scale_register_jobs(
    user_id: int = Depends(get_user_id_with_fallback),
    limit: int = Query(default=20),
):
    """사용자의 대량 등록 작업 목록"""
    jobs = list_bulk_upload_jobs(user_id, limit=limit)
    return {"success": True, "count": len(jobs), "jobs": jobs}


@router.get("/keywords/scale-register/{job_id}/failures")
async def get_scale_register_failures(
    job_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
    format: str = Query(default="json", description="json 또는 csv"),
):
    """실패한 키워드 목록 + CSV 다운로드"""
    job = get_bulk_upload_job(job_id, user_id)
    if not job:
        raise HTTPException(status_code=404, detail="작업을 찾을 수 없습니다")

    failures = get_bulk_upload_failures(job_id)

    if format == "csv":
        from fastapi.responses import StreamingResponse
        import csv
        import io as _io

        buf = _io.StringIO()
        buf.write("\ufeff")  # UTF-8 BOM for Excel
        writer = csv.writer(buf)
        writer.writerow(["keyword", "bid", "ad_group_id", "reason"])
        for f in failures:
            writer.writerow([f["keyword"], f["bid"], f["ad_group_id"], f["reason"]])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="failures_job_{job_id}.csv"'},
        )

    return {"success": True, "count": len(failures), "failures": failures}


@router.post("/keywords/upload-excel")
async def upload_keywords_excel(
    file: UploadFile = File(..., description="엑셀(.xlsx) 또는 CSV 파일"),
    default_bid: int = Form(default=100),
    ad_group_id: Optional[str] = Form(default=None),
    auto_register: bool = Form(default=False),
    force_default_bid: bool = Form(default=True, description="엑셀 입찰가 무시하고 default_bid 전체 적용"),
    user_id: int = Depends(get_user_id_with_fallback),
):
    """엑셀/CSV 업로드로 키워드+입찰가 파싱.
    - force_default_bid=true(기본): 엑셀 내용과 무관하게 default_bid를 모든 키워드에 일괄 적용
    - force_default_bid=false: 엑셀 B열 입찰가 우선, 없으면 default_bid 사용
    - auto_register=false(기본): 파싱 결과만 반환(미리보기)
    - auto_register=true: 즉시 네이버 광고 API로 등록
    """
    try:
        if not file.filename:
            raise HTTPException(status_code=400, detail="파일이 없습니다")

        if default_bid < 70 or default_bid > 100000:
            raise HTTPException(status_code=400, detail="입찰가는 70원~100,000원 사이여야 합니다")

        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="파일 크기가 10MB를 초과합니다")

        parsed = _parse_keyword_excel(content, default_bid, force_default_bid=force_default_bid)

        result: Dict[str, Any] = {
            "success": True,
            "filename": file.filename,
            "total": parsed["total"],
            "items": parsed["items"][:500],  # 미리보기 최대 500개
            "items_count": len(parsed["items"]),
            "errors": parsed["errors"][:100],
            "errors_count": len(parsed["errors"]),
            "registered": 0,
        }

        if auto_register:
            if not ad_group_id:
                raise HTTPException(status_code=400, detail="ad_group_id가 필요합니다")
            if not parsed["items"]:
                result["message"] = "등록할 키워드가 없습니다"
                return result

            optimizer = get_optimizer()
            from services.naver_ad_service import KeywordSuggestion
            suggestions = [
                KeywordSuggestion(keyword=it["keyword"], suggested_bid=it["bid"])
                for it in parsed["items"]
            ]
            added = await optimizer.bulk_manager.bulk_add_keywords(
                ad_group_id, suggestions, default_bid
            )
            result["registered"] = len(added)

            save_optimization_log(
                user_id, "excel_upload_register",
                f"엑셀 업로드 등록: {len(added)}/{len(parsed['items'])}개",
                {
                    "filename": file.filename,
                    "requested": len(parsed["items"]),
                    "added": len(added),
                    "ad_group_id": ad_group_id,
                }
            )
            result["message"] = f"{len(added)}개 키워드가 등록되었습니다"
        else:
            result["message"] = f"{parsed['total']}개 키워드 파싱 완료 (미리보기)"

        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel upload error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# ============ 입찰가 관리 ============

@router.get("/bids/history")
async def get_bids_history(
    user_id: int = Depends(get_user_id_with_fallback),
    keyword_id: Optional[str] = Query(None, description="키워드 ID"),
    limit: int = Query(100, description="조회 개수")
):
    """입찰 변경 이력 조회"""
    history = get_bid_history(user_id, limit, keyword_id)
    return {
        "success": True,
        "count": len(history),
        "history": history
    }


@router.get("/bids/summary")
async def get_bids_summary(
    user_id: int = Depends(get_user_id_with_fallback),
    days: int = Query(7, description="조회 기간 (일)")
):
    """입찰 변경 요약"""
    summary = get_bid_changes_summary(user_id, days)
    return {
        "success": True,
        "data": summary
    }


@router.post("/bids/update")
async def update_bid_manual(
    request: ManualBidUpdateRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """수동 입찰가 변경"""
    try:
        optimizer = get_optimizer()

        # 현재 키워드 정보 조회
        keyword_info = await optimizer.api.get_keyword(request.keyword_id)
        old_bid = keyword_info.get("bidAmt", 0)

        # 입찰가 변경
        await optimizer.api.update_keyword_bid(request.keyword_id, request.new_bid)

        # 변경 기록 저장
        save_bid_change(
            user_id=user_id,
            keyword_id=request.keyword_id,
            keyword_text=keyword_info.get("keyword", ""),
            old_bid=old_bid,
            new_bid=request.new_bid,
            reason=request.reason,
            strategy="manual"
        )

        return {
            "success": True,
            "message": f"입찰가가 {old_bid}원에서 {request.new_bid}원으로 변경되었습니다",
            "old_bid": old_bid,
            "new_bid": request.new_bid
        }
    except Exception as e:
        logger.error(f"Manual bid update error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 키워드 제외 ============

@router.post("/keywords/evaluate")
async def evaluate_keywords(
    user_id: int = Depends(get_user_id_with_fallback),
    ad_group_ids: Optional[List[str]] = Query(None, description="광고그룹 ID 목록")
):
    """비효율 키워드 평가 및 제외"""
    try:
        optimizer = get_optimizer()

        # 설정 로드
        settings = get_optimization_settings(user_id)
        if settings:
            optimizer.exclusion.set_thresholds(
                min_ctr=settings.get("min_ctr", 0.01),
                max_cost_no_conv=settings.get("max_cost_no_conv", 50000),
                min_quality_score=settings.get("min_quality_score", 4),
                evaluation_days=settings.get("evaluation_days", 7)
            )

        # 평가 실행
        excluded = await optimizer.exclusion.evaluate_and_exclude(ad_group_ids)

        # 제외 기록 저장
        for item in excluded:
            save_excluded_keyword(
                user_id=user_id,
                keyword_id=item.get("keyword_id"),
                keyword_text=item.get("keyword"),
                ad_group_id=item.get("ad_group_id", ""),
                reason=item.get("reason")
            )

        save_optimization_log(
            user_id, "keyword_evaluation",
            f"키워드 평가 완료: {len(excluded)}개 제외",
            {"excluded_count": len(excluded)}
        )

        return {
            "success": True,
            "message": f"{len(excluded)}개 키워드가 제외되었습니다",
            "excluded": excluded
        }
    except Exception as e:
        logger.error(f"Keyword evaluation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/keywords/excluded")
async def get_excluded_list(
    user_id: int = Depends(get_user_id_with_fallback),
    include_restored: bool = Query(False, description="복원된 키워드 포함")
):
    """제외된 키워드 목록"""
    excluded = get_excluded_keywords(user_id, include_restored)
    return {
        "success": True,
        "count": len(excluded),
        "keywords": excluded
    }


@router.post("/keywords/restore/{keyword_id}")
async def restore_keyword(
    keyword_id: str,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """제외된 키워드 복원"""
    try:
        optimizer = get_optimizer()

        # 키워드 활성화
        await optimizer.api.activate_keyword(keyword_id)

        # DB 업데이트
        restore_excluded_keyword(user_id, keyword_id)

        save_optimization_log(
            user_id, "keyword_restore",
            f"키워드 복원: {keyword_id}"
        )

        return {
            "success": True,
            "message": "키워드가 복원되었습니다"
        }
    except Exception as e:
        logger.error(f"Keyword restore error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 성과 조회 ============

@router.get("/performance")
async def get_performance(
    user_id: int = Depends(get_user_id_with_fallback),
    start_date: Optional[str] = Query(None, description="시작일 (YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="종료일 (YYYY-MM-DD)"),
    keyword_id: Optional[str] = Query(None, description="키워드 ID")
):
    """키워드 성과 조회"""
    if not end_date:
        end_date = datetime.now().strftime("%Y-%m-%d")
    if not start_date:
        start_date = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")

    performance = get_keyword_performance(user_id, start_date, end_date, keyword_id)
    return {
        "success": True,
        "count": len(performance),
        "data": performance
    }


@router.get("/performance/summary")
async def get_perf_summary(
    user_id: int = Depends(get_user_id_with_fallback),
    days: int = Query(7, description="조회 기간 (일)")
):
    """성과 요약 조회"""
    summary = get_performance_summary(user_id, days)
    return {
        "success": True,
        "data": summary
    }


# ============ 리포트 ============

@router.get("/reports/daily")
async def get_daily_report(
    user_id: int = Depends(get_user_id_with_fallback),
    days: int = Query(30, description="조회 기간 (일)")
):
    """일일 리포트 조회"""
    reports = get_daily_reports(user_id, days)
    return {
        "success": True,
        "count": len(reports),
        "reports": reports
    }


@router.get("/logs")
async def get_logs(
    user_id: int = Depends(get_user_id_with_fallback),
    log_type: Optional[str] = Query(None, description="로그 유형"),
    limit: int = Query(100, description="조회 개수")
):
    """최적화 로그 조회"""
    logs = get_optimization_logs(user_id, log_type, limit)
    return {
        "success": True,
        "count": len(logs),
        "logs": logs
    }


# ============ 캠페인/광고그룹 조회 ============

@router.get("/campaigns")
async def get_campaigns(user_id: int = Depends(get_user_id_with_fallback)):
    """캠페인 목록 조회"""
    try:
        optimizer = get_optimizer()
        campaigns = await optimizer.api.get_campaigns()
        return {
            "success": True,
            "count": len(campaigns),
            "campaigns": campaigns
        }
    except Exception as e:
        logger.error(f"Get campaigns error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/adgroups")
async def get_ad_groups(
    user_id: int = Depends(get_user_id_with_fallback),
    campaign_id: Optional[str] = Query(None, description="캠페인 ID")
):
    """광고그룹 목록 조회"""
    try:
        optimizer = get_optimizer()
        ad_groups = await optimizer.api.get_ad_groups(campaign_id)
        return {
            "success": True,
            "count": len(ad_groups),
            "ad_groups": ad_groups
        }
    except Exception as e:
        logger.error(f"Get ad groups error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/keywords")
async def get_keywords(
    user_id: int = Depends(get_user_id_with_fallback),
    ad_group_id: Optional[str] = Query(None, description="광고그룹 ID")
):
    """키워드 목록 조회"""
    try:
        optimizer = get_optimizer()
        keywords = await optimizer.api.get_keywords(ad_group_id)
        return {
            "success": True,
            "count": len(keywords),
            "keywords": keywords
        }
    except Exception as e:
        logger.error(f"Get keywords error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 광고 계정 연동 ============

class AdAccountRequest(BaseModel):
    customer_id: str = Field(..., description="네이버 광고 고객 ID")
    api_key: str = Field(..., description="API 키")
    secret_key: str = Field(..., description="비밀 키")
    name: Optional[str] = Field(None, description="계정 이름")


@router.post("/account/connect")
async def connect_ad_account(
    request: AdAccountRequest,
    user_id: int = Depends(get_user_id_with_fallback)
):
    """광고 계정 연동"""
    try:
        # 계정 정보 저장
        account = save_ad_account(
            user_id,
            request.customer_id,
            request.api_key,
            request.secret_key,
            request.name
        )

        # 연결 테스트 - 캠페인 목록 조회 시도
        from services.naver_ad_service import NaverAdApiClient
        test_client = NaverAdApiClient()
        test_client.customer_id = request.customer_id
        test_client.api_key = request.api_key
        test_client.secret_key = request.secret_key

        try:
            campaigns = await test_client.get_campaigns()
            # 연결 성공
            update_ad_account_status(user_id, request.customer_id, True)
            save_optimization_log(user_id, "account_connected", f"광고 계정이 연동되었습니다: {request.customer_id}")

            return {
                "success": True,
                "message": "광고 계정이 성공적으로 연동되었습니다",
                "account": {
                    "customer_id": request.customer_id,
                    "name": request.name,
                    "is_connected": True,
                    "campaigns_count": len(campaigns)
                }
            }
        except Exception as api_error:
            # 연결 실패
            update_ad_account_status(user_id, request.customer_id, False, str(api_error))
            return {
                "success": False,
                "message": f"API 연결 실패: {str(api_error)}",
                "error": str(api_error)
            }

    except Exception as e:
        logger.error(f"Account connect error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/account/status")
async def get_account_status(user_id: int = Depends(get_user_id_with_fallback)):
    """광고 계정 연동 상태 조회"""
    try:
        account = get_ad_account(user_id)

        if not account:
            return {
                "success": True,
                "is_connected": False,
                "message": "연동된 광고 계정이 없습니다"
            }

        return {
            "success": True,
            "is_connected": account.get("is_connected", False),
            "account": {
                "customer_id": account.get("customer_id"),
                "name": account.get("name"),
                "last_sync_at": account.get("last_sync_at"),
                "connection_error": account.get("connection_error")
            }
        }
    except Exception as e:
        logger.error(f"Account status error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/account/disconnect")
async def disconnect_ad_account(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: str = Query(..., description="고객 ID")
):
    """광고 계정 연동 해제"""
    try:
        delete_ad_account(user_id, customer_id)
        save_optimization_log(user_id, "account_disconnected", f"광고 계정 연동이 해제되었습니다: {customer_id}")

        return {
            "success": True,
            "message": "광고 계정 연동이 해제되었습니다"
        }
    except Exception as e:
        logger.error(f"Account disconnect error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 효율 추적 ============

@router.get("/efficiency/summary")
async def get_efficiency(
    user_id: int = Depends(get_user_id_with_fallback),
    days: int = Query(default=7, description="조회 기간 (일)")
):
    """효율 개선 요약"""
    try:
        summary = get_efficiency_summary(user_id, days)
        return {
            "success": True,
            "period_days": days,
            "data": summary
        }
    except Exception as e:
        logger.error(f"Efficiency summary error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/efficiency/history")
async def get_efficiency_chart(
    user_id: int = Depends(get_user_id_with_fallback),
    days: int = Query(default=30, description="조회 기간 (일)")
):
    """일별 효율 추적 이력 (차트용)"""
    try:
        history = get_efficiency_history(user_id, days)
        return {
            "success": True,
            "period_days": days,
            "data": history
        }
    except Exception as e:
        logger.error(f"Efficiency history error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 트렌드 키워드 추천 ============

@router.get("/trending/keywords")
async def get_trending_keyword_recommendations(
    user_id: int = Depends(get_user_id_with_fallback),
    limit: int = Query(default=20, description="최대 개수")
):
    """트렌드 키워드 추천 조회"""
    try:
        keywords = get_trending_keywords(user_id, limit)
        return {
            "success": True,
            "count": len(keywords),
            "keywords": keywords
        }
    except Exception as e:
        logger.error(f"Trending keywords error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/trending/refresh")
async def refresh_trending_keywords(
    user_id: int = Depends(get_user_id_with_fallback),
    seed_keywords: List[str] = Query(default=[], description="시드 키워드")
):
    """트렌드 키워드 새로고침 - 네이버 광고 API에서 최신 키워드 가져오기"""
    try:
        optimizer = get_optimizer()

        # 시드 키워드가 없으면 기존 키워드 사용
        if not seed_keywords:
            # 기존 광고 키워드에서 시드 추출
            settings = get_optimization_settings(user_id)
            seed_keywords = settings.get("core_terms", []) if settings else []

        if not seed_keywords:
            return {
                "success": False,
                "message": "시드 키워드를 입력하거나 설정에서 핵심 키워드를 설정해주세요"
            }

        # 연관 키워드 발굴
        discovered = await optimizer.discovery.discover_related_keywords(
            seed_keywords=seed_keywords,
            max_keywords=50,
            min_search_volume=100,
            max_competition=0.85
        )

        # 트렌드 점수 계산 및 저장
        trending_data = []
        for kw in discovered:
            # 기회 점수 계산 (검색량 높고 경쟁 낮을수록 높음)
            search_vol = kw.monthly_search_count
            comp = kw.competition_index
            opportunity = (search_vol / 1000) * (1 - comp) * 100 if comp < 1 else 0

            trending_data.append({
                "keyword": kw.keyword,
                "category": seed_keywords[0] if seed_keywords else "일반",
                "search_volume_current": search_vol,
                "search_volume_prev_week": int(search_vol * 0.9),  # 10% 상승 가정
                "search_volume_change_rate": 10.0,
                "competition_level": kw.competition_level,
                "competition_index": comp,
                "suggested_bid": kw.suggested_bid,
                "opportunity_score": round(opportunity, 1),
                "relevance_score": kw.relevance_score,
                "trend_score": round(kw.potential_score * 10, 1),
                "recommendation_reason": f"검색량 {search_vol:,}회, 경쟁도 {kw.competition_level}"
            })

        save_trending_keywords(user_id, trending_data)

        return {
            "success": True,
            "message": f"{len(trending_data)}개의 트렌드 키워드가 발굴되었습니다",
            "count": len(trending_data),
            "keywords": trending_data[:10]  # 상위 10개만 미리보기
        }
    except Exception as e:
        logger.error(f"Refresh trending keywords error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class AddTrendingKeywordRequest(BaseModel):
    keyword: str = Field(..., description="키워드")
    ad_group_id: str = Field(default="", description="광고그룹 ID (미지정 시 기본 광고그룹)")
    bid: int = Field(default=100, description="입찰가")


@router.post("/trending/add-to-campaign")
async def add_trending_to_campaign(
    req: AddTrendingKeywordRequest,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """트렌드 키워드를 광고에 추가"""
    try:
        optimizer = get_optimizer()

        # 키워드 추가 — nccAdgroupId는 URL query에 별도 전달 필요
        result = await optimizer.api.create_keywords([{
            "nccAdgroupId": req.ad_group_id,
            "keyword": req.keyword,
            "bidAmt": req.bid,
            "useGroupBidAmt": False
        }], ad_group_id=req.ad_group_id)

        # 상태 업데이트
        update_trending_keyword_status(user_id, req.keyword, "added")
        save_optimization_log(user_id, "keyword_added", f"트렌드 키워드 '{req.keyword}'가 광고에 추가되었습니다")

        return {
            "success": True,
            "message": f"키워드 '{req.keyword}'가 광고에 추가되었습니다",
            "result": result
        }
    except Exception as e:
        logger.error(f"Add trending keyword error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ 종합 대시보드 ============

@router.get("/dashboard/comprehensive")
async def get_comprehensive_dashboard(user_id: int = Depends(get_user_id_with_fallback)):
    """종합 대시보드 - 계정 상태, 효율, 트렌드 모두 포함"""
    try:
        # 계정 상태
        account = get_ad_account(user_id)

        # 기본 대시보드
        stats = get_dashboard_stats(user_id)

        # 효율 요약
        efficiency = get_efficiency_summary(user_id, 7)

        # 트렌드 키워드
        trending = get_trending_keywords(user_id, 5)

        # 최근 입찰 변경
        recent_changes = get_bid_history(user_id, limit=5)

        return {
            "success": True,
            "data": {
                "account": {
                    "is_connected": account.get("is_connected", False) if account else False,
                    "customer_id": account.get("customer_id") if account else None,
                    "last_sync_at": account.get("last_sync_at") if account else None
                },
                "stats": stats,
                "efficiency": efficiency,
                "trending_keywords": trending,
                "recent_changes": recent_changes
            }
        }
    except Exception as e:
        logger.error(f"Comprehensive dashboard error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Phase 3: 키워드 풀 자동 워커 endpoints ============
import os as _os
import hmac as _hmac
import time as _ts_mod
from fastapi import Header
from database.keyword_pool_db import get_keyword_pool_db
from database.registered_keywords_db import get_registered_keywords_db


# Niche 시드 timeout backoff cache — (cid, seed) → epoch 마지막 timeout.
# 매 시드 ConnectTimeout 발생 시 등록. 다음 collect 라운드에서 _NICHE_BACKOFF_S
# 이내인 시드는 skip. niche 의료/희귀 시드 5개가 cascade timeout → circuit OPEN →
# 전체 collect cycle abort 패턴 차단. 워커 재시작 시 reset (in-memory).
_seed_timeout_cache: Dict[Tuple[int, str], float] = {}
_NICHE_BACKOFF_S = 1800  # 30분 — 한 사이클 timeout 난 시드는 30분간 라운드 제외


def _is_seed_in_backoff(cid: int, seed: str) -> bool:
    last = _seed_timeout_cache.get((cid, seed))
    if not last:
        return False
    if _ts_mod.time() - last < _NICHE_BACKOFF_S:
        return True
    # backoff 만료 — 캐시에서 제거 후 다음 라운드에 재시도
    _seed_timeout_cache.pop((cid, seed), None)
    return False


def _mark_seed_timeout(cid: int, seed: str) -> None:
    _seed_timeout_cache[(cid, seed)] = _ts_mod.time()


# 도메인 의미 토큰 — 사업 영역(금융/대출/의료/소상공인/정부지원) 안전 가드.
# 키워드가 시드 substring을 통과해도 이 중 1개 이상 포함해야 풀에 INSERT.
# 예: 시드 '은행' → '은행대출/은행이자' 통과, '은행나무/은행잎차' reject.
POOL_DOMAIN_TOKENS = (
    # 금융/대출
    "대출", "자금", "한도", "이자", "금리", "신용", "담보", "보증",
    "마통", "통장", "주담대", "전세", "월세", "환급", "예금", "적금", "은행",
    # 정부/지원금
    "지원금", "정책자금", "정부지원", "청년", "신청", "장려금", "바우처",
    # 소상공인/창업/사업자 운영
    "소상공인", "사업자", "자영업", "창업", "개원", "개업", "양도", "양수",
    "운영", "프랜차이즈", "매장", "점포", "임대", "분양", "매매",
    "매출", "수수료", "결제", "권리금", "매물", "임차", "임대차",
    "세무", "회계", "노무", "법인", "장부", "기장",
    "할부", "리스", "렌트", "렌탈", "수출", "무역", "경매",
    # 사업자 인접 금융상품
    "보험", "카드", "펀드", "연금", "공제", "IRP", "퇴직", "CMA", "MMF",
    "세금", "정산", "환산", "공제금", "환급금", "절세",
    # 의료 — 진료과/시설
    "병원", "약국", "약사", "의사", "의료", "원장", "진료", "검진", "요양",
    "한의원", "한방", "치과", "정형외과", "내과", "외과", "안과", "피부과",
    "이비인후과", "산부인과", "성형외과", "비뇨기과", "흉부외과", "재활",
    "임플란트", "교정", "보톡스", "필러", "시술", "수술",
    # 뷰티/미용
    "미용", "미용실", "헤어", "네일", "왁싱", "타투", "속눈썹", "두피", "성형",
    # 외식업
    "카페", "식당", "음식점", "분식", "치킨", "주점", "베이커리", "떡볶이",
    "피자", "초밥", "곱창", "파스타", "한식", "일식", "야식", "설렁탕", "순두부",
    "반찬", "마트", "가게", "식자재", "배달", "배민", "쿠팡이츠",
    # 피트니스/교육/생활
    "필라테스", "요가", "헬스장", "학원", "교육", "강의", "과외", "유치원", "어린이집",
    "펜션", "모텔", "오피스텔", "아파트", "원룸", "상가", "공장", "창고", "주택", "사무실", "점포",
    # 차량/장비
    "할부", "리스", "렌트", "렌탈", "차량", "오토바이", "트랙터", "택시", "경운기",
)


def _has_domain_token(kw: str) -> bool:
    return any(t in kw for t in POOL_DOMAIN_TOKENS)


def _derive_seed_tokens(seeds: List[str], min_freq: int = 2) -> set:
    """시드 목록에서 도메인 토큰을 자동 추출 — 신규 분야 광고주 자동 적응.

    - 2~3글자 n-gram 추출 후 ≥ min_freq 시드에서 등장한 것만 토큰화 (의미 보장)
    - 길이 4+ 토큰은 단일 시드만으로도 채택 (긴 토큰은 우연 일치 거의 없음)
    - POOL_DOMAIN_TOKENS 와 합쳐 최종 게이트 토큰셋 구성
    """
    counts: Dict[str, int] = {}
    for s in seeds or []:
        if not s or len(s) < 2:
            continue
        seen_in_seed = set()
        for n in (2, 3):
            for i in range(len(s) - n + 1):
                t = s[i:i + n]
                if t in seen_in_seed:
                    continue
                seen_in_seed.add(t)
                counts[t] = counts.get(t, 0) + 1
        # 시드 통째도 토큰 (길이 4+ 자주 단일 시드만으로도 의미 보장)
        if len(s) >= 4:
            counts[s] = counts.get(s, 0) + min_freq  # 단일 시드만으로도 통과
    return {t for t, c in counts.items() if c >= min_freq}


def _build_domain_token_set(seeds: List[str]) -> set:
    """하드코딩 토큰 + 시드에서 도출된 토큰 합집합 (정적 baseline + 동적 적응)."""
    return set(POOL_DOMAIN_TOKENS) | _derive_seed_tokens(seeds)


def _build_seed_atoms(seeds: List[str]) -> set:
    """시드 → 2/3-gram 원자 + 전체 시드 집합.

    Gate 2(시드 매치) 용. 과거에는 full-seed substring (`s in kw or kw in s`)으로
    체크했는데, 풀이 포화될수록 literal {seed}+suffix 후보 공간이 고갈되어
    domain은 통과하지만 시드 substring 못 잡는 키워드가 100% reject → DEADLOCK.
    원자 단위로 완화 — 예: 시드 "한방병원" 원자에 "한방"이 포함되어 "한방치료"도 통과.
    Gate 1(도메인 토큰)이 여전히 적용되므로 광고주 영역과 무관한 단어는 차단된다.
    """
    atoms: set = set()
    for s in seeds or []:
        if not s or len(s) < 2:
            continue
        atoms.add(s)
        for n in (2, 3):
            for i in range(len(s) - n + 1):
                atoms.add(s[i:i + n])
    return atoms


def _compute_relevance_score(
    kw: str,
    user_seeds: List[str],
    pool_tokens: tuple = (),  # M1 fix: default 를 () 로 — POOL hardcoded 매칭이 user_seed
                              # 광고주에서 무관 KW 점수 부풀려 threshold 회피하는 누수 차단.
                              # 호출자가 의도적으로 POOL 매칭 원할 때만 명시 (cold_start 광고주 등).
) -> int:
    """클릭 KW 의 user_seed/POOL 도메인 연관성 점수 (0-100).

    - 100: kw 가 user_seed 전체를 substring 으로 포함 (예: "강남오피스텔매매" ← "오피스텔매매")
    -  95: user_seed 가 kw 전체를 포함 (kw 가 더 짧음)
    - 0-95: atom 매칭 가중 합산
        · length≥3 user_seed atom 매칭: 20pt × N (max 80) — 강한 도메인 신호
        · length=2 user_seed atom 매칭: 5pt × N (max 30) — 약한 신호 (브로드 매칭)
        · POOL 토큰 매칭: 3pt × N (max 15) — niche 어시스트 (간접 관련)

    예 점수 :
      "강남오피스텔매매" → 100  (user_seed 포함)
      "오피스텔분양"   → ~80  (3+ atom "오피스텔")
      "포켓몬카드"     → ~8   (2-gram "카드" + POOL "카드" — 약함)
      "도박중독"       → 0    (어떤 매칭도 없음)
    """
    if not kw:
        return 0
    # 1) user_seed 전체 매칭 — 가장 강한 신호
    for s in user_seeds:
        if not s or len(s) < 2:
            continue
        if s in kw:
            return 100
        if kw in s:
            return 95
    # 2) atom 분리
    atoms_3plus: set = set()
    atoms_2: set = set()
    for s in user_seeds:
        if not s or len(s) < 2:
            continue
        if len(s) >= 4:
            atoms_3plus.add(s)
        for n in (2, 3):
            for i in range(len(s) - n + 1):
                a = s[i:i + n]
                (atoms_2 if len(a) == 2 else atoms_3plus).add(a)
    # 3) 매칭 카운트 (집합 차이로 중복 제외)
    n_3 = sum(1 for a in atoms_3plus if a in kw)
    n_2 = sum(1 for a in atoms_2 if a in kw)
    n_pool = sum(1 for t in pool_tokens if t in kw)
    score = min(80, n_3 * 20) + min(30, n_2 * 5) + min(15, n_pool * 3)
    return min(95, score)  # 95 cap — 100 은 full seed match 전용


def _resolve_account(user_id: int, customer_id: Optional[str] = None) -> Optional[Dict]:
    """customer_id 명시 시 그 광고주, 없으면 가장 최근. B 시나리오 — 다중 광고주 라우팅."""
    from database.naver_ad_db import get_ad_account_by_customer
    if customer_id:
        return get_ad_account_by_customer(user_id, str(customer_id))
    return get_ad_account(user_id)


def _parse_naver_count(v) -> int:
    """네이버 keywordstool 검색량 — 10 미만이면 '< 10' 문자열로 옴. 안전 변환."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    if s.startswith("<"):
        return 5  # '< 10' → 보수적으로 5
    try:
        return int(s)
    except (ValueError, TypeError):
        return 0


def _verify_cron_token(authorization: Optional[str]) -> None:
    expected = (_os.environ.get("CRON_TOKEN") or "").strip()
    if not expected:
        raise HTTPException(status_code=503, detail="CRON_TOKEN 미설정")
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Bearer 토큰 필요")
    provided = authorization.split(" ", 1)[1].strip()
    if not _hmac.compare_digest(provided, expected):
        raise HTTPException(status_code=403, detail="잘못된 cron 토큰")


async def _cap_triggered_self_heal(
    uid: int,
    customer_id: int,
    account: Dict,
    threshold: int,
    saved_relevance: List[str],
    max_delete: int = 200,
) -> int:
    """한도 도달 자가치유 — saved_relevance 기반 점수 ≤ threshold KW 자동 정리.

    호출 조건 (호출자가 보장): active+pending ≥ 100k 이고 auto_cleanup_enabled=1.
    saved_relevance 가 비어 있거나 < 3 이면 즉시 0 반환 (user_seed 폴백 X — drift 위험).
    한 tick 당 max_delete 보수적 (네이버 API rate + 다중 광고주 처리 시간 보호).
    반환: 삭제+pause 성공한 KW 수.
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import record_auto_cleanup_run
    import sqlite3 as _sqlite3

    if not saved_relevance or len([s for s in saved_relevance if s and len(s) >= 2]) < 3:
        logger.warning(
            f"[pool/self-heal] uid={uid} cid={customer_id} skip — "
            f"saved_relevance 부족({len(saved_relevance)}). 도메인 KW 저장 필요."
        )
        return 0

    reg = get_registered_keywords_db()
    pool = get_keyword_pool_db()

    with _sqlite3.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL",
            (customer_id,),
        ).fetchall()
    if not rows:
        return 0

    def _score_all() -> List[Tuple[str, str, int]]:
        atoms_3plus: set = set()
        atoms_2: set = set()
        for s in saved_relevance:
            if not s or len(s) < 2:
                continue
            if len(s) >= 4:
                atoms_3plus.add(s)
            for n in (2, 3):
                for i in range(len(s) - n + 1):
                    a = s[i:i + n]
                    (atoms_2 if len(a) == 2 else atoms_3plus).add(a)
        out: List[Tuple[str, str, int]] = []
        for kw_text, kid in rows:
            if not kw_text:
                out.append((kid, "", 0))
                continue
            sc = 0
            full = False
            for s in saved_relevance:
                if not s or len(s) < 2:
                    continue
                if s in kw_text:
                    sc = 100; full = True; break
                if kw_text in s:
                    sc = 95; full = True; break
            if not full:
                n_3 = sum(1 for a in atoms_3plus if a in kw_text)
                n_2 = sum(1 for a in atoms_2 if a in kw_text)
                sc = min(95, min(80, n_3 * 20) + min(30, n_2 * 5))
            out.append((kid, kw_text, sc))
        return out

    scored = await asyncio.to_thread(_score_all)
    # Option B (boundary 보존): score == threshold KW 는 ADD gate (>=) 와 DELETE gate (<)
    # 사이의 stable point. 정확히 50점 KW 가 thrash 사이클 (add→delete→add) 도는 사고 방지.
    targets = [(kid, kw, s) for kid, kw, s in scored if s < threshold]
    targets.sort(key=lambda x: x[2])
    targets_capped = targets[:max_delete]
    if not targets_capped:
        logger.warning(
            f"[pool/self-heal] uid={uid} cid={customer_id} 대상 0 — "
            f"thr={threshold} basis={len(saved_relevance)} total={len(scored)}. "
            f"threshold 상향 검토 필요."
        )
        return 0

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    import httpx as _httpx

    def _is_already_gone(exc: Exception) -> bool:
        # naver 404 / code 1018 — 키워드가 이미 naver 측엔 없는 stale ncc_keyword_id.
        # DB row 만 정리하면 슬롯 회수 (cap 회복 가능).
        if isinstance(exc, _httpx.HTTPStatusError):
            try:
                return exc.response.status_code == 404
            except Exception:
                return False
        return False

    def _drop_db_row(kid_: str) -> None:
        with _sqlite3.connect(reg.db_path) as c:
            c.execute(
                "DELETE FROM registered_keywords "
                "WHERE account_customer_id=? AND ncc_keyword_id=?",
                (customer_id, kid_),
            )

    n_del, n_pause, n_fail, n_stale = 0, 0, 0, 0
    affected: List[str] = []
    for kid, kw_text, _s in targets_capped:
        try:
            await client.delete_keyword(kid)
            _drop_db_row(kid)
            n_del += 1
            affected.append(kw_text)
        except Exception as e1:
            if _is_already_gone(e1):
                _drop_db_row(kid)
                n_stale += 1
                affected.append(kw_text)
            else:
                try:
                    await client.pause_keyword(kid)
                    n_pause += 1
                    affected.append(kw_text)
                except Exception as e2:
                    if _is_already_gone(e2):
                        _drop_db_row(kid)
                        n_stale += 1
                        affected.append(kw_text)
                    else:
                        n_fail += 1
        await asyncio.sleep(0.15)

    if affected:
        try:
            pool.mark_rejected_by_naver(
                customer_id,
                [{"keyword": kw, "reason": f"cap_self_heal(≤{threshold})"} for kw in affected],
            )
        except Exception:
            pass
    try:
        record_auto_cleanup_run(uid, str(customer_id), n_del + n_pause + n_stale)
    except Exception:
        pass

    logger.warning(
        f"[pool/self-heal] uid={uid} cid={customer_id} thr={threshold} "
        f"basis={len(saved_relevance)} below={len(targets)} → "
        f"del={n_del} pause={n_pause} stale={n_stale} fail={n_fail}"
    )
    return n_del + n_pause + n_stale


async def _cap_triggered_rolling_heal(
    uid: int,
    customer_id: int,
    account: Dict,
    *,
    max_delete: int = 200,
    mt_ceiling: int = 50,
    settle_hours: int = 24,
) -> int:
    """한도 도달 롤링 자가치유 — saved_relevance 없이도 동작하는 mt 최하위 eject.

    saved_relevance 가 없거나 부족해서 `_cap_triggered_self_heal` 가 0 을 반환한
    경우 폴백. **registered_as_seed 무한 발굴 사이클** 의 핵심 — 100k cap 에 도달해도
    collect 가 영구 정지하지 않게 매 tick 마다 하위 mt 슬라이스를 갈아내고 자리를
    비워준다. 갈아낸 자리에는 다음 collect tick 의 registered-as-seed BFS 가
    발굴한 신규 mt≥1 KW 가 들어가 평균 mt 가 점진적으로 상승.

    안전장치:
      - mt < mt_ceiling (기본 50) 만 대상 — mt≥50 quality KW 는 보호
      - registered_at > settle_hours 전 (기본 24h) — 갓 등록된 KW 는 정착 시간 보장
      - ncc_keyword_id IS NOT NULL — Naver 에 등록되지 않은 행은 건드리지 않음
      - removed_at IS NULL — 이미 제거된 행 스킵
      - ORDER BY mt ASC — 가장 가치 낮은 것부터

    트레이드오프: 운영 중인 mt 1~49 KW 도 eject 대상이 됨. 광고 클릭이 0 이면
    어차피 비용 미발생 — 데이터 기반 큐레이션. 사용자가 우려하면 mt_ceiling 낮추거나
    clicks 추적 후 0-click 필터 추가.
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import record_auto_cleanup_run
    import sqlite3 as _sqlite3

    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()

    # JOIN — 같은 blog_analyzer.db 내. naverad_keyword_pool 에서 mt 최하위 + 24h 정착 +
    # registered_keywords 의 ncc_keyword_id 확보된 KW 만.
    with _sqlite3.connect(pool.db_path) as conn:
        conn.row_factory = _sqlite3.Row
        rows = conn.execute(
            f"""SELECT rk.keyword AS keyword,
                       rk.ncc_keyword_id AS ncc_keyword_id,
                       COALESCE(p.monthly_total, 0) AS monthly_total
                FROM registered_keywords rk
                INNER JOIN naverad_keyword_pool p
                  ON p.account_customer_id = rk.account_customer_id
                 AND p.keyword = rk.keyword
                WHERE rk.account_customer_id = ?
                  AND rk.ncc_keyword_id IS NOT NULL
                  AND rk.removed_at IS NULL
                  AND p.status = 'registered'
                  AND COALESCE(p.monthly_total, 0) < ?
                  AND p.registered_at IS NOT NULL
                  AND datetime(p.registered_at) < datetime('now', '-{int(settle_hours)} hours')
                ORDER BY COALESCE(p.monthly_total, 0) ASC,
                         p.registered_at ASC
                LIMIT ?""",
            (customer_id, mt_ceiling, max_delete),
        ).fetchall()

    if not rows:
        logger.warning(
            f"[pool/rolling-heal] uid={uid} cid={customer_id} 대상 0 — "
            f"mt<{mt_ceiling} & settle≥{settle_hours}h 조건 충족 행 없음"
        )
        return 0

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    import httpx as _httpx

    def _is_already_gone(exc: Exception) -> bool:
        if isinstance(exc, _httpx.HTTPStatusError):
            try:
                return exc.response.status_code == 404
            except Exception:
                return False
        return False

    def _drop_db_row(kid_: str) -> None:
        with _sqlite3.connect(reg.db_path) as c:
            c.execute(
                "DELETE FROM registered_keywords "
                "WHERE account_customer_id=? AND ncc_keyword_id=?",
                (customer_id, kid_),
            )

    n_del, n_pause, n_fail, n_stale = 0, 0, 0, 0
    affected: List[str] = []
    for r in rows:
        kid = r["ncc_keyword_id"]
        kw_text = r["keyword"]
        try:
            await client.delete_keyword(kid)
            _drop_db_row(kid)
            n_del += 1
            affected.append(kw_text)
        except Exception as e1:
            if _is_already_gone(e1):
                _drop_db_row(kid)
                n_stale += 1
                affected.append(kw_text)
            else:
                try:
                    await client.pause_keyword(kid)
                    n_pause += 1
                    affected.append(kw_text)
                except Exception as e2:
                    if _is_already_gone(e2):
                        _drop_db_row(kid)
                        n_stale += 1
                        affected.append(kw_text)
                    else:
                        n_fail += 1
        await asyncio.sleep(0.15)

    if affected:
        try:
            pool.mark_rejected_by_naver(
                customer_id,
                [{"keyword": kw, "reason": f"rolling_heal(mt<{mt_ceiling})"} for kw in affected],
            )
        except Exception:
            pass
    try:
        record_auto_cleanup_run(uid, str(customer_id), n_del + n_pause + n_stale)
    except Exception:
        pass

    logger.warning(
        f"[pool/rolling-heal] uid={uid} cid={customer_id} "
        f"mt<{mt_ceiling} candidates={len(rows)} → "
        f"del={n_del} pause={n_pause} stale={n_stale} fail={n_fail}"
    )
    return n_del + n_pause + n_stale


# 등록-KW atom 계산 캐시 — collect tick 마다 5000 KW × 수천 atom 정규식 스캔(주석상
# "30M ops 동기 블록")이 단일 프로세스 event loop 를 수초 점유 → 그동안 /health 같은
# 초경량 요청도 10~27초 멈춤 (페이지 로딩 답답함의 주범). 등록 KW 는 천천히 변하므로
# customer 별로 결과를 캐시하고 TTL 안에는 재계산을 건너뛴다. 미스 시엔 to_thread 로
# 오프로드해 계산 중에도 event loop 가 API 요청을 계속 처리하게 한다.
_REG_ATOM_CACHE: Dict[int, Dict[str, Any]] = {}
_REG_ATOM_TTL_S = 900  # 15분 — atom 학습은 soft 휴리스틱이라 이 정도 staleness 무해


async def _run_pool_collect(uid: int, customer_id: Optional[int] = None, max_new: int = 5000, min_volume: int = 1):
    """수집 1회 — keywordstool로 새 키워드 발굴해 풀에 추가.
    customer_id 명시 시 그 광고주만 처리, 없으면 사용자의 가장 최근 광고주."""
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import get_ad_account_by_customer, get_ad_account_auto_cleanup
    import time as _time

    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    if customer_id is not None:
        account = get_ad_account_by_customer(uid, str(customer_id))
    else:
        account = get_ad_account(uid)
    if not account or not account.get("is_connected"):
        pool.record_run(uid, customer_id, "collect", "no_account",
                        error_message="광고 계정 미연결",
                        duration_ms=int((_time.monotonic()-t0)*1000))
        return
    customer_id = int(account.get("customer_id"))
    customer_id_for_log = customer_id

    # 채우기 에스컬레이션 — 관련성 floor/level 을 읽어 이번 라운드 게이트/조합주입에 반영.
    # level<5 이면 floor=50(기본) → ADD/DELETE 게이트가 오늘과 동일. 상위 레벨에서 floor 가
    # 내려가면 collect ADD gate 와 cap self-heal DELETE gate 를 함께 낮춰(대칭) thrash 없이
    # 관련성 낮은 롱테일까지 순서대로 흡수 → 10만 채우기. level≥2 면 조합 시드 주입 발동.
    try:
        _esc = pool.get_escalation(customer_id)
        _esc_floor = int(_esc.get("relevance_floor") or 50)
        _esc_level = int(_esc.get("level") or 0)
    except Exception:
        _esc_floor, _esc_level = 50, 0

    # STALL 방지 — floor 를 낮춰 약관련 롱테일을 흡수할 때는 등록 가능(volume≥10) 후보만 admit.
    # register 는 volume≥10 만 가져가므로(claim_pending min_volume=10), volume<10 을 admit 하면
    # pending 에 영구 적체 → headroom(=100k−active−pending) 잠식 → active 가 10만 못 미치고 정지.
    # floor<50 (티어 하강 중) 이면 min_volume 을 10 으로 올려 "등록되는 롱테일"만 채운다.
    # floor=50 (level<5) 이면 오늘과 동일(min_volume 그대로).
    if _esc_floor < 50:
        min_volume = max(min_volume, 10)

    reg = get_registered_keywords_db()
    pool_pending = (pool.stats(customer_id).get("by_status") or {}).get("pending", 0)
    active_reg = int((reg.stats(customer_id) or {}).get("active") or 0)
    headroom = 100_000 - active_reg - pool_pending
    # saturation 가드 — ≥95% (headroom ≤ 5000) 부터 self_heal 발동.
    # cleanup 으로 슬롯 회수되면 같은 tick 에서 곧바로 collect 이어 진행 (early return X) →
    # 다음 5분 tick 대기 제거. 100% 도달 + cleanup 0 인 dead state 만 cap_reached 로 skip.
    # autocomplete saturation guard (≥98%) 와 사이클: cleanup 으로 98% 미만 →
    # autocomplete 발굴 재개 → 다시 98% → cleanup → 평형.
    if headroom <= 5_000:
        cleaned_total = 0
        cleanup_label_parts: List[str] = []
        cfg = get_ad_account_auto_cleanup(uid, str(customer_id)) or {}
        if cfg.get("enabled"):
            cleaned = await _cap_triggered_self_heal(
                uid, customer_id, account,
                # DELETE 임계를 ADD floor(_esc_floor) 이하로 고정 — floor 가 25 로 내려가도
                # DELETE<floor, ADD≥floor 대칭 유지 → add→delete→add thrash 차단.
                threshold=min(int(cfg.get("threshold") or 30), _esc_floor),
                saved_relevance=list(cfg.get("relevance_keywords") or []),
                max_delete=500,
            )
            cleaned_total += cleaned
            if cleaned > 0:
                cleanup_label_parts.append(f"self_heal={cleaned}")

        # 폴백 — saved_relevance 가 없거나 self-heal 이 0 이면 mt 최하위 롤링 eject.
        # registered-as-seed 무한 발굴 사이클이 cap 에서 멈추지 않도록 보장.
        if cleaned_total == 0:
            rolled = await _cap_triggered_rolling_heal(
                uid, customer_id, account,
                max_delete=500,
                mt_ceiling=50,
                settle_hours=24,
            )
            cleaned_total += rolled
            if rolled > 0:
                cleanup_label_parts.append(f"rolling={rolled}")

        # cleanup 결과 별도 record_run — UI visibility 유지. early return 폐기:
        # 100% 도달이어도 cleanup 으로 회수된 슬롯 있으면 같은 tick 에서 즉시 collect 이어서
        # 진행 → 물갈이 속도 ↑ (다음 5분 tick 대기 제거).
        if cleaned_total > 0:
            try:
                pool.record_run(
                    uid, customer_id, "collect", "self_heal_cleanup",
                    pending_after=pool_pending,
                    error_message=(
                        f"cap_cleanup {' '.join(cleanup_label_parts)} → "
                        f"같은 tick 에서 collect {cleaned_total} 슬롯 회수"
                    )[:300],
                    duration_ms=int((_time.monotonic()-t0)*1000),
                )
            except Exception:
                pass
            logger.warning(
                f"[pool/collect] user={uid} saturation pre-cleanup "
                f"({100_000 - headroom}/100k) — {' '.join(cleanup_label_parts)} → collect 계속"
            )
            # cleanup 으로 회수된 슬롯 반영 (active_reg 가 줄었지만 stats 재조회는 부담 →
            # cleaned_total 만큼 headroom 가산).
            headroom = min(100_000, headroom + cleaned_total)
        elif headroom <= 0:
            # 100% 도달 + cleanup 0 → 진행 불가. 다음 tick 대기.
            logger.warning(f"[pool/collect] user={uid} 한도 도달 — skip (active={active_reg}, pending={pool_pending})")
            pool.record_run(uid, customer_id, "collect", "cap_reached",
                            pending_after=pool_pending,
                            error_message=f"active={active_reg}+pending={pool_pending}≥100000",
                            duration_ms=int((_time.monotonic()-t0)*1000))
            return
    target = min(max_new, headroom)

    # 동적 도메인 토큰셋 — 우선순위: saved relevance_keywords > user_seed > POOL baseline.
    # 2026-05-08 추가: relevance_keywords 가 저장돼 있으면 그것만으로 도메인 게이트 빌드.
    # Why: user_seed 풀이 한약재 (두릅/황기/천문동/용골/행인/지모/백합) 등으로 오염된 경우
    #      그 atom 이 도메인 토큰에 합류해 식물·원예 KW (나무수국/꽃산딸나무/고사리종자)
    #      가 cross-domain 통과. relevance_keywords 가 사용자 진짜 의도라면 그것만 사용해
    #      collect 게이트를 좁게 유지 → 풀에 잡음 시드 있어도 drift 차단.
    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel
    saved_relevance = _get_rel(uid, str(customer_id))
    initial_seeds = pool.list_seed_whitelist(customer_id)
    initial_user_seeds_only = pool.list_user_seeds(customer_id)
    cold_start = not initial_user_seeds_only

    if saved_relevance and len(saved_relevance) >= 1:
        # 사용자가 명시한 relevance_keywords 만으로 도메인 게이트 — 풀 오염 무시.
        domain_token_set = _derive_seed_tokens(saved_relevance) | _build_seed_atoms(saved_relevance)
        domain_basis = f"saved_relevance({len(saved_relevance)})"
    elif cold_start:
        # 시드 0 광고주 — POOL baseline (사용자 시드 추가까지 진입 가능).
        domain_token_set = _build_domain_token_set(initial_seeds)
        domain_basis = f"cold_start_pool_baseline({len(POOL_DOMAIN_TOKENS)})"
    else:
        # 일반 — user_seed atom. POOL hardcoded + auto_promoted_seed 게이트 atom 에서 배제.
        domain_token_set = _derive_seed_tokens(initial_user_seeds_only) | _build_seed_atoms(initial_user_seeds_only)
        domain_basis = f"user_seed_pool({len(initial_user_seeds_only)})"
    derived_count = len(domain_token_set) - (len(POOL_DOMAIN_TOKENS) if cold_start and not saved_relevance else 0)

    # collect ADD score gate — atom 화이트리스트 + 점수 ≥ threshold 양쪽 충족 요구.
    # atom-only 게이트는 우연한 2-gram 매칭 (예: "탈" → "테이블렌탈") 으로 점수 30짜리도
    # 통과 → 15분 cleanup 이 다시 삭제 → API/네이버 호출 낭비. ADD/DELETE 양쪽 모두
    # 같은 threshold 사용해 전체 사이클 일관성 유지.
    # saved_relevance < 3 (도메인 시그널 부족) 인 경우 게이트 비활성화 (cold_start 광고주 보호).
    from database.naver_ad_db import get_ad_account_auto_cleanup as _get_collect_thr_cfg
    _collect_thr_cfg = _get_collect_thr_cfg(uid, str(customer_id)) or {}
    _collect_score_thr = int(_collect_thr_cfg.get("threshold") or 50)
    # 에스컬레이션 floor 하강 반영 — floor 는 낮추기만(min), 절대 올리지 않음.
    # level<5 이면 floor=50 이라 기본 threshold(50) 와 동일. 상위 레벨에서 floor 가
    # 45→40→…→25 로 내려가면 도메인 토큰은 통과했으나 점수 낮은 롱테일까지 순서대로 흡수.
    _collect_score_thr = min(_collect_score_thr, _esc_floor)
    _collect_score_seeds = list(saved_relevance) if saved_relevance and len([s for s in saved_relevance if s and len(s) >= 2]) >= 3 else []

    # 등록 KW atom — cleanup/collect 게이트와 동일 기준.
    # ANCHOR: user_seed atom (length≥3) 을 포함하는 등록 KW 만 학습.
    # Why: 무필터 학습 시 POOL 토큰("교육"/"강의" 등)으로만 통과한 cross-domain KW 가
    #      자기 atom 을 토큰셋에 주입 → 다음 라운드에 cascade drift.
    #      예: 시드 "내일배움카드" → POOL "교육" 매치로 "블렌더교육" 등록 →
    #          "블렌더" atom 학습 → "블렌더VFX/2D/모션" 모두 통과 → 도메인 점프.
    # anchor 로 user_seed 라인 KW 만 atom 기여 → drift 전파 차단.
    # 학습 atom 도 length≥3 — 2-letter (RM/AI/IT) 가 영문 KW 전체를 통과시키는 폴루션 방지.
    # 등록-KW atom (anchor_set + registered_atoms) — 캐시 우선, 미스 시 to_thread 오프로드.
    # anchor: user_seed 의 length≥3 atom 만 — 짧은 2-gram atom (간/염/의/원 등) 이
    # cross-domain 통과시키는 누수 차단. niche 의료 시드 ("A형 간염" → "간염" atom)
    # 가 무관 KW (예: "간장/감자/은염생산") 의 2-gram 매칭으로 anchor 통과해
    # registered_atom 학습 → cascade drift 발생 위험.
    _cache_hit = _REG_ATOM_CACHE.get(customer_id)
    if _cache_hit and (_time.monotonic() - _cache_hit["ts"]) < _REG_ATOM_TTL_S:
        anchor_set = _cache_hit["anchor_set"]
        registered_atoms = _cache_hit["registered_atoms"]
        _reg_raw_n = _cache_hit["reg_raw_n"]
        _reg_learned_n = _cache_hit["reg_learned_n"]
    else:
        def _compute_reg_atoms():
            try:
                reg_raw = pool.list_top_registered(customer_id, limit=5000, min_volume=30)
            except Exception as e:
                logger.warning(f"[pool/collect] 등록 KW atom 조회 실패: {e}")
                reg_raw = []
            a_set = {a for a in _build_seed_atoms(pool.list_user_seeds(customer_id)) if len(a) >= 3}
            if a_set:
                # PERF: 5000 KW × 6000+ atom Python loop = 30M ops. 정규식 multi-pattern (~100배 빠름).
                import re as _re_a
                _anchor_re = _re_a.compile("|".join(_re_a.escape(a) for a in a_set))
                reg_for = [kw for kw in reg_raw if _anchor_re.search(kw)]
            else:
                # user_seed 0개인 신규 광고주 — anchor 비어있으면 학습 안 함 (drift 위험 큼).
                reg_for = []
            # 학습 atom 만 length≥3 — RM/AI/IT 같은 2-letter 영문 폴루션 차단.
            reg_atoms = {a for a in _build_seed_atoms(reg_for) if len(a) >= 3}
            return a_set, reg_atoms, len(reg_raw), len(reg_for)

        anchor_set, registered_atoms, _reg_raw_n, _reg_learned_n = await asyncio.to_thread(_compute_reg_atoms)
        _REG_ATOM_CACHE[customer_id] = {
            "ts": _time.monotonic(),
            "anchor_set": anchor_set,
            "registered_atoms": registered_atoms,
            "reg_raw_n": _reg_raw_n,
            "reg_learned_n": _reg_learned_n,
        }

    logger.warning(
        f"[pool/collect] user={uid} 도메인 토큰 {len(domain_token_set)}개 "
        f"basis={domain_basis} "
        f"+ 등록 atom {len(registered_atoms)}개 "
        f"({_reg_learned_n}/{_reg_raw_n} KW anchor 통과, "
        f"anchor {len(anchor_set)}개{' [cached]' if _cache_hit else ''}) cold_start={cold_start}"
    )

    # 도메인 미포함 키워드 자동 cleanup (registered 제외) — 매 라운드 시작 시
    # 게이트 = domain_token_set ∪ initial_seed_atoms ∪ registered_atoms (collect 게이트와 동일)
    # initial_seed_atoms 추가: 사용자가 풀에 직접 넣은 시드의 atom 도 cleanup 통과 보장
    # (predictably user_seed 가 변경된 직후 한 라운드만 이 단계에서 cleanup 게이트가 살아남)
    # PERF: cleanup 은 sync DB 작업 — token 수 × pending row 수 substring check 가
    #       async 이벤트 루프를 블록 (12k tokens × 50k rows → 헬스체크 timeout).
    #       to_thread 로 worker thread 에서 실행해 이벤트 루프 보호 + token 수 가
    #       임계 초과 시 skip (다음 라운드에서 좁아지면 재시도).
    # cleanup atom 도 user_seed 만 — auto_promoted_seed 가 cascade drift 로 무관 KW 였을
    # 가능성 차단. 시드 atom 누락된 KW 는 어차피 domain_token_set 에서도 매치 안 됨.
    initial_seed_atoms = _build_seed_atoms(initial_user_seeds_only) if not cold_start else _build_seed_atoms(initial_seeds)
    cleanup_tokens = domain_token_set | initial_seed_atoms | registered_atoms
    if len(cleanup_tokens) <= 3000:
        try:
            cleaned = await asyncio.to_thread(
                pool.cleanup_offdomain, customer_id, list(cleanup_tokens)
            )
            if cleaned > 0:
                logger.warning(f"[pool/cleanup] off-domain row 자동 삭제 {cleaned}개")
        except Exception as e:
            logger.warning(f"[pool/cleanup] 실패: {e}")
    else:
        logger.warning(
            f"[pool/cleanup] skip — 토큰 {len(cleanup_tokens)}개 > 3000 임계 "
            f"(이벤트 루프 보호). registered_atoms 학습이 안정화되면 재진입."
        )

    # 자동 승격 시드 중 자식 0 + 30분 경과 자력 삭제 (user_seed는 면제)
    try:
        childless = pool.cleanup_childless_auto_seeds(customer_id, min_age_minutes=30)
        if childless > 0:
            logger.warning(f"[pool/cleanup] 자식 0 자동 시드 자력 삭제 {childless}개")
    except Exception as e:
        logger.warning(f"[pool/cleanup-childless] 실패: {e}")

    # 시드 자가확장 — anchor (user_seed atom) 포함 KW 만 promote.
    # POOL bridge 로 등록된 cross-niche KW (예: "블렌더강의") 가 promote → seed_atoms
    # 합류 → 다음 라운드 그 niche cascade drift 발생을 차단. 그 niche 는 bridge 가
    # 매 라운드 재호출하므로 promote 없어도 새 KW 발굴 계속됨.
    try:
        promoted = pool.promote_seeds(
            customer_id, limit=50, min_volume=30, max_total_seeds=500,
            domain_tokens=list(anchor_set),
        )
        if promoted:
            logger.warning(
                f"[pool/collect] user={uid} 시드 자동 승격 {len(promoted)}개: "
                + ", ".join(f"{p['keyword']}({p['monthly_total']})" for p in promoted)
            )
            # 승격 후 토큰셋 재계산 — saved_relevance 우선 (사용자 의도 고정).
            if saved_relevance and len(saved_relevance) >= 1:
                # saved_relevance 사용 중이면 promote 가 토큰셋을 흔들면 안 됨 — 그대로 유지.
                pass
            elif cold_start:
                domain_token_set = _build_domain_token_set(
                    pool.list_seed_whitelist(customer_id)
                )
            else:
                fresh_user_seeds = pool.list_user_seeds(customer_id) or initial_user_seeds_only
                domain_token_set = _derive_seed_tokens(fresh_user_seeds) | _build_seed_atoms(fresh_user_seeds)
    except Exception as e:
        logger.warning(f"[pool/collect] promote_seeds 실패: {e}")
        promoted = []

    # C1 fix: get_recent_seeds 는 source 필터 없이 모든 row 의 seed 를 반환 → legacy POOL
    # bridge 시드 ("대출/렌탈/배달/미용" 등) 가 풀 row 자식으로 살아있는 한 매 라운드
    # keywordstool 호출에 사용 → API quota 낭비 + 잔재 재활성화 risk. list_seed_whitelist
    # 는 source IN ('user_seed', 'auto_promoted_seed') 만 반환해 legacy POOL bridge 차단.
    seeds = pool.list_seed_whitelist(customer_id)
    if not seeds:
        # 자가치유 (a): 등록 키워드 중 검색량 상위 10개를 user_seed 로 자동 reseed.
        # 시드가 비면 collection 영구 정지 → 등록 키워드에서 핵심어 자동 추출.
        try:
            top_kw = pool.list_top_registered(customer_id, limit=10, min_volume=100)
        except Exception as e:
            logger.warning(f"[pool/collect] auto-reseed 후보 조회 실패: {e}")
            top_kw = []
        # 도메인 게이트 — saved_relevance 있으면 score>30 만 reseed. 등록 풀이 drift 로
        # 오염된 계정 (소잠한의원 차 KW 사고) 에서 top-mt 가 차 KW 라면 그게 user_seed 로
        # 재주입되는 catastrophic loop 차단.
        if top_kw and saved_relevance and len([s for s in saved_relevance if s and len(s) >= 2]) >= 3:
            before = len(top_kw)
            top_kw = [k for k in top_kw if _compute_relevance_score(k, saved_relevance) > 30]
            if before != len(top_kw):
                logger.warning(
                    f"[pool/collect] auto-reseed 도메인 게이트 — {before} → {len(top_kw)}"
                )
        if top_kw:
            items = [{"keyword": k, "seed": k, "source": "user_seed", "monthly_total": 0} for k in top_kw]
            try:
                pool.add_candidates(uid, customer_id, items)
                logger.warning(
                    f"[pool/collect] user={uid} 시드 자동 복구 {len(top_kw)}개: "
                    + ", ".join(top_kw[:5]) + (" ..." if len(top_kw) > 5 else "")
                )
                seeds = top_kw
            except Exception as e:
                logger.warning(f"[pool/collect] auto-reseed insert 실패: {e}")
        if not seeds:
            logger.warning(f"[pool/collect] user={uid} 시드 없음 + 등록 키워드 없음 — UI에서 초기 시드 제공 필요")
            pool.record_run(uid, customer_id, "collect", "no_seed",
                            pending_after=pool_pending,
                            error_message="UI에서 초기 시드 추가 필요",
                            duration_ms=int((_time.monotonic()-t0)*1000))
            return

    # 화이트리스트 (keywordstool 호출용): user_seed + auto_promoted_seed.
    # 발굴 다양성은 유지하되, 게이트 atom 은 user_seed 만으로 좁힘 (cascade drift 차단).
    whitelist = pool.list_seed_whitelist(customer_id)
    if not whitelist:
        whitelist = seeds  # 폴백
    # 게이트 seed_atoms 는 user_seed 만 — promoted 가 발굴해온 KW 도 user_seed atom 매치 필수.
    user_seed_now = pool.list_user_seeds(customer_id) or initial_user_seeds_only
    seed_atoms = _build_seed_atoms(user_seed_now) if user_seed_now else _build_seed_atoms(whitelist)

    # 통합 게이트 — domain_token_set ∪ seed_atoms ∪ registered_atoms.
    # 모두 user_seed lineage. cold_start 광고주는 domain_token_set 에 POOL baseline 포함.
    unified_tokens = set(domain_token_set) | seed_atoms | registered_atoms

    # Loose mode 자동 진입 — cold_start 광고주만 사용. user_seed ≥ 1 면 절대 비활성.
    # Why: niche user_seed (예: 의료 희귀병) 가 5회 연속 발굴 0 + reject 폭주하면 loose_mode 가
    #      "길이 2+ 모든 KW 통과" 로 게이트를 무력화 → user_seed 무관 KW 무차별 INSERT →
    #      registered_atoms 학습 → 영구적 cascade drift. 사용자 의도 ("내 시드 외 무관 KW reject")
    #      정면 위반. user_seed 가 niche 라 발굴 못 해도 무관 도메인으로 점프해선 안 됨 —
    #      이 경우 "발굴 0" 으로 두는 게 옳다 (사용자가 시드 추가하거나 niche 포기).
    loose_mode = False
    if cold_start:
        try:
            recent = pool.recent_runs(customer_id, limit=5)
            collects = [r for r in recent if r.get("kind") == "collect"]
            if len(collects) >= 3:
                high_reject = sum(
                    1 for r in collects
                    if (r.get("added") or 0) == 0 and (r.get("skipped") or 0) >= 500
                )
                if high_reject >= 3:
                    loose_mode = True
        except Exception:
            pass

    logger.warning(
        f"[pool/collect] user={uid} 시작 target={target} seeds={len(seeds)} "
        f"whitelist={len(whitelist)} unified_tokens={len(unified_tokens)} "
        f"seed_atoms={len(seed_atoms)} reg_atoms={len(registered_atoms)} "
        f"reg_learned_from={_reg_learned_n} "
        f"promoted={len(promoted)} loose={loose_mode}"
    )

    if loose_mode:
        logger.warning(
            f"[pool/collect] user={uid} LOOSE MODE — 최근 collect 3+ 회 연속 high-reject. "
            f"min_volume {min_volume}→1, 게이트 완화."
        )
        min_volume = max(1, min_volume // 5)

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    # PERF: unified_tokens 가 14k+ 이면 per-kw `any(t in kw for t in tokens)` =
    # 1.3B+ Python ops/round → async 이벤트 루프 30-60s 블록 → fly 헬스체크 실패.
    # 정규식 컴파일로 C 구현 멀티패턴 매칭 (~100배 빠름) 으로 전환.
    import re as _re
    _whitelist_re = _re.compile(
        "|".join(_re.escape(t) for t in unified_tokens)
    ) if unified_tokens else None

    def _matches_whitelist(kw: str) -> str:
        # 단일 게이트 — unified_tokens 중 하나라도 매치하면 통과.
        # 반환: "" = 통과, "domain" = 어떤 토큰도 안 맞음.
        # loose_mode 면 길이 ≥ 2 인 한국어/영문/숫자 키워드는 통과 (도메인 무관 fallback).
        if _whitelist_re and _whitelist_re.search(kw):
            return ""
        if loose_mode and len(kw) >= 2:
            return ""
        return "domain"

    added = 0
    rejected = 0
    reject_no_domain = 0
    reject_no_seed_match = 0
    sample_no_domain: List[str] = []
    sample_no_seed: List[str] = []
    # 도메인미스 reject 처리 2-layer:
    #   [Top tier mt≥100]  reject_for_ai     → GPT 분류 통과만 자식 합류 (보수)
    #   [Mid tier mt 30~99] reject_direct    → GPT 우회 자식 풀 직접 추가 (drift 감수)
    # 사용자 명시: "drift 위험 감수해도 빠르게 채우기".
    # Drift 안전 장치 (이미 cron):
    #   - 등록 후 검수 거부 KW = inspect cron 10분마다 자동 삭제
    #   - 클릭 발생한 무관 KW = click cleanup cron 15분마다 점수 ≤ 30 자동 삭제
    #   - 도메인 안 맞는 KW = domain cleanup cron 매시 자동 삭제
    # 풀/분류 이력 KW 제외 — INSERT OR IGNORE 사고 방지.
    reject_for_ai: List[Dict] = []
    reject_for_ai_seen: Set[str] = set()
    reject_direct: List[Dict] = []
    reject_direct_seen: Set[str] = set()
    classified_reject_set: Set[str] = set()
    pool_kw_set: Set[str] = set()
    try:
        classified_reject_set = set(pool.list_classified_reject_keywords(customer_id))
    except Exception as e:
        logger.warning(f"[pool/collect] classified set 로드 실패: {e}")
    try:
        pool_kw_set = pool.list_pool_keyword_set(customer_id)
    except Exception as e:
        logger.warning(f"[pool/collect] pool_kw_set 로드 실패: {e}")
    api_errors: List[str] = []
    seeds_processed = 0
    bfs_calls = 0

    # 자가치유 (b): 포화 감지 — 최근 5회 모두 added=0 + skipped<500 면 풀이 saturate.
    # keywordstool 결과가 모두 중복이라 같은 시드로는 새 발굴 불가능. 시드 확장 주입.
    saturated = False
    try:
        sat = pool.detect_saturation(customer_id, n_recent=5)
        saturated = bool(sat.get("is_saturated"))
    except Exception:
        pass

    # L2 조합 생성 — 지역·의도 수식어 × 머리어(시드) 로 새 hint 생성해 keywordstool
    # 폐포를 깬다. 같은 시드 반복 호출이 같은 결과만 돌려주는 유리천장 돌파의 핵심.
    # 발동: 포화 감지(기존) 또는 에스컬레이션 floor 하강 중(_esc_floor<50). 미포화·floor50이면 오늘과 동일.
    if saturated or _esc_floor < 50:
        # 전국 시/도·주요 시 + 구매의도 수식어 — level≥2 면 조합 폭을 넓힌다.
        EXPANSION_AFFIXES = [
            "강남", "서울", "부산", "대구", "인천", "경기", "대전", "광주", "울산",
            "수원", "성남", "고양", "용인", "창원", "청주", "전주", "천안", "제주",
            "추천", "후기", "비교", "가격", "잘하는곳", "잘하는", "찾기", "순위",
            "전문", "전문점", "무료", "상담", "신청", "예약", "할인", "이벤트",
            "문의", "근처", "당일", "야간", "주말",
        ]
        # level 이 높을수록 시드당 조합 수·주입 상한을 키운다.
        per_seed = 3 if _esc_level < 2 else (4 if _esc_level < 4 else 5)
        take = 60 if _esc_level < 2 else (120 if _esc_level < 4 else 200)
        seed_span = 30 if _esc_level < 2 else 60
        logger.warning(
            f"[pool/collect] user={uid} L2 조합 주입 — saturated={saturated} level={_esc_level} "
            f"(per_seed={per_seed} take={take})"
        )
        existing = set(seeds)
        injected: List[str] = []
        for s in seeds[:seed_span]:
            for aff in random.sample(EXPANSION_AFFIXES, min(per_seed, len(EXPANSION_AFFIXES))):
                for combo in (aff + s, s + aff):
                    if combo not in existing and len(combo) <= 25:
                        injected.append(combo)
                        existing.add(combo)
        if injected:
            seeds = list(seeds) + injected[:take]
            logger.warning(
                f"[pool/collect] user={uid} 확장 시드 {len(injected[:take])}개 주입 (sample: "
                + ", ".join(injected[:5]) + ")"
            )

    # NICHE BRIDGE — cold start (user_seed 0) 일 때만 POOL bridge 사용.
    # 사용자 의도 변경 ("내가 시드 넣은거에서만 추천, 무관한거 싹 삭제") 으로
    # user_seed ≥ 1 광고주는 BRIDGE 비활성 — 시드 도메인 외 niche 자동 점프 차단.
    # 의료 광고주(소잠한의원)에 "대출/렌탈/리스" 가 매 라운드 강제 시드로 들어가던 누수 차단.
    if cold_start:
        bridge_pool = [t for t in POOL_DOMAIN_TOKENS if len(t) >= 2]
        random.shuffle(bridge_pool)
        bridge_round = bridge_pool[:15]
    else:
        bridge_round = []

    # 시드 셔플 — 매 라운드 다른 60개 처리 (200 시드 다양성 확보).
    seed_pool = list(seeds)
    random.shuffle(seed_pool)
    # niche timeout backoff — 최근 30분 내 ConnectTimeout 난 시드는 라운드 제외.
    # 의료 희귀병 시드(피부 전이암/혈관각화증 등) cascade timeout → circuit OPEN →
    # 전체 collect 정지 패턴 차단. 같은 시드를 매 5분마다 재시도하지 않고 30분 backoff.
    seeds_in_backoff = [s for s in seed_pool if _is_seed_in_backoff(customer_id, s)]
    seeds_eligible = [s for s in seed_pool if not _is_seed_in_backoff(customer_id, s)]
    # 슬롯: cold_start 면 user(45) + bridge(15) = 60, 아니면 user 120 + registered 120.
    # user_seed 충분한 광고주는 라운드당 시드 2배 → 시간당 신규 발굴량 ~2배.
    user_quota = 45 if cold_start else 120
    # REGISTERED-AS-SEED — saturation 돌파용 (사용자 명시 의도: 검색량 있는 등록 KW 의
    # 연관 KW 끝까지 발굴). user_seed 2~5k 만으로는 keywordstool 응답 saturate →
    # +0~+10/배치 정체. 90k+ 등록 KW 자체를 매 라운드 다른 120개씩 시드 투입.
    # unified_tokens 에 registered_atoms 가 이미 포함 → 자식 게이트 통과 보장.
    registered_round: List[str] = []
    if not cold_start:
        try:
            registered_round = pool.list_registered_random_seeds(
                customer_id, limit=120, min_volume=10,
            )
        except Exception as e:
            logger.warning(f"[pool/collect] registered-as-seed 로드 실패: {e}")
            registered_round = []
    seed_round = seeds_eligible[:user_quota] + registered_round + bridge_round
    logger.warning(
        f"[pool/collect] user={uid} 시드 라운드 — user/promoted={min(user_quota, len(seeds_eligible))} "
        f"+ registered={len(registered_round)} + POOL bridge={len(bridge_round)} "
        f"(총 {len(seed_round)}) cold_start={cold_start} backoff_skip={len(seeds_in_backoff)}"
    )
    # circuit breaker 인스턴스 — 시드 라운드 중 OPEN 감지 시 남은 시드 fail-fast 차단.
    # naver_ad_service 모듈 레벨 singleton 공유.
    from services.naver_ad_service import _naver_api_breaker, NaverApiCircuitOpenError
    circuit_aborted = False

    # ==========================================================================
    # BATCHED keywordstool 호출 — 시드 5개씩 묶어 1콜로 처리 (2026-05-07).
    # 종전: 시드 60개 = 60콜 → ConnectTimeout 10회 누적 → circuit OPEN → 50시드 abort.
    # 변경: 시드 60개 = 12콜 (배치당 5 hint) → API 호출 5배 감소 → timeout 폭주 차단.
    # 네이버 /keywordstool 은 hintKeywords 콤마구분 5개까지 허용 (naver_ad_service.py:629).
    # ==========================================================================
    SEED_BATCH = 5

    def _process_keyword_items(
        items: List[Dict],
        seed_label: str,
        sub_candidates_out: List[Dict],
        bfs_pool_out: List[tuple],
    ) -> None:
        """keywordList 처리 — primary/BFS 양쪽에서 동일 로직 공유."""
        for item in items:
            kw = (item.get("relKeyword") or "").strip()
            if not kw:
                continue
            pc = _parse_naver_count(item.get("monthlyPcQcCnt"))
            mob = _parse_naver_count(item.get("monthlyMobileQcCnt"))
            mt = pc + mob
            if mt < min_volume:
                continue
            reason = _matches_whitelist(kw)
            if reason:
                # rejected 카운트는 outer scope (nonlocal)
                nonlocal rejected, reject_no_domain, reject_no_seed_match
                rejected += 1
                if reason == "domain":
                    reject_no_domain += 1
                    if len(sample_no_domain) < 10:
                        sample_no_domain.append(kw)
                    # AI-first 게이트 — mt≥1 모든 도메인미스 KW 를 GPT 분류 대기열로.
                    # 한의원 13만 drift 사고 (mt 30~99 GPT 우회) + AI cleanup 끈 결정과 일관.
                    # GPT 통과만 풀 합류 → 사후 cleanup DELETE 가 필요 없는 구조.
                    if (
                        kw in classified_reject_set
                        or kw in pool_kw_set
                    ):
                        pass  # 이미 처리된 KW
                    elif mt >= 1 and kw not in reject_for_ai_seen:
                        reject_for_ai_seen.add(kw)
                        reject_for_ai.append({"keyword": kw, "monthly_total": mt})
                else:
                    reject_no_seed_match += 1
                    if len(sample_no_seed) < 10:
                        sample_no_seed.append(kw)
                continue

            # 점수 게이트 — atom 화이트리스트 통과해도 _compute_relevance_score < threshold
            # 이면 풀 합류 차단. ADD/DELETE 양쪽 모두 같은 threshold (옵션 B 경계 보존).
            # saved_relevance 시드가 < 3 이면 게이트 비활성 (cold_start 광고주 보호).
            # nonlocal 은 line 3181 에서 이미 함수 스코프로 선언됨 — 여기서 재선언 X.
            if _collect_score_seeds:
                _sc = _compute_relevance_score(kw, _collect_score_seeds)
                if _sc < _collect_score_thr:
                    rejected += 1
                    reject_no_seed_match += 1  # 점수 미달 — 시드미스 카테고리로 기록
                    if len(sample_no_seed) < 10:
                        sample_no_seed.append(kw)
                    continue

            sub_candidates_out.append({
                "keyword": kw, "monthly_total": mt,
                "monthly_pc": pc, "monthly_mobile": mob,
                "comp_idx": item.get("compIdx"),
                "seed": seed_label,
            })
            if mt >= 100 and len(kw) >= 2:
                bfs_pool_out.append((kw, mt))

    # 배치 단위로 청크
    for batch_start in range(0, len(seed_round), SEED_BATCH):
        if added >= target:
            break
        if _naver_api_breaker.is_open():
            circuit_aborted = True
            logger.warning(
                f"[pool/collect] user={uid} circuit OPEN — 남은 시드 abort (처리 {seeds_processed}/{len(seed_round)})"
            )
            break

        chunk_raw = seed_round[batch_start:batch_start + SEED_BATCH]
        # 각 시드 sanitize — 빈/짧은 시드는 chunk 에서 제외
        chunk_sanitized: List[Tuple[str, str]] = []  # [(seed_raw, seed_clean)]
        for s_raw in chunk_raw:
            s_clean = (s_raw or "").replace(" ", "").strip()
            if not s_clean or len(s_clean) < 2:
                continue
            chunk_sanitized.append((s_raw, s_clean))
        if not chunk_sanitized:
            continue

        seeds_processed += len(chunk_sanitized)
        hints = [c[1] for c in chunk_sanitized]
        hint_str = ",".join(hints)
        # seed 컬럼 attribution — batch hint 전체를 콤마구분으로 저장하면 UI 시드별 표가
        # "척수성근위축증,마자인,…" 같이 묶여 보이고 그 row 의 자식 카운트가 0 으로 잡힘.
        # batch 의 첫 시드를 대표 라벨로. 5 시드 중 1로 attribution 부풀려지지만 UI 명확.
        seed_label = chunk_sanitized[0][0]

        try:
            related = await client.get_related_keywords(hint_str, show_detail=True)
        except NaverApiCircuitOpenError:
            circuit_aborted = True
            logger.warning(
                f"[pool/collect] user={uid} batch '{hint_str[:40]}' 처리 중 circuit OPEN — abort"
            )
            break
        except Exception as e:
            # 배치 전체 실패 — 5개 시드 모두 backoff 처리 (어느 시드가 원인인지 알 수 없음)
            err_name = type(e).__name__
            err_msg_full = str(e)[:200]
            msg = f"batch[{hint_str[:40]}]: {err_name}: {err_msg_full[:80]}"
            logger.warning(f"[pool/collect] BATCH API 실패 {msg}")
            if "Timeout" in err_name or "ConnectTimeout" in err_msg_full:
                for s_raw, _ in chunk_sanitized:
                    _mark_seed_timeout(customer_id, s_raw)
            if "11001" in err_msg_full or "400" in err_msg_full:
                logger.warning(
                    f"[pool/collect/11001] hints={hints!r} — 네이버 keywordstool 거부"
                )
            api_errors.append(msg)
            continue

        items = related.get("keywordList", []) if isinstance(related, dict) else []
        candidates: List[Dict] = []
        bfs_pool: List[tuple] = []
        _process_keyword_items(items, seed_label, candidates, bfs_pool)
        added += pool.add_candidates(uid, customer_id, candidates)
        await asyncio.sleep(0.3)

        # BFS 2nd-level — 배치당 검색량 상위 4개 (발굴 면적 2배, 비용 0 — keywordstool 무료).
        bfs_pool.sort(key=lambda x: -x[1])
        for bfs_kw, _ in bfs_pool[:4]:
            if _naver_api_breaker.is_open():
                break
            try:
                bfs_calls += 1
                related2 = await client.get_related_keywords(bfs_kw, show_detail=True)
                items2 = related2.get("keywordList", []) if isinstance(related2, dict) else []
                sub_candidates: List[Dict] = []
                _process_keyword_items(items2, seed_label, sub_candidates, [])
                added += pool.add_candidates(uid, customer_id, sub_candidates)
                await asyncio.sleep(0.3)
            except Exception as e:
                logger.warning(f"[pool/collect/BFS] {bfs_kw} 실패: {e}")
    logger.warning(
        f"[pool/collect] user={uid} 새 키워드 {added}개 "
        f"(rejected {rejected} = 도메인미스 {reject_no_domain} / 시드미스 {reject_no_seed_match}, "
        f"시드 {seeds_processed}개, BFS {bfs_calls}회)"
    )
    if sample_no_domain:
        logger.warning(f"[pool/collect] 도메인미스 샘플: {', '.join(sample_no_domain)}")
    if sample_no_seed:
        logger.warning(f"[pool/collect] 시드미스 샘플: {', '.join(sample_no_seed)}")
    # AI 분류용 reject 누적 batch INSERT — 라운드당 최대 1000개로 cap (DB 비대화 방지).
    # 백업 cron (_ai_classify_tick) 의 입력 풀 + UI 카운터 표시용.
    if reject_for_ai:
        try:
            saved = pool.add_rejects(customer_id, reject_for_ai[:1000])
            if saved:
                logger.warning(
                    f"[pool/collect] AI 분류 후보 reject {saved}개 누적 (검색량≥100)"
                )
        except Exception as e:
            logger.warning(f"[pool/collect] add_rejects 실패: {e}")

    # ============ reject_direct (GPT 우회 자식 합류) — 비활성 ============
    # 한의원 광고주 13만 drift 사고의 근본 — mt 30~99 KW 가 GPT 검증 없이 자식 합류 →
    # 무관 KW 등록 → 사후 AI cleanup 이 다시 대량 DELETE 하는 위험 사이클.
    # 현재 흐름: 모든 도메인미스 KW (mt≥1) 가 reject_for_ai 로 모이고
    # inline AI 게이트가 의미적으로 일치하는 것만 풀 합류 → drift 사전 차단.
    # 이 블록은 호환성 위해 남겨두지만 reject_direct 는 항상 빈 리스트.
    direct_added = 0
    # ===========================================================================

    # ============ Inline AI 자식 게이트 — 매 collect 직후 즉시 GPT 분류 ============
    # reject 풀의 fresh 후보 (검색량 상위 200개) 를 GPT-4o-mini 가 시드 도메인 일치 분류.
    # approved 는 시드가 아니라 자식 KW 로 직접 등록 풀에 추가 (source='ai_inline').
    # 게이트 atom 우회 — 다음 register cron (2분 후) 즉시 광고그룹 등록.
    # 비용: collect 1회당 ~$0.001 (시간당 ~$0.012, 월 ~$10). reject_for_ai 비면 skip.
    inline_ai_added = 0
    inline_ai_approved = 0
    inline_ai_discarded = 0
    try:
        from config import settings as _settings
        if reject_for_ai and _settings.OPENAI_API_KEY:
            from services.ai_seed_suggester import classify_rejects as _classify_rejects
            # AI-first 빠른 채움 — 라운드당 GPT 분류 처리량 10배 (200 → 2000).
            # classify_rejects 내부 cap 200 / 호출 → 200 batch × 10 = 2000 audit.
            # asyncio.gather + Semaphore(4) 병렬화 — GPT round-trip 4~5초 × 10 sequential =
            # 50초 → 병렬 ~12초. collect cron 5분 안에 충분히 끝남.
            top_rejects = sorted(
                reject_for_ai, key=lambda r: -int(r.get("monthly_total") or 0)
            )[:2000]
            ai_seeds_input = pool.list_user_seeds(customer_id) or whitelist
            if ai_seeds_input and top_rejects:
                ai_t0 = _time.monotonic()
                BATCH = 200
                _sem = asyncio.Semaphore(4)

                async def _classify_batch(batch_items: List[Dict]) -> Dict[str, Any]:
                    async with _sem:
                        return await _classify_rejects(
                            ai_seeds_input, batch_items, seed_sample_size=50,
                            saved_relevance=saved_relevance,
                        )

                batches = [
                    top_rejects[i:i + BATCH]
                    for i in range(0, len(top_rejects), BATCH)
                ]
                batch_results = await asyncio.gather(
                    *[_classify_batch(b) for b in batches],
                    return_exceptions=True,
                )
                ai_ms = int((_time.monotonic() - ai_t0) * 1000)

                approved: List[str] = []
                discarded: List[str] = []
                batch_ok = 0
                batch_fail = 0
                last_rationale = ""
                for r in batch_results:
                    if isinstance(r, Exception):
                        batch_fail += 1
                        logger.warning(f"[pool/collect/ai-inline] batch 예외: {r}")
                        continue
                    if not r.get("success"):
                        batch_fail += 1
                        logger.warning(
                            f"[pool/collect/ai-inline] batch 실패: {r.get('message')}"
                        )
                        continue
                    batch_ok += 1
                    approved.extend(r.get("approved") or [])
                    discarded.extend(r.get("discarded") or [])
                    if r.get("rationale"):
                        last_rationale = r["rationale"]

                # dedup (병렬 batch 동일 KW 가 중복 분류될 일은 없지만 안전장치)
                approved = list(dict.fromkeys(approved))
                discarded = list(dict.fromkeys(discarded))
                inline_ai_approved = len(approved)
                inline_ai_discarded = len(discarded)

                # GPT 통과 < 50 일 때 fallback — niche 시드에서 GPT 가 거의 다 컷하면
                # 풀이 영구 안 채워지는 사고 차단. 검색량 상위 + 점수 ≥ threshold 만 합류.
                # 옛 fallback 은 검색량만 봐서 drift 발생 → cleanup 무한 회전. 이제 점수 컷
                # 적용해서 풀 점수 분포가 사용자 threshold 이상으로 직접 수렴.
                ai_inline_fallback = False
                if len(approved) < 50 and batch_ok > 0 and top_rejects:
                    from database.naver_ad_db import get_ad_account_auto_cleanup as _get_thr_inline
                    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel_inline
                    _thr_cfg_inline = _get_thr_inline(uid, str(customer_id)) or {}
                    _inline_thr = int(_thr_cfg_inline.get("threshold") or 50)
                    _rel_basis = _get_rel_inline(uid, str(customer_id)) or []
                    if not _rel_basis:
                        _rel_basis = [s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2]
                    ai_inline_fallback = True
                    existing = set(approved)
                    # 검색량 상위 중 점수 ≥ threshold 만 보충 — drift 차단
                    extras: List[str] = []
                    for r in top_rejects:
                        kw_ = r["keyword"]
                        if kw_ in existing:
                            continue
                        if _rel_basis and _compute_relevance_score(kw_, _rel_basis) < _inline_thr:
                            continue
                        extras.append(kw_)
                    boost = 2000 - len(approved)
                    approved = list(approved) + extras[:max(0, boost)]
                    inline_ai_approved = len(approved)
                    logger.warning(
                        f"[pool/collect/ai-inline] user={uid} GPT 통과 적음 — "
                        f"검색량 상위 + 점수≥{_inline_thr} 만 +{len(extras[:boost])}개 fallback "
                        f"보충 → 총 {len(approved)}"
                    )

                if approved:
                    mt_map = {
                        r["keyword"]: int(r.get("monthly_total") or 0)
                        for r in top_rejects
                    }
                    inline_items = [
                        {
                            "keyword": k,
                            "monthly_total": mt_map.get(k, 0),
                            "monthly_pc": 0,
                            "monthly_mobile": 0,
                            "comp_idx": None,
                            "source": "ai_inline",
                            "seed": "ai_classified",
                        }
                        for k in approved
                    ]
                    try:
                        inline_ai_added = pool.add_candidates(
                            uid, customer_id, inline_items
                        )
                        added += inline_ai_added  # 헤드룸 카운트 정확화
                    except Exception as e:
                        logger.warning(f"[pool/collect/ai-inline] add 실패: {e}")
                    try:
                        pool.mark_rejects_classified(customer_id, approved, "promoted")
                    except Exception:
                        pass
                if discarded:
                    try:
                        pool.mark_rejects_classified(customer_id, discarded, "discarded")
                    except Exception:
                        pass
                logger.warning(
                    f"[pool/collect/ai-inline] user={uid} cid={customer_id} "
                    f"분류 {len(top_rejects)}개 "
                    f"(batch {batch_ok}OK/{batch_fail}fail{', fallback' if ai_inline_fallback else ''}) → "
                    f"통과 {inline_ai_approved} (자식 추가 {inline_ai_added}) / "
                    f"컷 {inline_ai_discarded} ({ai_ms}ms) "
                    f"rationale={last_rationale[:80]}"
                )
        elif reject_for_ai and not _settings.OPENAI_API_KEY:
            logger.warning("[pool/collect/ai-inline] OPENAI_API_KEY 미설정 — skip")
    except Exception as e:
        logger.warning(f"[pool/collect/ai-inline] 예외: {type(e).__name__}: {e}", exc_info=True)
    # ===========================================================================
    pending_after = (pool.stats(customer_id).get("by_status") or {}).get("pending", 0)
    err_parts = list(api_errors)
    if circuit_aborted:
        err_parts.append(
            "네이버 API circuit breaker OPEN — niche 시드 keywordstool timeout 누적. "
            "다음 cron tick (5분 후) 자동 재시도. broader 시드 추가 권장."
        )
    if rejected > 0:
        err_parts.append(
            f"화이트리스트 reject {rejected}개 "
            f"(도메인미스 {reject_no_domain} / 시드미스 {reject_no_seed_match})"
        )
    if inline_ai_added > 0 or inline_ai_approved > 0 or inline_ai_discarded > 0:
        err_parts.append(
            f"AI inline 자식 +{inline_ai_added} (통과 {inline_ai_approved} / 컷 {inline_ai_discarded})"
        )
    if direct_added > 0:
        err_parts.append(f"reject-direct 자식 +{direct_added} (mt 30~99, GPT 우회)")
    pool.record_run(
        uid, customer_id, "collect",
        "success" if not api_errors else ("partial" if added > 0 else "failed"),
        added=added, skipped=rejected, seeds_count=seeds_processed,
        pending_after=pending_after,
        error_message=" | ".join(err_parts)[:500] if err_parts else None,
        duration_ms=int((_time.monotonic()-t0)*1000),
    )

    # 데드락 감지 — 최근 5회 collect 가 전부 added=0 + reject ≥ 500 이면 alert.
    # 사용자가 며칠 동안 0건인 걸 모르고 지나치는 사고 방지.
    try:
        deadlock = pool.detect_collect_deadlock(customer_id, n_recent=5, min_rejected=500)
        if deadlock.get("is_deadlock"):
            logger.error(
                f"[pool/collect] DEADLOCK user={uid} customer={customer_id} "
                f"— 최근 {deadlock['consecutive_zero_runs']}회 연속 0건 + "
                f"누적 reject {deadlock['total_rejected']}. "
                f"시드/도메인 토큰 점검 필요."
            )
            pool.record_run(
                uid, customer_id, "collect", "alert",
                error_message=(
                    f"[DEADLOCK] {deadlock['consecutive_zero_runs']}회 연속 0건. "
                    f"시드/도메인 점검 필요."
                ),
                duration_ms=0,
            )
    except Exception as e:
        logger.warning(f"[pool/collect] deadlock 감지 실패: {e}")


async def _run_pool_ai_classify(
    uid: int,
    customer_id: int,
    *,
    cooldown_minutes: int = 30,
    candidates_limit: int = 200,
    min_volume: int = 100,
    force: bool = False,
) -> Dict[str, Any]:
    """AI 분류 1회 — reject 풀 → user_seed 자동 promote.

    흐름:
      1. 쿨다운 체크 (force=True 면 무시)
      2. user_seed + 미분류 reject 가져옴
      3. classify_rejects (GPT-4o-mini) 호출
      4. approved → source='user_seed' 로 add_candidates → 시드 합류
      5. reject 풀 status 갱신 + cooldown stamp
      6. record_run

    트리거 조건은 호출자(스케줄러/라우트) 가 결정. 이 함수는 단순 1회 실행.
    """
    from services.ai_seed_suggester import classify_rejects
    from datetime import datetime, timedelta
    import time as _time

    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    # 쿨다운 — 매번 GPT 호출 비용/품질 안정 위해 광고주별 30분 간격 강제
    if not force:
        last = pool.get_classify_cooldown(customer_id)
        if last:
            try:
                last_dt = datetime.fromisoformat(str(last).replace("T", " ").split(".")[0])
                elapsed = datetime.utcnow() - last_dt
                wait_min = cooldown_minutes - int(elapsed.total_seconds() / 60)
                if elapsed < timedelta(minutes=cooldown_minutes):
                    logger.warning(
                        f"[pool/ai-classify] user={uid} cid={customer_id} 쿨다운 잔여 {wait_min}분 — skip"
                    )
                    return {
                        "success": False, "reason": "cooldown",
                        "wait_minutes": max(wait_min, 1),
                    }
            except Exception:
                pass  # 파싱 실패 시 그냥 진행

    user_seeds = pool.list_user_seeds(customer_id)
    if not user_seeds:
        logger.warning(f"[pool/ai-classify] user={uid} cid={customer_id} user_seed 없음 — skip")
        return {"success": False, "reason": "no_user_seed"}

    rejects = pool.list_unclassified_rejects(
        customer_id, limit=candidates_limit, min_volume=min_volume
    )
    if not rejects:
        logger.warning(
            f"[pool/ai-classify] user={uid} cid={customer_id} 미분류 reject 없음 (검색량≥{min_volume}) — skip"
        )
        return {"success": False, "reason": "no_rejects"}

    # strict 모드 — saved_relevance 있으면 classify 가 보수적으로 전환됨 (drift 차단)
    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel
    saved_relevance_local = _get_rel(uid, str(customer_id)) or []

    logger.warning(
        f"[pool/ai-classify] user={uid} cid={customer_id} 시작 — "
        f"seeds={len(user_seeds)} rejects={len(rejects)} (검색량≥{min_volume}) "
        f"relevance={len(saved_relevance_local)} {'STRICT' if len(saved_relevance_local) >= 3 else 'lenient'}"
    )

    result = await classify_rejects(
        user_seeds, rejects, seed_sample_size=50,
        saved_relevance=saved_relevance_local,
    )
    if not result.get("success"):
        msg = result.get("message", "unknown")
        logger.warning(f"[pool/ai-classify] user={uid} 분류 실패: {msg}")
        pool.record_run(
            uid, customer_id, "ai_classify", "failed",
            error_message=msg[:300],
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "ai_failed", "message": msg}

    approved = result.get("approved") or []
    discarded = result.get("discarded") or []

    # approved → user_seed 로 합류 (검색량 그대로 보존, 즉시 다음 collect 게이트 atom 합류)
    promoted = 0
    if approved:
        mt_map = {r["keyword"]: r.get("monthly_total", 0) for r in rejects}
        items = [
            {
                "keyword": k,
                "seed": k,
                "source": "user_seed",
                "monthly_total": mt_map.get(k, 0),
            }
            for k in approved
        ]
        try:
            promoted = pool.add_candidates(uid, customer_id, items)
        except Exception as e:
            logger.warning(f"[pool/ai-classify] promote 실패: {e}")

    try:
        if approved:
            pool.mark_rejects_classified(customer_id, approved, "promoted")
        if discarded:
            pool.mark_rejects_classified(customer_id, discarded, "discarded")
    except Exception as e:
        logger.warning(f"[pool/ai-classify] mark 실패: {e}")

    pool.stamp_classify_cooldown(customer_id)

    duration_ms = int((_time.monotonic() - t0) * 1000)
    logger.warning(
        f"[pool/ai-classify] user={uid} cid={customer_id} 완료 — "
        f"approved={len(approved)} promoted={promoted} discarded={len(discarded)} "
        f"({duration_ms}ms) rationale={(result.get('rationale') or '')[:100]}"
    )
    pool.record_run(
        uid, customer_id, "ai_classify",
        "success" if approved else "no_match",
        added=promoted, skipped=len(discarded),
        seeds_count=len(user_seeds),
        error_message=(result.get("rationale") or "")[:300] or None,
        duration_ms=duration_ms,
    )
    return {
        "success": True,
        "approved": len(approved),
        "promoted": promoted,
        "discarded": len(discarded),
        "rationale": result.get("rationale", ""),
        "model": result.get("model"),
        "duration_ms": duration_ms,
    }


async def _run_pool_ai_cleanup_registered(
    uid: int,
    customer_id: int,
    *,
    dry_run: bool = True,
    batch_size: int = 200,
    max_kws: int = 1000,
    max_delete: int = 2000,
    incremental_minutes: Optional[int] = None,
    sample_seeds: int = 50,
) -> Dict[str, Any]:
    """등록된 KW 를 GPT 가 user_seed 와 비교 분류 → 무관 KW 실제 네이버 DELETE.

    기존 점수 기반 cleanup (`_run_domain_cleanup_for_account`) 의 atoms_2 인플레
    문제 우회 — 시드 1500개일 때 한국어 2-gram 합집합이 한국어 음절 거의 다
    포함해 무관 KW 도 30+ 점이라 ≤30 임계 통과 못함.

    GPT 가 의미적으로 도메인 일치 여부 판정 → 점수 인플레 무관, 정확.

    Args:
        dry_run: True 면 GPT 분류만 + 결과 통계 반환 (실제 DELETE 안 함, 실측용).
        batch_size: GPT 분류 batch 크기 (200 권장)
        max_kws: 한 번에 audit 할 최대 KW 수 (1000 권장, GPT 5 batch)
        incremental_minutes:
            None = 최근 등록순 max_kws 개 audit
            N = 최근 N분 등록 KW 만 (cron 인크리멘탈)
    """
    from services.ai_seed_suggester import classify_rejects
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import get_ad_account_by_customer
    from config import settings
    import time as _time
    import sqlite3 as _sql

    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()
    t0 = _time.monotonic()

    if not settings.OPENAI_API_KEY:
        return {"success": False, "reason": "no_api_key"}

    account = get_ad_account_by_customer(uid, str(customer_id))
    if not account or not account.get("is_connected"):
        return {"success": False, "reason": "no_account"}

    user_seeds = pool.list_user_seeds(customer_id)
    if not user_seeds:
        return {"success": False, "reason": "no_user_seed"}

    # 1) 등록 KW 조회
    with _sql.connect(reg.db_path) as conn:
        if incremental_minutes:
            rows = conn.execute(
                "SELECT keyword, ncc_keyword_id FROM registered_keywords "
                "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL "
                "AND datetime(registered_at) > datetime('now', ?) "
                "ORDER BY registered_at DESC LIMIT ?",
                (customer_id, f"-{int(incremental_minutes)} minutes", int(max_kws)),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT keyword, ncc_keyword_id FROM registered_keywords "
                "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL "
                "ORDER BY id DESC LIMIT ?",
                (customer_id, int(max_kws)),
            ).fetchall()

    if not rows:
        pool.record_run(
            uid, customer_id, "ai_cleanup", "no_new",
            error_message="audit 대상 0",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "no_registered"}

    # 2) batch GPT 분류 — classify_rejects 재활용. saved_relevance → strict 모드 (drift 정리용).
    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel
    saved_relevance_local = _get_rel(uid, str(customer_id)) or []
    candidates = [{"keyword": kw, "monthly_total": 0} for kw, _kid in rows]
    kid_map: Dict[str, str] = {kw: kid for kw, kid in rows if kid}
    all_approved: Set[str] = set()
    all_discarded: Set[str] = set()
    batch_count = 0
    gpt_ms_total = 0

    for i in range(0, len(candidates), batch_size):
        batch = candidates[i:i + batch_size]
        ai_t0 = _time.monotonic()
        try:
            ai = await classify_rejects(
                user_seeds, batch, seed_sample_size=sample_seeds,
                saved_relevance=saved_relevance_local,
            )
        except Exception as e:
            logger.warning(f"[ai-cleanup] batch {i} 예외: {type(e).__name__}: {e}")
            continue
        gpt_ms_total += int((_time.monotonic() - ai_t0) * 1000)
        if not ai.get("success"):
            logger.warning(f"[ai-cleanup] batch {i} GPT 실패: {ai.get('message')}")
            continue
        all_approved.update(ai.get("approved") or [])
        all_discarded.update(ai.get("discarded") or [])
        batch_count += 1

    discarded_list = sorted(all_discarded)
    approved_list = sorted(all_approved)

    result: Dict[str, Any] = {
        "success": True,
        "dry_run": dry_run,
        "user_id": uid,
        "customer_id": customer_id,
        "incremental_minutes": incremental_minutes,
        "total_audited": len(rows),
        "batches": batch_count,
        "approved_count": len(approved_list),
        "discarded_count": len(discarded_list),
        "approved_samples": approved_list[:10],
        "discarded_samples": discarded_list[:20],
        "deleted": 0,
        "delete_failed": 0,
        "gpt_ms_total": gpt_ms_total,
    }

    # 3) dry_run=False — 실제 네이버 API DELETE
    if not dry_run and discarded_list:
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]

        n_del, n_fail = 0, 0
        # 한 round 의 DELETE 상한 — naver rate (0.12s/call) + 사고 방지.
        # 2000 = 약 4분 소요. 한의원 광고주 13만 drift KW 정리 시 약 65 라운드 = 11시간.
        for kw in discarded_list[:max(0, min(max_delete, 5000))]:
            kid = kid_map.get(kw)
            if not kid:
                continue
            try:
                await client.delete_keyword(kid)
                with _sql.connect(reg.db_path) as conn:
                    conn.execute(
                        "DELETE FROM registered_keywords "
                        "WHERE account_customer_id=? AND ncc_keyword_id=?",
                        (customer_id, kid),
                    )
                with _sql.connect(pool.db_path) as conn:
                    conn.execute(
                        "UPDATE naverad_keyword_pool SET status='deleted' "
                        "WHERE account_customer_id=? AND keyword=?",
                        (customer_id, kw),
                    )
                n_del += 1
                await asyncio.sleep(0.12)
            except Exception as e:
                n_fail += 1
                logger.warning(f"[ai-cleanup] DELETE 실패 {kw}({kid}): {e}")
        result["deleted"] = n_del
        result["delete_failed"] = n_fail

    duration_ms = int((_time.monotonic() - t0) * 1000)
    result["duration_ms"] = duration_ms

    pool.record_run(
        uid, customer_id, "ai_cleanup",
        "success" if (dry_run or result["deleted"] > 0) else "no_match",
        added=0, skipped=result["deleted"], seeds_count=len(rows),
        error_message=(
            f"AI cleanup {'dry-run' if dry_run else 'EXEC'} — "
            f"audit {len(rows)} → 통과 {len(approved_list)} / 컷 {len(discarded_list)} → "
            f"DELETE {result['deleted']} (fail {result['delete_failed']})"
        )[:300],
        duration_ms=duration_ms,
    )
    logger.warning(
        f"[ai-cleanup] user={uid} cid={customer_id} {'dry-run' if dry_run else 'EXEC'} "
        f"({duration_ms}ms) — audit {len(rows)} → 통과 {len(approved_list)} / "
        f"컷 {len(discarded_list)} → DELETE {result['deleted']} fail {result['delete_failed']}"
    )
    return result


async def _run_pool_seed_amplify(
    uid: int,
    customer_id: int,
    *,
    seed_sample_size: int = 100,
    target_count: int = 300,
    min_volume: int = 1,
    chunks_cap: int = 200,
) -> Dict[str, Any]:
    """user_seed 자가 amplify — GPT 가 시드 패턴 분석해 새 시드 후보 생성 →
    검색량 검증 → user_seed 합류. 게이트 atom 다양성↑ → collect/autocomplete 발굴↑.

    keywordstool BFS / 자동완성과 다른 channel — LLM 의 학습 지식 기반
    semantic 시드 펼침. niche 도메인에서도 cartesian product 으로 시드 ↑.

    흐름:
      1) user_seed 100개 sample → amplify_seeds (GPT 패턴 펼침, target=300)
      2) 풀 dedup (이미 있는 KW 제외)
      3) keywordstool 검색량 batch (5개씩)
      4) ≥ min_volume 인 것만 source='user_seed', seed=keyword 로 add_candidates
      5) → 다음 collect 부터 게이트 atom + BFS hint 로 사용
    """
    from services.ai_seed_suggester import amplify_seeds
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import get_ad_account_by_customer
    from config import settings
    import time as _time
    import random

    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    if not settings.OPENAI_API_KEY:
        return {"success": False, "reason": "no_api_key"}

    account = get_ad_account_by_customer(uid, str(customer_id))
    if not account or not account.get("is_connected"):
        return {"success": False, "reason": "no_account"}

    user_seeds = pool.list_user_seeds(customer_id)
    if not user_seeds:
        return {"success": False, "reason": "no_user_seed"}

    # 시드 cap 도달 시 amplify 의미 없음 — 500 시드 이상이면 skip (성능 보호)
    if len(user_seeds) >= 5000:
        return {"success": False, "reason": "seeds_capped", "current": len(user_seeds)}

    seed_sample = (
        random.sample(user_seeds, seed_sample_size)
        if len(user_seeds) > seed_sample_size else list(user_seeds)
    )

    logger.warning(
        f"[pool/amplify] user={uid} cid={customer_id} 시작 — "
        f"시드 {len(seed_sample)}/{len(user_seeds)} → amplify target {target_count}"
    )

    # 1) GPT amplify
    am_t0 = _time.monotonic()
    am_result = await amplify_seeds(seed_sample, target_count=target_count)
    am_ms = int((_time.monotonic() - am_t0) * 1000)

    if not am_result.get("success"):
        msg = am_result.get("message", "unknown")
        pool.record_run(
            uid, customer_id, "seed_amplify", "failed",
            seeds_count=len(seed_sample),
            error_message=f"amplify 실패: {msg[:200]}",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "amplify_failed", "message": msg}

    raw_seeds = am_result.get("seeds") or []
    # 원본 시드 제외 (이미 user_seed 로 있음)
    user_seed_set = set(user_seeds)
    new_seeds = [s for s in raw_seeds if isinstance(s, str) and s.strip() and s not in user_seed_set]

    # 2) 풀 dedup (이미 어떤 status 로든 풀에 있는 KW 제외)
    pool_set = pool.list_pool_keyword_set(customer_id)
    fresh_seeds = [s for s in new_seeds if s not in pool_set]

    if not fresh_seeds:
        pool.record_run(
            uid, customer_id, "seed_amplify", "no_new",
            seeds_count=len(seed_sample),
            error_message=f"amplify {len(raw_seeds)} → fresh 0 (전부 dedup)",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "all_known"}

    # 2-b) 도메인 게이트 — saved_relevance 있는 계정만. drift 증폭기 차단.
    # 한의원 계정에 차 KW (2024쏘나타) 가 amplify cartesian 으로 폭발해 user_seed 합류 →
    # 다음 collect 라운드 anchor → 자식 KW 도메인 게이트 무력화 → 100k drift 사고.
    # saved_relevance 비어있으면 (cold start) skip — 시드 0 광고주 진입 봉쇄 방지.
    # 컷 점수 — 사용자 auto_cleanup_threshold (광고주별 30~75) 와 동기화. 옛 hardcoded 30
    # 으로는 풀에 31~49 점수 KW drift → 다음 cleanup tick 에서 정리 → 무한 회전.
    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel
    from database.naver_ad_db import get_ad_account_auto_cleanup as _get_thr
    saved_relevance = _get_rel(uid, str(customer_id)) or []
    _thr_cfg = _get_thr(uid, str(customer_id)) or {}
    _domain_gate_thr = int(_thr_cfg.get("threshold") or 50)
    domain_filtered_count = 0
    if saved_relevance and len([s for s in saved_relevance if s and len(s) >= 2]) >= 3:
        before = len(fresh_seeds)
        kept = [s for s in fresh_seeds if _compute_relevance_score(s, saved_relevance) >= _domain_gate_thr]
        domain_filtered_count = before - len(kept)
        fresh_seeds = kept
        if not fresh_seeds:
            pool.record_run(
                uid, customer_id, "seed_amplify", "no_new",
                seeds_count=len(seed_sample),
                error_message=(
                    f"amplify {len(raw_seeds)} → fresh {before} → 도메인필터 0 "
                    f"(relevance={len(saved_relevance)} 와 매칭 0)"
                )[:300],
                duration_ms=int((_time.monotonic() - t0) * 1000),
            )
            return {"success": False, "reason": "domain_filter_all_out"}

    logger.warning(
        f"[pool/amplify] amplify ({am_ms}ms) — raw {len(raw_seeds)} → "
        f"fresh {len(fresh_seeds)}"
        + (f" (도메인필터 컷 {domain_filtered_count})" if domain_filtered_count else "")
    )

    # 3) keywordstool 검색량 batch
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    vol_t0 = _time.monotonic()
    vol_map: Dict[str, dict] = {}
    CHUNK = 5
    chunks = [fresh_seeds[i:i + CHUNK] for i in range(0, len(fresh_seeds), CHUNK)][:chunks_cap]
    for chunk in chunks:
        try:
            r = await client.get_keywords_volume_batch(chunk)
            vol_map.update(r)
        except Exception as e:
            logger.debug(f"[pool/amplify] volume batch 실패: {e}")
        await asyncio.sleep(0.1)
    vol_ms = int((_time.monotonic() - vol_t0) * 1000)

    # 4) ≥ min_volume 만 user_seed 합류
    qualified_items = []
    for s in fresh_seeds:
        v = vol_map.get(s) or vol_map.get(s.replace(" ", ""))
        if not v:
            continue
        mt = int(v.get("monthly_total") or 0)
        if mt < min_volume:
            continue
        qualified_items.append({
            "keyword": s,
            "monthly_total": mt,
            "monthly_pc": int(v.get("monthly_pc") or 0),
            "monthly_mobile": int(v.get("monthly_mobile") or 0),
            "comp_idx": v.get("comp_idx"),
            "source": "user_seed",
            "seed": s,  # 시드 자기 자신 = 시드 row
        })

    promoted = 0
    if qualified_items:
        try:
            promoted = pool.add_candidates(uid, customer_id, qualified_items)
        except Exception as e:
            logger.warning(f"[pool/amplify] add 실패: {e}")

    duration_ms = int((_time.monotonic() - t0) * 1000)
    logger.warning(
        f"[pool/amplify] user={uid} cid={customer_id} 완료 ({duration_ms}ms) — "
        f"amplify {len(raw_seeds)} → fresh {len(fresh_seeds)} → "
        f"검색량≥{min_volume} {len(qualified_items)} → user_seed +{promoted} "
        f"(GPT {am_ms}ms, vol {vol_ms}ms, pattern={am_result.get('detected_pattern', '')[:60]})"
    )
    pool.record_run(
        uid, customer_id, "seed_amplify",
        "success" if promoted > 0 else "no_match",
        added=promoted,
        seeds_count=len(seed_sample),
        error_message=(
            f"amplify {len(raw_seeds)} → fresh {len(fresh_seeds)} → "
            f"검색량≥{min_volume} {len(qualified_items)} → user_seed +{promoted}"
        )[:300],
        duration_ms=duration_ms,
    )
    return {
        "success": True,
        "amplify_total": len(raw_seeds),
        "fresh": len(fresh_seeds),
        "volume_qualified": len(qualified_items),
        "promoted": promoted,
        "duration_ms": duration_ms,
    }


async def _run_pool_autocomplete_mining(
    uid: int,
    customer_id: int,
    *,
    # ⚠️ 총 질의 수를 원래 수준(~200)으로 유지해야 한다.
    # 2026-07-29 실측: 60시드×15변형=900질의로 올렸더니 네이버/Bing 이 Fly IP 를
    # 스로틀링해 처리율이 초당 0.06~0.63 질의로 떨어졌다(로컬은 33.6). 그 결과
    # 100초 예산 안에 900개 중 45개만 처리, **수확이 109개 → 1개로 회귀**했다.
    # 12시드 × 15변형 = 180질의 ≈ 원래 200질의. 자모 이득은 유지하면서 스로틀은 피한다.
    seed_sample_size: int = 12,
    per_seed: int = 30,
    min_volume: int = 1,
    # 후보가 늘어 250개 검증(=50청크)은 대부분을 버린다. 150청크(750KW)로 확대.
    # sleep 0.3 유지 → 45초, keywordstool 429 안전구간.
    chunks_cap: int = 150,
    # Bing 은 Fly IP 스로틀로 프로덕션 수확이 거의 없다 → 기본 OFF (위 주석 참조)
    use_bing: bool = False,
) -> Dict[str, Any]:
    """naver 검색 자동완성으로 시드 인접 KW 발굴 → GPT 분류 → 자식 풀 직접 추가.

    keywordstool BFS 와 별도 발굴 채널. 시드별 자동완성 ~10 KW = 1500~2000 후보.
    검색량 검증 + GPT 도메인 분류 통과 KW 만 자식 풀 진입.

    흐름:
      1) user_seed N개 random sample → naver 자동완성 batch (concurrency 10)
      2) 풀/reject 분류 KW dedup
      3) keywordstool 검색량 batch (5개씩, 250 chunks cap)
      4) 검색량 ≥ min_volume → GPT 분류 (검색량 상위 200개)
      5) 통과 KW source='ai_autocomplete' 로 add_candidates → 자식 풀 즉시 진입
      6) 분류 결과 reject 풀에 INSERT+mark → 다음 cron 재호출 차단
    """
    from services.naver_autocomplete import collect_autocomplete_expanded, collect_bing_expanded
    from services.naver_ad_service import NaverAdApiClient
    from services.ai_seed_suggester import classify_rejects
    from database.naver_ad_db import get_ad_account_by_customer
    from config import settings
    import time as _time
    import random

    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    if not settings.OPENAI_API_KEY:
        logger.warning(f"[pool/autocomplete] OPENAI_API_KEY 미설정 — skip")
        return {"success": False, "reason": "no_api_key"}

    account = get_ad_account_by_customer(uid, str(customer_id))
    if not account or not account.get("is_connected"):
        return {"success": False, "reason": "no_account"}

    user_seeds = pool.list_user_seeds(customer_id)
    if not user_seeds:
        return {"success": False, "reason": "no_user_seed"}

    # saturation 가드 — 풀 사용량 ≥ 98% 일 때 skip.
    # domain_cleanup(30분 주기) + cap_self_heal 가 무관 KW 빼면 풀이 98% 미만에서 평형 →
    # autocomplete 가 95~98% 구간에서도 새 KW 발굴 지속 → 물갈이 사이클 자연 가동.
    # 98% 도달 시엔 skip — niche 도메인은 mt=0 long-tail 만 토하므로 API 비용 낭비.
    reg_db = get_registered_keywords_db()
    pool_pending = (pool.stats(customer_id).get("by_status") or {}).get("pending", 0)
    active_reg = int((reg_db.stats(customer_id) or {}).get("active") or 0)
    used = active_reg + pool_pending
    # 98% → 99.5% 로 완화. 옛 가드는 cleanup ≈ register 평형 시 영구 skip 사고 발생.
    # autocomplete 는 발굴 자체에 비용 작음 (Naver autocomplete API 만 호출, GPT 불사용).
    # 발굴된 후보가 풀에 들어갈 슬롯 없으면 pool.add_keywords 가 알아서 skip 처리.
    if used >= 99_500:
        pool.record_run(
            uid, customer_id, "autocomplete", "no_new",
            seeds_count=0,
            error_message=(
                f"saturation guard — used {used}/100k ≥99.5% → autocomplete skip "
                f"(cleanup 후 재진입)"
            )[:300],
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        logger.warning(
            f"[pool/autocomplete] user={uid} cid={customer_id} saturation guard "
            f"({used}/100k ≥99.5%) — skip"
        )
        return {"success": False, "reason": "saturation_guard", "used": used}

    if len(user_seeds) > seed_sample_size:
        seed_sample = random.sample(user_seeds, seed_sample_size)
    else:
        seed_sample = list(user_seeds)

    logger.warning(
        f"[pool/autocomplete] user={uid} cid={customer_id} 시작 — "
        f"시드 {len(seed_sample)}/{len(user_seeds)} 자동완성 mining "
        f"(used {used}/100k)"
    )

    # 1) 자동완성 batch 수집 — 자모 접두사까지 확장
    # 시드를 그대로만 물으면 시드당 7~10개에서 끝난다. 초성 자모를 붙여 물으면
    # 네이버가 전혀 다른 결과 집합을 준다(2026-07-29 실측: "리프팅" 7개 → 140개, 20배).
    # "완전포화" 판정이 이 표면 하나를 안 훑어서 나온 오판인 경우가 많다.
    ac_t0 = _time.monotonic()
    try:
        # adaptive=False: tier2(시드당 42변형)를 끈다. 켜면 질의가 다시 500+ 로
        # 불어나 스로틀을 유발한다. tier1(자모 14변형)이 질의당 신규 KW 의 주력이다.
        ac_result = await collect_autocomplete_expanded(
            seed_sample, per_seed=max(per_seed, 30), concurrency=4, timeout=5.0,
            adaptive=False, budget_seconds=200.0,
        )
    except Exception as e:
        logger.error(f"[pool/autocomplete] 자동완성 호출 실패: {e}", exc_info=True)
        pool.record_run(
            uid, customer_id, "autocomplete", "failed",
            seeds_count=len(seed_sample),
            error_message=f"자동완성 호출 실패: {type(e).__name__}",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "autocomplete_failed"}
    ac_ms = int((_time.monotonic() - ac_t0) * 1000)

    all_kws: Set[str] = set()
    for kws in ac_result.values():
        for k in kws:
            all_kws.add(k)

    # Bing 서제스트 — 로컬에서는 네이버 위에 +40%를 얹는다(1,516 → 2,128).
    # 그러나 프로덕션(Fly IP)에서는 처리율이 초당 0.18~0.63 질의로 스로틀돼
    # 60초 예산에 900개 중 11~38개만 처리하고 KW 0~36개를 얻었다. 그 60초는
    # 네이버 질의에 쓰는 편이 낫다 → 기본 OFF. 스로틀이 풀리면 다시 켠다.
    naver_only = len(all_kws)
    if use_bing:
        try:
            all_kws |= await collect_bing_expanded(seed_sample, concurrency=3)
        except Exception as e:
            logger.warning(f"[pool/autocomplete] Bing 채널 실패 — 네이버만 사용: {type(e).__name__}")
        logger.warning(
            f"[pool/autocomplete] 표면 합집합 — 네이버 {naver_only} → +Bing {len(all_kws)}"
        )
    else:
        logger.warning(f"[pool/autocomplete] 네이버 표면 수확 {naver_only}개 (Bing OFF)")

    if not all_kws:
        pool.record_run(
            uid, customer_id, "autocomplete", "no_new",
            seeds_count=len(seed_sample),
            error_message="자동완성 결과 0개",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "no_autocomplete"}

    logger.warning(
        f"[pool/autocomplete] 자동완성 ({ac_ms}ms) — KW {len(all_kws)}개"
    )

    # 2) 분류 이력 dedup (풀 중복은 add_candidates 가 INSERT OR IGNORE 처리)
    classified_set = set(pool.list_classified_reject_keywords(customer_id))
    fresh_kws: List[str] = [kw for kw in all_kws if kw not in classified_set]

    if not fresh_kws:
        pool.record_run(
            uid, customer_id, "autocomplete", "no_new",
            seeds_count=len(seed_sample),
            error_message=f"자동완성 {len(all_kws)}개 모두 dedup",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "all_known"}

    # 3) keywordstool 검색량 batch — 5개씩, 250 chunks (1250 KW) cap
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    vol_t0 = _time.monotonic()
    vol_map: Dict[str, dict] = {}
    CHUNK = 5
    # chunks_cap 50 × CHUNK 5 = 최대 250 KW 검증. sleep 0.3 — keywordstool 429 rate 회피.
    # 실측 (cid 1858907): chunks 250 × sleep 0.1 = 90~230초 + 429 retry → 4분 소요 + 결과 0.
    # chunks 50 × sleep 0.3 = 15초, 429 회피 + 정상 결과.
    # 후보가 cap 을 넘으면 앞에서부터 자른다. 순서를 고정하면 뒤쪽 후보는 매 라운드
    # 똑같이 잘려 영영 검증되지 않는다(검색량 미달 KW 는 기록되지 않아 다음 라운드에도
    # 다시 앞자리를 차지한다). 섞어서 라운드마다 다른 표본이 뽑히게 한다.
    if len(fresh_kws) > chunks_cap * CHUNK:
        random.shuffle(fresh_kws)
    chunks = [fresh_kws[i:i + CHUNK] for i in range(0, len(fresh_kws), CHUNK)][:chunks_cap]
    for chunk in chunks:
        try:
            r = await client.get_keywords_volume_batch(chunk)
            vol_map.update(r)
        except Exception as e:
            logger.debug(f"[pool/autocomplete] volume batch 실패 {chunk[:1]}: {e}")
        await asyncio.sleep(0.3)
    vol_ms = int((_time.monotonic() - vol_t0) * 1000)

    qualified: List[Dict] = []
    for kw in fresh_kws:
        v = vol_map.get(kw) or vol_map.get(kw.replace(" ", ""))
        if not v:
            continue
        mt = int(v.get("monthly_total") or 0)
        if mt < min_volume:
            continue
        qualified.append({
            "keyword": kw,
            "monthly_total": mt,
            "monthly_pc": int(v.get("monthly_pc") or 0),
            "monthly_mobile": int(v.get("monthly_mobile") or 0),
            "comp_idx": v.get("comp_idx"),
        })

    logger.warning(
        f"[pool/autocomplete] 검색량 ({vol_ms}ms) — "
        f"{len(fresh_kws)} → 검색량≥{min_volume} {len(qualified)}"
    )

    if not qualified:
        # mt=0 zerovol fallback 제거 (2026-05-12) — 사용자 명시 거부:
        # "검색량 있는 키워드로 연관된 키워드 싹다 잡아야". mt=0 KW 는 광고비/노출 0 이라
        # 등록 가치 없음. 또한 claim_pending min_volume=1 필터에 영구 걸려 register 워커가
        # 처리 못 함 → pending 풀에 dead row 누적 → register cron 이 "등록가능 0" 으로 정지.
        # niche 시드라 keywordstool mt=0 만 나오면 그 시드는 자식 발굴 못 함 — 시드 자체
        # 부적합. 이 라운드는 no_new 로 끝내고 registered-as-seed (collect 측) 가 발굴 담당.
        duration_ms = int((_time.monotonic() - t0) * 1000)
        logger.warning(
            f"[pool/autocomplete] user={uid} cid={customer_id} zero-vol skip — "
            f"자동완성 {len(all_kws)} → 검색량≥{min_volume} 통과 0 → mt=0 fallback 차단 "
            f"({duration_ms}ms)"
        )
        pool.record_run(
            uid, customer_id, "autocomplete", "no_new",
            added=0, seeds_count=len(seed_sample),
            error_message=(
                f"자동완성 {len(all_kws)} → 검색량 통과 0 → mt=0 fallback 차단 (정책)"
            )[:300],
            duration_ms=duration_ms,
        )
        return {"success": False, "reason": "zero_vol_skipped", "promoted": 0}

    # 4) GPT 분류 — 검색량 상위 1000개 (200 batch × 5 병렬, AI-first 빠른 채움)
    qualified.sort(key=lambda x: -x["monthly_total"])
    classify_input = qualified[:1000]
    ai_t0 = _time.monotonic()
    BATCH = 200
    _sem = asyncio.Semaphore(3)

    # saved_relevance → strict 모드 (autocomplete 도 drift 차단 게이트)
    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel
    autocomplete_relevance = _get_rel(uid, str(customer_id)) or []

    async def _classify_one(batch: List[Dict]) -> Dict[str, Any]:
        async with _sem:
            return await classify_rejects(
                user_seeds, batch, seed_sample_size=50,
                saved_relevance=autocomplete_relevance,
            )

    batches = [classify_input[i:i + BATCH] for i in range(0, len(classify_input), BATCH)]
    batch_results = await asyncio.gather(
        *[_classify_one(b) for b in batches], return_exceptions=True,
    )
    ai_ms = int((_time.monotonic() - ai_t0) * 1000)

    approved: List[str] = []
    discarded: List[str] = []
    batch_ok = 0
    batch_fail = 0
    for r in batch_results:
        if isinstance(r, Exception) or not (isinstance(r, dict) and r.get("success")):
            batch_fail += 1
            continue
        batch_ok += 1
        approved.extend(r.get("approved") or [])
        discarded.extend(r.get("discarded") or [])
    approved = list(dict.fromkeys(approved))
    discarded = list(dict.fromkeys(discarded))

    if batch_ok == 0:
        msg = "all batches failed"
        pool.record_run(
            uid, customer_id, "autocomplete", "failed",
            seeds_count=len(seed_sample),
            error_message=f"GPT 분류 실패: {msg}",
            duration_ms=int((_time.monotonic() - t0) * 1000),
        )
        return {"success": False, "reason": "ai_failed", "message": msg}

    # GPT 통과 0 fallback — 검색량 상위 200개를 풀 직접 합류 (drift 감수).
    # niche 시드 (의료/희귀) 에서 GPT 가 모두 컷 판정해도 풀이 마르지 않게 보장.
    # 무관 KW 가 풀에 들어가도 네이버 검수 → 노출제한 → inspect cron 자동 삭제로 자정.
    # cap 30 → 200: 시간당 autocomplete 12회 × +200 = +2400 풀 합류.
    # 도메인 점수 컷 — 사용자 auto_cleanup_threshold (default 50) 이상만 통과.
    # autocomplete GPT 가 도메인 분류했어도 점수 ≥ thr 보장 안 됨 → 직접 컷.
    from database.naver_ad_db import get_ad_account_auto_cleanup as _get_thr_ac
    from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel_ac
    _thr_cfg_ac = _get_thr_ac(uid, str(customer_id)) or {}
    _ac_thr = int(_thr_cfg_ac.get("threshold") or 50)
    _ac_rel = _get_rel_ac(uid, str(customer_id)) or []
    if not _ac_rel:
        _ac_rel = [s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2]

    ai_fallback = False
    if not approved and classify_input:
        ai_fallback = True
        # 검색량 상위 fallback 도 점수 ≥ thr 만 통과
        approved = [
            q["keyword"] for q in classify_input[:500]
            if not _ac_rel or _compute_relevance_score(q["keyword"], _ac_rel) >= _ac_thr
        ][:200]
        logger.warning(
            f"[pool/autocomplete] user={uid} GPT 통과 0 — 검색량 상위 + 점수≥{_ac_thr} "
            f"만 fallback {len(approved)}개 합류"
        )

    # 5) 통과 KW → 자식 풀 직접 추가 (점수 컷 후)
    promoted = 0
    if approved:
        approved_set = set(approved)
        # GPT 통과 + 점수 컷 — drift 차단
        items = [
            {
                "keyword": q["keyword"],
                "monthly_total": q["monthly_total"],
                "monthly_pc": q["monthly_pc"],
                "monthly_mobile": q["monthly_mobile"],
                "comp_idx": q.get("comp_idx"),
                "source": "ai_autocomplete",
                "seed": "ai_autocomplete",
            }
            for q in classify_input
            if q["keyword"] in approved_set
            and (not _ac_rel or _compute_relevance_score(q["keyword"], _ac_rel) >= _ac_thr)
        ]
        try:
            promoted = pool.add_candidates(uid, customer_id, items)
        except Exception as e:
            logger.warning(f"[pool/autocomplete] add 실패: {e}")

    # 6) 분류 결과를 reject 풀에 INSERT + mark → 다음 cron classified_set 에 잡힘
    try:
        all_classified_items = [
            {"keyword": q["keyword"], "monthly_total": q["monthly_total"]}
            for q in classify_input
        ]
        pool.add_rejects(customer_id, all_classified_items)
        if approved:
            pool.mark_rejects_classified(customer_id, approved, "promoted")
        if discarded:
            pool.mark_rejects_classified(customer_id, discarded, "discarded")
    except Exception as e:
        logger.debug(f"[pool/autocomplete] reject mark: {e}")

    duration_ms = int((_time.monotonic() - t0) * 1000)
    logger.warning(
        f"[pool/autocomplete] user={uid} cid={customer_id} 완료 ({duration_ms}ms) — "
        f"자동완성 {len(all_kws)} → 검증 {len(fresh_kws)} → "
        f"검색량≥{min_volume} {len(qualified)} → 분류 {len(classify_input)} "
        f"(batch {batch_ok}OK/{batch_fail}fail{', fallback' if ai_fallback else ''}) → "
        f"통과 {len(approved)} (자식 +{promoted}) / 컷 {len(discarded)} "
        f"(GPT {ai_ms}ms)"
    )
    pool.record_run(
        uid, customer_id, "autocomplete",
        "success" if promoted > 0 else "no_match",
        added=promoted, skipped=len(discarded),
        seeds_count=len(seed_sample),
        error_message=(
            f"자동완성 {len(all_kws)} → 검색량≥{min_volume} {len(qualified)} → "
            f"GPT 통과 {len(approved)}{' (fallback)' if ai_fallback else ''} "
            f"(자식 +{promoted}) / 컷 {len(discarded)}"
        )[:300],
        duration_ms=duration_ms,
    )
    return {
        "success": True,
        "autocomplete_kws": len(all_kws),
        "volume_qualified": len(qualified),
        "approved": len(approved),
        "promoted": promoted,
        "discarded": len(discarded),
        "duration_ms": duration_ms,
    }


# 의료/요양 기관·의료인 토큰 — category_split 모드에서 키워드 분류용.
_MEDICAL_TOKENS = (
    "병원", "약국", "한의원", "한방병원", "요양원", "요양병원", "동물병원", "산후조리원",
    "재활병원", "정신병원", "치과", "검진센터", "노인요양", "요양시설", "의원", "의료기관",
    "메디컬", "의료", "의사", "약사", "한의사", "수의사", "간호사", "전공의", "개원의",
    "봉직의", "페이닥터", "전문의", "개원", "원장",
)


def _classify_medical(keyword: str) -> bool:
    """키워드가 의료·요양 도메인이면 True (의료대출 캠페인), 아니면 False (비의료대출)."""
    kw = (keyword or "").replace(" ", "")
    return any(t in kw for t in _MEDICAL_TOKENS)


# ── 피부과 테마 분류 (한글 캠페인 등록용) ──────────────────────────────
# 등록 시 키워드를 테마별 한글 캠페인 '[피부]<라벨>'으로 라우팅. customer 스코프됨.
_SKIN_CAT_CUSTOMERS = {4422132}  # 리베리의원 — 테마 등록 + 화장품 허용 적용 대상
# 화장품 키워드 허용 — register 하드게이트 _NEG_TOKENS 에서 이 토큰들만 (스코프된 계정 한정) 컷 해제.
_COSMETIC_ALLOW_TOKENS = {"화장품", "향수", "립스틱", "컨실러", "쿠션", "파운데이션", "비비크림", "마스카라"}
# (key, 한글라벨, 매칭토큰) — 우선순위 순. 위에서부터 첫 매칭 카테고리로 분류.
_SKIN_CATEGORIES = [
    ("질환", "피부질환", (
        "아토피", "건선", "습진", "한포진", "두드러기", "백반증", "대상포진", "포진", "헤르페스",
        "무좀", "백선", "완선", "어루러기", "진균", "조갑", "사마귀", "농가진", "모낭염", "봉와직염",
        "한선염", "다한증", "액취증", "비립종", "쥐젖", "한관종", "켈로이드", "모공각화", "어린선",
        "소양증", "주사피부염", "딸기코", "두피염", "비듬", "탈모", "피부염", "피부질환", "피부병",
        "발진", "티눈", "지방종", "결절종", "표피낭종",
    )),
    ("주사", "주사·수액·부스터", (
        "주사", "수액", "리쥬란", "쥬베룩", "스킨부스터", "물광", "연어", "백옥", "글루타치온", "태반",
        "비타민", "엑소좀", "nctf", "필러", "윤곽주사", "스컬트라", "콜라겐주사", "부스터", "앰플",
    )),
    ("리프팅", "리프팅·탄력", (
        "울쎄라", "써마지", "슈링크", "인모드", "올리지오", "볼뉴머", "덴서티", "소프웨이브", "온다",
        "리프테라", "튠", "리프팅", "탄력", "처짐", "거상", "실리프팅", "민트실", "고주파", "초음파",
        "hifu", "이중턱", "브이라인",
    )),
    ("색소", "색소·기미·미백", (
        "기미", "잡티", "주근깨", "검버섯", "색소", "미백", "토닝", "피코", "점빼기", "점제거",
        "멜라닌", "흑자", "ipl", "잡티제거",
    )),
    ("여드름", "여드름·트러블", (
        "여드름", "트러블", "피지", "화농", "좁쌀", "압출", "아그네스", "pdt", "여드름흉터", "여드름자국",
    )),
    ("모공흉터", "모공·흉터", (
        "모공", "흉터", "프락셀", "포텐자", "co2", "크로스", "서브시전", "패인", "피부결",
    )),
    ("주름", "주름·안티에이징", (
        "주름", "팔자", "눈가", "이마", "미간", "보톡스", "안티에이징", "동안",
    )),
    ("화장품", "화장품·성분", (
        "화장품", "세럼", "에센스", "크림", "로션", "토너", "클렌징", "클렌저", "선크림", "마스크팩",
        "더마코스메틱", "코스메슈티컬", "스킨케어", "히알루론산", "세라마이드", "나이아신아마이드",
        "레티놀", "판테놀", "마데카", "시카", "펩타이드", "성분",
    )),
]


def _classify_skin_category(keyword: str):
    """키워드 → (key, 한글라벨). 매칭 없으면 ('기타', '기타피부')."""
    kw = (keyword or "").replace(" ", "")
    for key, label, toks in _SKIN_CATEGORIES:
        if any(t in kw for t in toks):
            return key, label
    return "기타", "기타피부"


# ── 두비전(DOVISION) 테마 분류 — 창의융합 뇌교육/아동 사고력 교육 프랜차이즈 ──────
# 등록 시 키워드를 테마별 한글 캠페인 '[두비전]<라벨>'으로 라우팅. customer 스코프됨.
# (key, 한글라벨, 매칭토큰) — 우선순위 순. 위에서부터 첫 매칭 카테고리로 분류.
# 가맹·창업 의도를 최상단에 둬 "유아교육창업" 류가 주제 캠페인 대신 가맹으로 모이게 함.
_DOVISION_CAT_CUSTOMERS = {4403292}  # 두비전 — 테마 등록 + 교육토큰 허용 적용 대상
# 교육 도메인 키워드 허용 — register 하드게이트 _NEG_TOKENS 에서 이 토큰들만 (스코프 계정 한정)
# 컷 해제. 교육업이라 학원/과외/강의/인강 이 핵심 도메인인데 기본 게이트는 이를 상업 컷으로 막음.
_DOVISION_ALLOW_TOKENS = {"학원", "과외", "강의", "인강"}
# 키네스(441986) — 무관도메인 강제등록 (2026-07-10 사용자 결정: 84k 클린천장 넘어 확장).
# register 하드게이트에서 학부모/아동 도메인(육아·교육·영양제·아동가구)을 막던 상업 neg-token 해제
# + relevance 점수컷(≥30) 우회 → seed_explode 앵커(required_tokens 210)만 통과하면 등록.
# 자동차/코인/보험/여행/가전 등 진짜 잡동사니 neg-token 은 그대로 유지(순수쓰레기 차단).
# ⚠️ 오염 감수. 노출 전 캠페인 분리 필수. 되돌리려면 이 계정을 set 에서 제거 후 재배포.
_KINESS_FORCE_CUSTOMERS = {441986}
_KINESS_ALLOW_TOKENS = {"학원", "과외", "강의", "인강", "영양제", "유산균", "오메가3", "홍삼",
                        "젤리", "비타민제", "프로틴", "콜라겐젤리", "책상", "의자", "가구",
                        "침대", "매트", "매트리스", "선반", "침구", "이불", "베개", "소파",
                        "수납", "옷장"}
# 계층 분류: (mid_key, 대분류, 중분류, [(소분류, (토큰...)), ...]). 위에서부터 첫 매칭.
# 캠페인명 = '[두비전] 대분류 - 중분류', 광고그룹명 = 소분류. 창업(B2B)을 먼저 둬
# '유아교육창업' 류가 교육 주제 대신 창업으로 모이게 함. 교육 토큰은 창업 미매칭분만 도달.
_DOVISION_TAXONOMY = [
    # ── 대분류 B. 창업·수익 (B2B, 먼저 매칭) ──
    ("biz_edu", "창업·수익", "교육창업", [
        ("공부방·교습소창업", ("공부방창업", "공부방차리기", "공부방프랜차이즈", "공부방운영",
                          "공부방부업", "교습소창업", "교습소차리기", "1인교습소")),
        ("학원창업", ("학원창업", "학원차리기", "보습학원창업", "학원인수", "소형학원창업", "공부방인수")),
        ("교육프랜차이즈", ("교육프랜차이즈", "교육창업", "교육사업", "교육가맹", "아동교육창업",
                       "유아교육창업", "초등교육창업", "사고력수학창업", "뇌교육창업", "코딩학원창업",
                       "독서논술창업", "영어공부방창업", "수학공부방창업", "방문학습지창업", "홈스쿨창업")),
        ("교사·원장모집", ("교사모집", "원장모집", "선생님모집", "강사모집", "방문교사", "학습지교사", "지사장")),
    ]),
    ("biz_online", "창업·수익", "무인·온라인", [
        ("무인매장", ("무인",)),
        ("온라인·스마트스토어", ("스마트스토어", "온라인창업", "온라인판매", "쇼핑몰창업", "위탁판매",
                          "공동구매", "체험단", "전자상거래")),
        ("무점포", ("무점포",)),
    ]),
    ("biz_side", "창업·수익", "소자본·부업", [
        ("부업·투잡", ("부업", "투잡", "부수입", "재택근무", "재택부업", "재택알바", "부업사이트", "부업거리")),
        ("소자본·1인창업", ("소자본", "1인창업", "소액창업")),
        ("전업맘·주부창업", ("전업맘", "주부창업", "여성창업", "경력단절")),
        ("돈버는아이템", ("돈버는", "돈되는", "부업추천")),
    ]),
    ("biz_industry", "창업·수익", "업종창업", [
        ("카페·디저트", ("카페", "커피", "베이커리", "디저트", "탕후루", "붕어빵", "아이스크림")),
        ("외식", ("치킨", "분식", "떡볶이", "피자", "햄버거", "샐러드", "밀키트", "반찬", "도시락",
                "마라탕", "핫도그", "곱창", "국밥", "김밥", "샌드위치", "포장마차", "술집", "주점")),
        ("서비스업", ("세탁", "빨래방", "애견", "헬스장", "필라테스", "요가", "pt샵", "공유주방",
                   "사진관", "네컷", "노래방", "청소", "심부름", "문구", "다이소", "인형뽑기")),
    ]),
    ("biz_franchise", "창업·수익", "프랜차이즈·창업일반", [
        ("프랜차이즈", ("프랜차이즈", "가맹")),
        ("창업일반", ("창업", "개업", "사업자", "자영업", "소상공인", "예비창업자")),
    ]),
    # ── 대분류 A. 교육 (B2C, 창업 미매칭분) ──
    ("edu_infant", "교육", "유아", [
        ("유아수학·한글", ("유아수학", "유아한글", "한글떼기", "유아학습")),
        ("오감·놀이", ("오감", "놀이교구", "놀이학교")),
        ("교구·가베", ("가베", "프뢰벨", "몬테소리", "교구")),
        ("누리·프리스쿨", ("누리", "프리스쿨", "유치원", "어린이집", "영유아", "유아교육", "유아", "전집")),
    ]),
    ("edu_math", "교육", "수학", [
        ("사고력수학", ("사고력수학", "창의수학", "교구수학")),
        ("연산", ("연산", "구구단", "주산", "암산", "셈")),
        ("도형", ("도형", "기하")),
        ("심화·경시", ("심화수학", "경시", "올림피아드", "미적분", "확률과통계")),
        ("수학일반", ("수학",)),
    ]),
    ("edu_korean", "교육", "국어·독서논술", [
        ("독서", ("독서", "책읽기")),
        ("논술·글쓰기", ("논술", "글쓰기", "작문")),
        ("독해·문해력", ("독해", "문해력")),
        ("어휘·한자", ("어휘", "한자")),
        ("토론·발표", ("토론", "스피치", "발표")),
        ("국어일반", ("국어", "비문학", "문학")),
    ]),
    ("edu_english", "교육", "영어", [
        ("파닉스", ("파닉스",)),
        ("영어회화", ("영어회화", "화상영어", "전화영어", "회화")),
        ("영문법·어휘", ("영문법", "영어단어", "리스닝")),
        ("영어독서", ("영어독서", "영어원서", "영어도서")),
        ("영어유치원", ("영어유치원", "영어유아")),
        ("영어일반", ("영어",)),
    ]),
    ("edu_science", "교육", "과학·사회", [
        ("과학실험", ("과학", "실험")),
        ("한국사·역사", ("한국사", "역사", "세계사")),
        ("사회·경제", ("사회", "지리", "경제")),
    ]),
    ("edu_coding", "교육", "코딩·SW", [
        ("코딩", ("코딩", "소프트웨어", "sw교육")),
        ("로봇코딩", ("로봇",)),
        ("스크래치·엔트리", ("스크래치", "엔트리")),
        ("파이썬·앱", ("파이썬", "앱개발", "게임개발", "웹개발")),
        ("AI·메이커", ("ai교육", "인공지능", "메이커", "드론코딩", "아두이노")),
    ]),
    ("edu_brain", "교육", "두뇌·사고력", [
        ("뇌교육", ("뇌교육", "두뇌", "전두엽", "브레인")),
        ("집중력·기억력", ("집중력", "기억력", "워킹메모리", "메타인지")),
        ("창의력", ("창의력", "창의융합", "창의")),
        ("영재교육", ("영재",)),
        ("융합·STEAM", ("steam", "스팀", "융합", "논리", "문제해결", "추론", "사고력")),
    ]),
    ("edu_art", "교육", "예체능", [
        ("미술", ("미술", "그림", "드로잉")),
        ("음악", ("피아노", "바이올린", "첼로", "드럼", "보컬", "성악", "실용음악", "작곡")),
        ("체육", ("태권도", "줄넘기", "수영", "축구", "농구", "체육")),
        ("무용·댄스", ("발레", "무용", "방송댄스", "댄스")),
        ("웅변·연기", ("웅변", "연기", "뮤지컬")),
    ]),
    ("edu_elem", "교육", "초등·학습관리", [
        ("초등전과목", ("초등", "전과목")),
        ("방과후", ("방과후",)),
        ("학습지", ("학습지",)),
        ("자기주도학습", ("자기주도", "학습습관", "학습코칭")),
        ("방문·공부방·과외", ("방문학습", "방문수업", "공부방", "교습소", "과외", "그룹과외", "홈스쿨", "엄마표")),
    ]),
    # ↓ 아래 2개는 기존 중분류 뒤(폴백 직전)에 배치 — 기존 분류를 뺏지 않고 '교육일반'에서만 끌어옴.
    ("edu_dev", "교육", "발달·학습고민", [
        ("집중력·산만", ("산만", "집중못하는", "주의력", "충동적인아이", "가만히못있는")),
        ("느린학습·경계선", ("느린학습자", "경계선지능", "발달지연", "학습부진", "학습장애")),
        ("난독·언어", ("난독증", "언어지연")),
        ("발달·적성검사", ("인지검사", "발달검사", "웩슬러", "학습유형", "영재성검사", "창의성검사", "레벨테스트")),
        ("학습태도·습관", ("공부습관", "학습흥미", "이해력", "응용력", "수포자", "자존감", "사회성", "감각통합")),
    ]),
    ("edu_method", "교육", "학습법·교구", [
        ("교육방법론", ("하브루타", "발도르프", "레지오", "프로젝트수업", "놀이중심", "그림책육아", "거꾸로수업")),
        ("놀이학습", ("놀이수학", "놀이한글", "블록놀이", "교육보드게임")),
        ("교구", ("원목교구", "자석교구", "패턴블록", "소마큐브", "칠교", "쌓기나무", "탱그램", "실물교구", "수학교구", "한글교구")),
        ("취학준비", ("취학전", "초등입학준비", "예비초등준비", "학교적응", "7세고시")),
        ("디지털학습", ("태블릿학습", "스마트학습", "에듀테크", "온라인학습지", "디지털교과서")),
        ("학부모·가정학습", ("초등맘", "유치원맘", "엄마표놀이", "가정학습", "자녀교육", "우리아이교육")),
    ]),
]
_DOVISION_FALLBACK = ("edu_etc", "교육", "교육일반", "교육일반")


def _classify_dovision_category(keyword: str):
    """키워드 → (mid_key, 캠페인라벨 '대분류 - 중분류', 소분류라벨). 매칭 없으면 교육일반."""
    kw = (keyword or "").replace(" ", "").lower()
    for mid_key, major, mid, subs in _DOVISION_TAXONOMY:
        for sub_label, toks in subs:
            if any(t.lower() in kw for t in toks):
                return mid_key, f"{major} - {mid}", sub_label
    mk, major, mid, sub = _DOVISION_FALLBACK
    return mk, f"{major} - {mid}", sub


async def _run_pool_register(uid: int, customer_id: Optional[int] = None, batch: int = 3000, bid: Optional[int] = None):
    """등록 1회 — pending → orchestrator로 일괄.
    customer_id 명시 시 그 광고주만 처리, 없으면 사용자의 가장 최근 광고주.
    bid=None 이면 광고주 default_bid (없으면 100원) 사용 — 광고주마다 다른 값 가능."""
    from services.bulk_upload_orchestrator import BulkUploadOrchestrator, BulkJobConfig
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import get_ad_account_by_customer
    import time as _time

    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    if customer_id is not None:
        account = get_ad_account_by_customer(uid, str(customer_id))
    else:
        account = get_ad_account(uid)
    if not account or not account.get("is_connected"):
        pool.record_run(uid, customer_id, "register", "no_account",
                        error_message="광고 계정 미연결",
                        duration_ms=int((_time.monotonic()-t0)*1000))
        return
    customer_id = int(account.get("customer_id"))

    # 입찰가 — 호출자 명시값 > 광고주 default_bid > 100원 (legacy fallback)
    if bid is None:
        bid = max(70, int(account.get("default_bid") or 100))
    else:
        bid = max(70, int(bid))

    pending = pool.claim_pending(customer_id, limit=batch, min_volume=10)  # 검색량 10 미만 등록 절대 차단
    if not pending:
        s = pool.stats(customer_id)
        pending_total = (s.get("by_status") or {}).get("pending", 0)
        pending_registerable = int(s.get("pending_registerable") or 0)
        seed_rows = max(0, pending_total - pending_registerable)
        logger.warning(
            f"[pool/register] user={uid} pending 없음 "
            f"(등록가능={pending_registerable} / 시드={seed_rows} / 전체pending={pending_total})"
        )
        pool.record_run(
            uid, customer_id, "register", "no_pending",
            pending_after=pending_registerable,
            error_message=(
                f"등록가능 0 (시드 {seed_rows}, 전체pending {pending_total}) — "
                f"새 키워드 수집 대기"
            ) if pending_total else "pending 0",
            duration_ms=int((_time.monotonic()-t0)*1000),
        )
        return
    keywords = [p["keyword"] for p in pending]
    logger.warning(f"[pool/register] user={uid} 시작 batch={len(keywords)}")

    # ───── 도메인 하드 게이트 (register 단계, 2026-05-22) ─────
    # collect/seed_amplify 가 pending 에 off-domain(다낭·유산균 등)을 넣어도 여기서 컷.
    # saved relevance_keywords 로 점수 < 30 인 pending 은 'domain_skipped' 로 빼서
    # pending 에서 제거(재claim 방지), 통과분만 네이버 등록. relevance 미설정 시 스킵(=구동작 유지).
    try:
        from database.naver_ad_db import get_ad_account_relevance_keywords as _grk_gate
        _rel_gate = [s for s in (_grk_gate(uid, str(customer_id)) or []) if s and len(s) >= 2]
        if len(_rel_gate) >= 3:
            _ga3, _ga2 = set(), set()
            for _s in _rel_gate:
                if len(_s) >= 4:
                    _ga3.add(_s)
                for _n in (2, 3):
                    for _i in range(len(_s) - _n + 1):
                        _a = _s[_i:_i + _n]
                        (_ga2 if len(_a) == 2 else _ga3).add(_a)
            _on, _off_ids, _junk_n, _neg_n = [], [], 0, 0
            _JUNK_TOKENS = ("후기", "추천", "비용", "상담", "전문", "정보", "비교", "잘하는곳")
            # negative-token: 한의원 진료가 아닌 상업/비의료 단어. 짧은 질환명(기침·건선·태선·
            # 모반 등)이 무관 단어에 substring 으로 박히는 오매칭 차단 (예: 아기'침대'·물'건선'반).
            _NEG_TOKENS = (
                "침대", "매트", "매트리스", "선반", "가구", "대여", "렌탈", "렌트", "침구", "이불",
                "베개", "소파", "책상", "의자", "수납", "옷장", "주택", "분양", "아파트", "오피스텔",
                "인테리어", "조명", "커튼", "벽지", "그릇", "용기", "포장", "택배", "자동차", "중고차",
                "타이어", "보험", "대출", "적금", "예금", "주식", "펀드", "코인", "비트코인", "재테크",
                "여행", "호텔", "펜션", "리조트", "항공권", "강의", "학원", "인강", "과외", "토익",
                "토플", "자격증", "공무원", "게임", "영화", "드라마", "웹툰", "만화", "레시피", "맛집",
                "식당", "배달", "쇼핑몰", "직구", "운동화", "신발", "가방", "지갑", "선글라스", "안경",
                "화장품", "향수", "립스틱", "컨실러", "쿠션", "파운데이션", "비비크림", "마스카라",
                "유산균", "젤리", "홍삼", "영양제", "비타민제", "콜라겐젤리", "오메가3", "프로틴",
                "노트북", "휴대폰", "에어컨", "냉장고", "세탁기", "청소기", "공기청정기",
            )
            # 화장품 허용 계정(피부과 등)은 화장품 관련 토큰 컷 해제 — 도메인 게이트(relevance)로만 판정.
            if customer_id in _SKIN_CAT_CUSTOMERS:
                _NEG_TOKENS = tuple(t for t in _NEG_TOKENS if t not in _COSMETIC_ALLOW_TOKENS)
            # 교육 광고주(두비전)는 학원/과외/강의/인강 이 핵심 도메인 — 상업 컷 해제.
            elif customer_id in _DOVISION_CAT_CUSTOMERS:
                _NEG_TOKENS = tuple(t for t in _NEG_TOKENS if t not in _DOVISION_ALLOW_TOKENS)
            # 키네스 강제등록 — 학부모/아동 도메인 neg-token 대량 해제.
            elif customer_id in _KINESS_FORCE_CUSTOMERS:
                _NEG_TOKENS = tuple(t for t in _NEG_TOKENS if t not in _KINESS_ALLOW_TOKENS)
            # 키네스 강제등록 — relevance 점수컷 우회(floor=0). junk/neg-token 만 남기고 통과.
            _score_floor = 0 if customer_id in _KINESS_FORCE_CUSTOMERS else 30
            for _p in pending:
                _kw = _p["keyword"] or ""
                _kwc = _kw.replace(" ", "")
                # 정크 컷 (GPT 패딩: 반복토큰 2회+ 또는 과길이) — Naver mt=10 floor 통과하는 무의미 문자열.
                if len(_kwc) >= 20 or any(_kwc.count(_t) >= 2 for _t in _JUNK_TOKENS):
                    _off_ids.append(_p["id"]); _junk_n += 1
                    continue
                # negative-token 컷 (상업/비의료 — substring 오매칭 방지)
                if any(_neg in _kwc for _neg in _NEG_TOKENS):
                    _off_ids.append(_p["id"]); _neg_n += 1
                    continue
                _sc = 0
                _full = False
                for _s in _rel_gate:
                    if _s in _kw:
                        _sc = 100; _full = True; break
                    if _kw and _kw in _s:
                        _sc = 95; _full = True; break
                if not _full:
                    _n3 = sum(1 for _a in _ga3 if _a in _kw)
                    _n2 = sum(1 for _a in _ga2 if _a in _kw)
                    _sc = min(95, min(80, _n3 * 20) + min(30, _n2 * 5))
                if _sc >= _score_floor:
                    _on.append(_p)
                else:
                    _off_ids.append(_p["id"])
            if _off_ids:
                pool.mark_status(_off_ids, "domain_skipped")
            logger.warning(
                f"[pool/register] 도메인게이트: {len(pending)} → 통과 {len(_on)} / 컷 {len(_off_ids)} (정크 {_junk_n} / 상업컷 {_neg_n})"
            )
            pending = _on
            keywords = [p["keyword"] for p in pending]
            if not pending:
                pool.record_run(
                    uid, customer_id, "register", "no_pending",
                    error_message="도메인게이트 통과 0 — claim 된 pending 전부 off-domain",
                    duration_ms=int((_time.monotonic() - t0) * 1000),
                )
                return
    except Exception as _e:
        logger.warning(f"[pool/register] 도메인게이트 예외(무시, 전체등록 진행): {type(_e).__name__}: {_e}")

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    # 호출 전 등록 set 캐시 — 호출 후 차집합으로 진짜 신규만 success 판정.
    reg = get_registered_keywords_db()
    existing_before = set(reg.get_existing_set(customer_id, keywords) or set())

    # 900 (네이버 캠페인당 그룹 한도 1000 아래 마진). 기존 50 은 조기 캠페인 분할을 유발:
    # register tick 마다 테마 버킷당 새 그룹을 만들어(부분 그룹 이어채우기 미지원) 그룹이 희소해지고
    # (그룹당 ~15/1000), 50 그룹 차면 테마별 새 캠페인 → 캠페인 폭증(리베리 180개) → 계정당 캠페인
    # 한도(code 1014)에 막혀 pending 등록 실패. 900 으로 올리면 기존 테마 캠페인을 재사용해 신규
    # 캠페인 생성 없이 등록 지속 (2026-07-06 패킹 fix).
    AD_GROUPS_PER_POOL_CAMPAIGN = 900
    reuse_id: Optional[str] = None

    # category_split 모드 여부 (의료/비의료 한글 캠페인 분리). 비-split 계정(소잠 등)은 기존 'auto' 경로.
    category_mode = False
    _cat_budgets = (3000, 1000)
    try:
        from database.naver_ad_db import get_domain_profile as _gdp_cs
        _profcs = _gdp_cs(uid, str(customer_id)) or {}
        category_mode = bool(_profcs.get("category_split"))
        _cat_budgets = (int(_profcs.get("daily_budget") or 3000), int(_profcs.get("nonmedical_budget") or 1000))
    except Exception:
        pass

    if customer_id in _SKIN_CAT_CUSTOMERS:
        # ── 피부 테마별 한글 캠페인 등록 — '[피부]<라벨>' 캠페인(재사용) + 그룹명에 대표 키워드 ──
        result = {"success": True, "campaign_ids": []}
        _skin_budget = 10000
        try:
            _skin_budget = int(_profcs.get("daily_budget") or 10000)
        except Exception:
            pass
        # 키워드를 테마별로 분류 (분류 순서는 _SKIN_CATEGORIES 우선순위)
        _buckets: Dict[str, List[str]] = {}
        _labels: Dict[str, str] = {}
        for _kw in keywords:
            _key, _lab = _classify_skin_category(_kw)
            _buckets.setdefault(_key, []).append(_kw)
            _labels[_key] = _lab
        for _key, _kws in _buckets.items():
            if not _kws:
                continue
            _label = _labels[_key]
            _campname = f"[피부]{_label}"
            _new_grp = (len(_kws) + 999) // 1000
            _st = pool.get_active_pool_campaign_cat(customer_id, _key)
            _reuse = None; _sidx = 0
            if _st and _st.get("ad_groups_count", 0) + _new_grp <= AD_GROUPS_PER_POOL_CAMPAIGN:
                _reuse = _st["campaign_id"]; _sidx = _st["ad_groups_count"]
            _jid = create_bulk_upload_job(
                user_id=uid, filename=f"pool_skin_{_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                campaign_prefix=_campname, keywords_per_group=1000, bid=bid,
                daily_budget=_skin_budget, total_keywords=len(_kws),
            )
            _cfg = BulkJobConfig(
                job_id=_jid, user_id=uid, campaign_prefix=_campname, keywords_per_group=1000,
                bid=bid, daily_budget=_skin_budget, campaign_tp="WEB_SITE",
                reuse_campaign_id=_reuse, start_ad_group_index=_sidx,
                descriptive_group_names=True,
            )
            try:
                _r = await BulkUploadOrchestrator(client).run(_cfg, _kws)
            except Exception as e:
                logger.error(f"[pool/register-skin] {_campname} orchestrator 실패: {e}", exc_info=True)
                result["success"] = False; result["error"] = f"{type(e).__name__}: {str(e)[:160]}"
                continue
            _cids = _r.get("campaign_ids") or []
            result["campaign_ids"].extend(_cids)
            if not _r.get("success"):
                result["success"] = False; result["error"] = _r.get("error")
            try:
                if _reuse:
                    pool.set_active_pool_campaign_cat(customer_id, _key, _reuse, _sidx + _new_grp)
                elif _cids:
                    pool.set_active_pool_campaign_cat(customer_id, _key, _cids[0], _new_grp)
            except Exception as e:
                logger.warning(f"[pool/register-skin] {_campname} cat-state 갱신 실패: {e}")
            logger.warning(f"[pool/register-skin] {_campname} {len(_kws)}개 (budget {_skin_budget}, reuse={bool(_reuse)})")
    elif customer_id in _DOVISION_CAT_CUSTOMERS:
        # ── 두비전 계층 한글 캠페인 등록 — 중분류=캠페인 '[두비전] 대 - 중'(재사용), 소분류=광고그룹명 ──
        result = {"success": True, "campaign_ids": []}
        _dovi_budget = 10000
        try:
            _dovi_budget = int(_profcs.get("daily_budget") or 10000)
        except Exception:
            pass
        # 이중 버킷: mid_key → {label: '대 - 중', subs: {소분류: [kw]}}
        _mid_buckets: Dict[str, Dict[str, Any]] = {}
        for _kw in keywords:
            _mk, _clabel, _slabel = _classify_dovision_category(_kw)
            _b = _mid_buckets.setdefault(_mk, {"label": _clabel, "subs": {}})
            _b["subs"].setdefault(_slabel, []).append(_kw)
        for _mk, _mb in _mid_buckets.items():
            _campname = f"[두비전] {_mb['label']}"
            # 이 틱에서 만들 그룹 수 = 소분류별 ceil(n/1000) 합
            _new_grp = sum((len(_v) + 999) // 1000 for _v in _mb["subs"].values())
            _st = pool.get_active_pool_campaign_cat(customer_id, _mk)
            _reuse = None; _sidx = 0
            if _st and _st.get("ad_groups_count", 0) + _new_grp <= AD_GROUPS_PER_POOL_CAMPAIGN:
                _reuse = _st["campaign_id"]; _sidx = _st["ad_groups_count"]
            _created_cid = _reuse          # 첫 소분류 등록이 새 캠페인을 만들면 채워짐 → 이후 소분류 재사용
            _grp_cursor = _sidx
            for _slabel, _skws in _mb["subs"].items():
                if not _skws:
                    continue
                _sub_grp = (len(_skws) + 999) // 1000
                _jid = create_bulk_upload_job(
                    user_id=uid, filename=f"pool_dovi_{_mk}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                    campaign_prefix=_campname, keywords_per_group=1000, bid=bid,
                    daily_budget=_dovi_budget, total_keywords=len(_skws),
                )
                _cfg = BulkJobConfig(
                    job_id=_jid, user_id=uid, campaign_prefix=_campname, keywords_per_group=1000,
                    bid=bid, daily_budget=_dovi_budget, campaign_tp="WEB_SITE",
                    reuse_campaign_id=_created_cid, start_ad_group_index=_grp_cursor,
                    group_label=_slabel,
                )
                try:
                    _r = await BulkUploadOrchestrator(client).run(_cfg, _skws)
                except Exception as e:
                    logger.error(f"[pool/register-dovi] {_campname}/{_slabel} orchestrator 실패: {e}", exc_info=True)
                    result["success"] = False; result["error"] = f"{type(e).__name__}: {str(e)[:160]}"
                    continue
                _cids = _r.get("campaign_ids") or []
                if not _created_cid and _cids:
                    _created_cid = _cids[0]
                    result["campaign_ids"].extend(_cids)
                if not _r.get("success"):
                    result["success"] = False; result["error"] = _r.get("error")
                _grp_cursor += _sub_grp
                logger.warning(f"[pool/register-dovi] {_campname} / {_slabel} {len(_skws)}개 (reuse={bool(_created_cid)})")
                # 새 캠페인 확보 실패(첫 소분류 등록 실패) 시 중단 — 다음 소분류가 동명 캠페인을
                # 또 만드는 중복 방지. 미등록 키워드는 pending 유지되어 다음 틱에 재시도.
                if not _created_cid:
                    logger.warning(f"[pool/register-dovi] {_campname} 캠페인 미확보 → 잔여 소분류 이번 틱 보류")
                    break
            try:
                if _created_cid:
                    pool.set_active_pool_campaign_cat(customer_id, _mk, _created_cid, _grp_cursor)
            except Exception as e:
                logger.warning(f"[pool/register-dovi] {_campname} cat-state 갱신 실패: {e}")
    elif category_mode:
        # ── 의료/비의료 분리 등록 — 각 카테고리별 한글 캠페인(재사용) + 차등 예산 ──
        result = {"success": True, "campaign_ids": []}
        med_kws = [k for k in keywords if _classify_medical(k)]
        non_kws = [k for k in keywords if not _classify_medical(k)]
        for _cat, _label, _kws, _bud in (
            ("medical", "의료대출", med_kws, _cat_budgets[0]),
            ("nonmedical", "비의료대출", non_kws, _cat_budgets[1]),
        ):
            if not _kws:
                continue
            _new_grp = (len(_kws) + 999) // 1000
            _st = pool.get_active_pool_campaign_cat(customer_id, _cat)
            _reuse = None; _sidx = 0
            if _st and _st.get("ad_groups_count", 0) + _new_grp <= AD_GROUPS_PER_POOL_CAMPAIGN:
                _reuse = _st["campaign_id"]; _sidx = _st["ad_groups_count"]
            _jid = create_bulk_upload_job(
                user_id=uid, filename=f"pool_{_cat}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                campaign_prefix=_label, keywords_per_group=1000, bid=bid,
                daily_budget=_bud, total_keywords=len(_kws),
            )
            _cfg = BulkJobConfig(
                job_id=_jid, user_id=uid, campaign_prefix=_label, keywords_per_group=1000,
                bid=bid, daily_budget=_bud, campaign_tp="WEB_SITE",
                reuse_campaign_id=_reuse, start_ad_group_index=_sidx,
            )
            try:
                _r = await BulkUploadOrchestrator(client).run(_cfg, _kws)
            except Exception as e:
                logger.error(f"[pool/register-cat] {_label} orchestrator 실패: {e}", exc_info=True)
                result["success"] = False; result["error"] = f"{type(e).__name__}: {str(e)[:160]}"
                continue
            _cids = _r.get("campaign_ids") or []
            result["campaign_ids"].extend(_cids)
            if not _r.get("success"):
                result["success"] = False; result["error"] = _r.get("error")
            try:
                if _reuse:
                    pool.set_active_pool_campaign_cat(customer_id, _cat, _reuse, _sidx + _new_grp)
                elif _cids:
                    pool.set_active_pool_campaign_cat(customer_id, _cat, _cids[0], _new_grp)
            except Exception as e:
                logger.warning(f"[pool/register-cat] {_label} cat-state 갱신 실패: {e}")
            logger.warning(f"[pool/register-cat] {_label} {len(_kws)}개 (budget {_bud}, reuse={bool(_reuse)})")
    else:
        # ── 기존 경로 (auto_ 단일 캠페인, 예산 1만) — 변경 없음 ──
        pool_state = pool.get_active_pool_campaign(customer_id)
        start_idx = 0
        new_groups_in_round = (len(keywords) + 999) // 1000  # 1000개당 광고그룹 1개
        if pool_state and pool_state.get("ad_groups_count", 0) + new_groups_in_round <= AD_GROUPS_PER_POOL_CAMPAIGN:
            reuse_id = pool_state["campaign_id"]
            start_idx = pool_state["ad_groups_count"]
            logger.warning(
                f"[pool/register] 캠페인 재사용 cid={reuse_id} groups={pool_state['ad_groups_count']}+{new_groups_in_round}"
            )
        job_id = create_bulk_upload_job(
            user_id=uid,
            filename=f"pool_auto_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            campaign_prefix="auto",
            keywords_per_group=1000,
            bid=bid,
            daily_budget=10000,
            total_keywords=len(keywords),
        )
        cfg = BulkJobConfig(
            job_id=job_id, user_id=uid,
            campaign_prefix="auto", keywords_per_group=1000,
            bid=bid, daily_budget=10000, campaign_tp="WEB_SITE",
            reuse_campaign_id=reuse_id,
            start_ad_group_index=start_idx,
        )
        orchestrator = BulkUploadOrchestrator(client)
        try:
            result = await orchestrator.run(cfg, keywords)
        except Exception as e:
            logger.error(f"[pool/register] orchestrator 실패: {e}", exc_info=True)
            pool.mark_status([p["id"] for p in pending], "failed",
                             error_message=f"{type(e).__name__}: {str(e)[:200]}")
            pool.record_run(uid, customer_id, "register", "failed",
                            failed=len(pending),
                            error_message=f"{type(e).__name__}: {str(e)[:300]}",
                            duration_ms=int((_time.monotonic()-t0)*1000))
            return

    existing_after = set(reg.get_existing_set(customer_id, keywords) or set())
    new_in_naver = existing_after - existing_before  # 진짜 신규 등록

    # 풀 state 업데이트 — 캠페인 재사용 또는 새 캠페인 등록
    # (category_mode / 피부 테마 모드는 위에서 cat-state 갱신 완료 → 비-cat state 갱신 skip)
    if not category_mode and customer_id not in _SKIN_CAT_CUSTOMERS and customer_id not in _DOVISION_CAT_CUSTOMERS:
        try:
            result_campaign_ids = result.get("campaign_ids") or []
            ad_groups_in_round = (len(keywords) + 999) // 1000
            if reuse_id:
                # 같은 캠페인에 광고그룹 추가됨
                pool.increment_pool_ad_groups(customer_id, ad_groups_in_round)
            elif result_campaign_ids:
                # 새 캠페인 → state 갱신
                pool.set_active_pool_campaign(customer_id, result_campaign_ids[0], ad_groups_in_round)
        except Exception as e:
            logger.warning(f"[pool/register] state 갱신 실패: {e}")

    succeeded_ids = [p["id"] for p in pending if p["keyword"] in new_in_naver]
    skipped_ids = [p["id"] for p in pending if p["keyword"] in existing_before]
    failed_ids = [
        p["id"] for p in pending
        if p["keyword"] not in new_in_naver and p["keyword"] not in existing_before
    ]
    pool.mark_status(succeeded_ids, "registered")
    pool.mark_status(skipped_ids, "skipped_existing",
                     error_message="이미 네이버 광고에 등록된 키워드 — orchestrator dedup")
    err_msg = str(result.get("error", "did not register"))[:300] if not result.get("success") else None

    # Pre-flight 실패 (channel lookup, no business channel, account cap 등) — 키워드
    # 자체엔 문제 없음. failed 영구 마킹 시 다음 tick 재시도 불가 → 117k 누적 실패
    # 사고. transient 식별 시 pending 유지하여 다음 tick 자동 재시도.
    _transient_markers = ("channel lookup", "no business channel", "account keyword cap")
    is_transient_preflight = (
        not result.get("success")
        and err_msg
        and any(m in err_msg for m in _transient_markers)
    )
    if is_transient_preflight and failed_ids:
        logger.warning(
            f"[pool/register] user={uid} pre-flight 실패 ({err_msg[:80]}) — "
            f"{len(failed_ids)}개 pending 유지 (다음 tick 재시도)"
        )
        # mark_status skip — pending 그대로 둠. 다음 register tick 이 다시 claim.
    else:
        pool.mark_status(failed_ids, "failed",
                         error_message=err_msg or "orchestrator did not register")
    logger.warning(
        f"[pool/register] user={uid} 신규={len(succeeded_ids)} "
        f"이미있음={len(skipped_ids)} fail={len(failed_ids)}"
    )
    pending_after = (pool.stats(customer_id).get("by_status") or {}).get("pending", 0)
    pool.record_run(
        uid, customer_id, "register",
        "success" if len(succeeded_ids) > 0 and len(failed_ids) == 0
            else ("partial" if len(succeeded_ids) > 0 else ("failed" if len(failed_ids) > 0 else "no_new")),
        registered=len(succeeded_ids), failed=len(failed_ids), skipped=len(skipped_ids),
        pending_after=pending_after,
        error_message=err_msg,
        duration_ms=int((_time.monotonic()-t0)*1000),
    )

    # 노출제한 검사는 register 끝이 아닌 _run_pool_inspect_only 단독 실행으로 이전됨.
    # Why: register 의 pending=0 / orchestrator 실패 early return 시 inspect 자체가 호출
    # 안 됨 → 노출제한 자동 삭제 영구 미실행 누수. cron tick 마다 register 와 독립적으로
    # _run_pool_inspect_only 호출되도록 _run_pool_workers_for_accounts 에서 보장.


async def _inspect_ad_groups(
    uid: int,
    customer_id: int,
    client,
    ad_group_ids: List[str],
    delete_from_naver: bool = True,
):
    """광고그룹들 키워드 검토 상태 조회 → 노출제한:
       1) 풀에 mark
       2) 네이버 광고에서 키워드 DELETE
       3) registered_keywords DB에서 row 제거 (active 카운트 정확화)"""
    if not ad_group_ids:
        return 0
    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()
    rejected_items: List[Dict] = []
    rejected_naver_ids: List[Tuple[str, str]] = []  # (ncc_keyword_id, keyword)
    debug_samples_logged = False  # 첫 광고그룹 첫 KW dict 한 번만 logging
    total_kws_checked = 0
    for ag_id in ad_group_ids:
        try:
            kws = await client.get_keywords(ad_group_id=ag_id) or []
        except Exception as e:
            logger.warning(f"[pool/inspect] {ag_id} 조회 실패: {e}")
            continue
        # 디버그 — 첫 광고그룹 첫 KW dict 의 keys + review/inspect/status/statusReason 값 logging.
        # fly logs 에서 실제 네이버 응답 형태 확인용. 한 cron tick 당 1회만.
        if not debug_samples_logged and kws:
            sample = kws[0]
            logger.warning(
                f"[pool/inspect/DEBUG] ag={ag_id} sample_keys={list(sample.keys())[:20]} "
                f"review={sample.get('reviewStatus')!r} inspect={sample.get('inspectStatus')!r} "
                f"status={sample.get('status')!r} statusReason={sample.get('statusReason')!r}"
            )
            # 거부 후보 sample 5개 — review/inspect/statusReason 에 PENDING 외 값 가진 첫 5개
            cand_samples = []
            for k in kws[:50]:
                rv = (k.get("reviewStatus") or "")
                ip = (k.get("inspectStatus") or "")
                rs = (k.get("statusReason") or "")
                if rv or ip or rs:
                    cand_samples.append(
                        f"{k.get('keyword','')}={rv}/{ip}/{rs}"
                    )
                    if len(cand_samples) >= 5:
                        break
            if cand_samples:
                logger.warning(f"[pool/inspect/DEBUG] non-empty samples: {' || '.join(cand_samples)}")
            debug_samples_logged = True

        for kw in kws:
            kw_text = kw.get("keyword")
            if not kw_text:
                continue
            total_kws_checked += 1
            review = (kw.get("reviewStatus") or "").upper()
            inspect = (kw.get("inspectStatus") or "").upper()
            status = (kw.get("status") or "").upper()
            stat_reason = (kw.get("statusReason") or "").upper()
            user_lock = kw.get("userLock", False)

            # statusReason 의 영구 거부 코드는 PENDING 가드보다 먼저 잡는다.
            # 네이버 거부 statusReason 토큰 (확인된 값들):
            #   KEYWORD_DISAPPROVED, BUSINESS_PROHIBITED, REVIEW_NOT_PASSED, INSPECT_FAIL,
            #   BAD_BUSINESS, BLOCKLISTED, PROHIBITED.
            # NOT_PASSED / FAIL 추가 — 이전엔 누락되어 REVIEW_NOT_PASSED / INSPECT_FAIL KW 가
            # PENDING 가드에 막혀 영구 미삭제 (사용자 보고 사례).
            REASON_REJECT_TOKENS = (
                "DISAPPROVED", "REJECTED", "PROHIBITED", "BLOCKLISTED",
                "NOT_PASSED", "FAIL", "BAD_BUSINESS", "INELIGIBLE",
            )
            reason_rejected = any(t in stat_reason for t in REASON_REJECT_TOKENS)
            # inspectStatus 자체에 거부 토큰 있어도 PENDING 가드보다 먼저 잡는다.
            # 네이버는 inspectStatus="REVIEW_FAILED" + statusReason="" 케이스도 있음.
            INSPECT_REJECT_TOKENS = (
                "DISAPPROVED", "REJECTED", "PROHIBITED", "BLOCKLISTED",
                "NOT_PASSED", "FAIL", "DENIED", "BAD_BUSINESS",
            )
            inspect_rejected_early = any(t in inspect for t in INSPECT_REJECT_TOKENS)
            review_rejected_early = ("REJECT" in review) or ("DISAPPROVE" in review) or ("FAIL" in review) or ("NOT_PASSED" in review)
            if reason_rejected or inspect_rejected_early or review_rejected_early:
                rejected_items.append({
                    "keyword": kw_text,
                    "reason": f"review={review} inspect={inspect} status={status} reason={stat_reason} userLock={user_lock}",
                })
                kid = kw.get("nccKeywordId")
                if kid:
                    rejected_naver_ids.append((kid, kw_text))
                continue

            # ============ 하드 가드 — 검수 완료 전 절대 건드리지 않는다 ============
            # 신규 키워드는 Naver 검수 완료 전까지 review/inspect 가 WAIT/UNDER/PENDING 계열.
            # 이 단계에서 어떤 판정도 하면 안 됨 (대량 삭제 사고 영구 차단).
            # review 와 inspect 둘 중 하나라도 'pending' 으로 보이면 즉시 skip.
            PENDING_TOKENS = (
                "WAIT", "UNDER", "PENDING", "PROGRESS",
                "IN_REVIEW", "AUTO_INSPECT", "INSPECT_REQ",
                "PRE_REVIEW", "BEFORE_REVIEW",
            )
            review_pending = any(t in review for t in PENDING_TOKENS)
            inspect_pending = any(t in inspect for t in PENDING_TOKENS)
            # review/inspect 모두 비어있으면 정보 없음 → 안전 측면 미완료 취급
            no_info = (review == "" and inspect == "" and stat_reason == "")
            if review_pending or inspect_pending or no_info:
                continue
            # =============================================================

            # 검수 완료 가정 하에 영구 거부 신호만 잡는다.
            # 일시 상태(PAUSED/userLock/EXPIRED_BUDGET 등)는 트리거 안 됨.
            review_rejected = ("REJECT" in review) or ("DISAPPROVE" in review)
            inspect_rejected = inspect in (
                "PROHIBIT", "BUSINESS_PROHIBIT", "REVIEW_REJECTED",
                "REJECTED", "DISAPPROVED", "FAIL", "FAILED",
            )
            if review_rejected or inspect_rejected:
                rejected_items.append({
                    "keyword": kw_text,
                    "reason": f"review={review} inspect={inspect} status={status} reason={stat_reason} userLock={user_lock}",
                })
                kid = kw.get("nccKeywordId")
                if kid:
                    rejected_naver_ids.append((kid, kw_text))
        await asyncio.sleep(0.15)

    logger.warning(
        f"[pool/inspect] user={uid} cid={customer_id} ag_groups={len(ad_group_ids)} "
        f"kws_checked={total_kws_checked} rejected_found={len(rejected_items)}"
    )
    if rejected_items:
        # 거부 KW 첫 5개 sample logging — 어떤 토큰이 매치됐는지 검증용
        sample_reasons = [f"{it['keyword']}:{it['reason'][:80]}" for it in rejected_items[:5]]
        logger.warning(f"[pool/inspect] rejected samples: {' | '.join(sample_reasons)}")
    n_mark = pool.mark_rejected_by_naver(customer_id, rejected_items)

    # 네이버에서 실제 DELETE — 실패 시 PUT pause로 fallback (광고 노출만 차단)
    n_deleted = 0
    n_paused = 0
    if delete_from_naver and rejected_naver_ids:
        for kid, kw_text in rejected_naver_ids:
            ok = False
            try:
                await client.delete_keyword(kid)
                # registered_keywords DB에서도 row 삭제 (한도 카운트 정확화)
                try:
                    with __import__("sqlite3").connect(reg.db_path) as conn:
                        conn.execute(
                            "DELETE FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id=?",
                            (customer_id, kid),
                        )
                except Exception:
                    pass
                n_deleted += 1
                ok = True
            except Exception as e:
                # DELETE 권한 1018 등 실패 시 — pause(userLock)로 광고 노출만 차단
                try:
                    await client.pause_keyword(kid)
                    n_paused += 1
                    ok = True
                except Exception as e2:
                    logger.warning(f"[pool/inspect] DELETE+PAUSE 모두 실패 {kw_text}({kid}): del={e} pause={e2}")
            if ok:
                await asyncio.sleep(0.15)

    if n_mark > 0 or n_deleted > 0:
        logger.warning(
            f"[pool/inspect] user={uid} mark={n_mark} 네이버삭제={n_deleted} ({len(ad_group_ids)} 그룹)"
        )
    # 실행 이력 기록 — 화면에 보이게
    try:
        pool.record_run(
            uid, customer_id, "inspect",
            "success" if n_deleted > 0 or n_mark > 0 else "no_new",
            registered=0, failed=0, skipped=n_deleted,  # skipped 컬럼에 삭제 카운트
            seeds_count=len(ad_group_ids),
            error_message=f"광고그룹 {len(ad_group_ids)}개 검사 — mark {n_mark} / 네이버 DELETE {n_deleted}" if (n_mark or n_deleted) else f"검사 {len(ad_group_ids)}개 그룹 — 노출제한 0",
        )
    except Exception:
        pass
    return n_mark


async def _run_pool_inspect_only(uid: int, customer_id: int) -> None:
    """노출제한 검사 단독 실행 — register 의 pending=0 early return 으로 inspect 가
    영구 미실행되는 누수 차단. 매 cron tick 마다 register 와 독립적으로 호출.
    cascade drift 정리 후 / collect circuit OPEN 광고주에서도 노출제한 자동 삭제 보장.
    """
    from services.naver_ad_service import NaverAdApiClient, _naver_api_breaker
    from database.naver_ad_db import get_ad_account_by_customer
    import sqlite3 as _sqlite3
    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()

    if _naver_api_breaker.is_open():
        # circuit OPEN 인 동안은 inspect 도 skip — 다음 tick 에 재시도
        pool.record_run(uid, customer_id, "inspect", "no_new",
                        error_message="circuit OPEN — inspect skip, 다음 tick 재시도")
        return

    account = get_ad_account_by_customer(uid, str(customer_id))
    if not account or not account.get("is_connected"):
        return  # 비연결 광고주 — record 도 노이즈 방지로 안 함

    with _sqlite3.connect(reg.db_path) as _conn:
        ag_ids = [r[0] for r in _conn.execute(
            "SELECT DISTINCT ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ad_group_id IS NOT NULL",
            (customer_id,),
        ).fetchall()]
    if not ag_ids:
        return  # 등록 KW 없음 — inspect 대상 없음 (record 노이즈 방지)

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]
    await _inspect_ad_groups(uid, customer_id, client, ag_ids, delete_from_naver=True)


async def _run_pool_ai_seed_topup(uid: int, customer_id: int) -> Dict[str, Any]:
    """collect 가 마른 우물 (added < N) 일 때 LLM 으로 새 시드 자동 주입.

    - saved relevance_keywords (없으면 user_seed Top80) 를 base 로 GPT-4o-mini 호출
    - 도메인 토큰셋 1차 필터 → keywordstool 5개 배치 검증 → user_seed 로 INSERT
    - 24시간 내 6회 cap (OpenAI cost 보호) + cooldown 25분
    - record_run(kind='ai_topup') 으로 frontend 표시
    """
    import json as _json
    from config import settings as _settings
    from database.naver_ad_db import (
        get_ad_account_by_customer,
        get_ad_account_relevance_keywords,
    )
    import time as _time
    import sqlite3 as _sqlite3

    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    # OpenAI 키 / 광고주 / 도메인 의도 확인
    if not getattr(_settings, "OPENAI_API_KEY", ""):
        logger.warning(f"[pool/ai-topup] uid={uid} cid={customer_id} OPENAI_API_KEY 미설정 — skip")
        return {"skipped": "no_openai_key"}
    account = get_ad_account_by_customer(uid, str(customer_id))
    if not account or not account.get("is_connected"):
        return {"skipped": "no_account"}
    saved = get_ad_account_relevance_keywords(uid, str(customer_id)) or []
    if len(saved) < 3:
        # 폴백 — user_seed 풀 Top 60 (오염 가능하므로 saved_relevance 가 진짜 의도).
        seeds_fb = pool.list_user_seeds(customer_id) or []
        if len(seeds_fb) < 3:
            return {"skipped": "insufficient_base_seeds"}
        base = seeds_fb[:60]
        basis = "user_seed_fallback"
    else:
        base = saved
        basis = "saved_relevance"

    # cooldown / daily cap — naverad_pool_runs 직접 조회 (전용 메서드 추가 회피).
    AI_TOPUP_COOLDOWN_S = 25 * 60
    AI_TOPUP_DAILY_CAP = 6
    try:
        with _sqlite3.connect(pool.db_path) as conn:
            row = conn.execute(
                """SELECT started_at, COUNT(*) FROM naverad_pool_runs
                   WHERE account_customer_id=? AND kind='ai_topup'
                     AND started_at > datetime('now','-24 hours')""",
                (customer_id,),
            ).fetchone()
            recent_cnt = int(row[1] or 0) if row else 0
            last_started = row[0] if row else None
            if recent_cnt >= AI_TOPUP_DAILY_CAP:
                logger.warning(f"[pool/ai-topup] uid={uid} cid={customer_id} 24h cap {AI_TOPUP_DAILY_CAP} 도달 — skip")
                return {"skipped": "daily_cap", "recent_24h": recent_cnt}
            if last_started:
                last_row = conn.execute(
                    """SELECT (julianday('now') - julianday(started_at)) * 86400 AS sec_ago
                       FROM naverad_pool_runs
                       WHERE account_customer_id=? AND kind='ai_topup'
                       ORDER BY id DESC LIMIT 1""",
                    (customer_id,),
                ).fetchone()
                sec_ago = float(last_row[0]) if last_row and last_row[0] else 999999
                if sec_ago < AI_TOPUP_COOLDOWN_S:
                    return {"skipped": "cooldown", "sec_ago": int(sec_ago)}
    except Exception as e:
        logger.warning(f"[pool/ai-topup] cooldown 조회 실패 (계속): {e}")

    # 도메인 토큰셋 — base seeds 만 (풀 오염 무시)
    domain_tokens = _build_domain_token_set(base) | _build_seed_atoms(base)
    def _matches_domain(kw: str) -> bool:
        k = (kw or "").replace(" ", "")
        if len(k) < 2:
            return False
        return any(t in k for t in domain_tokens)

    # 1) LLM 호출 — 80개 후보 생성
    prompt_seeds = ", ".join(base[:60])
    prompt = (
        f"다음 한국어 키워드들과 동일 도메인의 검색 가능성 있는 한국어 키워드를 정확히 80개 생성해줘.\n\n"
        f"입력 키워드:\n{prompt_seeds}\n\n"
        f"규칙:\n- 입력 키워드와 동일 분야 안에서만 (예: 의료면 의료)\n"
        f"- 다른 도메인 단어 절대 금지\n- 띄어쓰기 가능\n- 한 줄당 1개\n- JSON array"
    )
    candidates: List[str] = []
    try:
        async with httpx.AsyncClient(timeout=60.0) as oai:
            resp = await oai.post(
                "https://api.openai.com/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {_settings.OPENAI_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "gpt-4o-mini",
                    "messages": [
                        {"role": "system", "content": "한국어 검색광고 키워드 전문가. 도메인 일관성 절대 위반 금지. JSON array 만 반환."},
                        {"role": "user", "content": prompt},
                    ],
                    "temperature": 0.7,
                    "max_tokens": 3000,
                },
            )
            resp.raise_for_status()
            content = resp.json()["choices"][0]["message"]["content"]
        cb = re.search(r"```(?:json)?\s*(.*?)\s*```", content, re.DOTALL)
        if cb:
            content = cb.group(1)
        gen = _json.loads(content.strip())
        if isinstance(gen, list):
            seen = set()
            for k in gen:
                if isinstance(k, str):
                    kw = k.strip()
                    if kw and len(kw) >= 2 and kw not in seen:
                        seen.add(kw)
                        candidates.append(kw)
    except Exception as e:
        pool.record_run(uid, customer_id, "ai_topup", "failed",
                        error_message=f"LLM 실패: {type(e).__name__}: {str(e)[:200]}",
                        duration_ms=int((_time.monotonic()-t0)*1000))
        logger.warning(f"[pool/ai-topup] LLM 실패 uid={uid} cid={customer_id}: {e}")
        return {"skipped": "llm_failed", "error": str(e)[:200]}

    # 2) 도메인 1차 필터 + 도메인 토큰셋 보강
    domain_pass = [k for k in candidates if _matches_domain(k)]
    domain_fail = len(candidates) - len(domain_pass)
    # LLM 후보 atom 을 도메인 토큰에 합류 — keywordstool 응답 검증 시 broader 매칭 보장.
    # Why: LLM 이 "콜린성/박탈성/카포시" 같은 보강 도메인어 만들어도 그 atom 없으면
    #      응답 KW 가 도메인 게이트 reject. atom 합류 → broader 통과.
    if domain_pass:
        domain_tokens = domain_tokens | _build_seed_atoms(domain_pass)

    # 3) keywordstool 배치 — hint 는 base (saved_relevance) 일반어 사용.
    # Why: LLM 후보가 niche 의학 용어 (카포시육종/유암종증후군/포피염 등) 면 keywordstool
    #      응답이 빈 list (Naver index 에 데이터 없음) → 검증 0. 일반어 hint 면 1000+ KW
    #      반환 → 도메인+검색량 통과한 KW 풍부. LLM 의 역할은 도메인 토큰셋 확장 (위).
    hint_source = list(base)  # saved_relevance 또는 user_seed Top60
    from services.naver_ad_service import NaverAdApiClient
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    validated: List[Dict] = []
    seen_validated: Set[str] = set()
    MIN_VOL = 5
    MAX_INSERT = 200  # 한 ai_topup 호출당 최대 user_seed INSERT 수
    for i in range(0, len(hint_source), 5):
        if len(validated) >= MAX_INSERT:
            break
        chunk = hint_source[i:i + 5]
        # 빈 / 짧은 hint 제외 — Naver keywordstool 거부 패턴
        chunk = [s.replace(" ", "").strip() for s in chunk if s and len(s.strip()) >= 2]
        if not chunk:
            continue
        hint_str = ",".join(chunk)
        try:
            related = await client.get_related_keywords(hint_str, show_detail=True)
        except Exception as e:
            logger.warning(f"[pool/ai-topup] keywordstool 실패 {chunk}: {e}")
            continue
        items = related.get("keywordList", []) if isinstance(related, dict) else []
        for it in items:
            kw = (it.get("relKeyword") or "").strip()
            if not kw or kw in seen_validated:
                continue
            pc = _parse_naver_count(it.get("monthlyPcQcCnt"))
            mob = _parse_naver_count(it.get("monthlyMobileQcCnt"))
            mt = pc + mob
            if mt < MIN_VOL:
                continue
            # 응답 KW 도 도메인 게이트 통과해야 함 (LLM 이 trigger 했어도 Naver 가 cross-
            # domain 연관 키워드 끼워 반환할 수 있음).
            if not _matches_domain(kw):
                continue
            seen_validated.add(kw)
            validated.append({
                "keyword": kw,
                "monthly_total": mt,
                "monthly_pc": pc,
                "monthly_mobile": mob,
                "comp_idx": it.get("compIdx"),
                "seed": "ai_topup",
                "source": "user_seed",  # collect 가 다음 라운드에 자동 사용
            })
        await asyncio.sleep(0.3)

    # 4) user_seed 로 INSERT
    seed_items = [{**v, "source": "user_seed"} for v in validated]
    added = pool.add_candidates(uid, customer_id, seed_items)

    duration_ms = int((_time.monotonic() - t0) * 1000)
    pool.record_run(
        uid, customer_id, "ai_topup",
        "success" if added > 0 else "no_new",
        added=added, seeds_count=len(base),
        error_message=(
            f"AI 시드 확장 ({basis}) — LLM {len(candidates)} → 도메인 {len(domain_pass)} "
            f"(컷 {domain_fail}) → 검증≥{MIN_VOL} {len(validated)} → INSERT {added}"
        ),
        duration_ms=duration_ms,
    )
    logger.warning(
        f"[pool/ai-topup] uid={uid} cid={customer_id} basis={basis} base={len(base)} "
        f"LLM={len(candidates)} domain={len(domain_pass)} validated={len(validated)} added={added} ({duration_ms}ms)"
    )
    return {"added": added, "llm": len(candidates), "validated": len(validated), "duration_ms": duration_ms}


# ==========================================================================
# 10만 자동채우기 에스컬레이션 컨트롤러 (#1 골격)
# 오퍼레이터 사다리를 영속 상태(naverad_pool_escalation)로 관리. collect 가 마르면
# 레벨을 올려 더 공격적으로 시드/앵글을 주입하고, 생산적이면 값싼 BFS 로 복귀한다.
# level 0 = 오늘과 100% 동일 동작. 상위 레벨의 신규 오퍼레이터(조합 생성·표면 채굴·
# 관련성 티어 하강)는 후속 단계에서 이 골격 위에 부착한다.
# ==========================================================================
_FILL_CAP = 100_000
_FILL_FRESH_OK = 30            # collect added ≥ 이면 생산적 라운드로 판정
_FILL_DRY_ESCALATE_AFTER = 2   # 마른 라운드 연속 N회 → 레벨 +1
_FILL_MAX_LEVEL = 5
_FILL_FLOOR_START = 50
_FILL_FLOOR_MIN = 25
_FILL_LEVEL_LABELS = {
    0: "BFS 기본 발굴",
    1: "LLM 앵글 시드 주입(공격)",
    2: "조합 생성(지역·수식어×머리어)",
    3: "표면 채굴(자동완성·연관검색어)",
    4: "LLM 앵글 강화",
    5: "관련성 티어 하강(롱테일·약관련 흡수)",
}


def _fill_next_lever(level: int) -> str:
    return _FILL_LEVEL_LABELS.get(min(level + 1, _FILL_MAX_LEVEL), "최대 단계")


async def _fill_escalation_decide(uid: int, cid: int) -> Dict[str, Any]:
    """직전 collect 결과를 보고 에스컬레이션 레벨을 갱신하고 이번 tick 의 결정을 반환.
    부작용은 상태 upsert 뿐. 실패 시 안전 기본값(run_topup=None → 호출부 폴백) 반환."""
    pool = get_keyword_pool_db()
    try:
        reg = get_registered_keywords_db()
        active_reg = int((reg.stats(cid) or {}).get("active") or 0)
        pool_pending = int((pool.stats(cid).get("by_status") or {}).get("pending", 0) or 0)
        headroom = _FILL_CAP - active_reg - pool_pending

        recent = pool.recent_runs(cid, limit=6) or []
        last_collect = next((r for r in recent if r.get("kind") == "collect"), None)
        last_added = int(last_collect.get("added") or 0) if last_collect else 0

        st = pool.get_escalation(cid)
        level = int(st.get("level") or 0)
        dry_streak = int(st.get("dry_streak") or 0)
        floor = int(st.get("relevance_floor") or _FILL_FLOOR_START)

        if headroom <= 0:
            # 이미 10만 도달 — floor 유지, 물갈이는 collect 내부 self-heal 담당.
            note = f"cap_reached active={active_reg} pending={pool_pending}"
            level = (_FILL_FLOOR_START - floor) // 5
            pool.set_escalation(cid, level, dry_streak, floor, last_added, note)
            return {"level": level, "dry_streak": dry_streak, "relevance_floor": floor,
                    "headroom": headroom, "run_topup": False, "note": note,
                    "next_lever": _fill_next_lever(level)}

        # 압축 사다리 — 레벨 5칸 대기 없이 관련성 floor 를 직접 조절.
        # L2 조합·L3 자동완성·seed_amplify 는 이미 독립 cron 으로 상시 실행되므로,
        # 컨트롤러의 고유 역할은 "그래도 마르면 관련성 floor 를 낮춰 롱테일 흡수".
        # collect 는 계정당 ~30분에 한 번 도므로, 매 dry collect 마다 floor 1스텝 하강해야
        # 몇 시간 안에 floor 25 까지 도달(레벨 climb 대기 제거).
        if last_added >= _FILL_FRESH_OK:
            # 생산적 — floor 를 서서히 복원(관련성 회복), dry 리셋.
            dry_streak = 0
            floor = min(_FILL_FLOOR_START, floor + 5)
            note = f"productive added={last_added} floor={floor}"
        else:
            dry_streak += 1
            if dry_streak >= _FILL_DRY_ESCALATE_AFTER and floor > _FILL_FLOOR_MIN:
                # 2회 이상 연속 마름 → 매 dry 라운드 floor 5씩 하강(50→25).
                floor = max(_FILL_FLOOR_MIN, floor - 5)
                note = f"dry x{dry_streak} → floor↓ {floor}"
            else:
                note = f"dry added={last_added} streak={dry_streak} floor={floor}"

        level = (_FILL_FLOOR_START - floor) // 5   # 0(floor50)..5(floor25) 표시용
        pool.set_escalation(cid, level, dry_streak, floor, last_added, note)
        # 마르기 시작하면(또는 floor 하강 중) LLM 앵글 topup 발동.
        run_topup = (dry_streak >= 1) or (floor < _FILL_FLOOR_START)
        return {"level": level, "dry_streak": dry_streak, "relevance_floor": floor,
                "headroom": headroom, "run_topup": run_topup, "note": note,
                "next_lever": _fill_next_lever(level)}
    except Exception as e:
        logger.warning(f"[pool/fill] escalation decide 실패 cid={cid}: {e}")
        return {"level": 0, "dry_streak": 0, "relevance_floor": _FILL_FLOOR_START,
                "headroom": None, "run_topup": None, "note": f"error:{e}",
                "next_lever": _fill_next_lever(0)}


async def _run_pool_workers_for_accounts(pairs: List[Tuple[int, int]]):
    """B 시나리오 — (user_id, customer_id) 페어별로 collect+register+inspect.
    한 사용자 여러 광고주를 가진 경우 광고주마다 독립적으로 워커 실행.
    inspect 는 register 와 분리 실행 — register 의 pending=0 early return 누수 차단."""
    pool = get_keyword_pool_db()
    AI_TOPUP_TRIGGER_THRESHOLD = 10  # 컨트롤러 실패 시 폴백용 — 직전 collect added < N 이면 topup
    for uid, cid in pairs:
        try:
            await _run_pool_collect(uid, customer_id=cid)
        except Exception as e:
            logger.error(f"[pool/run] collect 실패 user={uid} cid={cid}: {e}", exc_info=True)
            try:
                pool.record_run(uid, cid, "collect", "failed",
                                error_message=f"{type(e).__name__}: {str(e)[:300]}")
            except Exception:
                pass

        # 채우기 에스컬레이션 컨트롤러 — 레벨 갱신 + 이번 tick 오퍼레이터 결정.
        # LLM 앵글 topup(cooldown/daily cap 내장)을 레벨 기반으로 발동. 컨트롤러 실패 시
        # 기존 "마른 우물(<10)" 휴리스틱으로 폴백해 안전망 유지.
        decision = await _fill_escalation_decide(uid, cid)
        do_topup = decision.get("run_topup")
        if do_topup is None:
            try:
                recent = pool.recent_runs(cid, limit=3) or []
                lc = next((r for r in recent if r.get("kind") == "collect"), None)
                do_topup = (int(lc.get("added") or 0) if lc else 0) < AI_TOPUP_TRIGGER_THRESHOLD
            except Exception:
                do_topup = False
        if do_topup:
            try:
                logger.warning(
                    f"[pool/fill] uid={uid} cid={cid} level={decision.get('level')} "
                    f"floor={decision.get('relevance_floor')} headroom={decision.get('headroom')} "
                    f"→ LLM 앵글 topup ({decision.get('note')})"
                )
                await _run_pool_ai_seed_topup(uid, cid)
            except Exception as e:
                logger.error(f"[pool/run] ai-topup 실패 user={uid} cid={cid}: {e}", exc_info=True)

        try:
            await _run_pool_register(uid, customer_id=cid)
        except Exception as e:
            logger.error(f"[pool/run] register 실패 user={uid} cid={cid}: {e}", exc_info=True)
            try:
                pool.record_run(uid, cid, "register", "failed",
                                error_message=f"{type(e).__name__}: {str(e)[:300]}")
            except Exception:
                pass
        # inspect 단독 실행 — register 의 early return 경로 (pending=0 / orchestrator 실패)
        # 와 무관하게 매 tick 마다 노출제한 검사 보장.
        try:
            await _run_pool_inspect_only(uid, cid)
        except Exception as e:
            logger.error(f"[pool/run] inspect 실패 user={uid} cid={cid}: {e}", exc_info=True)
            try:
                pool.record_run(uid, cid, "inspect", "failed",
                                error_message=f"{type(e).__name__}: {str(e)[:300]}")
            except Exception:
                pass


# 호환 wrapper — user_id 만 받는 옛 호출자 (외부 cron 등) 위해 유지.
async def _run_pool_workers_for_users(user_ids: List[int]):
    """Deprecated — 가장 최근 광고주만 처리. _run_pool_workers_for_accounts 사용 권장."""
    from database.naver_ad_db import list_ad_accounts_for_user
    pairs: List[Tuple[int, int]] = []
    for uid in user_ids:
        try:
            accounts = list_ad_accounts_for_user(uid) or []
            for a in accounts:
                if a.get("is_connected"):
                    pairs.append((uid, int(a["customer_id"])))
        except Exception as e:
            logger.error(f"[pool/run] list_ad_accounts_for_user 실패 user={uid}: {e}")
    await _run_pool_workers_for_accounts(pairs)


@router.post("/keyword-pool/ai-cleanup-registered")
async def keyword_pool_ai_cleanup_registered(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = Query(None),
    dry_run: bool = Query(True, description="True 면 GPT 분류만, False 면 실제 DELETE"),
    batch_size: int = Query(200),
    max_kws: int = Query(1000),
    incremental_minutes: Optional[int] = Query(None),
):
    """등록 KW AI 의미 분류 cleanup — 점수 인플레 우회.

    GPT-4o-mini 가 user_seed 와 등록 KW 를 도메인 비교 → 무관 KW 만 실제 네이버 DELETE.
    """
    if not customer_id:
        raise HTTPException(status_code=400, detail="customer_id 필요")
    try:
        cid = int(customer_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="customer_id 정수 필요")
    return await _run_pool_ai_cleanup_registered(
        user_id, cid,
        dry_run=dry_run, batch_size=batch_size,
        max_kws=max_kws, incremental_minutes=incremental_minutes,
    )


@router.get("/keyword-pool/diagnostics/off-domain-audit")
async def keyword_pool_off_domain_audit(
    customer_id: int = Query(..., description="광고주 customer_id"),
    sample: int = Query(40, ge=0, le=200000),
):
    """소잠 등 피부/한방 광고주의 off-domain(무관) 키워드 감지 (인증 없음, 진단).
    판정: 피부/한방/의료 on-domain 토큰이 하나도 없거나, 명백한 off-domain 토큰 포함 → off-domain."""
    import sqlite3 as _sq
    from database.registered_keywords_db import get_registered_keywords_db
    ON = ["피부","여드름","아토피","건선","탈모","습진","두드러기","한포진","지루","모공","각질","색소",
        "홍조","기미","주근깨","사마귀","무좀","비듬","두피","손톱","발톱","비립종","쥐젖","흉터","튼살",
        "임신선","셀룰라이트","피지","트러블","각화","백반","대상포진","수족구","농가진","모낭","켈로이드",
        "주사비","딸기코","자반","백선","완선","어루러기","티눈","혈관종","다한","액취","질환","피부염",
        "가려","발진","진물","딱지","한의원","한방","한약","약침","봉독","뜸","부항","체질","면역","해독",
        "어혈","혈허","음허","의원","피부과","치료","증상","원인","완치","아토피","두드러기","건선","탈모"]
    OFF = ["대출","자금","융자","대부","캐피탈","저축은행","포스터","인쇄","창고","전세","월세","디톡스",
        "축구","야구","가수","배우","연예인","주식","코인","부동산","자동차","보험","여행","호텔","맛집",
        "레시피","게임","영화","드라마","웹툰","경락대출","사업자대출","법인","제조","공장","프랜차이즈",
        "상담센터","심리상담","언어치료","마음상담","분양","오피스텔","상가주택","채용","알바","자격증","명함"]
    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword FROM registered_keywords WHERE account_customer_id=? AND removed_at IS NULL",
            (customer_id,),
        ).fetchall()
    total = len(rows)
    off = []
    off_by_neg = 0
    off_no_on = 0
    for (kw,) in rows:
        t = (kw or "").replace(" ", "")
        if any(n in t for n in OFF):
            off.append(kw); off_by_neg += 1
        elif not any(o in t for o in ON):
            off.append(kw); off_no_on += 1
    return {
        "success": True, "customer_id": customer_id, "total": total,
        "off_domain_total": len(off),
        "off_by_negative": off_by_neg, "off_no_ondomain_token": off_no_on,
        "samples": off[:sample],
    }


_SOJAM_ON_TOKENS = ["피부","여드름","아토피","건선","탈모","습진","두드러기","한포진","지루","모공","각질",
    "색소","홍조","기미","주근깨","사마귀","무좀","비듬","두피","손톱","발톱","비립종","쥐젖","흉터","튼살",
    "임신선","셀룰라이트","피지","트러블","각화","백반","대상포진","수족구","농가진","모낭","켈로이드","주사비",
    "딸기코","자반","백선","완선","어루러기","티눈","혈관종","다한","액취","질환","피부염","가려","발진","진물",
    "딱지","한의원","한방","한약","약침","봉독","뜸","부항","체질","면역","해독","어혈","혈허","음허","의원",
    "피부과","치료","증상","원인","완치"]
_SOJAM_OFF_TOKENS = ["대출","자금","융자","대부","캐피탈","저축은행","포스터","인쇄","창고","전세","월세","디톡스",
    "축구","야구","가수","배우","연예인","아이돌","주식","코인","선물","재테크","부동산","청약","자동차","중고차",
    "오토바이","바이크","노트북","핸드폰","가전","쇼핑몰","의류","신발","운동화","가방","시계","세이코","롤렉스",
    "귀금속","명품","보험","여행","항공권","항공","호텔","리조트","펜션","게스트하우스","홋카이도","북해도","도쿄",
    "오사카","보라카이","골프","낚시","캠핑","등산","맛집","레시피","술안주","와인","맥주","게임","영화","드라마",
    "웹툰","만화","노래","이용권","웨이브","넷플릭스","경락대출","사업자대출","법인","제조","공장","프랜차이즈",
    "상담센터","심리상담","언어치료","마음상담","분양","오피스텔","상가주택","채용","알바","자격증","명함",
    "수국","우엉차","홍삼","발포비타민","츄어블","보조제","드링크",
    "건조기","화상회의","스카이레이크","배구공","레드페퍼","LPG가스","힐스테이트","파크드림","위시카드","패스카드",
    "노니환","숀리","보자기코리아","월풀","K패스","옥수수수염차","마이위시","상업용건조","업소용건조",
    "숙소","숙박","모텔","민박","글램핑","캠핑장","풀빌라","스테이","개선문","에펠탑","관광","관광지",
    "여행지","가볼만한곳","투어","코스요리","데이트코스","해수욕장","스키장","워터파크",
    "놀이공원","박물관","전시회","공연","아파트","빌라","원룸","월세방","자취방","고시원","독서실","렌터카",
    "주차장","대리운전","택배","꽃배달","에어컨","냉장고","세탁기","청소기","근사숙소"]


class OffDomainCleanupRequest(BaseModel):
    dry_run: bool = Field(True, description="true: 개수만, false: 실제 삭제")
    extra_on: Optional[List[str]] = Field(None, description="on-domain 으로 칠 추가 토큰")
    extra_off: Optional[List[str]] = Field(None, description="off-domain 으로 칠 추가 토큰")
    keep_tokens: Optional[List[str]] = Field(None, description="설정 시: 이 토큰을 '하나도' 포함 안 한 키워드를 삭제(소잠 OFF 베이스 무시). 메디론 드리프트 정밀 제거용")
    max_delete: int = Field(30000, description="안전 상한")
    strict_off: bool = Field(False, description="true: 소잠 OFF base 무시하고 extra_off 만 사용 (다른 한방 광고주의 정밀 정리용 — 골프/건조기 등 도메인별 false positive 차단)")


@router.post("/keyword-pool/off-domain-cleanup")
async def keyword_pool_off_domain_cleanup(
    request: OffDomainCleanupRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """off-domain(무관) 키워드 일괄 삭제 — on-domain 토큰 없거나 off 토큰 포함 키워드를 네이버에서 삭제.
    dry_run=true: 대상 개수. false: 백그라운드 bulk delete (DELETE /ncc/keywords?ids=)."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    ON = _SOJAM_ON_TOKENS + (request.extra_on or [])
    # strict_off=True: 소잠 OFF base 무시(다른 한방 광고주의 도메인별 false positive 차단).
    # 예: 위례해오름 "골프엘보" 한방치료는 소잠 OFF 의 "골프" 토큰에 잘못 걸림.
    OFF = (list(request.extra_off or []) if request.strict_off
           else _SOJAM_OFF_TOKENS + (request.extra_off or []))
    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND removed_at IS NULL",
            (cid,),
        ).fetchall()
    # keep_tokens 모드: anchor 토큰을 하나도 안 가진 키워드 삭제(드리프트 정밀 제거). 소잠 OFF 베이스 무시.
    _keep = request.keep_tokens or []
    targets = []  # (keyword, ncc_id)
    if _keep:
        for kw, nid in rows:
            t = (kw or "").replace(" ", "")
            if not any(k in t for k in _keep):
                targets.append((kw, nid))
    else:
        # 안전: 명시적 OFF 토큰 포함만 삭제 (no-on-domain 휴리스틱은 희귀질환 오삭제 위험 → 미사용).
        for kw, nid in rows:
            t = (kw or "").replace(" ", "")
            if any(n in t for n in OFF):
                targets.append((kw, nid))
    targets = targets[:request.max_delete]
    if request.dry_run:
        return {"success": True, "dry_run": True, "customer_id": cid,
                "scanned": len(rows), "delete_targets": len(targets),
                "samples": [t[0] for t in targets[:40]]}

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        ids = [nid for _, nid in targets if nid]
        logger.warning(f"[off-domain-cleanup] 삭제 시작 — {len(ids)}개")
        done = 0; failed = 0
        for i in range(0, len(ids), 100):
            batch = ids[i:i+100]
            try:
                await client.delete_keywords_bulk(batch)
                done += len(batch)
            except Exception as e:
                failed += len(batch)
                logger.warning(f"[off-domain-cleanup] batch 실패: {str(e)[:80]}")
            if done % 2000 < 100:
                logger.warning(f"[off-domain-cleanup] 진행 — 삭제 {done}/{len(ids)} (실패 {failed})")
            await asyncio.sleep(0.1)
        try:
            reg.mark_removed(cid, [kw for kw, _ in targets])
        except Exception:
            pass
        logger.warning(f"[off-domain-cleanup] 완료 — 삭제 {done} / 실패 {failed}")
    background_tasks.add_task(_run)
    return {"success": True, "started": True, "delete_targets": len(targets),
            "message": f"off-domain {len(targets)}개 삭제 백그라운드 시작 (로그 확인)"}


class BulkPauseOffdomainRequest(BaseModel):
    loan_tokens: List[str] = Field(..., description="금융/대출 토큰. keep 후보는 이 중 하나 이상 포함해야 함")
    domain_tokens: List[str] = Field(..., description="의료/소상공인 주체 토큰. keep 후보는 이 중 하나 이상 포함해야 함")
    exclude_tokens: List[str] = Field(default_factory=list, description="부정 토큰. 이 중 하나라도 포함하면 keep에서 제외(off). 소상공인/지원금(grant)/브랜드 오매칭 제거용")
    keywords: Optional[List[str]] = Field(None, description="명시적 키워드 리스트. 설정 시 토큰 로직 무시하고 '정확히 이 키워드들'만 off(activate면 재개). 정밀 정리용")
    dry_run: bool = Field(True, description="true: off 대상 개수+샘플, false: 실제 bulk userLock=true")
    max_pause: int = Field(200000, description="안전 상한")
    activate: bool = Field(False, description="true: keep(loan AND domain AND NOT exclude)을 userLock=false 로 재개(복구). 비keep은 건드리지 않음")
    keep_campaign_prefixes: Optional[List[str]] = Field(None, description="이 prefix 로 시작하는 캠페인명의 키워드는 토큰 판정과 무관하게 절대 off/재개 대상에서 제외(보호). 예: ['제휴_'] — PG/지급대행 별도 라인 보호용")


@router.post("/keyword-pool/registered/bulk-pause-offdomain")
async def keyword_pool_bulk_pause_offdomain(
    request: BulkPauseOffdomainRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """도메인 밖 키워드 **bulk 일시정지(off)** — keep=(loan_tokens 중 1+ AND domain_tokens 중 1+).
    keep 아닌 키워드를 userLock=true 로 일괄 off (삭제 아님, 되돌릴 수 있음). bulk PUT 100개/콜.
    activate=true 면 반대로 keep 을 userLock=false 로 재개(복구). dry_run 으로 대상 미리보기."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    loan = [t for t in (request.loan_tokens or []) if t and t.strip()]
    dom = [t for t in (request.domain_tokens or []) if t and t.strip()]
    excl = [t for t in (request.exclude_tokens or []) if t and t.strip()]
    if not loan or not dom:
        raise HTTPException(status_code=400, detail="loan_tokens 와 domain_tokens 둘 다 필요")
    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path, timeout=30.0) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id, campaign_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL "
            "AND removed_at IS NULL",
            (cid,),
        ).fetchall()

    # 캠페인 prefix 보호: 지정 prefix 로 시작하는 캠페인의 키워드는 대상에서 완전 제외
    protected_campaign_ids: set = set()
    prefixes = [p for p in (request.keep_campaign_prefixes or []) if p and p.strip()]
    if prefixes:
        _pc = NaverAdApiClient()
        _pc.customer_id = account["customer_id"]; _pc.api_key = account["api_key"]; _pc.secret_key = account["secret_key"]
        try:
            camps = await _pc.get_campaigns()
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"캠페인 조회 실패(prefix 보호 불가): {str(e)[:120]}")
        for c in (camps or []):
            nm = (c.get("name") or "")
            if any(nm.startswith(p) for p in prefixes):
                protected_campaign_ids.add(c.get("nccCampaignId"))
    if protected_campaign_ids:
        rows = [r for r in rows if r[3] not in protected_campaign_ids]
    rows = [(r[0], r[1], r[2]) for r in rows]

    def _is_keep(kw: str) -> bool:
        t = (kw or "").replace(" ", "")
        if excl and any(x in t for x in excl):
            return False
        return any(l in t for l in loan) and any(d in t for d in dom)

    explicit = None
    if request.keywords:
        explicit = {k.strip() for k in request.keywords if k and k.strip()}

    if explicit is not None:
        # 명시적 리스트 모드 — 정확히 이 키워드들만 대상(토큰 로직 무시)
        sel = [(kw, nid, gid) for kw, nid, gid in rows if (kw or "").strip() in explicit]
        if request.activate:
            verb = "재개(userLock=false)"; lock_val = False
        else:
            verb = "일시정지(userLock=true)"; lock_val = True
    elif request.activate:
        # 복구 모드 — keep 만 재개
        sel = [(kw, nid, gid) for kw, nid, gid in rows if _is_keep(kw)]
        verb = "재개(userLock=false)"
        lock_val = False
    else:
        # off 모드 — 비keep 일시정지
        sel = [(kw, nid, gid) for kw, nid, gid in rows if not _is_keep(kw)]
        verb = "일시정지(userLock=true)"
        lock_val = True
    sel = sel[: request.max_pause]

    if request.dry_run:
        return {
            "success": True, "dry_run": True, "customer_id": cid, "mode": verb,
            "scanned": len(rows), "targets": len(sel),
            "protected_campaigns": len(protected_campaign_ids),
            "samples": [s[0] for s in sel[:40]],
        }

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        logger.warning(f"[bulk-pause-offdomain] {verb} 시작 — {len(sel)}개")
        done = 0; failed = 0
        for i in range(0, len(sel), 100):
            batch = sel[i:i + 100]
            items = [{"nccKeywordId": nid, "nccAdgroupId": gid, "userLock": lock_val}
                     for _, nid, gid in batch]
            try:
                await client.set_keywords_userlock_bulk(items)
                done += len(items)
            except Exception as e:
                failed += len(items)
                logger.warning(f"[bulk-pause-offdomain] batch 실패: {str(e)[:80]}")
            if done % 5000 < 100:
                logger.warning(f"[bulk-pause-offdomain] 진행 — {done}/{len(sel)} (실패 {failed})")
            await asyncio.sleep(0.1)
        logger.warning(f"[bulk-pause-offdomain] 완료 — {verb} {done} / 실패 {failed}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "mode": verb, "targets": len(sel),
            "message": f"{len(sel)}개 {verb} 백그라운드 시작 (bulk userLock 100/콜, 로그 확인)"}


class SetUserLockByIdsRequest(BaseModel):
    """ncc ID 를 직접 지정해 userLock 변경. registered_keywords DB 에 없는
    **수동 등록 키워드**를 다루기 위한 경로 — bulk-pause-offdomain 은 DB 기반이라
    레거시 캠페인(소잠 기준 광고비의 다수)에 닿지 않는다."""
    items: List[Dict[str, str]] = Field(..., description="[{keyword_id, group_id}, ...]")
    lock: bool = Field(True, description="true=일시정지(userLock), false=재개")
    dry_run: bool = Field(True)


@router.post("/keyword-pool/registered/set-userlock-by-ids")
async def keyword_pool_set_userlock_by_ids(
    request: SetUserLockByIdsRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """지정한 ncc 키워드 ID 들을 일시정지/재개. 삭제 아님 — lock=false 로 되돌린다."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    pairs = [(it.get("keyword_id"), it.get("group_id")) for it in (request.items or [])]
    pairs = [(k, g) for k, g in pairs if k and g]
    if not pairs:
        raise HTTPException(status_code=400, detail="keyword_id/group_id 쌍이 필요")

    verb = "일시정지(userLock=true)" if request.lock else "재개(userLock=false)"
    if request.dry_run:
        return {"success": True, "dry_run": True, "customer_id": cid,
                "mode": verb, "targets": len(pairs),
                "samples": [p[0] for p in pairs[:10]]}

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]
        done = failed = 0
        logger.warning(f"[set-userlock-by-ids] {verb} 시작 — {len(pairs)}개 (cid={cid})")
        for i in range(0, len(pairs), 100):
            batch = pairs[i:i + 100]
            items = [{"nccKeywordId": k, "nccAdgroupId": g, "userLock": request.lock}
                     for k, g in batch]
            try:
                await client.set_keywords_userlock_bulk(items)
                done += len(items)
            except Exception as e:
                failed += len(items)
                logger.warning(f"[set-userlock-by-ids] batch 실패: {str(e)[:120]}")
            await asyncio.sleep(0.1)
        logger.warning(f"[set-userlock-by-ids] 완료 — {verb} {done} / 실패 {failed}")
        try:
            await client.close()
        except Exception:
            pass

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "customer_id": cid,
            "mode": verb, "targets": len(pairs),
            "message": f"{len(pairs)}개 {verb} 백그라운드 시작"}


@router.get("/keyword-pool/diagnostics/keyword-inspect")
async def keyword_pool_keyword_inspect(
    group_ids: str = Query(..., description="광고그룹 ID 쉼표구분 (최대 20)"),
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """광고그룹의 키워드 원본(bidAmt/useGroupBidAmt/userLock)을 그대로 조회 — 변경 검증용.
    기존 /api/naver-ad/keywords 는 get_optimizer() 의 전역 자격증명을 써서 광고주 계정에
    접근하지 못한다(500). 여기서는 _resolve_account 로 해당 광고주 키를 쓴다."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    gids = [g.strip() for g in (group_ids or "").split(",") if g.strip()][:20]
    if not gids:
        raise HTTPException(status_code=400, detail="group_ids 필요")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]
    out: Dict[str, Any] = {}
    for gid in gids:
        try:
            res = await client._request("GET", "/ncc/keywords", {"nccAdgroupId": gid})
            rows = res if isinstance(res, list) else (res.get("data") or [])
            out[gid] = [{"keyword": k.get("keyword"), "bidAmt": k.get("bidAmt"),
                         "useGroupBidAmt": k.get("useGroupBidAmt"),
                         "userLock": k.get("userLock"),
                         "status": k.get("status")} for k in rows]
        except Exception as e:
            out[gid] = {"error": f"{type(e).__name__}: {str(e)[:150]}"}
    try:
        await client.close()
    except Exception:
        pass
    return {"success": True, "customer_id": int(account.get("customer_id")), "groups": out}


class SetBidByIdsRequest(BaseModel):
    """ncc ID 직접 지정 입찰 변경 — registered_keywords DB 에 없는 **수동 등록 키워드**용.
    bulk-bid-by-tokens 는 DB 기반이라 레거시 캠페인에 닿지 않는다."""
    items: List[Dict[str, str]] = Field(..., description="[{keyword_id, group_id}, ...]")
    bid: int = Field(..., ge=70, le=100000, description="설정할 입찰가(원)")
    dry_run: bool = Field(True)


@router.post("/keyword-pool/registered/set-bid-by-ids")
async def keyword_pool_set_bid_by_ids(
    request: SetBidByIdsRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """지정한 ncc 키워드 ID 들의 입찰가를 설정. useGroupBidAmt=False 를 함께 걸어
    그룹 기본입찰 대신 키워드 입찰이 실제로 적용되게 한다."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    pairs = [(it.get("keyword_id"), it.get("group_id")) for it in (request.items or [])]
    pairs = [(k, g) for k, g in pairs if k and g]
    if not pairs:
        raise HTTPException(status_code=400, detail="keyword_id/group_id 쌍이 필요")
    new_bid = max(70, int(round(request.bid / 10.0)) * 10)

    if request.dry_run:
        return {"success": True, "dry_run": True, "customer_id": cid,
                "bid": new_bid, "targets": len(pairs),
                "samples": [p[0] for p in pairs[:10]]}

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]
        done = failed = 0
        logger.warning(f"[set-bid-by-ids] bid={new_bid} 시작 — {len(pairs)}개 (cid={cid})")
        for i in range(0, len(pairs), 100):
            batch = pairs[i:i + 100]
            items = [{"nccKeywordId": k, "nccAdgroupId": g,
                      "bidAmt": new_bid, "useGroupBidAmt": False} for k, g in batch]
            try:
                await client.update_keywords_bid_bulk(items)
                done += len(items)
            except Exception as e:
                failed += len(items)
                logger.warning(f"[set-bid-by-ids] batch 실패: {str(e)[:120]}")
            await asyncio.sleep(0.1)
        logger.warning(f"[set-bid-by-ids] 완료 — bid={new_bid} {done} / 실패 {failed}")
        try:
            await client.close()
        except Exception:
            pass

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "customer_id": cid,
            "bid": new_bid, "targets": len(pairs),
            "message": f"{len(pairs)}개 입찰가 {new_bid}원 설정 백그라운드 시작"}


class KeepVolumeAuditRequest(BaseModel):
    loan_tokens: List[str] = Field(..., description="금융/대출 토큰. keep=이 중 1+ 포함")
    domain_tokens: List[str] = Field(..., description="도메인(의료 등) 토큰. keep=이 중 1+ 포함")
    exclude_tokens: List[str] = Field(default_factory=list, description="부정 토큰. 이 중 1+ 포함 시 keep 제외")
    min_volume: int = Field(1, description="실검색량 하한. monthly_total >= 이 값이면 '검색량 있음'")
    check_live: bool = Field(True, description="true: 네이버에서 adgroup별 keyword 조회해 userLock(on/off) 실측")
    max_groups: int = Field(2000, description="실측 시 조회할 adgroup 상한(안전)")
    include_keywords: bool = Field(False, description="true: keep_on_with_volume(또는 check_live=false면 keep_with_volume) 키워드 목록을 [{keyword, monthly_total}]로 반환")


@router.post("/keyword-pool/registered/keep-volume-audit")
async def keyword_pool_keep_volume_audit(
    request: KeepVolumeAuditRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """keep(loan AND domain) 키워드 중 **실검색량 보유 + 현재 ON** 개수 집계.
    검색량=naverad_keyword_pool.monthly_total, ON=네이버 userLock=false(실측, check_live)."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    from database.keyword_pool_db import get_keyword_pool_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    loan = [t for t in (request.loan_tokens or []) if t and t.strip()]
    dom = [t for t in (request.domain_tokens or []) if t and t.strip()]
    excl = [t for t in (request.exclude_tokens or []) if t and t.strip()]
    if not loan or not dom:
        raise HTTPException(status_code=400, detail="loan_tokens 와 domain_tokens 둘 다 필요")

    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path, timeout=30.0) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL "
            "AND removed_at IS NULL",
            (cid,),
        ).fetchall()

    def _is_keep(kw: str) -> bool:
        t = (kw or "").replace(" ", "")
        if excl and any(x in t for x in excl):
            return False
        return any(l in t for l in loan) and any(d in t for d in dom)

    keep = [(kw, nid, gid) for kw, nid, gid in rows if _is_keep(kw)]

    # 검색량 — pool 테이블에서 keep 키워드 monthly_total 로드
    pool = get_keyword_pool_db()
    vol = {}
    keep_kws = list({kw for kw, _, _ in keep})
    with _sq.connect(pool.db_path, timeout=30.0) as conn:
        for i in range(0, len(keep_kws), 500):
            chunk = keep_kws[i:i + 500]
            ph = ",".join("?" * len(chunk))
            for kw, mt in conn.execute(
                f"SELECT keyword, monthly_total FROM naverad_keyword_pool "
                f"WHERE account_customer_id=? AND keyword IN ({ph})",
                [cid, *chunk],
            ).fetchall():
                vol[kw] = mt or 0

    keep_total = len(keep)
    keep_with_vol = sum(1 for kw, _, _ in keep if vol.get(kw, 0) >= request.min_volume)
    no_vol_data = sum(1 for kw, _, _ in keep if kw not in vol)

    result = {
        "success": True, "customer_id": cid,
        "keep_total": keep_total,
        "keep_with_volume": keep_with_vol,
        "min_volume": request.min_volume,
        "keep_no_pool_volume_data": no_vol_data,
        "live_checked": False,
    }

    if request.include_keywords and not request.check_live:
        lst = [{"keyword": kw, "monthly_total": vol.get(kw, 0)}
               for kw, _, _ in keep if vol.get(kw, 0) >= request.min_volume]
        lst.sort(key=lambda x: x["monthly_total"], reverse=True)
        result["keywords"] = lst

    if request.check_live:
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        groups = list({gid for _, _, gid in keep})[: request.max_groups]
        lock_by_id = {}
        scanned = 0; gerr = 0
        for gid in groups:
            try:
                kws = await client.get_keywords(ad_group_id=gid)
                for k in (kws or []):
                    nid = k.get("nccKeywordId")
                    if nid:
                        lock_by_id[nid] = bool(k.get("userLock"))
                scanned += 1
            except Exception:
                gerr += 1
            await asyncio.sleep(0.05)
        keep_on = 0; keep_on_vol = 0; unknown = 0
        on_vol_list = []
        for kw, nid, _ in keep:
            st = lock_by_id.get(nid)
            if st is None:
                unknown += 1; continue
            if st is False:  # userLock=false → ON
                keep_on += 1
                if vol.get(kw, 0) >= request.min_volume:
                    keep_on_vol += 1
                    if request.include_keywords:
                        on_vol_list.append({"keyword": kw, "monthly_total": vol.get(kw, 0)})
        result.update({
            "live_checked": True,
            "groups_total": len({gid for _, _, gid in keep}),
            "groups_scanned": scanned, "groups_error": gerr,
            "keep_on": keep_on,
            "keep_on_with_volume": keep_on_vol,
            "keep_status_unknown": unknown,
        })
        if request.include_keywords:
            on_vol_list.sort(key=lambda x: x["monthly_total"], reverse=True)
            result["keywords"] = on_vol_list
    return result


class LiveOffdomainScanRequest(BaseModel):
    off_tokens: List[str] = Field(default_factory=list, description="명백한 무관 토큰. 키워드가 이 중 1+ 포함하면 무관 후보")
    keep_tokens: List[str] = Field(default_factory=list, description="보호 토큰. 무관 후보라도 이 중 1+ 포함하면 제외(오탐 방지)")
    no_ondomain_mode: bool = Field(False, description="true: off_tokens 대신 'keep_tokens(온도메인)을 하나도 안 가진' 키워드를 후보로. keep_tokens=온도메인 토큰")
    min_volume: int = Field(0, description="pool monthly_total 하한. 이 값 이상만 반환/집계")
    max_groups: int = Field(4000, description="라이브 조회할 adgroup 상한(안전)")
    sample: int = Field(80, ge=0, le=5000)


@router.post("/keyword-pool/registered/live-offdomain-scan")
async def keyword_pool_live_offdomain_scan(
    request: LiveOffdomainScanRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """**현재 ON(userLock=false) 상태인 무관 키워드**만 라이브 실측으로 골라낸다 (읽기 전용, 아무것도 안 바꿈).
    후보 선정: off_tokens 포함(기본) 또는 no_ondomain_mode=true 면 keep_tokens 하나도 없음.
    keep_tokens 는 오탐 방지 보호. 후보가 든 adgroup만 네이버 라이브 조회 → userLock 실측 → ON 만 반환.
    검색량(pool monthly_total) 조인해 큰 것부터 정렬."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    from database.keyword_pool_db import get_keyword_pool_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    off = [t for t in (request.off_tokens or []) if t and t.strip()]
    keep = [t for t in (request.keep_tokens or []) if t and t.strip()]
    if request.no_ondomain_mode:
        if not keep:
            raise HTTPException(status_code=400, detail="no_ondomain_mode 는 keep_tokens(온도메인) 필요")
    elif not off:
        raise HTTPException(status_code=400, detail="off_tokens 필요 (또는 no_ondomain_mode=true)")

    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path, timeout=30.0) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL "
            "AND removed_at IS NULL",
            (cid,),
        ).fetchall()

    def _is_candidate(kw: str) -> bool:
        t = (kw or "").replace(" ", "")
        if keep and any(k in t for k in keep):
            return False
        if request.no_ondomain_mode:
            return True  # keep(온도메인) 토큰 하나도 없음 = 후보
        return any(o in t for o in off)

    cand = [(kw, nid, gid) for kw, nid, gid in rows if _is_candidate(kw)]

    # 검색량 로드
    pool = get_keyword_pool_db()
    vol = {}
    cand_kws = list({kw for kw, _, _ in cand})
    with _sq.connect(pool.db_path, timeout=30.0) as conn:
        for i in range(0, len(cand_kws), 500):
            chunk = cand_kws[i:i + 500]
            ph = ",".join("?" * len(chunk))
            for kw, mt in conn.execute(
                f"SELECT keyword, monthly_total FROM naverad_keyword_pool "
                f"WHERE account_customer_id=? AND keyword IN ({ph})",
                [cid, *chunk],
            ).fetchall():
                vol[kw] = mt or 0

    # 후보가 든 adgroup만 라이브 조회 → userLock 실측
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
    groups = list({gid for _, _, gid in cand})[: request.max_groups]
    lock_by_id = {}
    scanned = 0; gerr = 0
    for gid in groups:
        try:
            kws = await client.get_keywords(ad_group_id=gid)
            for k in (kws or []):
                nid = k.get("nccKeywordId")
                if nid:
                    lock_by_id[nid] = bool(k.get("userLock"))
            scanned += 1
        except Exception:
            gerr += 1
        await asyncio.sleep(0.03)

    on_list = []; off_cnt = 0; unknown = 0
    for kw, nid, _ in cand:
        st = lock_by_id.get(nid)
        if st is None:
            unknown += 1; continue
        if st is False:  # userLock=false → ON
            on_list.append({"keyword": kw, "monthly_total": vol.get(kw, 0)})
        else:
            off_cnt += 1
    # 검색량 하한 적용 + 정렬
    on_flt = [x for x in on_list if x["monthly_total"] >= request.min_volume]
    on_flt.sort(key=lambda x: x["monthly_total"], reverse=True)

    return {
        "success": True, "customer_id": cid,
        "registered_total": len(rows),
        "candidates": len(cand),
        "groups_total": len({gid for _, _, gid in cand}),
        "groups_scanned": scanned, "groups_error": gerr,
        "candidate_on": len(on_list),
        "candidate_off_already": off_cnt,
        "candidate_status_unknown": unknown,
        "min_volume": request.min_volume,
        "on_with_min_volume": len(on_flt),
        "samples": on_flt[: request.sample],
    }


class BulkRankBidRequest(BaseModel):
    """키워드별 중요도 점수 → 개별 PC 목표순위 → 순위별 estimate 입찰. 가중치는 옵션(기본값 내장)."""
    geo_top: List[str] = Field(default_factory=list, description="최상위 지역(강남권). 비우면 기본")
    geo_adj: List[str] = Field(default_factory=list, description="인접 지역(서초권). 비우면 기본")
    geo_other: List[str] = Field(default_factory=list, description="타지역(감점). 비우면 기본")
    diseases: List[str] = Field(default_factory=list)
    brand: List[str] = Field(default_factory=list)
    device: str = "PC"
    bid_cap: int = Field(2000, description="입찰 상한(원) — 초경쟁 키워드가 예산 독식 방지. estimate가 이보다 크면 cap")
    dry_run: bool = True
    max_keywords: int = 200000


@router.post("/keyword-pool/registered/bulk-rank-bid")
async def keyword_pool_bulk_rank_bid(
    request: BulkRankBidRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """키워드마다 중요도 점수를 매겨 개별 PC 목표순위를 정하고, 그 순위 노출 입찰가(estimate)를 설정.
    점수 = 지역(강남+40/인접+25/타지역-25) + 의도(예약/상담+30, 비용/가격+25, 추천/후기/명의+20,
    한의원/병원+15, 치료/한약+8, 정보성-25) + 질환+10 + 브랜드+50.
    점수→순위: ≥75→1, 60→2, 48→3, 38→5, 28→7, 18→10, 미만→floor(미적용). dry_run=점수/순위 분포."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    _dev = (request.device or "PC").upper()
    if _dev not in ("PC", "MOBILE"):
        _dev = "PC"

    GEO_TOP = request.geo_top or ["강남", "역삼", "논현", "신사", "청담", "압구정", "선릉", "대치", "학동", "신논현", "강남구", "양재", "도곡"]
    GEO_ADJ = request.geo_adj or ["서초", "방배", "잠원", "개포", "수서", "일원", "세곡", "우면"]
    GEO_OTHER = request.geo_other or [
        "위례", "동탄", "분당", "판교", "수지", "기흥", "일산", "평촌", "산본", "범계", "부천", "안산", "상록수", "본오동", "광명", "시흥",
        "인천", "부평", "송도", "부산", "해운대", "서면", "대구", "수성", "범어", "만촌", "광주", "대전", "울산", "세종", "수원", "성남",
        "용인", "고양", "천안", "아산", "청주", "흥덕", "전주", "익산", "군산", "포항", "경주", "구미", "창원", "김해", "양산", "진주",
        "목포", "여수", "순천", "제주", "춘천", "원주", "강릉", "노원", "송파", "마포", "은평", "구로", "관악", "동작", "성북", "정릉",
        "논산", "백암", "석우", "범계", "산본", "안양", "의정부", "남양주", "구리", "하남", "김포", "파주",
    ]
    DIS = request.diseases or [
        "아토피", "건선", "여드름", "두드러기", "습진", "지루성", "탈모", "무좀", "대상포진", "사마귀", "백반증", "기미", "모낭염",
        "한포진", "주사비", "다한증", "켈로이드", "피부", "한방", "피부염", "피부질환", "가려움", "곤지름", "헤르페스", "티눈",
        "뾰루지", "구내염", "색소침착", "흉터", "모공", "땀띠", "주근깨", "비듬",
    ]
    BRAND = request.brand or ["소잠"]
    I_BOOK = ["예약", "상담", "문의", "예약문의", "전화상담", "예약하기", "당일"]
    I_COST = ["비용", "가격", "얼마"]
    I_CHOICE = ["추천", "후기", "잘하는곳", "명의", "유명"]
    I_CLINIC = ["한의원", "한방병원", "병원", "의원", "피부과", "클리닉"]
    I_TREAT = ["치료", "한약", "약침", "한방치료", "봉독", "완치", "낫는법"]
    I_INFO = ["증상", "원인", "사진", "이미지", "뜻", "종류", "에좋은", "음식", "민간요법", "전염", "옮나", "초기증상"]

    def _score(kw: str) -> int:
        t = (kw or "").replace(" ", "")
        s = 0
        if any(b in t for b in BRAND): s += 50
        if any(g in t for g in GEO_TOP): s += 40
        elif any(g in t for g in GEO_ADJ): s += 25
        elif any(g in t for g in GEO_OTHER): s -= 25
        if any(i in t for i in I_BOOK): s += 30
        if any(i in t for i in I_COST): s += 25
        if any(i in t for i in I_CHOICE): s += 20
        if any(i in t for i in I_CLINIC): s += 15
        elif any(i in t for i in I_TREAT): s += 8
        if any(i in t for i in I_INFO): s -= 25
        if any(d in t for d in DIS): s += 10
        return s

    def _pos(score: int) -> Optional[int]:
        if score >= 75: return 1
        if score >= 60: return 2
        if score >= 48: return 3
        if score >= 38: return 5
        if score >= 28: return 7
        if score >= 18: return 10
        return None  # floor 유지

    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path, timeout=30.0) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL AND removed_at IS NULL",
            (cid,),
        ).fetchall()
    # 키워드별 점수+목표순위
    by_pos: Dict[int, List[Tuple[str, str, str]]] = {}
    dist: Dict[str, int] = {}
    for kw, nid, gid in rows[: request.max_keywords]:
        p = _pos(_score(kw))
        key = f"pos{p}" if p else "floor"
        dist[key] = dist.get(key, 0) + 1
        if p:
            by_pos.setdefault(p, []).append((kw, nid, gid))

    if request.dry_run:
        samp = {}
        for p, lst in sorted(by_pos.items()):
            samp[f"pos{p}"] = [x[0] for x in lst[:8]]
        return {"success": True, "dry_run": True, "customer_id": cid, "device": _dev,
                "distribution": dict(sorted(dist.items())), "samples": samp}

    # 상위밴드(1/2/3/5위)=실제 estimate 입찰(cap), 하위밴드(7/10위)=저가 flat(estimate stall 회피).
    EST_POS = {1, 2, 3, 5}
    FLAT_POS = {7: 120, 10: 80}
    CAP = max(70, int(request.bid_cap))

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        from collections import defaultdict as _dd
        total_done = 0; total_fail = 0
        for pos in sorted(by_pos.keys()):
            items_kw = by_pos[pos]
            by_gid = _dd(list)
            if pos in EST_POS:
                # 키워드별 해당순위 estimate → cap. 미커버(데이터없음)는 floor 70.
                texts = list({k[0] for k in items_kw if k[0]})
                bidmap: Dict[str, int] = {}
                for cas in [pos, 5, 3, 1, 2]:
                    remaining = [t for t in texts if t not in bidmap]
                    if not remaining:
                        break
                    for i in range(0, len(remaining), 15):
                        try:
                            r = await client.get_avg_position_bids(remaining[i:i + 15], cas, device=_dev)
                            for e in (r.get("estimate") or []):
                                kt, bd = (e.get("keyword") or "").strip(), e.get("bid")
                                if kt and bd and kt not in bidmap:
                                    bidmap[kt] = bd
                        except Exception:
                            pass
                        await asyncio.sleep(0.2)
                for kw, nid, gid in items_kw:
                    bd = bidmap.get((kw or "").strip())
                    nb = min(CAP, max(70, round(int(bd) / 10) * 10)) if bd else 70
                    by_gid[gid].append({"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": nb, "useGroupBidAmt": False})
            else:
                nb = FLAT_POS.get(pos, 70)
                for kw, nid, gid in items_kw:
                    by_gid[gid].append({"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": nb, "useGroupBidAmt": False})
            done = 0; fail = 0
            for gid, its in by_gid.items():
                for i in range(0, len(its), 100):
                    try:
                        await client.update_keywords_bid_bulk(its[i:i + 100]); done += len(its[i:i + 100])
                    except Exception:
                        fail += len(its[i:i + 100])
                    await asyncio.sleep(0.1)
            total_done += done; total_fail += fail
            mode = f"estimate(cap{CAP})" if pos in EST_POS else f"flat{FLAT_POS.get(pos)}"
            logger.warning(f"[bulk-rank-bid] pos={pos} 완료({mode}) — {done}개 / 실패 {fail} (대상 {len(items_kw)})")
        logger.warning(f"[bulk-rank-bid] 전체 완료 — {total_done}개 적용 / 실패 {total_fail}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "customer_id": cid, "device": _dev,
            "distribution": dict(sorted(dist.items())),
            "message": "키워드별 중요도점수→개별 PC목표순위 입찰 백그라운드 시작 (로그 [bulk-rank-bid])"}


class BulkBidByTokensRequest(BaseModel):
    domain_tokens: List[str] = Field(..., description="주체/질환 토큰 — 이 중 1+ 포함해야 대상")
    intent_tokens: List[str] = Field(default_factory=list, description="의향 토큰 — 비우면 domain만, 지정 시 domain AND intent")
    exclude_tokens: List[str] = Field(default_factory=list, description="제외 토큰 — 1+ 포함 시 제외(상위티어 중복 방지)")
    bid: int = Field(..., ge=70, le=100000, description="설정할 입찰가(원). 10원단위 반올림")
    dry_run: bool = Field(True)
    max_keywords: int = Field(200000)


@router.post("/keyword-pool/registered/bulk-bid-by-tokens")
async def keyword_pool_bulk_bid_by_tokens(
    request: BulkBidByTokensRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """토큰 매칭(domain AND intent, NOT exclude) 키워드의 입찰가를 flat 일괄 설정 — 중요도 티어링용.
    estimate 무관하게 전체 매칭분에 확실히 적용. bulk PUT(update_keywords_bid_bulk) 100/콜. dry_run 미리보기."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    dom = [t for t in (request.domain_tokens or []) if t and t.strip()]
    intent = [t for t in (request.intent_tokens or []) if t and t.strip()]
    exc = [t for t in (request.exclude_tokens or []) if t and t.strip()]
    if not dom:
        raise HTTPException(status_code=400, detail="domain_tokens 필요")
    new_bid = max(70, min(100000, round(int(request.bid) / 10) * 10))
    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path, timeout=30.0) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL "
            "AND removed_at IS NULL",
            (cid,),
        ).fetchall()

    def _is_target(kw: str) -> bool:
        t = (kw or "").replace(" ", "")
        if exc and any(x in t for x in exc):
            return False
        if not any(d in t for d in dom):
            return False
        if intent and not any(i in t for i in intent):
            return False
        return True

    sel = [(kw, nid, gid) for kw, nid, gid in rows if _is_target(kw)][: request.max_keywords]
    if request.dry_run:
        return {"success": True, "dry_run": True, "customer_id": cid, "bid": new_bid,
                "scanned": len(rows), "targets": len(sel), "samples": [s[0] for s in sel[:40]]}

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        logger.warning(f"[bulk-bid-by-tokens] bid={new_bid} 시작 — {len(sel)}개")
        done = 0; failed = 0
        for i in range(0, len(sel), 100):
            batch = sel[i:i + 100]
            items = [{"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": new_bid, "useGroupBidAmt": False}
                     for _, nid, gid in batch]
            try:
                await client.update_keywords_bid_bulk(items)
                done += len(items)
            except Exception as e:
                failed += len(items)
                logger.warning(f"[bulk-bid-by-tokens] batch 실패: {str(e)[:80]}")
            await asyncio.sleep(0.1)
        logger.warning(f"[bulk-bid-by-tokens] 완료 — bid={new_bid} {done}/{len(sel)} (실패 {failed})")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "customer_id": cid, "bid": new_bid, "targets": len(sel),
            "message": f"{len(sel)}개 키워드 입찰가 {new_bid}원 일괄설정 백그라운드 시작"}


class PrunePoolBacklogRequest(BaseModel):
    statuses: List[str] = Field(
        default_factory=lambda: ["failed", "rejected_by_naver", "deleted", "domain_skipped", "skipped_existing"],
        description="삭제할 풀 status (비-라이브 발굴이력). 기본값=실광고 아닌 backlog 전체. registered/pending 은 안전상 제외 권장",
    )
    dry_run: bool = Field(True, description="true: status별 행수, false: 배치 삭제")
    max_rows: int = Field(3000000, description="안전 상한")
    vacuum: bool = Field(False, description="삭제 후 VACUUM 시도(공간 부족 시 실패해도 무시) — 파일 축소로 OS 디스크 반환")


@router.post("/keyword-pool/admin/prune-pool-backlog")
async def keyword_pool_admin_prune_pool_backlog(
    request: PrunePoolBacklogRequest,
    background_tasks: BackgroundTasks,
    customer_id: int = Query(..., description="대상 광고주 customer_id"),
):
    """비-라이브 풀 backlog(failed/rejected/deleted/skipped) 일괄 삭제 — 디스크 확보용.
    naverad_keyword_pool 의 발굴이력만 삭제(라이브 광고=registered_keywords/네이버는 안 건드림).
    배치 5000 + commit 으로 디스크풀에서도 점진 진행. dry_run 으로 status별 행수 먼저 확인."""
    import sqlite3 as _sq
    pool = get_keyword_pool_db()
    allowed = {"failed", "rejected_by_naver", "deleted", "domain_skipped", "skipped_existing", "pending"}
    statuses = [s for s in (request.statuses or []) if s in allowed]
    if not statuses:
        raise HTTPException(status_code=400, detail=f"유효 status 없음. 허용: {sorted(allowed)}")
    ph = ",".join("?" * len(statuses))
    with _sq.connect(pool.db_path, timeout=60.0) as conn:
        by_status = {}
        for s in statuses:
            by_status[s] = conn.execute(
                "SELECT COUNT(*) FROM naverad_keyword_pool WHERE account_customer_id=? AND status=?",
                (customer_id, s),
            ).fetchone()[0]
    total = sum(by_status.values())
    if request.dry_run:
        return {"success": True, "dry_run": True, "customer_id": customer_id,
                "by_status": by_status, "total_to_delete": total}

    async def _run():
        import sqlite3 as _sq2
        deleted = 0
        try:
            with _sq2.connect(pool.db_path, timeout=120.0) as c:
                while deleted < request.max_rows:
                    cur = c.execute(
                        f"""DELETE FROM naverad_keyword_pool WHERE rowid IN (
                                SELECT rowid FROM naverad_keyword_pool
                                WHERE account_customer_id=? AND status IN ({ph}) LIMIT 5000)""",
                        (customer_id, *statuses),
                    )
                    n = cur.rowcount or 0
                    c.commit()
                    if n == 0:
                        break
                    deleted += n
                    if deleted % 50000 < 5000:
                        logger.warning(f"[prune-pool-backlog] cid={customer_id} 삭제 {deleted}/{total}")
            logger.warning(f"[prune-pool-backlog] cid={customer_id} 완료 — 삭제 {deleted}")
            if request.vacuum:
                try:
                    with _sq2.connect(pool.db_path, timeout=600.0) as cv:
                        cv.execute("VACUUM")
                    logger.warning("[prune-pool-backlog] VACUUM 완료 — 파일 축소")
                except Exception as ve:
                    logger.warning(f"[prune-pool-backlog] VACUUM 실패(공간부족 추정): {str(ve)[:100]}")
        except Exception as e:
            logger.error(f"[prune-pool-backlog] 실패 — 삭제 {deleted}: {type(e).__name__}: {str(e)[:150]}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "customer_id": customer_id,
            "by_status": by_status, "total_to_delete": total,
            "message": f"비-라이브 backlog {total}행 삭제 백그라운드 시작 (배치5000, vacuum={request.vacuum})"}


class LowVolumeCleanupRequest(BaseModel):
    max_volume: int = Field(10, description="이 monthly_total 이하 등록키워드 삭제. '< 10'은 5+5=10으로 저장되므로 10이면 실검색<10 쓰레기 제거")
    dry_run: bool = Field(True, description="true: 대상수+샘플, false: 실제 삭제")
    max_delete: int = Field(200000, description="안전 상한")
    off_only: bool = Field(False, description="true: 삭제 대신 userLock=true 로 off (복구 가능). DB 유지, 노출만 차단")
    require_tokens: Optional[List[str]] = Field(None, description="설정 시: 이 토큰 중 하나라도 포함한 키워드만 대상 (예: 의료대출 정크만 off). None=전체 저검색")


@router.post("/keyword-pool/low-volume-cleanup")
async def keyword_pool_low_volume_cleanup(
    request: LowVolumeCleanupRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """저검색량(monthly_total ≤ max_volume) 등록 키워드 삭제 — '< 10'(실검색 없음) 쓰레기 정리.
    registered_keywords(ncc_id) ⋈ keyword_pool(monthly_total) 조인으로 대상 선별."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    from database.keyword_pool_db import get_keyword_pool_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    reg = get_registered_keywords_db()
    pool = get_keyword_pool_db()
    with _sq.connect(reg.db_path) as conn:
        regrows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND removed_at IS NULL",
            (cid,),
        ).fetchall()
    # off_only 는 userLock bulk 에 nccAdgroupId 필요 → (nid, gid) 보관.
    regmap = {kw: (nid, gid) for kw, nid, gid in regrows}
    with _sq.connect(pool.db_path) as conn:
        # 정확한 쓰레기 정의: PC AND 모바일 둘 다 < 10 (한쪽이라도 실제 ≥10이면 보존).
        # '< 10'은 5로 저장되므로 monthly_pc<10 AND monthly_mobile<10 = 양쪽 다 '< 10' 플레이스홀더.
        lowrows = conn.execute(
            "SELECT keyword, monthly_total FROM naverad_keyword_pool "
            "WHERE account_customer_id=? AND monthly_pc < 10 AND monthly_mobile < 10",
            (cid,),
        ).fetchall()
    lowmap = {kw: mt for kw, mt in lowrows}
    req_tokens = [t for t in (request.require_tokens or []) if t and t.strip()]
    def _wanted(kw: str) -> bool:
        if not req_tokens:
            return True
        t = (kw or "").replace(" ", "")
        return any(tok in t for tok in req_tokens)
    # targets: (keyword, nid, gid, vol)
    targets = [(kw, regmap[kw][0], regmap[kw][1], lowmap[kw])
               for kw in regmap if kw in lowmap and _wanted(kw)]
    targets = targets[:request.max_delete]
    verb = "off(userLock)" if request.off_only else "삭제"
    if request.dry_run:
        return {"success": True, "dry_run": True, "customer_id": cid, "mode": verb,
                "registered_total": len(regmap), "pool_lowvol": len(lowmap),
                "require_tokens": req_tokens or None,
                "delete_targets": len(targets), "max_volume": request.max_volume,
                "samples": [{"kw": t[0], "vol": t[3]} for t in targets[:40]]}

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        logger.warning(f"[lowvol-cleanup] {verb} 시작 — {len(targets)}개 (vol≤{request.max_volume})")
        done = 0; failed = 0
        for i in range(0, len(targets), 100):
            batch = targets[i:i+100]
            try:
                if request.off_only:
                    items = [{"nccKeywordId": nid, "nccAdgroupId": gid, "userLock": True}
                             for _, nid, gid, _ in batch if nid and gid]
                    await client.set_keywords_userlock_bulk(items)
                else:
                    await client.delete_keywords_bulk([nid for _, nid, _, _ in batch if nid])
                done += len(batch)
            except Exception as e:
                failed += len(batch)
                logger.warning(f"[lowvol-cleanup] batch 실패: {str(e)[:80]}")
            if done % 5000 < 100:
                logger.warning(f"[lowvol-cleanup] 진행 — {verb} {done}/{len(targets)} (실패 {failed})")
            await asyncio.sleep(0.1)
        if not request.off_only:
            # 삭제 모드만 DB 에서 제거 표시. off 는 DB 유지.
            try:
                reg.mark_removed(cid, [kw for kw, _, _, _ in targets])
            except Exception:
                pass
        logger.warning(f"[lowvol-cleanup] 완료 — {verb} {done} / 실패 {failed}")
    background_tasks.add_task(_run)
    return {"success": True, "started": True, "mode": verb, "delete_targets": len(targets),
            "message": f"저검색량 {len(targets)}개 {verb} 백그라운드 시작 (로그 확인)"}


class EmptyAdgroupCleanupRequest(BaseModel):
    dry_run: bool = Field(True, description="true: 빈 그룹 개수만, false: 실제 삭제")
    max_delete: int = Field(5000, description="안전 상한")


@router.post("/keyword-pool/empty-adgroup-cleanup")
async def keyword_pool_empty_adgroup_cleanup(
    request: EmptyAdgroupCleanupRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """키워드 0개인 빈 광고그룹 삭제. **네이버 실시간 키워드 수가 정확히 0인 그룹만** 삭제.
    키워드 1개라도 있으면 절대 삭제 안 함. 조회 실패(에러/타임아웃) 그룹도 안전하게 skip(보존)."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    async def _collect_empty():
        all_camps = _as_list(await client.get_campaigns() or [])
        camps = [c for c in all_camps if (c.get("campaignTp") or "") == "WEB_SITE"]
        gids = []
        for c in camps:
            for _att in range(3):
                try:
                    groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                    for g in groups:
                        gid = g.get("nccAdgroupId")
                        if gid:
                            gids.append((gid, g.get("name") or ""))
                    break
                except Exception:
                    await asyncio.sleep(1.5)
            await asyncio.sleep(0.08)
        empty = []  # (gid, name)
        scanned = 0; errs = 0
        for gid, name in gids:
            try:
                kws = _as_list(await client.get_keywords(ad_group_id=gid) or [])
                scanned += 1
                if len(kws) == 0:   # ★ 정확히 0개만 — 1개라도 있으면 보존
                    empty.append((gid, name))
            except Exception:
                errs += 1   # 조회 실패 → 안전하게 보존(삭제 안 함)
            await asyncio.sleep(0.1)
        return gids, empty, scanned, errs

    if request.dry_run:
        gids, empty, scanned, errs = await _collect_empty()
        return {"success": True, "dry_run": True, "customer_id": cid,
                "groups_total": len(gids), "scanned": scanned, "scan_errors": errs,
                "empty_groups": len(empty), "samples": [n or g for g, n in empty[:30]]}

    async def _run():
        gids, empty, scanned, errs = await _collect_empty()
        ids = [g for g, _ in empty][:request.max_delete]
        logger.warning(f"[empty-adgroup] 삭제 시작 — 빈그룹 {len(ids)} / 전체 {len(gids)} / 조회실패 {errs}")
        done = 0; failed = 0
        for i in range(0, len(ids), 100):
            batch = ids[i:i+100]
            try:
                await client.delete_ad_groups_bulk(batch)
                done += len(batch)
            except Exception as e:
                failed += len(batch)
                logger.warning(f"[empty-adgroup] batch 실패: {str(e)[:80]}")
            await asyncio.sleep(0.15)
        logger.warning(f"[empty-adgroup] 완료 — 삭제 {done} / 실패 {failed} / 조회실패(보존) {errs}")
    background_tasks.add_task(_run)
    return {"success": True, "started": True, "message": "빈 광고그룹(키워드 0개) 삭제 백그라운드 시작 (로그 확인)"}


class EmptyCampaignCleanupRequest(BaseModel):
    dry_run: bool = Field(True, description="true: 빈 캠페인 개수만, false: 실제 삭제")
    max_delete: int = Field(5000, description="안전 상한")


@router.post("/keyword-pool/empty-campaign-cleanup")
async def keyword_pool_empty_campaign_cleanup(
    request: EmptyCampaignCleanupRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """키워드 0개인 빈 캠페인 삭제. **캠페인의 모든 광고그룹 키워드 합이 정확히 0인 캠페인만** 삭제.
    키워드 1개라도 있으면 절대 삭제 안 함. 조회 실패(에러/타임아웃)한 캠페인은 안전하게 skip(보존)."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    async def _campaigns_retry():
        for _att in range(5):
            try:
                cs = _as_list(await client.get_campaigns() or [])
                if cs:
                    return cs
            except Exception:
                pass
            await asyncio.sleep(2.5)
        return []

    async def _collect_empty():
        camps = await _campaigns_retry()
        camps = [c for c in camps if (c.get("campaignTp") or "") == "WEB_SITE"]
        empty = []  # (cid, name)
        scanned = 0; errs = 0
        for c in camps:
            cpid = c.get("nccCampaignId")
            if not cpid:
                continue
            cp_err = False
            total_kw = 0
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=cpid) or [])
            except Exception:
                errs += 1
                continue  # 그룹 조회 실패 → 캠페인 보존
            for g in groups:
                gid = g.get("nccAdgroupId")
                if not gid:
                    continue
                try:
                    kws = _as_list(await client.get_keywords(ad_group_id=gid) or [])
                    total_kw += len(kws)
                except Exception:
                    cp_err = True   # 키워드 조회 실패 → 이 캠페인 보존
                    break
                if total_kw > 0:
                    break  # 1개라도 있으면 즉시 보존
                await asyncio.sleep(0.05)
            scanned += 1
            if not cp_err and total_kw == 0:
                empty.append((cpid, c.get("name") or ""))
            await asyncio.sleep(0.08)
        return camps, empty, scanned, errs

    if request.dry_run:
        camps, empty, scanned, errs = await _collect_empty()
        return {"success": True, "dry_run": True, "customer_id": cid,
                "campaigns_total": len(camps), "scanned": scanned, "scan_errors": errs,
                "empty_campaigns": len(empty), "samples": [n or g for g, n in empty[:40]]}

    # 동기 삭제 — 백그라운드 task 가 fly 에서 완료 안 되는 문제로, 스캔+삭제를 요청 내에서 처리.
    camps, empty, scanned, errs = await _collect_empty()
    ids = [g for g, _ in empty][:request.max_delete]
    done = 0; failed = 0; fail_msgs: List[str] = []
    for i in range(0, len(ids), 100):
        batch = ids[i:i+100]
        try:
            await client.delete_campaigns_bulk(batch)
            done += len(batch)
        except Exception as e:
            failed += len(batch)
            if len(fail_msgs) < 3:
                fail_msgs.append(f"{type(e).__name__}: {str(e)[:120]}")
        await asyncio.sleep(0.2)
    logger.warning(f"[empty-campaign] 동기 완료 — 삭제 {done} / 실패 {failed} / 조회실패(보존) {errs}")
    return {"success": True, "deleted": done, "failed": failed, "preserved_scan_errors": errs,
            "empty_found": len(empty), "campaigns_total": len(camps), "fail_msgs": fail_msgs}


@router.get("/keyword-pool/diagnostics/accounts-list")
async def keyword_pool_diagnostics_accounts_list():
    """진단 — 모든 활성 광고주 + user_seed 샘플 (인증 없음).

    한의원 광고주 customer_id 식별용. cid 자체는 민감 정보 아님 (네이버 광고
    조회 가능). user_seed 샘플 5개로 도메인 식별 가능.
    """
    from database.naver_ad_db import list_connected_ad_accounts
    pool = get_keyword_pool_db()
    accts = list_connected_ad_accounts() or []
    out = []
    for a in accts:
        cid = int(a.get("customer_id") or 0)
        uid = int(a.get("user_id") or 0)
        if not cid:
            continue
        try:
            seeds = pool.list_user_seeds(cid)
        except Exception:
            seeds = []
        try:
            stats = pool.stats(cid) or {}
            by_status = stats.get("by_status") or {}
        except Exception:
            by_status = {}
        out.append({
            "user_id": uid,
            "customer_id": cid,
            "user_seed_count": len(seeds),
            "user_seed_samples": seeds[:8],
            "pool_by_status": by_status,
        })
    return {"success": True, "count": len(out), "accounts": out}


@router.get("/keyword-pool/diagnostics/recent-registered")
async def keyword_pool_diagnostics_recent_registered(
    customer_id: int = Query(..., description="광고주 customer_id"),
    limit: int = Query(100, ge=1, le=500),
    order: str = Query("recent", regex="^(recent|mt_asc|mt_desc)$"),
):
    """진단 — 등록 KW 최근/하위/상위 N개 리스트 (인증 없음).

    도메인 적합성 audit 용. naverad_keyword_pool 의 status='registered' 행만.
    seed attribution + monthly_total 포함 → 어떤 시드로부터 발굴됐는지 추적 가능.

    order:
      - recent: registered_at DESC (디폴트, 최근 등록 순)
      - mt_asc: monthly_total ASC (rolling_heal 대상 후보 확인)
      - mt_desc: monthly_total DESC (상위 KW 분포 확인)
    """
    import sqlite3 as _sqlite3
    pool = get_keyword_pool_db()

    order_sql = {
        "recent": "COALESCE(registered_at, discovered_at) DESC, id DESC",
        "mt_asc": "COALESCE(monthly_total, 0) ASC, registered_at ASC",
        "mt_desc": "COALESCE(monthly_total, 0) DESC, registered_at DESC",
    }[order]

    with _sqlite3.connect(pool.db_path) as conn:
        conn.row_factory = _sqlite3.Row
        rows = conn.execute(
            f"""SELECT keyword, seed, monthly_total, monthly_pc, monthly_mobile,
                       source, registered_at, discovered_at
                FROM naverad_keyword_pool
                WHERE account_customer_id = ?
                  AND status = 'registered'
                ORDER BY {order_sql}
                LIMIT ?""",
            (customer_id, limit),
        ).fetchall()

    items = [dict(r) for r in rows]
    # mt 분포 요약 (도메인 적합성 빠른 확인용)
    bucket = {"mt_0": 0, "mt_1_9": 0, "mt_10_99": 0, "mt_100_999": 0, "mt_1000_plus": 0}
    for r in items:
        mt = int(r.get("monthly_total") or 0)
        if mt <= 0: bucket["mt_0"] += 1
        elif mt < 10: bucket["mt_1_9"] += 1
        elif mt < 100: bucket["mt_10_99"] += 1
        elif mt < 1000: bucket["mt_100_999"] += 1
        else: bucket["mt_1000_plus"] += 1

    return {
        "success": True,
        "customer_id": customer_id,
        "order": order,
        "count": len(items),
        "mt_distribution": bucket,
        "items": items,
    }


@router.get("/keyword-pool/diagnostics/seed-audit")
async def keyword_pool_diagnostics_seed_audit(
    customer_id: int = Query(..., description="광고주 customer_id"),
    relevance_keywords: Optional[str] = Query(
        None,
        description="콤마구분 도메인 KW. 비우면 saved relevance_keywords 사용",
    ),
    score_threshold: int = Query(30, ge=0, le=95),
):
    """진단 — user_seed 풀의 도메인 점수 분포 (인증 없음).

    drift 근본 원인 추적: 오염된 user_seed (예: "2024쏘나타") 가 amplify cartesian
    폭발 → drift 100k. 이 endpoint 로 어떤 시드들이 score ≤ threshold 인지 식별 →
    purge-drift endpoint 로 일괄 정리.

    반환:
      score_distribution: {0: N, 10: N, ..., 100: N} 점수 구간별 시드 수
      contaminated_samples: score ≤ threshold 시드 30개 샘플 (정리 대상)
      clean_samples: score > threshold 시드 30개 샘플 (보존 대상)
    """
    from database.naver_ad_db import list_connected_ad_accounts, get_ad_account_relevance_keywords
    accts = list_connected_ad_accounts() or []
    matched = next((a for a in accts if int(a.get("customer_id") or 0) == int(customer_id)), None)
    if not matched:
        raise HTTPException(status_code=404, detail=f"customer_id {customer_id} 미연결")
    uid = int(matched.get("user_id") or 0)

    if relevance_keywords:
        score_basis = [s.strip() for s in relevance_keywords.replace("\n", ",").split(",") if s.strip() and len(s.strip()) >= 2]
        basis_source = "query"
    else:
        saved = get_ad_account_relevance_keywords(uid, str(customer_id)) or []
        if not saved:
            return {
                "success": False,
                "reason": "no_relevance",
                "message": "saved relevance_keywords 비어있음. ?relevance_keywords=... 명시 또는 화면에서 도메인 KW 저장 필요.",
            }
        score_basis = saved
        basis_source = "saved"

    pool = get_keyword_pool_db()
    user_seeds = pool.list_user_seeds(customer_id) or []

    # 점수 매김
    dist: Dict[int, int] = {}
    contaminated: List[Tuple[str, int]] = []
    clean: List[Tuple[str, int]] = []
    for s in user_seeds:
        sc = _compute_relevance_score(s, score_basis)
        bucket = (sc // 10) * 10
        dist[bucket] = dist.get(bucket, 0) + 1
        if sc < score_threshold:  # Option B: boundary 보존 — score == threshold 는 clean
            contaminated.append((s, sc))
        else:
            clean.append((s, sc))

    contaminated.sort(key=lambda x: x[1])  # 점수 낮은 것부터 (확실히 drift)
    clean.sort(key=lambda x: -x[1])  # 점수 높은 것부터 (확실히 도메인)

    return {
        "success": True,
        "customer_id": customer_id,
        "basis_source": basis_source,
        "basis_count": len(score_basis),
        "basis_sample": score_basis[:8],
        "user_seed_total": len(user_seeds),
        "score_threshold": score_threshold,
        "contaminated_count": len(contaminated),
        "clean_count": len(clean),
        "score_distribution": dict(sorted(dist.items())),
        "contaminated_samples": [{"seed": s, "score": sc} for s, sc in contaminated[:30]],
        "clean_samples": [{"seed": s, "score": sc} for s, sc in clean[:30]],
    }


@router.get("/keyword-pool/diagnostics/ai-cleanup-preview")
async def keyword_pool_diagnostics_ai_cleanup_preview(
    customer_id: int = Query(..., description="audit 대상 광고주 customer_id"),
    max_kws: int = Query(200),
):
    """진단 dry-run — customer_id 명시 (인증 없음). 등록 KW GPT 분류 미리보기.

    /diagnostics/accounts-list 로 cid 확인 → 이 endpoint 로 광고주별 audit.
    """
    from database.naver_ad_db import list_connected_ad_accounts
    accts = list_connected_ad_accounts() or []
    matched = next(
        (a for a in accts if int(a.get("customer_id") or 0) == int(customer_id)),
        None,
    )
    if not matched:
        return {"success": False, "reason": "customer_id_not_in_accounts"}
    uid = int(matched.get("user_id") or 0)
    cid = int(matched.get("customer_id") or 0)
    res = await _run_pool_ai_cleanup_registered(
        uid, cid, dry_run=True, max_kws=max_kws,
    )
    res["debug_meta"] = {"audited_uid": uid, "audited_cid": cid}
    return res


@router.get("/keyword-pool/diagnostics/ai-cleanup-preview-first")
async def keyword_pool_ai_cleanup_preview_first(
    max_kws: int = Query(200),
):
    """진단 dry-run — 인증 없음. 첫 활성 광고주의 등록 KW 를 GPT 분류 미리보기.

    실제 DELETE 안 함. 시스템 내부 상태 진단용 (실측 보고).
    """
    from database.naver_ad_db import list_connected_ad_accounts
    accts = list_connected_ad_accounts() or []
    if not accts:
        return {"success": False, "reason": "no_accounts"}
    a = accts[0]
    uid = int(a.get("user_id") or 0)
    cid = int(a.get("customer_id") or 0)
    if not uid or not cid:
        return {"success": False, "reason": "invalid_account"}
    res = await _run_pool_ai_cleanup_registered(
        uid, cid, dry_run=True, max_kws=max_kws,
    )
    res["debug_meta"] = {"audited_uid": uid, "audited_cid": cid}
    return res


@router.get("/keyword-pool/diagnostics/fill-progress")
async def keyword_pool_fill_progress(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = Query(None),
):
    """10만 자동채우기 진행/에스컬레이션 상태 — 계정별 레벨·마른streak·관련성floor·다음레버.
    '왜 5.2만에서 안 늘지?' 를 한눈에 보고 어떤 레버가 다음에 열리는지 확인하는 진단 패널."""
    from database.naver_ad_db import list_ad_accounts_for_user, get_ad_account_by_customer
    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()

    if customer_id:
        acct = get_ad_account_by_customer(user_id, str(customer_id))
        accounts = [acct] if acct else []
    else:
        accounts = list_ad_accounts_for_user(user_id) or []

    out: List[Dict[str, Any]] = []
    for a in accounts:
        if not a or not a.get("customer_id"):
            continue
        cid = int(a.get("customer_id"))
        try:
            active = int((reg.stats(cid) or {}).get("active") or 0)
        except Exception:
            active = 0
        by_status = (pool.stats(cid) or {}).get("by_status") or {}
        pending = int(by_status.get("pending") or 0)
        st = pool.get_escalation(cid)
        recent = pool.recent_runs(cid, limit=12) or []
        collect_added = [int(r.get("added") or 0) for r in recent if r.get("kind") == "collect"][:5]
        level = int(st.get("level") or 0)
        headroom = _FILL_CAP - active - pending
        out.append({
            "customer_id": str(cid),
            "name": a.get("name"),
            "active_registered": active,
            "pending": pending,
            "cap": _FILL_CAP,
            "headroom": headroom,
            "fill_pct": round((active + pending) / _FILL_CAP * 100, 1),
            "level": level,
            "level_label": _FILL_LEVEL_LABELS.get(level, str(level)),
            "dry_streak": int(st.get("dry_streak") or 0),
            "relevance_floor": int(st.get("relevance_floor") or _FILL_FLOOR_START),
            "recent_collect_added": collect_added,
            "next_lever": _fill_next_lever(level),
            "last_note": st.get("note"),
            "updated_at": st.get("updated_at"),
        })
    return {"success": True, "cap": _FILL_CAP, "accounts": out}


@router.get("/keyword-pool/diagnostics/scheduler-jobs")
def keyword_pool_scheduler_diagnostics():
    """APScheduler 등록 cron + 다음 실행 시각 — cron 살아있는지 즉시 확인. sync def."""
    try:
        from services.keyword_pool_scheduler import keyword_pool_scheduler
    except Exception as e:
        return {"success": False, "error": f"scheduler import 실패: {e}"}
    sched = keyword_pool_scheduler.scheduler
    if not keyword_pool_scheduler._running:
        # 2-프로세스 분리 — 스케줄러는 worker 프로세스(:8001)에만 산다. API 프로세스는
        # _running=False 이므로 worker 의 동일 엔드포인트로 프록시해 실제 상태를 보여준다.
        try:
            import httpx as _httpx
            r = _httpx.get(
                "http://127.0.0.1:8001/api/naver-ad/keyword-pool/diagnostics/scheduler-jobs",
                timeout=5.0,
            )
            if r.status_code == 200:
                data = r.json()
                data["via"] = "worker"
                return data
        except Exception:
            pass
        # 프록시 실패 (worker 가 무거운 cron 으로 바빠 :8001 응답 지연) — DB 최근 실행
        # 기록으로 판단. cross-process 안정 (worker API 응답성에 의존 안 함).
        try:
            import sqlite3 as _sq
            pool = get_keyword_pool_db()
            with _sq.connect(pool.db_path) as _c:
                row = _c.execute("SELECT MAX(started_at) FROM naverad_pool_runs").fetchone()
            last = row[0] if row else None
            if last:
                age_min = (datetime.utcnow() - datetime.fromisoformat(str(last))).total_seconds() / 60.0
                if age_min < 15:
                    return {
                        "success": True, "running": True, "jobs": [],
                        "via": "db_recency",
                        "message": f"스케줄러 정상 (worker, 최근 실행 {age_min:.0f}분 전)",
                        "now": datetime.now().isoformat(timespec="seconds"),
                    }
        except Exception:
            pass
        return {"success": False, "running": False, "message": "scheduler not running"}
    jobs_info = []
    for job in sched.get_jobs():
        jobs_info.append({
            "id": job.id,
            "name": job.name,
            "next_run_time": str(job.next_run_time) if job.next_run_time else None,
            "trigger": str(job.trigger),
        })
    return {
        "success": True,
        "running": True,
        "jobs": jobs_info,
        "now": datetime.now().isoformat(timespec="seconds"),
    }


@router.get("/keyword-pool/diagnostics/cleanup-audit")
async def keyword_pool_cleanup_audit(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = Query(None),
    sample_ad_groups: int = Query(3, description="네이버 실측용 샘플 광고그룹 수 (1~10)"),
):
    """cleanup cron 동작 진단 + 네이버 광고 콘솔 실측 비교.

    반환:
      cron_summary: 종류별 (inspect/click_cleanup/domain_cleanup/auto-cleanup) 마지막 실행 시각, 결과
      naver_audit: 광고그룹 N개 샘플의 키워드 status 조회 → 검수 거부됐는데
                   여전히 콘솔에 살아있는 KW (= cleanup 누수) 카운트
    """
    if not customer_id:
        raise HTTPException(status_code=400, detail="customer_id 필요")
    try:
        cid = int(customer_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="customer_id 정수 필요")
    sample_n = max(1, min(int(sample_ad_groups), 10))

    pool = get_keyword_pool_db()
    runs = pool.recent_runs(cid, limit=200)

    # cron 종류별 마지막 + 최근 24개 합산
    from collections import defaultdict
    by_kind: Dict[str, List[Dict]] = defaultdict(list)
    for r in runs:
        kind = r.get("kind") or ""
        by_kind[kind].append(r)

    # 종류별 '마지막 실행'은 최근 200행 창이 아니라 kind 별 MAX(id) 로 뽑는다.
    # register 가 30초마다 행을 써서 200행이 몇 분치밖에 안 되고, 45분 주기인
    # autocomplete 는 창 밖으로 밀려 조회할 때마다 나타났다 사라졌다 한다.
    last_by_kind = pool.last_run_by_kind(cid)

    summary: Dict[str, Dict[str, Any]] = {}
    for kind in set(list(by_kind.keys()) + list(last_by_kind.keys())):
        rs = by_kind.get(kind) or []
        last = last_by_kind.get(kind) or (rs[0] if rs else {})
        recent = rs[:24]
        summary[kind] = {
            "last_run_at": last.get("started_at"),
            "last_status": last.get("status"),
            "last_added": last.get("added"),
            "last_skipped": last.get("skipped"),
            "last_message": (last.get("error_message") or "")[:200],
            # 아래 3개는 '최근 200행 창' 기준이라 자주 도는 cron 만 의미가 있다
            "runs_count_recent": len(recent),
            "total_added_recent": sum(int(r.get("added") or 0) for r in recent),
            "total_skipped_recent": sum(int(r.get("skipped") or 0) for r in recent),
        }

    # 네이버 광고 콘솔 실측 — 광고그룹 N개 샘플의 키워드 검수 상태 조회
    naver_audit: Dict[str, Any] = {
        "sample_ad_groups": [],
        "total_kws_checked": 0,
        "stale_rejected_count": 0,
        "stale_rejected_samples": [],
    }
    try:
        from database.naver_ad_db import get_ad_account_by_customer
        from database.registered_keywords_db import get_registered_keywords_db
        from services.naver_ad_service import NaverAdApiClient
        import sqlite3

        account = get_ad_account_by_customer(user_id, str(cid))
        if not account or not account.get("is_connected"):
            naver_audit["error"] = "광고주 미연결"
        else:
            client = NaverAdApiClient()
            client.customer_id = account["customer_id"]
            client.api_key = account["api_key"]
            client.secret_key = account["secret_key"]

            reg = get_registered_keywords_db()
            with sqlite3.connect(reg.db_path) as conn:
                ag_ids = [r[0] for r in conn.execute(
                    "SELECT DISTINCT ad_group_id FROM registered_keywords "
                    "WHERE account_customer_id=? AND ad_group_id IS NOT NULL "
                    "ORDER BY id DESC LIMIT ?",
                    (cid, sample_n),
                ).fetchall()]

            REJECT_TOKENS = (
                "DISAPPROVED", "REJECTED", "PROHIBITED", "BLOCKLISTED",
                "NOT_PASSED", "FAIL", "BAD_BUSINESS", "INELIGIBLE", "DENIED",
            )
            for ag_id in ag_ids:
                try:
                    kws = await client.get_keywords(ad_group_id=ag_id) or []
                except Exception as e:
                    naver_audit["sample_ad_groups"].append({
                        "ad_group_id": ag_id,
                        "error": f"{type(e).__name__}: {str(e)[:120]}",
                    })
                    continue

                stale: List[Dict] = []
                for kw in kws:
                    review = (kw.get("reviewStatus") or "").upper()
                    inspect = (kw.get("inspectStatus") or "").upper()
                    stat_reason = (kw.get("statusReason") or "").upper()
                    is_rejected = (
                        any(t in review for t in REJECT_TOKENS)
                        or any(t in inspect for t in REJECT_TOKENS)
                        or any(t in stat_reason for t in REJECT_TOKENS)
                    )
                    if is_rejected:
                        stale.append({
                            "keyword": kw.get("keyword"),
                            "review": review,
                            "inspect": inspect,
                            "reason": stat_reason,
                        })

                naver_audit["sample_ad_groups"].append({
                    "ad_group_id": ag_id,
                    "total_kws": len(kws),
                    "stale_rejected": len(stale),
                })
                naver_audit["total_kws_checked"] += len(kws)
                naver_audit["stale_rejected_count"] += len(stale)
                # 전체 sample 누적 (앞에서 5개만 노출)
                if len(naver_audit["stale_rejected_samples"]) < 5:
                    naver_audit["stale_rejected_samples"].extend(
                        stale[: 5 - len(naver_audit["stale_rejected_samples"])]
                    )
    except Exception as e:
        naver_audit["error"] = f"{type(e).__name__}: {str(e)[:200]}"

    return {
        "success": True,
        "now": datetime.now().isoformat(timespec="seconds"),
        "customer_id": cid,
        "cron_summary": summary,
        "naver_audit": naver_audit,
        "interpretation": {
            "stale_rejected_OK": (
                "stale_rejected_count == 0 이면 inspect cron 이 검수 거부 KW 를 "
                "잘 청소하는 중. > 0 이면 누수 — fly logs 확인 필요."
            ),
            "domain_cleanup_OK": (
                "cron_summary['inspect'].last_run_at 가 30분 이내 + "
                "cron_summary 의 click cleanup / 도메인 cleanup 마지막 시각 "
                "각각 15분 / 1시간 이내면 정상."
            ),
        },
    }


@router.post("/keyword-pool/ai-classify-rejects")
async def keyword_pool_ai_classify_rejects(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = Query(None),
    force: bool = Query(False, description="True 면 30분 쿨다운 무시"),
):
    """AI reject 분류 1회 수동 발동 — reject 풀에서 시드 도메인 일치 KW 만 user_seed 로 promote.

    - 시드 1+ 광고주만 동작 (cold_start 광고주는 skip)
    - GPT-4o-mini 호출, 후보 200개당 약 5~8초
    - 30분 쿨다운 (force=True 시 무시)
    """
    if not customer_id:
        raise HTTPException(status_code=400, detail="customer_id 필요")
    try:
        cid = int(customer_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="customer_id 정수 필요")

    result = await _run_pool_ai_classify(user_id, cid, force=force)
    return {
        "success": result.get("success", False),
        **result,
    }


@router.get("/keyword-pool/reject-stats")
def keyword_pool_reject_stats(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = Query(None),
):
    """reject 풀 상태 — UI 분류 버튼 옆 카운터 표시용. sync def → threadpool."""
    if not customer_id:
        raise HTTPException(status_code=400, detail="customer_id 필요")
    try:
        cid = int(customer_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="customer_id 정수 필요")

    pool = get_keyword_pool_db()
    stats = pool.reject_stats(cid) or {}
    cooldown_iso = pool.get_classify_cooldown(cid)
    cooldown_remaining_min = 0
    if cooldown_iso:
        try:
            from datetime import datetime, timedelta
            last_dt = datetime.fromisoformat(str(cooldown_iso).replace("T", " ").split(".")[0])
            elapsed = datetime.utcnow() - last_dt
            if elapsed < timedelta(minutes=30):
                cooldown_remaining_min = max(1, 30 - int(elapsed.total_seconds() / 60))
        except Exception:
            pass
    return {
        "success": True,
        "pending": int(stats.get("pending", 0)),
        "promoted": int(stats.get("promoted", 0)),
        "discarded": int(stats.get("discarded", 0)),
        "cooldown_remaining_min": cooldown_remaining_min,
        "last_run_at": cooldown_iso,
    }


@router.post("/keyword-pool/trigger-now")
async def keyword_pool_trigger_now(
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = Query(None),
):
    """사용자 트리거 — cron 다음 tick 안 기다리고 본인 광고주의 collect+register 즉시 실행.
    시드 저장 직후 / 새 광고주 초기 발굴 등에서 "5분 후" 대기 없이 즉시 시작.
    Bearer 인증 불필요 (본인 광고주만 처리).
    """
    from database.naver_ad_db import list_ad_accounts_for_user
    pairs: List[Tuple[int, int]] = []
    if customer_id:
        try:
            pairs = [(user_id, int(customer_id))]
        except ValueError:
            raise HTTPException(status_code=400, detail="customer_id 정수 필요")
    else:
        accounts = list_ad_accounts_for_user(user_id) or []
        pairs = [(user_id, int(a["customer_id"])) for a in accounts if a.get("is_connected")]

    if not pairs:
        return {"success": True, "queued": 0, "message": "활성 광고 계정 없음"}

    background_tasks.add_task(_run_pool_workers_for_accounts, pairs)
    return {
        "success": True,
        "queued": len(pairs),
        "message": f"{len(pairs)}개 광고주 즉시 발굴 시작 — 1~3분 후 화면 갱신",
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


@router.post("/keyword-pool/admin/run")
async def keyword_pool_admin_run(
    background_tasks: BackgroundTasks,
    authorization: Optional[str] = Header(None),
    user_id: Optional[int] = Query(None),
    customer_id: Optional[str] = Query(None),
):
    """자동 워커 — collect + register 통합 트리거 (Bearer 인증).
    - user_id 만: 그 사용자의 모든 활성 광고주 (B 시나리오)
    - user_id + customer_id: 그 광고주 단건만
    - 둘 다 없음: 모든 사용자 × 모든 활성 광고주
    """
    _verify_cron_token(authorization)

    pairs: List[Tuple[int, int]] = []
    try:
        from database.naver_ad_db import list_connected_ad_accounts, list_ad_accounts_for_user
        if user_id and customer_id:
            pairs = [(user_id, int(customer_id))]
        elif user_id:
            accounts = list_ad_accounts_for_user(user_id) or []
            pairs = [(user_id, int(a["customer_id"])) for a in accounts if a.get("is_connected")]
        else:
            rows = list_connected_ad_accounts() or []
            pairs = [(int(r["user_id"]), int(r["customer_id"])) for r in rows if r.get("user_id") and r.get("customer_id")]
    except Exception as e:
        logger.error(f"[pool/admin/run] 광고주 조회 실패: {type(e).__name__}: {e}", exc_info=True)
        pairs = []

    if not pairs:
        return {"success": True, "queued": 0, "message": "활성 광고 계정 없음"}

    background_tasks.add_task(_run_pool_workers_for_accounts, pairs)
    return {
        "success": True,
        "queued": len(pairs),
        "pairs": [{"user_id": uid, "customer_id": cid} for uid, cid in pairs],
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


@router.get("/keyword-pool/accounts")
def keyword_pool_list_accounts(user_id: int = Depends(get_user_id_with_fallback)):
    """사용자의 모든 활성 광고주 list (B 시나리오 — 다중 광고주).

    sync def — 단순 sqlite read. async def 일 때 cron tick 이 event loop 점유 중이면
    30s timeout 발생 (frontend `광고주 목록 조회 실패` 사고). threadpool dispatch 로
    event loop 무관하게 응답.
    """
    from database.naver_ad_db import list_ad_accounts_for_user
    rows = list_ad_accounts_for_user(user_id) or []
    return {
        "success": True,
        "accounts": [
            {
                "customer_id": str(r.get("customer_id")),
                "name": r.get("name"),
                "is_connected": bool(r.get("is_connected")),
                "last_sync_at": r.get("last_sync_at"),
                "default_bid": int(r.get("default_bid") or 100),
            }
            for r in rows
        ],
    }


@router.get("/keyword-pool/stats")
def keyword_pool_stats(
    user_id: int = Depends(get_user_id_with_fallback),
    customer_id: Optional[str] = None,
    lite: bool = False,
):
    """본인 풀/등록 상태 — customer_id 명시 시 그 광고주.

    sync def — 모든 호출이 sqlite read (pool.stats/recent_runs/seed_breakdown 등).
    threadpool dispatch 로 cron 점유 event loop 와 격리.

    lite=true: 첫 페인트용 — pool.stats, reg.stats, recent_runs[:5] 만. seed_breakdown
    (시드 200+ 일 때 300ms+) / recent_keywords / deadlock 는 응답에서 제외 → 응답 시간
    1초 미만 보장. 풀 페이지가 useEffect 첫 호출에서 lite=true 쓰고, 그 다음 idle
    callback 에서 full 호출.
    """
    try:
        account = _resolve_account(user_id, customer_id)
        if not account:
            return {"success": False, "message": "광고 계정 미연결", "pool": {}, "registered": {}, "account_cap": 100_000}
        customer_id = int(account.get("customer_id"))
        pool = get_keyword_pool_db()
        reg = get_registered_keywords_db()
        try:
            pool_stats = pool.stats(customer_id)
        except Exception as e:
            logger.error(f"keyword-pool/stats pool.stats 실패: {e}", exc_info=True)
            pool_stats = {"error": f"{type(e).__name__}: {str(e)[:200]}"}
        try:
            reg_stats = reg.stats(customer_id)
        except Exception as e:
            logger.error(f"keyword-pool/stats reg.stats 실패: {e}", exc_info=True)
            reg_stats = {"error": f"{type(e).__name__}: {str(e)[:200]}"}
        try:
            recent = pool.recent_runs(customer_id, limit=5 if lite else 20)
        except Exception as e:
            logger.error(f"keyword-pool/stats recent_runs 실패: {e}", exc_info=True)
            recent = []

        # lite 모드 — 무거운 쿼리 3종 skip. 응답 1초 미만 보장.
        if lite:
            return {
                "success": True,
                "customer_id": customer_id,
                "pool": pool_stats,
                "registered": reg_stats,
                "account_cap": 100_000,
                "recent_runs": recent,
                "seed_breakdown": [],
                "recent_keywords": [],
                "collect_deadlock": {"is_deadlock": False, "consecutive_zero_runs": 0, "total_rejected": 0},
                "lite": True,
                "now": datetime.now().isoformat(timespec="seconds"),
            }

        try:
            seed_break = pool.seed_breakdown(customer_id)
        except Exception as e:
            logger.error(f"keyword-pool/stats seed_breakdown 실패: {e}", exc_info=True)
            seed_break = []
        try:
            recent_kw = pool.recent_keywords(customer_id, limit=30)
        except Exception as e:
            logger.error(f"keyword-pool/stats recent_keywords 실패: {e}", exc_info=True)
            recent_kw = []
        try:
            deadlock = pool.detect_collect_deadlock(customer_id, n_recent=5, min_rejected=500)
        except Exception as e:
            logger.error(f"keyword-pool/stats detect_collect_deadlock 실패: {e}", exc_info=True)
            deadlock = {"is_deadlock": False, "consecutive_zero_runs": 0, "total_rejected": 0}
        return {
            "success": True,
            "customer_id": customer_id,
            "pool": pool_stats,
            "registered": reg_stats,
            "account_cap": 100_000,
            "recent_runs": recent,
            "seed_breakdown": seed_break,
            "recent_keywords": recent_kw,
            "collect_deadlock": deadlock,
            "now": datetime.now().isoformat(timespec="seconds"),
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/stats 전체 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


class PoolSeedsRequest(BaseModel):
    seeds: List[str]


class AdminSeedsRequest(BaseModel):
    seeds: List[str]
    user_id: int


@router.post("/keyword-pool/admin/add-seeds")
async def keyword_pool_admin_add_seeds(
    request: AdminSeedsRequest,
    authorization: Optional[str] = Header(None),
):
    """Bearer 토큰으로 시드 일괄 추가 — workflow_dispatch / curl 용."""
    _verify_cron_token(authorization)
    try:
        account = get_ad_account(request.user_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail=f"user_id={request.user_id} 광고 계정 미연결")
        customer_id = int(account.get("customer_id"))
        pool = get_keyword_pool_db()
        items = [
            {"keyword": s.strip(), "seed": s.strip(), "source": "user_seed", "monthly_total": 0}
            for s in request.seeds if s and s.strip()
        ]
        added = pool.add_candidates(request.user_id, customer_id, items)
        return {"success": True, "added": added, "total_input": len(items),
                "user_id": request.user_id, "customer_id": customer_id}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/admin/add-seeds 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


class SeedExplodeRequest(BaseModel):
    seeds: List[str]
    min_volume: int = 100  # 월 총 검색량 최소치 — 등록 가치 있는 것만
    max_per_seed: int = 1000  # 시드당 최대 연관키워드 수
    min_score: int = 50  # 연관성 점수 최소치 — 자동삭제 크론(점수<50 삭제)과 일관
    customer_id: Optional[str] = None


async def _run_seed_explode(
    user_id: int, customer_id: int, account: Dict,
    seeds: List[str], min_volume: int, per_seed_cap: int, min_score: int = 50,
) -> None:
    """연관키워드 폭발 — 시드별 keywordstool 연관키워드 수집 → 검색량 + 연관성 점수 필터 →
    pending 직접 삽입.

    AI classify(LLM) 게이트는 안 거치되, 연관성 점수(_compute_relevance_score) ≥ min_score
    필터는 적용 — 자동삭제 크론이 점수<50 등록 KW 를 지우므로, 그 기준 이상만 등록해
    churn(등록→삭제) 을 막고 도메인 정밀도 유지. 점수 기준(saved_relevance→user_seed)은
    자동삭제 크론과 동일.
    """
    import time as _time
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import get_ad_account_relevance_keywords
    pool = get_keyword_pool_db()
    t0 = _time.monotonic()

    # 연관성 점수 기준 — 자동삭제 크론과 동일: saved relevance_keywords → user_seed 폴백.
    score_basis = get_ad_account_relevance_keywords(user_id, str(customer_id))
    if not score_basis:
        score_basis = [s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2]
    # negative_keywords (drift 차단) + required_tokens (핵심의도 앵커) — 프로파일에서 로드.
    negatives = []
    required_tokens = []
    try:
        from database.naver_ad_db import get_domain_profile as _get_prof
        _prof = _get_prof(user_id, str(customer_id)) or {}
        negatives = [n for n in _prof.get("negative_keywords", []) if n and len(n) >= 2]
        required_tokens = [t for t in _prof.get("required_tokens", []) if t and len(t) >= 2]
    except Exception:
        pass

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _to_int(v):
        if v is None:
            return 0
        if isinstance(v, (int, float)):
            return int(v)
        s = str(v).replace(",", "").strip()
        if s in ("< 10", "<10"):
            return 5
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return 0

    seen: Set[str] = set()
    items: List[Dict] = []
    total_related = 0
    n_vol_pass = 0   # 검색량 통과 수
    n_score_cut = 0  # 검색량 통과했으나 점수 미달로 컷
    n_neg_cut = 0    # negative_keywords 포함으로 컷 (drift 차단)
    for seed in seeds:
        try:
            resp = await client.get_related_keywords(seed, show_detail=True)
        except Exception as e:
            logger.warning(f"[pool/explode] seed='{seed}' 연관 조회 실패: {e}")
            await asyncio.sleep(0.3)
            continue
        rows = resp.get("keywordList", []) if isinstance(resp, dict) else (resp if isinstance(resp, list) else [])
        total_related += len(rows)
        added_this_seed = 0
        for it in rows:
            kw = (it.get("relKeyword") or "").strip()
            if not kw or kw in seen:
                continue
            pc = _to_int(it.get("monthlyPcQcCnt"))
            mo = _to_int(it.get("monthlyMobileQcCnt"))
            mt = pc + mo
            if mt < min_volume:
                continue
            n_vol_pass += 1
            # 연관성 점수 필터 — 자동삭제 크론(점수<min_score 삭제)과 일관. 점수 기준은
            # saved_relevance(없으면 user_seed). seen 은 점수 미달이어도 마킹해 재계산 방지.
            seen.add(kw)
            if negatives and any(nt in kw for nt in negatives):
                n_neg_cut += 1
                continue
            # 핵심의도 앵커 — 필수 토큰 중 하나도 없으면 컷 (시설토큰만으론 통과 못 함).
            if required_tokens and not any(rt in kw for rt in required_tokens):
                n_neg_cut += 1
                continue
            # 관련성 점수 게이트 — 앵커 모드(required_tokens)면 앵커+negative 가 도메인 테스트이므로
            # 점수 컷 skip (앵커있는 진짜 대출이 좁은 relevance 로 과삭제되는 것 방지). 비앵커 도메인만 점수.
            if not required_tokens and score_basis and _compute_relevance_score(kw, score_basis) < min_score:
                n_score_cut += 1
                continue
            items.append({
                "keyword": kw, "seed": seed, "source": "seed_explode",
                "monthly_total": mt, "monthly_pc": pc, "monthly_mobile": mo,
                "comp_idx": it.get("compIdx", ""),
            })
            added_this_seed += 1
            if added_this_seed >= per_seed_cap:
                break
        await asyncio.sleep(0.3)  # keywordstool 429 rate 회피

    added = pool.add_candidates(user_id, customer_id, items) if items else 0
    dur_ms = int((_time.monotonic() - t0) * 1000)
    logger.warning(
        f"[pool/explode] user={user_id} cid={customer_id} 시드 {len(seeds)} → "
        f"연관 {total_related} → 검색량≥{min_volume} {n_vol_pass} → 점수≥{min_score} {len(items)} "
        f"(점수컷 {n_score_cut}, neg컷 {n_neg_cut}) → pending +{added} ({dur_ms}ms)"
    )
    try:
        pool.record_run(
            user_id, customer_id, "seed_explode", "success" if added else "no_new",
            added=added, seeds_count=len(seeds),
            error_message=(
                f"연관 {total_related} → 검색량≥{min_volume} {n_vol_pass} → "
                f"점수≥{min_score} {len(items)} → pending +{added}"
            )[:300],
            duration_ms=dur_ms,
        )
    except Exception:
        pass


@router.post("/keyword-pool/seed-explode-register")
async def keyword_pool_seed_explode(
    request: SeedExplodeRequest,
    background_tasks: BackgroundTasks,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """연관키워드 폭발 등록 — 사용자가 고른 시드의 연관키워드를 대량 수집·검색량 필터만
    통과시켜 pending 직접 삽입 (AI 도메인 게이트 우회). register cron 이 네이버 등록.

    자동 발굴은 drift 방지 게이트가 빡빡해 통과율이 낮다. 사용자가 명시한 시드의
    연관키워드는 신뢰 가능하므로 classify/whitelist 없이 검색량만 보고 대량 등록한다.
    백그라운드 처리 (시드 다수 × keywordstool ~1s → fly 60s proxy 초과 방지). 결과는 실행 이력.
    """
    seeds = [s.strip() for s in (request.seeds or []) if s and s.strip()]
    if not seeds:
        raise HTTPException(status_code=400, detail="시드가 비어있습니다")
    seeds = seeds[:150]  # 1회 최대 150 시드 (저장된 도메인 키워드 전체 폭발 지원). 백그라운드 처리.
    min_volume = max(0, min(100_000, request.min_volume))
    per_seed_cap = max(1, min(1000, request.max_per_seed))
    min_score = max(0, min(100, request.min_score))

    account = _resolve_account(user_id, request.customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="네이버 광고 계정을 먼저 연동하세요")
    customer_id = int(account.get("customer_id"))

    # ⚠️ app 프로세스(SCHEDULERS_DISABLED=1)에서는 **절대 여기서 실행하지 않는다** —
    #    keywordstool 수백 콜이 API 이벤트루프를 점유해 로그인 hang 을 재발시킨다.
    #    예전엔 이걸 막으려고 `_WORKER_OFFLOAD_PATHS` 로 worker 에 HTTP 프록시했는데,
    #    그 프록시가 8s ReadTimeout 후 httpx 를 닫으면 **worker 요청이 끊겨 핸들러가 아예
    #    안 돌았다**(2026-07-30 실측: 4회 연속+드라이버 90회 전부 8.1초 실패, run 0건).
    #    → 제어를 공유 볼륨으로 넘긴다. app 은 큐 파일만 쓰고, worker 워치독이 집어 실행.
    #    (ceiling-backtest·backfill-creative 가 같은 이유로 쓰는 패턴)
    # entrypoint.sh: app = `SCHEDULERS_DISABLED=1 ROLE=app` / worker = `SCHEDULERS_DISABLED=0
    # ROLE=worker`. 로컬 단일 프로세스는 둘 다 미설정이라 direct 로 떨어진다(기존 동작 유지).
    if _os.getenv("SCHEDULERS_DISABLED") == "1":
        from services.seed_explode_queue import enqueue
        q = enqueue(user_id, customer_id, seeds, min_volume, per_seed_cap, min_score)
        if not q.get("queued"):
            raise HTTPException(status_code=503,
                                detail=f"실행 큐 적재 실패: {q.get('reason')}")
        return {
            "success": True,
            # ★ 여기서의 started 는 "잡이 디스크 큐에 **내구성 있게** 들어갔다"는 뜻이다.
            #   워커 워치독이 20초 내 반드시 집어가므로 드라이버는 커서를 전진시켜도 된다.
            #   (예전 합성 ack 는 아무 보장이 없어 시드가 허공에 소진됐다 — 그것과 다르다)
            "started": True,
            "via": "queue",
            "job_id": q.get("job_id"),
            "queue_len": q.get("queue_len"),
            "customer_id": customer_id,
            "seeds_used": len(seeds),
            "min_volume": min_volume,
            "min_score": min_score,
        }

    # worker(스케줄러 켜진 프로세스) 또는 로컬 단일 프로세스 → 즉시 실행
    background_tasks.add_task(
        _run_seed_explode, user_id, customer_id, account, seeds, min_volume, per_seed_cap, min_score,
    )
    return {
        "success": True,
        "started": True,
        "via": "direct",
        "customer_id": customer_id,
        "seeds_used": len(seeds),
        "min_volume": min_volume,
        "min_score": min_score,
        "message": (
            f"연관키워드 폭발 시작 — 시드 {len(seeds)}개의 연관키워드를 수집해 "
            f"검색량≥{min_volume} + 연관성 점수≥{min_score} 인 것을 pending 에 추가합니다. "
            f"진행/결과는 '최근 실행 이력'의 seed_explode 항목에서 확인하세요."
        ),
    }


class AdminInspectRequest(BaseModel):
    user_id: int


@router.post("/keyword-pool/admin/inspect-all")
async def keyword_pool_admin_inspect_all(
    request: AdminInspectRequest,
    authorization: Optional[str] = Header(None),
):
    """모든 풀 광고그룹 키워드 검토 상태 조회 → 노출제한 mark.

    DEBUG mode (request.user_id < 0): user_id=-N 으로 호출 시 N (절대값) 의 키워드 1개로
    /stats 호출 시도 + 빌드된 URL 반환 (디버깅 전용).
    """
    _verify_cron_token(authorization)
    # /stats endpoint 디버깅 — user_id=-1 같이 음수면 stats 호출 1회만 시도하고
    # 빌드된 URL + Naver 응답 반환.
    if request.user_id < 0:
        from services.naver_ad_service import NaverAdApiClient
        from datetime import datetime, timedelta
        import sqlite3 as _sq
        target_uid = abs(request.user_id)
        account = get_ad_account(target_uid)
        if not account or not account.get("is_connected"):
            return {"debug": True, "error": "광고 계정 미연결"}
        customer_id_dbg = int(account.get("customer_id"))
        reg_dbg = get_registered_keywords_db()
        with _sq.connect(reg_dbg.db_path) as conn:
            row = conn.execute(
                "SELECT ncc_keyword_id FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL LIMIT 1",
                (customer_id_dbg,),
            ).fetchone()
        if not row:
            return {"debug": True, "error": "키워드 없음"}
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]
        end_d = datetime.now().strftime("%Y-%m-%d")
        start_d = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
        try:
            stats = await client.get_stats("KEYWORD", [row[0]], start_d, end_d)
            return {"debug": True, "test_id": row[0], "result": stats}
        except Exception as e:
            return {"debug": True, "test_id": row[0], "error": f"{type(e).__name__}: {str(e)[:300]}"}
    try:
        from services.naver_ad_service import NaverAdApiClient
        account = get_ad_account(request.user_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        customer_id = int(account.get("customer_id"))
        reg = get_registered_keywords_db()
        # 광고그룹 list
        with __import__("sqlite3").connect(reg.db_path) as conn:
            ag_ids = [r[0] for r in conn.execute(
                "SELECT DISTINCT ad_group_id FROM registered_keywords WHERE account_customer_id=? AND ad_group_id IS NOT NULL",
                (customer_id,),
            ).fetchall()]
        if not ag_ids:
            return {"success": True, "ad_groups": 0, "rejected": 0}
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]
        n = await _inspect_ad_groups(request.user_id, customer_id, client, ag_ids)
        return {"success": True, "ad_groups": len(ag_ids), "rejected": n}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/admin/inspect-all 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


class AdminDeleteRequest(BaseModel):
    keywords: List[str]
    user_id: int


class DovisionMigrateRequest(BaseModel):
    customer_id: Optional[str] = None
    mid_key: Optional[str] = Field(default=None, description="대상 중분류 key (없으면 분포만)")
    dry_run: bool = Field(default=True, description="True면 재분류 분포만 반환(무변경)")
    limit: int = Field(default=2000, description="1회 처리 배치 — 반복 호출로 재개")


@router.post("/keyword-pool/admin/dovision-migrate-hierarchy")
async def dovision_migrate_hierarchy(
    request: DovisionMigrateRequest,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """두비전 계층 마이그레이션 — 중분류 단위 단계 실행 (blast radius 최소화).

    dry_run=True  : 등록분을 새 계층으로 재분류한 분포만 반환 (변경 없음).
    dry_run=False + mid_key : 그 중분류 키워드만 네이버에서 삭제 → registered/pool 을
      pending 으로 되돌림 → 배포된 계층 register 크론이 '[두비전] 대 - 중' 캠페인의
      소분류 그룹으로 자동 재등록. limit 단위로 끊어 실행하므로 반복 호출로 재개 가능.
      캠페인을 통째로 지우지 않아 다른 중분류 노출은 유지됨.
    """
    from services.naver_ad_service import NaverAdApiClient

    account = _resolve_account(user_id, request.customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    if cid not in _DOVISION_CAT_CUSTOMERS:
        raise HTTPException(status_code=400, detail=f"두비전 전용 엔드포인트 (cid {cid} 대상 아님)")

    pool = get_keyword_pool_db()
    rows = pool.list_registered_rows(cid)
    buckets: Dict[str, Dict[str, Any]] = {}
    for _r in rows:
        _mk, _clabel, _slabel = _classify_dovision_category(_r["keyword"])
        _b = buckets.setdefault(_mk, {"label": _clabel, "count": 0, "subs": {}, "rows": []})
        _b["count"] += 1
        _b["subs"][_slabel] = _b["subs"].get(_slabel, 0) + 1
        _b["rows"].append(_r)

    if request.dry_run or not request.mid_key:
        return {
            "success": True, "customer_id": cid, "dry_run": True,
            "total_registered": len(rows),
            "distribution": [
                {
                    "mid_key": _k, "campaign": f"[두비전] {_v['label']}",
                    "keywords": _v["count"],
                    "groups": sorted(_v["subs"].items(), key=lambda x: -x[1]),
                }
                for _k, _v in sorted(buckets.items(), key=lambda x: -x[1]["count"])
            ],
        }

    b = buckets.get(request.mid_key)
    if not b:
        raise HTTPException(status_code=400, detail=f"mid_key '{request.mid_key}' 해당 등록 키워드 없음")
    batch = b["rows"][: max(1, min(int(request.limit or 2000), 5000))]
    kws = [_r["keyword"] for _r in batch]
    ids = [_r["id"] for _r in batch]

    reg = get_registered_keywords_db()
    ncc_rows = reg.get_ncc_ids(cid, kws)
    ncc_ids = [x["ncc_keyword_id"] for x in ncc_rows if x.get("ncc_keyword_id")]

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    deleted, del_fail = 0, 0
    for i in range(0, len(ncc_ids), 100):
        chunk = ncc_ids[i:i + 100]
        try:
            await client.delete_keywords_bulk(chunk)
            deleted += len(chunk)
        except Exception as e:
            del_fail += len(chunk)
            logger.warning(f"[dovi-migrate] 키워드 삭제 실패 {len(chunk)}개: {e}")
        await asyncio.sleep(0.2)

    # 네이버 삭제가 전량 실패하면 되돌리지 않음 — 중복등록 방지.
    if ncc_ids and deleted == 0:
        raise HTTPException(status_code=502, detail=f"네이버 삭제 전량 실패 ({del_fail}개) — 상태 미변경")

    removed = reg.mark_removed(cid, kws)
    requeued = pool.mark_status(ids, "pending")
    logger.warning(
        f"[dovi-migrate] cid={cid} {request.mid_key} 배치 {len(batch)} → "
        f"naver삭제 {deleted}(실패 {del_fail}) / reg-removed {removed} / pool→pending {requeued}"
    )
    return {
        "success": True, "customer_id": cid, "dry_run": False,
        "mid_key": request.mid_key, "campaign": f"[두비전] {b['label']}",
        "batch": len(batch), "naver_deleted": deleted, "naver_delete_failed": del_fail,
        "reg_removed": removed, "pool_requeued": requeued,
        "remaining_in_mid": b["count"] - len(batch),
        "note": "register 크론이 새 계층 캠페인으로 재등록합니다. remaining_in_mid>0 이면 재호출로 이어서 진행.",
    }


@router.post("/keyword-pool/admin/delete-keywords")
async def keyword_pool_admin_delete_keywords(
    request: AdminDeleteRequest,
    authorization: Optional[str] = Header(None),
):
    """Bearer 토큰으로 풀에서 특정 키워드 일괄 삭제 — cleanup용."""
    _verify_cron_token(authorization)
    try:
        account = get_ad_account(request.user_id)
        if not account:
            raise HTTPException(status_code=400, detail=f"user_id={request.user_id} 광고 계정 없음")
        customer_id = int(account.get("customer_id"))
        pool = get_keyword_pool_db()
        deleted = pool.delete_keywords(customer_id, request.keywords)
        return {"success": True, "deleted": deleted, "user_id": request.user_id}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/admin/delete-keywords 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.get("/keyword-pool/clicked-keywords")
async def keyword_pool_clicked_keywords(
    days: int = 7,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """클릭 발생한 키워드 list — 사용자 검수용. 시드 매칭 여부 표시."""
    from services.naver_ad_service import NaverAdApiClient
    from datetime import datetime, timedelta
    import sqlite3 as _sqlite3

    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        customer_id = int(account.get("customer_id"))

        reg = get_registered_keywords_db()
        with _sqlite3.connect(reg.db_path) as conn:
            rows = conn.execute(
                "SELECT keyword, ncc_keyword_id FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL",
                (customer_id,),
            ).fetchall()
        if not rows:
            return {"success": True, "days": days, "total": 0, "items": []}
        keyword_map = {r[1]: r[0] for r in rows}

        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]

        end_date = datetime.now().strftime("%Y-%m-%d")
        start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

        # Naver /stats 는 단일 ID 호출만 안정적 (multi-id 11001 잘못된 파라미터).
        # 95k+ KW 모두 querying 은 비현실적 — 최근 등록 1500개만 sample.
        # 대부분 KW 는 click=0 이라 sample 제한해도 active KW 잡힘 (registered DESC).
        ids = list(keyword_map.keys())[:1500]
        # 동시 호출 20 → 3 으로 축소. Naver outbound 가 timeout 폭주할 때
        # circuit breaker (threshold=5) 가 빠르게 OPEN 되어 나머지 task 가 즉시 503 fail.
        # 정상 시에도 Sem=3 이면 1500개 ÷ 3 ≈ 500 round × ~200ms = 100s 내 완료.
        sem = asyncio.Semaphore(3)
        # /stats 전용 breaker — inspect/collect 와 격리.
        from services.naver_ad_service import _stats_breaker, NaverApiCircuitOpenError

        async def _fetch_one(kid: str) -> List[dict]:
            # stats circuit OPEN 상태면 task 진입 자체 skip — sem 점유 안 함
            if _stats_breaker.is_open():
                return []
            async with sem:
                try:
                    stats = await client.get_stats(
                        stat_type="KEYWORD", ids=[kid],
                        start_date=start_date, end_date=end_date,
                    )
                    return stats or []
                except NaverApiCircuitOpenError:
                    return []
                except Exception as e:
                    logger.warning(f"clicked-keywords {kid} 실패: {str(e)[:120]}")
                    return []

        results = await asyncio.gather(*[_fetch_one(kid) for kid in ids])
        all_stats: List[dict] = [s for batch in results for s in batch]

        pool = get_keyword_pool_db()
        user_seeds = [s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2]

        items = []
        for stat in all_stats:
            keyword_id = stat.get("id")
            if not keyword_id:
                continue
            clicks = int(stat.get("clkCnt", 0) or 0)
            if clicks <= 0:
                continue
            kw_text = keyword_map.get(keyword_id)
            if not kw_text:
                continue
            score = _compute_relevance_score(kw_text, user_seeds)
            items.append({
                "keyword_id": keyword_id,
                "keyword": kw_text,
                "impressions": int(stat.get("impCnt", 0) or 0),
                "clicks": clicks,
                "cost": int(stat.get("salesAmt", 0) or 0),
                "ctr": float(stat.get("ctr", 0) or 0),
                "cpc": int(stat.get("cpc", 0) or 0),
                "matches_seed": score >= 100,  # 호환성 유지 — full seed 매칭만 true
                "relevance_score": score,
            })
        # 점수 낮은 순 + 클릭 많은 순 — 가장 무관한 KW 먼저 (낭비 큰 것 우선 노출)
        items.sort(key=lambda x: (x["relevance_score"], -x["clicks"]))
        return {"success": True, "days": days, "total": len(items), "items": items}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"clicked-keywords 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.get("/keyword-pool/diagnostics/clicked-off-domain")
async def keyword_pool_clicked_off_domain(
    days: int = 7,
    customer_id: Optional[str] = None,
    min_clicks: int = 1,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """전체 계정 클릭 퍼널 — 캠페인→그룹→키워드로 좁혀 클릭 난 키워드를 전수 조회 후 소잠 on/off 토큰으로 분류.
    clicked-keywords 는 오래된 1500개만 sample 하는 편향이 있어 대량 등록된 off-domain 을 놓침.
    본 진단은 클릭이 소수 그룹에 몰리는 특성을 이용해 95k active 를 전수 커버(캠페인/그룹 stat 로 0클릭 가지치기)."""
    from services.naver_ad_service import NaverAdApiClient, _stats_breaker, NaverApiCircuitOpenError
    from datetime import datetime, timedelta

    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        cid = int(account.get("customer_id"))
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]

        end_date = datetime.now().strftime("%Y-%m-%d")
        start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

        def _as_list(x):
            if isinstance(x, list):
                return x
            if isinstance(x, dict):
                return x.get("data") or x.get("list") or []
            return []

        def _clk(stats):
            for s in (stats or []):
                try:
                    return int(s.get("clkCnt", 0) or 0)
                except Exception:
                    return 0
            return 0

        sem = asyncio.Semaphore(4)

        async def _stat(_id):
            if _stats_breaker.is_open():
                return []
            async with sem:
                try:
                    return await client.get_stats(stat_type="", ids=[_id],
                                                  start_date=start_date, end_date=end_date) or []
                except NaverApiCircuitOpenError:
                    return []
                except Exception:
                    return []

        # 1) 캠페인 전수 → 클릭 있는 것만
        camps = [c for c in _as_list(await client.get_campaigns() or []) if c.get("nccCampaignId")]
        camp_stats = await asyncio.gather(*[_stat(c["nccCampaignId"]) for c in camps])
        clicked_camps = [c for c, s in zip(camps, camp_stats) if _clk(s) >= min_clicks]

        # 2) 클릭 캠페인의 그룹 → 클릭 있는 것만
        all_groups = []
        for c in clicked_camps:
            for _att in range(3):
                try:
                    gs = _as_list(await client.get_ad_groups(campaign_id=c["nccCampaignId"]) or [])
                    all_groups.extend([g["nccAdgroupId"] for g in gs if g.get("nccAdgroupId")])
                    break
                except Exception:
                    await asyncio.sleep(1.0)
        grp_stats = await asyncio.gather(*[_stat(g) for g in all_groups])
        clicked_groups = [g for g, s in zip(all_groups, grp_stats) if _clk(s) >= min_clicks]

        # 3) 클릭 그룹의 키워드 → 전수 stat
        kw_pairs = []  # (nccKeywordId, keyword)
        for gid in clicked_groups:
            try:
                for k in _as_list(await client.get_keywords(ad_group_id=gid) or []):
                    nid = k.get("nccKeywordId"); txt = k.get("keyword")
                    if nid and txt:
                        kw_pairs.append((nid, txt))
            except Exception:
                pass
        kw_stats = await asyncio.gather(*[_stat(nid) for nid, _ in kw_pairs])

        ON = _SOJAM_ON_TOKENS; OFF = _SOJAM_OFF_TOKENS
        items = []
        for (nid, txt), s in zip(kw_pairs, kw_stats):
            st = (s or [None])[0] or {}
            try:
                clk = int(st.get("clkCnt", 0) or 0)
            except Exception:
                clk = 0
            if clk < min_clicks:
                continue
            t = (txt or "").replace(" ", "")
            if any(n in t for n in OFF):
                verdict = "off_by_negative"
            elif not any(o in t for o in ON):
                verdict = "off_no_ondomain"
            else:
                verdict = "on_domain"
            items.append({
                "keyword_id": nid, "keyword": txt, "clicks": clk,
                "impressions": int(st.get("impCnt", 0) or 0),
                "cost": int(st.get("salesAmt", 0) or 0),
                "ctr": float(st.get("ctr", 0) or 0),
                "cpc": int(st.get("cpc", 0) or 0),
                "verdict": verdict,
            })
        off_items = [i for i in items if i["verdict"] != "on_domain"]
        # off 토큰 명확한 것 먼저, 그 다음 비용 큰 순 (낭비 큰 것 우선)
        off_items.sort(key=lambda x: (x["verdict"] != "off_by_negative", -x["cost"]))
        return {
            "success": True, "customer_id": cid, "days": days,
            "campaigns_total": len(camps), "campaigns_clicked": len(clicked_camps),
            "groups_scanned": len(all_groups), "groups_clicked": len(clicked_groups),
            "clicked_keywords_total": len(items),
            "off_domain_total": len(off_items),
            "off_by_negative": sum(1 for i in items if i["verdict"] == "off_by_negative"),
            "off_no_ondomain": sum(1 for i in items if i["verdict"] == "off_no_ondomain"),
            "on_domain": sum(1 for i in items if i["verdict"] == "on_domain"),
            "off_domain": off_items,
            "breaker_open": _stats_breaker.is_open(),
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"clicked-off-domain 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


class ClickedBidAdjustRequest(BaseModel):
    days: int = Field(7, description="클릭 집계 기간(일)")
    min_clicks: int = Field(1, description="이 클릭수 이상만 대상")
    off_bid: int = Field(70, ge=70, description="off-domain(무관) 클릭 키워드 바닥 입찰가 — 싹 낮춤")
    bid_cap: int = Field(2000, description="on-domain estimate 입찰 상한(원)")
    device: str = Field("PC", description="PC 또는 MOBILE")
    dry_run: bool = Field(True, description="true: 계획만, false: 실제 입찰 적용(백그라운드)")


@router.post("/keyword-pool/registered/clicked-bid-adjust")
async def keyword_pool_clicked_bid_adjust(
    request: ClickedBidAdjustRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """**클릭 난 키워드만** 대상으로 중요도 기반 입찰 조정.
    on-domain(피부/한방) 클릭 KW → 중요도 점수(브랜드/지역/의도/질환)로 목표순위 입찰(estimate, cap).
    off-domain(무관) 클릭 KW → off_bid(기본 70) 바닥가로 싹 낮춤. 클릭 퍼널로 95k 전수 커버."""
    from services.naver_ad_service import NaverAdApiClient, _stats_breaker, NaverApiCircuitOpenError
    from datetime import datetime, timedelta
    from collections import defaultdict as _dd

    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        cid = int(account.get("customer_id"))
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        _dev = (request.device or "PC").upper()
        if _dev not in ("PC", "MOBILE"):
            _dev = "PC"
        end_date = datetime.now().strftime("%Y-%m-%d")
        start_date = (datetime.now() - timedelta(days=request.days)).strftime("%Y-%m-%d")

        def _as_list(x):
            if isinstance(x, list):
                return x
            if isinstance(x, dict):
                return x.get("data") or x.get("list") or []
            return []

        def _clk(stats):
            for s in (stats or []):
                try:
                    return int(s.get("clkCnt", 0) or 0)
                except Exception:
                    return 0
            return 0

        sem = asyncio.Semaphore(4)

        async def _stat(_id):
            if _stats_breaker.is_open():
                return []
            async with sem:
                try:
                    return await client.get_stats(stat_type="", ids=[_id],
                                                  start_date=start_date, end_date=end_date) or []
                except NaverApiCircuitOpenError:
                    return []
                except Exception:
                    return []

        # ── 클릭 퍼널: 캠페인 → 그룹 → 키워드 (gid 보존) ──
        camps = [c for c in _as_list(await client.get_campaigns() or []) if c.get("nccCampaignId")]
        camp_stats = await asyncio.gather(*[_stat(c["nccCampaignId"]) for c in camps])
        clicked_camps = [c for c, s in zip(camps, camp_stats) if _clk(s) >= request.min_clicks]

        all_groups = []
        for c in clicked_camps:
            for _att in range(3):
                try:
                    gs = _as_list(await client.get_ad_groups(campaign_id=c["nccCampaignId"]) or [])
                    all_groups.extend([g["nccAdgroupId"] for g in gs if g.get("nccAdgroupId")])
                    break
                except Exception:
                    await asyncio.sleep(1.0)
        grp_stats = await asyncio.gather(*[_stat(g) for g in all_groups])
        clicked_groups = [g for g, s in zip(all_groups, grp_stats) if _clk(s) >= request.min_clicks]

        kw_triples = []  # (nid, txt, gid)
        for gid in clicked_groups:
            try:
                for k in _as_list(await client.get_keywords(ad_group_id=gid) or []):
                    nid = k.get("nccKeywordId"); txt = k.get("keyword")
                    if nid and txt:
                        kw_triples.append((nid, txt, gid))
            except Exception:
                pass
        kw_stats = await asyncio.gather(*[_stat(nid) for nid, _, _ in kw_triples])

        # ── 중요도 스코어 (bulk-rank-bid 와 동일 체계) ──
        GEO_TOP = ["강남", "역삼", "논현", "신사", "청담", "압구정", "선릉", "대치", "학동", "신논현", "강남구", "양재", "도곡"]
        GEO_ADJ = ["서초", "방배", "잠원", "개포", "수서", "일원", "세곡", "우면"]
        DIS = ["아토피", "건선", "여드름", "두드러기", "습진", "지루성", "탈모", "무좀", "대상포진", "사마귀", "백반증", "기미",
               "모낭염", "한포진", "주사비", "다한증", "켈로이드", "피부", "한방", "피부염", "피부질환", "가려움", "곤지름",
               "헤르페스", "티눈", "뾰루지", "색소침착", "흉터", "모공", "땀띠", "주근깨", "비듬"]
        BRAND = ["소잠"]
        I_BOOK = ["예약", "상담", "문의", "예약문의", "전화상담", "예약하기", "당일"]
        I_COST = ["비용", "가격", "얼마"]
        I_CHOICE = ["추천", "후기", "잘하는곳", "명의", "유명"]
        I_CLINIC = ["한의원", "한방병원", "병원", "의원", "피부과", "클리닉"]
        I_TREAT = ["치료", "한약", "약침", "한방치료", "봉독", "완치", "낫는법"]
        I_INFO = ["증상", "원인", "사진", "이미지", "뜻", "종류", "에좋은", "음식", "민간요법", "전염", "옮나", "초기증상"]

        def _score(kw: str) -> int:
            t = (kw or "").replace(" ", "")
            s = 0
            if any(b in t for b in BRAND): s += 50
            if any(g in t for g in GEO_TOP): s += 40
            elif any(g in t for g in GEO_ADJ): s += 25
            if any(i in t for i in I_BOOK): s += 30
            if any(i in t for i in I_COST): s += 25
            if any(i in t for i in I_CHOICE): s += 20
            if any(i in t for i in I_CLINIC): s += 15
            elif any(i in t for i in I_TREAT): s += 8
            if any(i in t for i in I_INFO): s -= 25
            if any(d in t for d in DIS): s += 10
            return s

        def _pos(score: int):
            if score >= 75: return 1
            if score >= 60: return 2
            if score >= 48: return 3
            if score >= 38: return 5
            if score >= 28: return 7
            if score >= 18: return 10
            return None

        ON = _SOJAM_ON_TOKENS; OFF = _SOJAM_OFF_TOKENS
        # 대상 분류: off → 바닥, on → pos 밴드
        off_targets = []   # (nid, gid, kw, cost)
        pos_targets = _dd(list)  # pos -> [(kw, nid, gid)]
        floor_on = []      # on-domain 인데 점수 낮아 floor 유지
        clicked_n = 0; on_n = 0; off_n = 0
        dist = _dd(int)
        cost_off = 0; cost_on = 0
        for (nid, txt, gid), s in zip(kw_triples, kw_stats):
            st = (s or [None])[0] or {}
            try:
                clk = int(st.get("clkCnt", 0) or 0)
            except Exception:
                clk = 0
            if clk < request.min_clicks:
                continue
            clicked_n += 1
            cost = int(st.get("salesAmt", 0) or 0)
            t = (txt or "").replace(" ", "")
            is_off = any(n in t for n in OFF) or not any(o in t for o in ON)
            if is_off:
                off_n += 1; cost_off += cost
                off_targets.append((nid, gid, txt, cost))
                dist["slash_off"] += 1
            else:
                on_n += 1; cost_on += cost
                p = _pos(_score(txt))
                if p:
                    pos_targets[p].append((txt, nid, gid))
                    dist[f"pos{p}"] += 1
                else:
                    floor_on.append((nid, gid, txt))
                    dist["floor_on"] += 1

        if request.dry_run:
            off_targets.sort(key=lambda x: -x[3])
            samp_pos = {f"pos{p}": [k[0] for k in pos_targets[p][:8]] for p in sorted(pos_targets)}
            return {
                "success": True, "dry_run": True, "customer_id": cid, "device": _dev,
                "days": request.days,
                "campaigns_clicked": len(clicked_camps), "groups_clicked": len(clicked_groups),
                "clicked_keywords": clicked_n, "on_domain": on_n, "off_domain": off_n,
                "cost_on_domain": cost_on, "cost_off_domain": cost_off,
                "plan": {
                    "off_slash_to": request.off_bid, "off_count": off_n,
                    "on_by_position": dict(sorted(dist.items())),
                    "bid_cap": request.bid_cap,
                },
                "off_samples_by_cost": [{"keyword": k[2], "cost": k[3]} for k in off_targets[:30]],
                "on_pos_samples": samp_pos,
                "breaker_open": _stats_breaker.is_open(),
            }

        EST_POS = {1, 2, 3, 5}
        FLAT_POS = {7: 120, 10: 80}
        CAP = max(70, int(request.bid_cap))
        OFFB = max(70, int(request.off_bid))

        async def _run():
            total_done = 0; total_fail = 0
            # 1) off-domain → 바닥가
            by_gid = _dd(list)
            for nid, gid, _txt, _c in off_targets:
                by_gid[gid].append({"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": OFFB, "useGroupBidAmt": False})
            for nid, gid, _txt in floor_on:
                by_gid[gid].append({"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": 70, "useGroupBidAmt": False})
            for gid, its in by_gid.items():
                for i in range(0, len(its), 100):
                    try:
                        await client.update_keywords_bid_bulk(its[i:i + 100]); total_done += len(its[i:i + 100])
                    except Exception:
                        total_fail += len(its[i:i + 100])
                    await asyncio.sleep(0.1)
            logger.warning(f"[clicked-bid-adjust] off/floor 적용 {total_done} (실패 {total_fail})")
            # 2) on-domain → 순위밴드
            for pos in sorted(pos_targets.keys()):
                items_kw = pos_targets[pos]
                bg = _dd(list)
                if pos in EST_POS:
                    texts = list({k[0] for k in items_kw if k[0]})
                    bidmap = {}
                    for cas in [pos, 5, 3, 1, 2]:
                        remaining = [t for t in texts if t not in bidmap]
                        if not remaining:
                            break
                        for i in range(0, len(remaining), 15):
                            try:
                                r = await client.get_avg_position_bids(remaining[i:i + 15], cas, device=_dev)
                                for e in (r.get("estimate") or []):
                                    kt, bd = (e.get("keyword") or "").strip(), e.get("bid")
                                    if kt and bd and kt not in bidmap:
                                        bidmap[kt] = bd
                            except Exception:
                                pass
                            await asyncio.sleep(0.2)
                    for kw, nid, gid in items_kw:
                        bd = bidmap.get((kw or "").strip())
                        nb = min(CAP, max(70, round(int(bd) / 10) * 10)) if bd else 70
                        bg[gid].append({"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": nb, "useGroupBidAmt": False})
                else:
                    nb = FLAT_POS.get(pos, 70)
                    for kw, nid, gid in items_kw:
                        bg[gid].append({"nccKeywordId": nid, "nccAdgroupId": gid, "bidAmt": nb, "useGroupBidAmt": False})
                d = 0; f = 0
                for gid, its in bg.items():
                    for i in range(0, len(its), 100):
                        try:
                            await client.update_keywords_bid_bulk(its[i:i + 100]); d += len(its[i:i + 100])
                        except Exception:
                            f += len(its[i:i + 100])
                        await asyncio.sleep(0.1)
                total_done += d; total_fail += f
                logger.warning(f"[clicked-bid-adjust] pos={pos} 적용 {d} (실패 {f}, 대상 {len(items_kw)})")
            logger.warning(f"[clicked-bid-adjust] 전체 완료 — 적용 {total_done} / 실패 {total_fail}")

        background_tasks.add_task(_run)
        return {"success": True, "started": True, "customer_id": cid, "device": _dev,
                "clicked_keywords": clicked_n, "on_domain": on_n, "off_domain": off_n,
                "off_slash_to": OFFB, "distribution": dict(sorted(dist.items())),
                "message": "클릭 키워드 입찰 조정 백그라운드 시작 (로그 [clicked-bid-adjust])"}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"clicked-bid-adjust 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


# 클릭 전수 census 진행상태 (cid -> 상태/결과 dict). 워커 1개 가정.
_CLICK_CENSUS_STATUS: Dict[int, dict] = {}


def _census_path(cid: int) -> str:
    return f"/data/_click_census_{cid}.json"


def _census_load(cid: int) -> Optional[dict]:
    import json as _json
    try:
        with open(_census_path(cid), "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return None


def _census_summarize(doc: dict) -> dict:
    """저장 doc(items_full 포함) → 가벼운 status 요약 + 상위 샘플."""
    items = doc.get("items_full") or []
    off = [i for i in items if i.get("verdict") != "on_domain"]
    on = [i for i in items if i.get("verdict") == "on_domain"]
    off.sort(key=lambda x: (x.get("verdict") != "off_by_negative", -int(x.get("cost") or 0)))
    on.sort(key=lambda x: -int(x.get("cost") or 0))
    cb = list(doc.get("campaign_breakdown") or [])
    cb.sort(key=lambda x: -int(x.get("cost") or 0))
    return {
        "state": doc.get("state"), "phase": doc.get("phase"),
        "days": doc.get("days"), "min_clicks": doc.get("min_clicks"),
        "plan_built_at": doc.get("plan_built_at"),
        "campaigns_total": doc.get("campaigns_total"),
        "campaigns_clicked": doc.get("campaigns_clicked"),
        "groups_scanned": doc.get("groups_scanned"),
        "groups_clicked": doc.get("groups_clicked"),
        "groups_processed": len(doc.get("processed_group_ids") or []),
        "clicked_keywords_total": len(items),
        "on_domain": len(on), "off_domain_total": len(off),
        "off_by_negative": sum(1 for i in items if i.get("verdict") == "off_by_negative"),
        "off_no_ondomain": sum(1 for i in items if i.get("verdict") == "off_no_ondomain"),
        "total_cost": sum(int(i.get("cost") or 0) for i in items),
        "cost_off_domain": sum(int(i.get("cost") or 0) for i in off),
        "cost_on_domain": sum(int(i.get("cost") or 0) for i in on),
        "campaign_breakdown": cb[:200],
        "off_domain_keywords": off[:500],
        "on_domain_top_cost": on[:100],
        "error": doc.get("error"), "breaker_open": doc.get("breaker_open"),
    }


@router.post("/keyword-pool/diagnostics/clicked-census-bg")
async def keyword_pool_clicked_census_bg(
    background_tasks: BackgroundTasks,
    days: int = 7,
    min_clicks: int = 1,
    force: bool = False,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """clicked-off-domain 의 **백그라운드 + 재시작내성 + 재개** 버전.
    소잠(그룹 ~4천, 후보키워드 ~6만) 같은 큰 계정은 네이버 /stats 단일ID·저속 때문에
    per-ID 전수 stat 이 ~80분+ 걸려 (a) HTTP 타임아웃 (b) fly 재배포/재시작 시 유실된다.
    이 버전은:
      · 캠페인 stat 후 **예산 브레이크다운을 즉시 /data 파일에 저장**(20~30초 내 예산답 확정)
      · 클릭 그룹을 **비용 큰 순으로 정렬**해 낭비 큰 곳부터 키워드 stat
      · 그룹 배치마다 부분결과를 파일에 **증분 저장**(재시작해도 부분결과 상시 조회 가능)
      · 재호출 시 **처리한 그룹은 건너뛰고 이어서 진행(resume)** — plan 6h 캐시.
    force=true 면 처음부터. 진행/결과: GET .../clicked-census-bg/status?customer_id="""
    from services.naver_ad_service import NaverAdApiClient, _stats_breaker, NaverApiCircuitOpenError
    from datetime import datetime, timedelta
    from collections import defaultdict as _dd
    import json as _json

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    prev = _CLICK_CENSUS_STATUS.get(cid)
    if prev and prev.get("state") == "running":
        return {"success": True, "already_running": True, "customer_id": cid,
                "status": {k: prev.get(k) for k in ("state", "phase", "groups_clicked",
                           "groups_processed", "clicked_keywords_total")},
                "message": "이미 census 진행 중 — status 로 확인"}

    ON = _SOJAM_ON_TOKENS
    OFF = _SOJAM_OFF_TOKENS

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    def _clk(stats):
        for s in (stats or []):
            try:
                return int(s.get("clkCnt", 0) or 0)
            except Exception:
                return 0
        return 0

    # 재개용 저장 doc 로드
    doc = None if force else _census_load(cid)
    if doc and doc.get("state") == "done" and doc.get("days") == days \
            and doc.get("min_clicks") == min_clicks:
        return {"success": True, "already_done": True, "customer_id": cid,
                "message": "완료된 census 존재 — status 로 결과 확인 (force=true 로 재실행)"}

    st = {
        "state": "running", "phase": "init", "days": days, "min_clicks": min_clicks,
        "campaigns_total": 0, "campaigns_clicked": 0,
        "groups_scanned": 0, "groups_clicked": 0, "groups_processed": 0,
        "clicked_keywords_total": 0, "error": None, "breaker_open": False,
    }
    _CLICK_CENSUS_STATUS[cid] = st

    async def _do():
        try:
            client = NaverAdApiClient()
            client.customer_id = account["customer_id"]
            client.api_key = account["api_key"]
            client.secret_key = account["secret_key"]
            end_date = datetime.now().strftime("%Y-%m-%d")
            start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
            sem = asyncio.Semaphore(4)

            async def _stat(_id):
                # ★정합성: breaker 열림 시 skip(=키워드 누락) 하지 않고 닫힐 때까지 대기·재시도.
                # get_stats 는 예외/breaker 를 삼키고 [] 반환하므로, 빈 결과가 breaker 탓인지
                # 실제 무실적인지 breaker 상태로 구분해 재시도한다(진짜 무실적만 [] 확정).
                for _att in range(240):
                    while _stats_breaker.is_open():
                        await asyncio.sleep(3)
                    async with sem:
                        try:
                            r = await client.get_stats(stat_type="", ids=[_id],
                                                       start_date=start_date, end_date=end_date) or []
                        except NaverApiCircuitOpenError:
                            r = []
                        except Exception:
                            return []  # 이 id 실제 오류 — 포기(드묾)
                    if r:
                        return r
                    if _stats_breaker.is_open():
                        await asyncio.sleep(3)
                        continue
                    return []  # breaker 안 열렸는데 빈 결과 = 진짜 무실적
                return []

            # ── 재개 판단: 최근(<6h) plan 이 같은 days/min_clicks 면 재사용 ──
            reuse = None
            if doc and not force and doc.get("days") == days and doc.get("min_clicks") == min_clicks:
                try:
                    built = datetime.fromisoformat(doc.get("plan_built_at"))
                    if (datetime.now() - built).total_seconds() < 6 * 3600 and doc.get("plan"):
                        reuse = doc
                except Exception:
                    reuse = None

            if reuse:
                plan = reuse["plan"]
                clicked_groups_cost = [tuple(x) for x in plan["clicked_groups"]]  # [(gid,cost)]
                gid_to_camp = plan["gid_to_camp"]
                camp_meta = plan["camp_meta"]
                camp_break = reuse.get("campaign_breakdown") or []
                processed = set(reuse.get("processed_group_ids") or [])
                items = list(reuse.get("items_full") or [])
                plan_built_at = reuse.get("plan_built_at")
                st["campaigns_total"] = reuse.get("campaigns_total") or 0
                st["campaigns_clicked"] = reuse.get("campaigns_clicked") or 0
                st["groups_scanned"] = reuse.get("groups_scanned") or 0
                st["groups_clicked"] = len(clicked_groups_cost)
                st["phase"] = "resume_keyword_stats"
            else:
                # 1) 캠페인 전수 stat → 예산 브레이크다운
                st["phase"] = "campaigns"
                camps = [c for c in _as_list(await client.get_campaigns() or []) if c.get("nccCampaignId")]
                st["campaigns_total"] = len(camps)
                st["phase"] = "campaign_stats"
                camp_stats = await asyncio.gather(*[_stat(c["nccCampaignId"]) for c in camps])
                camp_break = []
                clicked_camps = []
                camp_meta = {}
                for c, s in zip(camps, camp_stats):
                    s0 = (s or [None])[0] or {}
                    try:
                        clk = int(s0.get("clkCnt", 0) or 0)
                    except Exception:
                        clk = 0
                    cost = int(s0.get("salesAmt", 0) or 0)
                    cmid = c["nccCampaignId"]
                    camp_meta[cmid] = {"name": c.get("name")}
                    camp_break.append({
                        "campaign_id": cmid, "name": c.get("name"),
                        "campaignTp": c.get("campaignTp"),
                        "dailyBudget": c.get("dailyBudget"), "useDailyBudget": c.get("useDailyBudget"),
                        "clicks": clk, "cost": cost, "impressions": int(s0.get("impCnt", 0) or 0),
                    })
                    if clk >= min_clicks:
                        clicked_camps.append(c)
                st["campaigns_clicked"] = len(clicked_camps)
                # 예산답은 여기서 이미 확정 — 즉시 저장(빠른 예산답 + 재시작내성)
                plan_built_at = datetime.now().isoformat()
                try:
                    with open(_census_path(cid), "w", encoding="utf-8") as _f:
                        _json.dump({"state": "running", "phase": "campaign_done",
                                    "days": days, "min_clicks": min_clicks,
                                    "plan_built_at": plan_built_at,
                                    "campaigns_total": st["campaigns_total"],
                                    "campaigns_clicked": st["campaigns_clicked"],
                                    "groups_scanned": 0, "groups_clicked": 0,
                                    "campaign_breakdown": camp_break,
                                    "processed_group_ids": [], "items_full": [],
                                    "breaker_open": _stats_breaker.is_open(),
                                    "error": None, "plan": None}, _f, ensure_ascii=False)
                except Exception:
                    pass

                async def _fg(cmid):
                    async with sem:
                        try:
                            gs = _as_list(await client.get_ad_groups(campaign_id=cmid) or [])
                            return cmid, [g["nccAdgroupId"] for g in gs if g.get("nccAdgroupId")]
                        except Exception:
                            return cmid, []

                st["phase"] = "groups"
                fg_res = await asyncio.gather(*[_fg(c["nccCampaignId"]) for c in clicked_camps])
                gid_to_camp = {}
                all_groups = []
                for cmid, gids in fg_res:
                    for g in gids:
                        gid_to_camp[g] = cmid
                        all_groups.append(g)
                st["groups_scanned"] = len(all_groups)
                st["phase"] = "group_stats"
                grp_stats = await asyncio.gather(*[_stat(g) for g in all_groups])
                clicked_groups_cost = sorted(
                    [(g, int(((s or [None])[0] or {}).get("salesAmt", 0) or 0))
                     for g, s in zip(all_groups, grp_stats) if _clk(s) >= min_clicks],
                    key=lambda x: -x[1])  # 비용 큰 그룹부터
                st["groups_clicked"] = len(clicked_groups_cost)
                processed = set()
                items = []

            def _save(state_val):
                st["state"] = state_val
                st["breaker_open"] = _stats_breaker.is_open()
                st["clicked_keywords_total"] = len(items)
                st["groups_processed"] = len(processed)
                doc_out = {
                    "state": state_val, "phase": st["phase"], "days": days, "min_clicks": min_clicks,
                    "plan_built_at": plan_built_at,
                    "campaigns_total": st["campaigns_total"], "campaigns_clicked": st["campaigns_clicked"],
                    "groups_scanned": st["groups_scanned"], "groups_clicked": st["groups_clicked"],
                    "campaign_breakdown": camp_break,
                    "processed_group_ids": list(processed),
                    "items_full": items,
                    "breaker_open": st["breaker_open"], "error": st.get("error"),
                    "plan": {"clicked_groups": [list(x) for x in clicked_groups_cost],
                             "gid_to_camp": gid_to_camp, "camp_meta": camp_meta},
                }
                try:
                    with open(_census_path(cid), "w", encoding="utf-8") as f:
                        _json.dump(doc_out, f, ensure_ascii=False)
                except Exception as fe:
                    logger.warning(f"[clicked-census-bg] 저장 실패: {fe}")

            # 예산 브레이크다운 즉시 저장
            st["phase"] = st.get("phase") or "keyword_stats"
            _save("running")

            async def _fk(gid):
                async with sem:
                    try:
                        ks = _as_list(await client.get_keywords(ad_group_id=gid) or [])
                        return [(k.get("nccKeywordId"), k.get("keyword"), gid)
                                for k in ks if k.get("nccKeywordId") and k.get("keyword")]
                    except Exception:
                        return []

            # 2) 남은 클릭 그룹을 비용순으로 배치 처리 (배치마다 증분 저장)
            st["phase"] = "keyword_stats"
            remaining = [gc for gc in clicked_groups_cost if gc[0] not in processed]
            BATCH = 8
            for bi in range(0, len(remaining), BATCH):
                chunk = [gc[0] for gc in remaining[bi:bi + BATCH]]
                fk_res = await asyncio.gather(*[_fk(g) for g in chunk])
                triples = [t for sub in fk_res for t in sub]
                kw_stats = await asyncio.gather(*[_stat(nid) for nid, _, _ in triples])
                for (nid, txt, gid), s in zip(triples, kw_stats):
                    s0 = (s or [None])[0] or {}
                    try:
                        clk = int(s0.get("clkCnt", 0) or 0)
                    except Exception:
                        clk = 0
                    if clk < min_clicks:
                        continue
                    t = (txt or "").replace(" ", "")
                    if any(n in t for n in OFF):
                        verdict = "off_by_negative"
                    elif not any(o in t for o in ON):
                        verdict = "off_no_ondomain"
                    else:
                        verdict = "on_domain"
                    cmid = gid_to_camp.get(gid)
                    items.append({
                        "keyword_id": nid, "keyword": txt, "clicks": clk,
                        "impressions": int(s0.get("impCnt", 0) or 0),
                        "cost": int(s0.get("salesAmt", 0) or 0),
                        "ctr": float(s0.get("ctr", 0) or 0), "cpc": int(s0.get("cpc", 0) or 0),
                        "campaign": (camp_meta.get(cmid) or {}).get("name"),
                        "verdict": verdict,
                    })
                processed.update(chunk)
                _save("running")

            _save("done")
            logger.warning(
                f"[clicked-census-bg] cid={cid} DONE camps={st['campaigns_total']}/{st['campaigns_clicked']} "
                f"grps={st['groups_clicked']} clicked_kw={len(items)}"
            )
        except Exception as e:
            st["state"] = "error"
            st["error"] = f"{type(e).__name__}: {str(e)[:200]}"
            logger.error(f"[clicked-census-bg] cid={cid} 실패: {e}", exc_info=True)

    background_tasks.add_task(_do)
    return {"success": True, "started": True, "customer_id": cid, "days": days,
            "min_clicks": min_clicks, "resuming": bool(doc and not force),
            "message": "클릭 census 백그라운드 시작(증분저장·재개) — status 로 진행/결과 확인"}


@router.get("/keyword-pool/diagnostics/clicked-census-bg/status")
async def keyword_pool_clicked_census_bg_status(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """클릭 census 진행상태/결과 조회. 인메모리 없으면(재시작 등) /data 파일에서 복구."""
    account = _resolve_account(user_id, customer_id)
    if not account:
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    mem = _CLICK_CENSUS_STATUS.get(cid)
    doc = _census_load(cid)
    if doc is not None:
        summary = _census_summarize(doc)
        # 인메모리가 running 이면 진행 phase 를 덮어써 최신 반영
        if mem and mem.get("state") == "running":
            summary["state"] = "running"
            summary["phase"] = mem.get("phase")
            summary["groups_processed"] = mem.get("groups_processed", summary.get("groups_processed"))
        return {"success": True, "customer_id": cid, "source": "file", "status": summary}
    return {"success": True, "customer_id": cid, "source": "memory",
            "status": mem or {"state": "none"}}


# ── 임의 기간 키워드 실적 전수 (batch multi-id /stats) ──────────────────────────
# 왜 census 와 별도인가: clicked-census 는 **키워드 1개당 /stats 1콜**이라 소잠(등록 12만)
# 전수에 1~3시간 걸리고 days 기준이라 "지난달" 같은 구간을 못 본다.
# ★2026-07-30 실측: /stats 는 `ids` 를 **반복 쿼리파라미터**로 넘기면 다건을 한 콜에 준다.
#   repeated(ids=a&ids=b) → 200 / comma → 200 / **JSON 배열 → 400 code 11001**.
#   기존 코드 주석의 "multi-id 11001" 은 JSON 배열로 보낸 탓이었다. 100개씩 묶으면 콜수 1/100.
# 키워드 ID 유니버스는 **계정에서 직접 열거**(campaign→adgroup→keyword). registered_keywords
# DB 에서 뽑으면 풀 자동등록분만 나와 '파워링크-대표키워드' 같은 수동 레거시 캠페인이 통째로
# 누락된다 — 문의를 실제로 만드는 머리어가 거기 있어서 치명적.
_KWWIN_STATUS: Dict[int, dict] = {}
_KWWIN_ID_BATCH = 100        # /stats 한 콜에 넣는 ID 수
_KWWIN_SEM = 4               # 네이버 동시 호출 — breaker OPEN 을 유발하지 않는 실측 상한


def _kwwin_path(cid: int) -> str:
    return f"/data/_kw_window_{cid}.json"


def _kwwin_map_path(cid: int) -> str:
    return f"/data/_kw_window_{cid}_map.json"


def _kwwin_load(cid: int) -> Optional[dict]:
    import json as _json
    try:
        with open(_kwwin_path(cid), "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return None


def _kwwin_save(cid: int, doc: dict) -> None:
    """원자적 저장 — 재시작/재배포 중 부분쓰기로 파일이 깨지지 않게."""
    import json as _json
    import os as _os
    p = _kwwin_path(cid)
    tmp = f"{p}.tmp"
    try:
        with open(tmp, "w", encoding="utf-8") as f:
            _json.dump(doc, f, ensure_ascii=False)
        _os.replace(tmp, p)
    except Exception as e:
        logger.warning(f"[kwwin] 저장 실패 cid={cid}: {str(e)[:120]}")


@router.post("/keyword-pool/diagnostics/keyword-window-stats")
async def keyword_window_stats(
    background_tasks: BackgroundTasks,
    start: str = Query(..., description="시작일 YYYY-MM-DD"),
    end: str = Query(..., description="종료일 YYYY-MM-DD"),
    min_clicks: int = 1,
    force: bool = False,
    enum_group_cap: int = Query(1500, description="레거시(수동) 캠페인 그룹 열거 상한"),
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """지정 기간의 **키워드 단위 노출/클릭/광고비 전수**를 백그라운드로 수집.

    · phase enumerate: 캠페인·광고그룹 열거 → 키워드는 registered_keywords DB 로드(API 0콜)
      + DB 미커버(수동 레거시) 캠페인의 그룹만 API 보충
    · phase daily: 일자별 캠페인 합계 → '클릭 일어난 일수' 시계열
    · phase keywords: ID 100개씩 묶어 /stats → clicks ≥ min_clicks 만 적재
    진행/결과: GET .../keyword-window-stats/status?customer_id=&limit=
    """
    from services.naver_ad_service import NaverAdApiClient, _stats_breaker, NaverApiCircuitOpenError
    import json as _json

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    prev = _KWWIN_STATUS.get(cid)
    if prev and prev.get("state") == "running" and not force:
        return {"success": True, "already_running": True, "customer_id": cid,
                "status": {k: prev.get(k) for k in ("state", "phase", "ids_total",
                           "batches_done", "batches_total", "clicked_total")},
                "message": "이미 수집 중 — status 로 확인"}

    st: dict = {
        "state": "running", "phase": "enumerate", "start": start, "end": end,
        "min_clicks": min_clicks, "started_at": datetime.now().isoformat(),
        "campaigns_total": 0, "groups_total": 0, "ids_total": 0,
        "batches_done": 0, "batches_total": 0, "clicked_total": 0,
        "total_clicks": 0, "total_cost": 0, "total_impressions": 0,
        "daily": [], "items": [], "error": None,
    }
    _KWWIN_STATUS[cid] = st

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    def _num(v) -> int:
        try:
            return int(float(v or 0))
        except Exception:
            return 0

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]
        sem = asyncio.Semaphore(_KWWIN_SEM)
        fields = _json.dumps(["impCnt", "clkCnt", "salesAmt", "ctr", "cpc", "avgRnk"])

        async def _stats(ids: List[str], since: str, until: str) -> List[dict]:
            """ids 를 반복 파라미터로 — httpx 가 list 값을 ids=a&ids=b 로 직렬화."""
            params = {
                "ids": list(ids),
                "fields": fields,
                "timeRange": _json.dumps({"since": since, "until": until}),
            }
            for _ in range(3):
                if _stats_breaker.is_open():
                    # breaker 열렸다고 skip 하면 그 ID 들이 clicks=0 으로 조용히 누락된다
                    # (census 1차본이 클릭 16% 만 포착한 원인). 닫힐 때까지 기다린다.
                    await asyncio.sleep(10)
                    continue
                async with sem:
                    try:
                        resp = await client._request("GET", "/stats", params)
                        return _as_list(resp)
                    except NaverApiCircuitOpenError:
                        await asyncio.sleep(10)
                    except Exception as e:
                        logger.warning(f"[kwwin] stats 실패 n={len(ids)}: {str(e)[:120]}")
                        return []
            return []

        try:
            # ── 1. 캠페인 → 그룹 → 키워드 열거 ────────────────────────────────
            camps = _as_list(await client._request("GET", "/ncc/campaigns", None))
            camp_name = {c.get("nccCampaignId"): c.get("name") for c in camps}
            camp_budget = {c.get("nccCampaignId"): _num(c.get("dailyBudget")) for c in camps}
            st["campaigns_total"] = len(camps)

            async def _groups(camp_id: str) -> List[dict]:
                async with sem:
                    try:
                        return _as_list(await client._request(
                            "GET", "/ncc/adgroups", {"nccCampaignId": camp_id}))
                    except Exception:
                        return []

            gres = await asyncio.gather(*[_groups(c.get("nccCampaignId")) for c in camps])
            groups = [g for batch in gres for g in batch]
            gid_to_camp = {g.get("nccAdgroupId"): g.get("nccCampaignId") for g in groups}
            st["groups_total"] = len(groups)
            st["phase"] = "group_stats"

            # ★키워드 열거 범위는 '이 기간에 돈을 쓴 그룹'으로 정한다.
            #  그룹당 1콜로 전체(5,948)를 훑으면 Fly→네이버 outbound 타임아웃으로 멈추고
            #  (2026-07-31 실측: 10분간 29,334개에서 정지), registered_keywords DB 로
            #  대체하면 **수동 등록 키워드가 통째로 빠진다**. DB 커버리지를 캠페인 단위로
            #  판정했을 때 광고비의 52.6%, 그룹 단위로 고쳐도 38.0% 가 미포착이었다
            #  (풀 키워드가 1개라도 섞인 그룹의 수동 키워드는 여전히 누락).
            #  → 그룹 stat 을 100개씩 배치로 먼저 받아(5,948그룹 = 60콜) 비용>0 인 그룹만
            #    API 로 전수 열거한다. 돈 안 쓴 그룹은 분석에 기여하지 않으므로 생략해도
            #    합계가 어긋나지 않는다.
            gids_all = [g for g in gid_to_camp if g]
            group_cost: Dict[str, int] = {}
            for i in range(0, len(gids_all), _KWWIN_ID_BATCH):
                for r in await _stats(gids_all[i:i + _KWWIN_ID_BATCH], start, end):
                    gid = r.get("id")
                    if gid:
                        group_cost[gid] = _num(r.get("salesAmt")) + _num(r.get("clkCnt"))
                st["groups_statted"] = len(group_cost)
                await asyncio.sleep(0)

            active_gids = [g for g, v in group_cost.items() if v > 0]
            st["active_groups"] = len(active_gids)
            st["phase"] = "enumerate_keywords"
            if len(active_gids) > enum_group_cap:
                logger.warning(f"[kwwin] 활성 그룹 {len(active_gids)} > cap {enum_group_cap} — 초과분 미열거")
                st["groups_skipped"] = len(active_gids) - enum_group_cap
                active_gids = active_gids[:enum_group_cap]
            legacy_gids = active_gids

            kw_map: Dict[str, dict] = {}
            enum_fail = 0

            async def _kws(gid: str) -> List[dict]:
                nonlocal enum_fail
                for _ in range(3):
                    async with sem:
                        try:
                            return _as_list(await client._request(
                                "GET", "/ncc/keywords", {"nccAdgroupId": gid}))
                        except NaverApiCircuitOpenError:
                            pass
                        except Exception as e:
                            logger.warning(f"[kwwin] 그룹 {gid} 열거 실패: {str(e)[:100]}")
                            enum_fail += 1
                            return []
                    await asyncio.sleep(5)   # breaker 대기 — skip 하면 조용히 누락된다
                enum_fail += 1
                return []

            CH = 100
            for i in range(0, len(legacy_gids), CH):
                chunk = legacy_gids[i:i + CH]
                res = await asyncio.gather(*[_kws(g) for g in chunk])
                for g, kws in zip(chunk, res):
                    for k in kws:
                        kid = k.get("nccKeywordId")
                        if not kid:
                            continue
                        kw_map[kid] = {
                            "keyword": k.get("keyword"),
                            "group_id": g,
                            "campaign_id": gid_to_camp.get(g),
                            "bid": _num(k.get("bidAmt")),
                            "user_lock": bool(k.get("userLock")),
                        }
                st["ids_total"] = len(kw_map)
                st["enum_failed_groups"] = enum_fail
                await asyncio.sleep(0)  # 이벤트 루프 양보 — login 등 다른 요청 보호
            try:
                with open(_kwwin_map_path(cid), "w", encoding="utf-8") as f:
                    _json.dump(kw_map, f, ensure_ascii=False)
            except Exception:
                pass

            # ── 2. 일자별 캠페인 합계 → '클릭 일어난 일수' ──────────────────────
            st["phase"] = "daily"
            camp_ids = [c.get("nccCampaignId") for c in camps if c.get("nccCampaignId")]
            d0 = datetime.strptime(start, "%Y-%m-%d")
            d1 = datetime.strptime(end, "%Y-%m-%d")
            daily = []
            cur = d0
            while cur <= d1:
                ds = cur.strftime("%Y-%m-%d")
                rows: List[dict] = []
                for i in range(0, len(camp_ids), _KWWIN_ID_BATCH):
                    rows += await _stats(camp_ids[i:i + _KWWIN_ID_BATCH], ds, ds)
                daily.append({
                    "date": ds,
                    "clicks": sum(_num(r.get("clkCnt")) for r in rows),
                    "cost": sum(_num(r.get("salesAmt")) for r in rows),
                    "impressions": sum(_num(r.get("impCnt")) for r in rows),
                })
                st["daily"] = daily
                cur += timedelta(days=1)
            _kwwin_save(cid, dict(st))

            # ── 3. 키워드 전수 stat (100개 배치) ──────────────────────────────
            st["phase"] = "keywords"
            all_ids = list(kw_map.keys())
            batches = [all_ids[i:i + _KWWIN_ID_BATCH]
                       for i in range(0, len(all_ids), _KWWIN_ID_BATCH)]
            st["batches_total"] = len(batches)
            items: List[dict] = []
            tot_c = tot_cost = tot_imp = 0
            for bi, batch in enumerate(batches):
                rows = await _stats(batch, start, end)
                for r in rows:
                    kid = r.get("id")
                    clk = _num(r.get("clkCnt"))
                    tot_c += clk
                    tot_cost += _num(r.get("salesAmt"))
                    tot_imp += _num(r.get("impCnt"))
                    if clk < min_clicks or not kid:
                        continue
                    meta = kw_map.get(kid) or {}
                    items.append({
                        "keyword_id": kid,
                        "keyword": meta.get("keyword"),
                        "campaign": camp_name.get(meta.get("campaign_id")),
                        "campaign_id": meta.get("campaign_id"),
                        "group_id": meta.get("group_id"),
                        "daily_budget": camp_budget.get(meta.get("campaign_id")),
                        "bid": meta.get("bid"),
                        "impressions": _num(r.get("impCnt")),
                        "clicks": clk,
                        "cost": _num(r.get("salesAmt")),
                        "ctr": float(r.get("ctr") or 0),
                        "cpc": _num(r.get("cpc")),
                        "avg_rank": float(r.get("avgRnk") or 0),
                    })
                st["batches_done"] = bi + 1
                st["clicked_total"] = len(items)
                st["total_clicks"], st["total_cost"], st["total_impressions"] = tot_c, tot_cost, tot_imp
                st["items"] = items
                if (bi + 1) % 20 == 0:
                    _kwwin_save(cid, dict(st))
                await asyncio.sleep(0)

            st["state"] = "done"
            st["phase"] = "done"
            st["finished_at"] = datetime.now().isoformat()
            _kwwin_save(cid, dict(st))
        except Exception as e:
            import traceback
            logger.error(f"[kwwin] 실패: {traceback.format_exc()}")
            st["state"] = "error"
            st["error"] = f"{type(e).__name__}: {str(e)[:300]}"
            _kwwin_save(cid, dict(st))
        finally:
            try:
                await client.close()
            except Exception:
                pass

    background_tasks.add_task(_run)
    return {"success": True, "customer_id": cid, "started": True,
            "start": start, "end": end,
            "message": "수집 시작 — GET .../keyword-window-stats/status 로 확인"}


@router.get("/keyword-pool/diagnostics/keyword-window-stats/status")
async def keyword_window_stats_status(
    customer_id: Optional[str] = None,
    limit: int = 5000,
    offset: int = 0,
    sort: str = Query("cost", description="cost|clicks"),
    user_id: int = Depends(get_user_id_with_fallback),
):
    """진행/결과. 인메모리 없으면(재시작) /data 파일에서 복구."""
    account = _resolve_account(user_id, customer_id)
    if not account:
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    doc = _KWWIN_STATUS.get(cid) or _kwwin_load(cid)
    if not doc:
        return {"success": True, "customer_id": cid, "status": {"state": "none"}}
    items = list(doc.get("items") or [])
    key = "clicks" if sort == "clicks" else "cost"
    items.sort(key=lambda x: -int(x.get(key) or 0))
    summary = {k: v for k, v in doc.items() if k != "items"}
    summary["items_returned"] = len(items[offset:offset + limit])
    return {"success": True, "customer_id": cid, "status": summary,
            "items": items[offset:offset + limit]}


class BulkDeleteKeywordsRequest(BaseModel):
    keyword_ids: List[str]


@router.post("/keyword-pool/clicked-keywords/bulk-delete")
async def keyword_pool_bulk_delete_clicked(
    request: BulkDeleteKeywordsRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """선택된 키워드 일괄 네이버 삭제 (실패 시 PAUSE) + 풀 mark + reg DB 제거."""
    from services.naver_ad_service import NaverAdApiClient
    import sqlite3 as _sqlite3

    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        customer_id = int(account.get("customer_id"))

        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]

        reg = get_registered_keywords_db()
        pool = get_keyword_pool_db()

        n_deleted = 0
        n_paused = 0
        n_failed = 0
        affected_keywords: List[str] = []

        for kid in request.keyword_ids:
            with _sqlite3.connect(reg.db_path) as conn:
                row = conn.execute(
                    "SELECT keyword FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id=?",
                    (customer_id, kid),
                ).fetchone()
            kw_text = row[0] if row else None
            try:
                await client.delete_keyword(kid)
                with _sqlite3.connect(reg.db_path) as conn:
                    conn.execute(
                        "DELETE FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id=?",
                        (customer_id, kid),
                    )
                n_deleted += 1
                if kw_text: affected_keywords.append(kw_text)
            except Exception:
                try:
                    await client.pause_keyword(kid)
                    n_paused += 1
                    if kw_text: affected_keywords.append(kw_text)
                except Exception:
                    n_failed += 1
            await asyncio.sleep(0.15)

        if affected_keywords:
            pool.mark_rejected_by_naver(
                customer_id,
                [{"keyword": kw, "reason": "사용자 일괄 삭제 (클릭 검수)"} for kw in affected_keywords],
            )
        return {"success": True, "deleted": n_deleted, "paused": n_paused, "failed": n_failed}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"bulk-delete-clicked 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


# ============ 연관성-점수 기반 자동 cleanup ============
# 사용자가 토글 ON + threshold(예: 30) 지정 → cron 이 매시 1회 클릭 발생 KW 중
# 점수 ≤ threshold 인 것 자동 DELETE (실패 시 PAUSE). 검수 없이 점수만 본다.
# 클릭 미발생 KW 는 건드리지 않음 — 노출 받기 전 KW 는 점수 낮아도 cost 0.

class AutoCleanupSettingsRequest(BaseModel):
    enabled: Optional[bool] = None
    threshold: Optional[int] = None  # 0~95
    relevance_keywords: Optional[List[str]] = None  # 도메인 기준 키워드 list (예: ["피부질환","피부","아토피"])


@router.get("/keyword-pool/auto-cleanup/settings")
def keyword_pool_auto_cleanup_get(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """광고주별 자동 cleanup 설정 + 마지막 실행 stamp. sync def → threadpool."""
    from database.naver_ad_db import get_ad_account_auto_cleanup
    account = _resolve_account(user_id, customer_id)
    if not account:
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid_str = str(account.get("customer_id"))
    s = get_ad_account_auto_cleanup(user_id, cid_str)
    return {"success": True, "customer_id": cid_str, **s}


@router.patch("/keyword-pool/auto-cleanup/settings")
async def keyword_pool_auto_cleanup_patch(
    request: AutoCleanupSettingsRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """자동 cleanup ON/OFF 또는 threshold 변경 — 부분 업데이트.
    enabled=true 로 변경 시 background 즉시 1회 실행 — 다음 cron 정각까지 대기 안 함.
    """
    from database.naver_ad_db import update_ad_account_auto_cleanup, get_ad_account_auto_cleanup
    account = _resolve_account(user_id, customer_id)
    if not account:
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid_str = str(account.get("customer_id"))
    logger.warning(
        f"[auto-cleanup/PATCH] uid={user_id} cid={cid_str} "
        f"enabled={request.enabled} threshold={request.threshold} "
        f"rel_kws_count={len(request.relevance_keywords) if request.relevance_keywords else 'None'} "
        f"rel_kws_sample={(request.relevance_keywords or [])[:5]}"
    )
    ok = update_ad_account_auto_cleanup(
        user_id, cid_str,
        enabled=request.enabled,
        threshold=request.threshold,
        relevance_keywords=request.relevance_keywords,
    )
    if not ok:
        raise HTTPException(status_code=400, detail="변경할 필드 없음 또는 광고주 미존재")
    s = get_ad_account_auto_cleanup(user_id, cid_str)
    logger.warning(
        f"[auto-cleanup/PATCH] uid={user_id} cid={cid_str} 저장 후 SELECT 결과 "
        f"rel_kws_count={len(s.get('relevance_keywords') or [])} "
        f"rel_kws_sample={(s.get('relevance_keywords') or [])[:5]}"
    )
    # enabled=true 로 변경 시 즉시 1회 실행 — 다음 cron 까지 대기 안 함.
    # asyncio.create_task fire-and-forget — fly.io 의 BackgroundTasks 가 worker 점유로
    # cancel 되는 케이스 회피. 시작 시 즉시 last_run_at stamp → 사용자가 "실행 중" 확인.
    triggered = False
    if request.enabled is True and s.get("enabled"):
        cid_int = int(cid_str)
        thr = int(s.get("threshold") or 30)
        # 1) 즉시 stamp — 사용자가 토글 ON 직후 "최근 실행: 방금 전" 즉시 확인 가능
        try:
            from database.naver_ad_db import record_auto_cleanup_run
            record_auto_cleanup_run(user_id, cid_str, 0)
            logger.warning(f"[auto-cleanup/PATCH] uid={user_id} cid={cid_str} 즉시 stamp (실행 시작 표시)")
        except Exception as _e:
            logger.warning(f"[auto-cleanup/PATCH] 즉시 stamp 실패: {_e}")

        # 2) fire-and-forget task — uvicorn 워커가 살아있는 동안 실행
        async def _trigger_now():
            try:
                logger.warning(f"[auto-cleanup/PATCH/trigger] uid={user_id} cid={cid_int} thr={thr} 시작")
                res = await _run_auto_cleanup_for_account(user_id, cid_int, thr)
                logger.warning(
                    f"[auto-cleanup/PATCH/trigger] uid={user_id} cid={cid_int} 실행 결과: {res}"
                )
            except Exception as e:
                logger.error(
                    f"[auto-cleanup/PATCH/trigger] uid={user_id} cid={cid_int} 실행 실패: "
                    f"{type(e).__name__}: {e}", exc_info=True
                )
                try:
                    from database.naver_ad_db import record_auto_cleanup_run
                    record_auto_cleanup_run(user_id, str(cid_int), 0)
                except Exception:
                    pass
        try:
            asyncio.create_task(_trigger_now())
            triggered = True
        except Exception as _e:
            # event loop 외 호출 시 — BackgroundTasks 폴백
            logger.warning(f"[auto-cleanup/PATCH] create_task 실패 → BackgroundTasks 폴백: {_e}")
            background_tasks.add_task(_trigger_now)
            triggered = True
    return {"success": True, "customer_id": cid_str, **s, "triggered_now": triggered}


async def _run_auto_cleanup_for_account(
    user_id: int, customer_id: int, threshold: int,
    days: int = 7, max_delete: int = 200,
) -> Dict:
    """한 광고주의 클릭 KW 중 점수 ≤ threshold 인 것 일괄 DELETE.
    - days: 최근 N 일 클릭 통계 (default 7)
    - max_delete: 한 tick 당 최대 삭제 수 (네이버 rate limit + 사고 방지)

    설계 (시작 stamp 제거):
    - 옛 코드는 시작 즉시 record_auto_cleanup_run(0) 으로 stamp 해서 hang 가드용
      "cron 살아있음" 표시 유지. 그러나 Naver stats circuit OPEN 시 click_cleanup
      이 0 처리 → last_deleted=0 stamp 가 domain_cleanup 의 실제 del=498 stamp 를
      overwrite → 사용자 화면 영구 "삭제 0" 으로 보이는 사고.
    - 새 정책: 처리 대상 0 이거나 circuit OPEN 이면 stamp 안 함. domain_cleanup 의
      실제 결과 stamp 만 보존. timeout 가드는 scheduler 단에서 처리.
    """
    from services.naver_ad_service import NaverAdApiClient, _stats_breaker
    from database.naver_ad_db import get_ad_account_by_customer, record_auto_cleanup_run
    from datetime import datetime, timedelta
    import sqlite3 as _sqlite3

    # Naver stats circuit OPEN 이면 click_cleanup 은 어차피 효과 0 — fly CPU 낭비 차단.
    # domain_cleanup (별도 cron) 이 circuit 무관하게 score 기반 정리하므로 누락 없음.
    if _stats_breaker.is_open():
        return {"customer_id": customer_id, "deleted": 0, "reason": "naver_stats_circuit_open"}

    account = get_ad_account_by_customer(user_id, str(customer_id))
    if not account or not account.get("is_connected"):
        return {"customer_id": customer_id, "deleted": 0, "reason": "not_connected"}

    reg = get_registered_keywords_db()
    pool = get_keyword_pool_db()
    with _sqlite3.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL",
            (customer_id,),
        ).fetchall()
    if not rows:
        record_auto_cleanup_run(user_id, str(customer_id), 0)
        return {"customer_id": customer_id, "deleted": 0, "reason": "no_registered_keywords"}
    keyword_map = {r[1]: r[0] for r in rows}

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    end_date = datetime.now().strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")

    # 1500 → 600. Naver stats API 응답 지연 (개당 1~5s) 으로 sem=3 직렬화 시 600s 도
    # 못 끝남 사례 다수 (cid=4362992 등). per-call timeout 으로 hang 차단 + sem=5 병렬도 ↑.
    # 600 ÷ 5 = 120 round × max 5s = 600s worst, 정상 시 120 × 200ms = 24s.
    ids = list(keyword_map.keys())[:600]
    sem = asyncio.Semaphore(5)
    from services.naver_ad_service import _stats_breaker, NaverApiCircuitOpenError

    PER_CALL_TIMEOUT = 5.0  # 한 stats 요청 5s 안 응답 → skip (KW 1개 잃음, hang 차단)
    n_stats_timeout = 0

    async def _fetch_one(kid: str) -> List[dict]:
        nonlocal n_stats_timeout
        # stats circuit OPEN 시 진입 즉시 skip — sem 점유 안 함
        if _stats_breaker.is_open():
            return []
        async with sem:
            try:
                stats = await asyncio.wait_for(
                    client.get_stats(
                        stat_type="KEYWORD", ids=[kid],
                        start_date=start_date, end_date=end_date,
                    ),
                    timeout=PER_CALL_TIMEOUT,
                )
                return stats or []
            except asyncio.TimeoutError:
                n_stats_timeout += 1
                return []
            except NaverApiCircuitOpenError:
                return []
            except Exception as e:
                logger.warning(f"[auto-cleanup] stats {kid} 실패: {str(e)[:120]}")
                return []

    results = await asyncio.gather(*[_fetch_one(kid) for kid in ids])
    all_stats = [s for batch in results for s in batch]
    if n_stats_timeout > 0:
        logger.warning(
            f"[auto-cleanup] uid={user_id} cid={customer_id} stats per-call timeout "
            f"{n_stats_timeout}/{len(ids)} (5s 초과 — Naver API 응답 지연)"
        )
    logger.warning(
        f"[auto-cleanup] uid={user_id} cid={customer_id} stats fetched ids={len(ids)} "
        f"non_empty={len(all_stats)} circuit_open={_stats_breaker.is_open()}"
    )

    # cron 자동 cleanup 도 cleanup-by-score 와 동일 우선순위:
    # ad_accounts.relevance_keywords (사용자 명시) → user_seed 폴백.
    from database.naver_ad_db import get_ad_account_relevance_keywords
    saved_basis = get_ad_account_relevance_keywords(user_id, str(customer_id))
    if saved_basis:
        user_seeds = saved_basis
    else:
        user_seeds = [s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2]
    targets: List[Tuple[str, str, int]] = []  # (kid, kw, score)
    for stat in all_stats:
        kid = stat.get("id")
        if not kid:
            continue
        clicks = int(stat.get("clkCnt", 0) or 0)
        if clicks <= 0:
            continue  # 클릭 미발생 KW 는 건드리지 않음
        kw = keyword_map.get(kid)
        if not kw:
            continue
        score = _compute_relevance_score(kw, user_seeds)
        if score < threshold:  # Option B: boundary 보존
            targets.append((kid, kw, score))

    # 가장 무관한 KW 부터 (낮은 점수 우선) — max_delete 캡 적용
    targets.sort(key=lambda x: x[2])
    targets = targets[:max_delete]

    n_deleted = 0
    n_paused = 0
    n_failed = 0
    n_stale_purged = 0  # 네이버 404 = 이미 사라진 KW → DB stale row 만 정리
    affected: List[str] = []
    import httpx as _httpx
    for kid, kw_text, _score in targets:
        try:
            await client.delete_keyword(kid)
            with _sqlite3.connect(reg.db_path) as conn:
                conn.execute(
                    "DELETE FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id=?",
                    (customer_id, kid),
                )
            n_deleted += 1
            affected.append(kw_text)
        except _httpx.HTTPStatusError as e:
            # 404 "No permission to access the resource" = 네이버 콘솔에서 이미 사라진 KW.
            # 옛 코드는 fail 카운트해서 DB row 영구 보존 → 한도 stale, register 가 cap 거부됨.
            # 이제 DB row 도 같이 제거 → 실제 한도 회수.
            if getattr(e, "response", None) is not None and e.response.status_code == 404:
                with _sqlite3.connect(reg.db_path) as conn:
                    conn.execute(
                        "DELETE FROM registered_keywords WHERE account_customer_id=? AND ncc_keyword_id=?",
                        (customer_id, kid),
                    )
                n_stale_purged += 1
                affected.append(kw_text)
            else:
                try:
                    await client.pause_keyword(kid)
                    n_paused += 1
                    affected.append(kw_text)
                except Exception:
                    n_failed += 1
        except Exception:
            try:
                await client.pause_keyword(kid)
                n_paused += 1
                affected.append(kw_text)
            except Exception:
                n_failed += 1
        await asyncio.sleep(0.15)

    if affected:
        pool.mark_rejected_by_naver(
            customer_id,
            [{"keyword": kw, "reason": f"자동 cleanup (점수≤{threshold})"} for kw in affected],
        )
    # 실행 이력 — 화면 '최근 실행 이력' 표에 노출
    total_purged = n_deleted + n_stale_purged
    try:
        pool.record_run(
            user_id, customer_id, "inspect",
            "success" if total_purged > 0 else "no_new",
            registered=0, failed=n_failed, skipped=total_purged,
            seeds_count=len(targets),
            error_message=(
                f"자동 cleanup (점수≤{threshold}) — DELETE {n_deleted} / 404 stale {n_stale_purged} / PAUSE {n_paused} / 실패 {n_failed}"
                if (total_purged or n_paused or n_failed) else f"자동 cleanup (점수≤{threshold}) — 대상 0"
            ),
        )
    except Exception:
        pass

    # 실제 정리한 게 있을 때만 stamp — 0 이면 domain_cleanup 의 이전 stamp 보존.
    # 사용자 화면 "최근 실행 N개" 가 의미있는 결과만 반영되도록.
    if total_purged + n_paused > 0:
        record_auto_cleanup_run(user_id, str(customer_id), total_purged + n_paused)

    # 임계 auto-promote — cleanup 직후 풀의 점수 분포 검사 → 90%+ 가 thr+10 이상이면
    # threshold 를 +10 상향. 점진 수렴: 30 → 40 → 50 → ... → 80 cap.
    # cleanup 으로 풀이 점수 ≥ thr 만 남으면 다음 단계로 자동 진입 → 사용자 개입 없이
    # "모든 KW 가 점수 N 이상" 목표에 수렴.
    try:
        promoted_to = await _maybe_promote_auto_cleanup_threshold(
            user_id, customer_id, threshold,
        )
        if promoted_to:
            logger.warning(
                f"[auto-cleanup/promote] uid={user_id} cid={customer_id} "
                f"threshold {threshold} → {promoted_to} (풀 90%+ ≥{threshold+10})"
            )
    except Exception as e:
        logger.warning(f"[auto-cleanup/promote] 실패 uid={user_id} cid={customer_id}: {e}")

    return {
        "customer_id": customer_id,
        "threshold": threshold,
        "candidates": len(targets),
        "deleted": n_deleted,
        "stale_purged": n_stale_purged,
        "paused": n_paused,
        "failed": n_failed,
    }


async def _maybe_promote_auto_cleanup_threshold(
    user_id: int, customer_id: int, current_threshold: int,
    *, sample_size: int = 1500, promote_step: int = 5,
    promote_ratio: float = 0.75, max_threshold: int = 75,
) -> Optional[int]:
    """풀의 점수 분포 검사 후 threshold 자동 상향.

    조건 (모두 충족 시 +promote_step):
      - current_threshold < max_threshold (75 이상이면 더 안 올림)
      - 등록 KW 수 ≥ 5000 (샘플 신뢰성)
      - 샘플 1500 random 중 ≥ promote_ratio(75%) 가 점수 ≥ current_threshold + promote_step

    why: cleanup 으로 점수≤thr KW 빠지면 풀 점수 분포가 thr 이상으로 수렴.
    분포의 75%가 thr+5 까지 도달했다면 다음 단계로 진입할 수 있다는 신호.
    구버전 (ratio 0.90, step 10) 은 promote 너무 보수적이라 풀 점수 31~39 분포에서
    영원히 promote 안 됨 → cleanup 대상 0 영구 정체 사고. 75% / +5 로 점진 적응.
    cap=75 — 너무 엄격해지면 빈 슬롯 못 채움 위험.
    """
    from database.naver_ad_db import (
        get_ad_account_relevance_keywords, update_ad_account_auto_cleanup,
    )
    import sqlite3 as _sqlite3
    import random as _random

    if current_threshold >= max_threshold:
        return None

    next_threshold = current_threshold + promote_step
    if next_threshold > max_threshold:
        next_threshold = max_threshold

    reg = get_registered_keywords_db()
    pool = get_keyword_pool_db()

    with _sqlite3.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL "
            "AND removed_at IS NULL",
            (customer_id,),
        ).fetchall()
    keywords = [r[0] for r in rows if r and r[0]]
    if len(keywords) < 5000:
        return None  # 풀 너무 작음 — 샘플 신뢰성 부족

    # 점수 기준 — saved_relevance > user_seed 폴백
    saved = get_ad_account_relevance_keywords(user_id, str(customer_id))
    if saved and len([s for s in saved if s and len(s) >= 2]) >= 3:
        score_basis = saved
    else:
        score_basis = [
            s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2
        ]
    if not score_basis:
        return None  # 점수 계산 불가

    sample = (
        _random.sample(keywords, sample_size) if len(keywords) > sample_size
        else keywords
    )
    pass_count = 0
    for kw in sample:
        if _compute_relevance_score(kw, score_basis) >= next_threshold:
            pass_count += 1
    ratio = pass_count / len(sample)
    if ratio < promote_ratio:
        return None

    ok = update_ad_account_auto_cleanup(
        user_id, str(customer_id), threshold=next_threshold,
    )
    if not ok:
        return None
    # 진행 이력 — 화면 '최근 실행 이력' 표에 노출
    try:
        pool.record_run(
            user_id, customer_id, "inspect", "success",
            error_message=(
                f"threshold auto-promote {current_threshold} → {next_threshold} "
                f"(샘플 {len(sample)} 중 {pass_count} ≥{next_threshold}, "
                f"ratio {ratio:.1%})"
            )[:300],
        )
    except Exception:
        pass
    return next_threshold


async def _run_domain_cleanup_for_account(
    user_id: int, customer_id: int, threshold: int = 30, max_delete: int = 750,
) -> Dict:
    """click 무관 — relevance_keywords 점수 ≤ threshold 인 등록 KW 일괄 DELETE.

    cron 으로 매 시간 실행되어 100k 풀에서 도메인 안 맞는 무관 KW 를 점진 정리.
    한 tick 당 max_delete (default 500) 제한 — Naver rate + 사고 방지.
    빈 자리는 collect/register cron 이 새 도메인 KW 로 채움 → 100k 가 점진적으로 도메인 KW 100% 로 수렴.
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import (
        get_ad_account_by_customer,
        get_ad_account_relevance_keywords,
        record_auto_cleanup_run,
    )
    import sqlite3 as _sqlite3
    import time as _t

    t0 = _t.monotonic()
    # start stamp 제거 — 0 stamp 가 다른 cleanup 의 실제 결과 stamp 를 overwrite
    # 하는 사고 차단. 의미있는 결과만 stamp (n_del + n_stale > 0 일 때).
    account = get_ad_account_by_customer(user_id, str(customer_id))
    if not account or not account.get("is_connected"):
        return {"customer_id": customer_id, "deleted": 0, "reason": "not_connected"}

    # 점수 기준 키워드 — saved relevance > user_seed 폴백
    saved = get_ad_account_relevance_keywords(user_id, str(customer_id))
    pool = get_keyword_pool_db()
    if saved and len(saved) >= 1:
        score_basis = saved
        basis = "saved_relevance"
    else:
        score_basis = [s for s in (pool.list_user_seeds(customer_id) or []) if s and len(s) >= 2]
        basis = "user_seed"
    if not score_basis:
        return {"customer_id": customer_id, "deleted": 0, "reason": "no_score_basis"}

    # 핵심의도 앵커 + negative — 앵커 모드면 (앵커없음 OR negative) 가 삭제 기준 (점수 무관).
    required_tokens = []
    neg_tokens = []
    try:
        from database.naver_ad_db import get_domain_profile as _gdp_ct
        _pf_ct = _gdp_ct(user_id, str(customer_id)) or {}
        required_tokens = [t for t in _pf_ct.get("required_tokens", []) if t and len(t) >= 2]
        neg_tokens = [n for n in _pf_ct.get("negative_keywords", []) if n and len(n) >= 2]
    except Exception:
        pass

    reg = get_registered_keywords_db()
    with _sqlite3.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL",
            (customer_id,),
        ).fetchall()
    if not rows:
        try: record_auto_cleanup_run(user_id, str(customer_id), 0)
        except Exception: pass
        return {"customer_id": customer_id, "deleted": 0, "reason": "no_registered"}

    # 점수 매김 — atoms precompute (95k KW × atoms 재빌드 차단)
    def _score_all() -> List[Tuple[str, str, int]]:
        atoms_3plus: set = set()
        atoms_2: set = set()
        for s in score_basis:
            if not s or len(s) < 2:
                continue
            if len(s) >= 4:
                atoms_3plus.add(s)
            for n in (2, 3):
                for i in range(len(s) - n + 1):
                    a = s[i:i + n]
                    (atoms_2 if len(a) == 2 else atoms_3plus).add(a)
        out: List[Tuple[str, str, int]] = []
        for kw_text, kid in rows:
            if not kw_text:
                out.append((kid, "", 0)); continue
            sc = 0
            full = False
            for s in score_basis:
                if not s or len(s) < 2:
                    continue
                if s in kw_text: sc = 100; full = True; break
                if kw_text in s: sc = 95; full = True; break
            if not full:
                n_3 = sum(1 for a in atoms_3plus if a in kw_text)
                n_2 = sum(1 for a in atoms_2 if a in kw_text)
                sc = min(95, min(80, n_3 * 20) + min(30, n_2 * 5))
            out.append((kid, kw_text, sc))
        return out

    scored = await asyncio.to_thread(_score_all)
    if required_tokens:
        # 앵커 모드 — (앵커 하나도 없음) 또는 (negative 포함) 만 삭제. 점수 무시 (앵커있는 진짜 대출 보존).
        targets = [
            (kid, kw, s) for kid, kw, s in scored
            if kw and (
                not any(rt in kw for rt in required_tokens)
                or (neg_tokens and any(nt in kw for nt in neg_tokens))
            )
        ]
    else:
        # 비앵커 도메인 — 점수<threshold 또는 negative 포함(substring 오매칭 off-domain 차단)
        targets = [
            (kid, kw, s) for kid, kw, s in scored
            if s < threshold or (neg_tokens and kw and any(nt in kw for nt in neg_tokens))
        ]
    targets.sort(key=lambda x: x[2])  # 무관한 것부터
    targets = targets[:max(0, min(max_delete, 5000))]
    if not targets:
        try: record_auto_cleanup_run(user_id, str(customer_id), 0)
        except Exception: pass
        return {"customer_id": customer_id, "deleted": 0, "reason": "no_below_threshold",
                "total_registered": len(scored), "basis": basis}

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    n_del, n_pause, n_fail = 0, 0, 0
    n_stale = 0  # 네이버 404 = 이미 사라진 KW. 옛 코드는 fail 처리 → DB stale 누적, 한도 영구 막힘.
    affected_kws: List[str] = []
    import httpx as _httpx
    def _purge_db(kid_: str, kw_: str):
        with _sqlite3.connect(reg.db_path) as c:
            c.execute(
                "DELETE FROM registered_keywords "
                "WHERE account_customer_id=? AND ncc_keyword_id=?",
                (customer_id, kid_),
            )
        if kw_:
            with _sqlite3.connect(pool.db_path) as c:
                c.execute(
                    "UPDATE naverad_keyword_pool SET status='deleted' "
                    "WHERE account_customer_id=? AND keyword=?",
                    (customer_id, kw_),
                )
    for kid, kw_text, _s in targets:
        try:
            await client.delete_keyword(kid)
            _purge_db(kid, kw_text)
            n_del += 1
            if kw_text: affected_kws.append(kw_text)
        except _httpx.HTTPStatusError as e:
            if getattr(e, "response", None) is not None and e.response.status_code == 404:
                _purge_db(kid, kw_text)
                n_stale += 1
                if kw_text: affected_kws.append(kw_text)
            else:
                try:
                    await client.pause_keyword(kid)
                    n_pause += 1
                    if kw_text: affected_kws.append(kw_text)
                except Exception:
                    n_fail += 1
        except Exception:
            try:
                await client.pause_keyword(kid)
                n_pause += 1
                if kw_text: affected_kws.append(kw_text)
            except Exception:
                n_fail += 1
        await asyncio.sleep(0.15)

    total_purged = n_del + n_stale
    try:
        # 의미있는 결과만 stamp — 0 stamp 가 이전 cleanup 결과 overwrite 방지.
        if total_purged + n_pause > 0:
            record_auto_cleanup_run(user_id, str(customer_id), total_purged + n_pause)
        pool.record_run(
            user_id, customer_id, "inspect",
            "success" if total_purged > 0 else "no_new",
            registered=0, failed=n_fail, skipped=total_purged,
            seeds_count=len(score_basis),
            error_message=(
                f"도메인 자동 정리 ({basis}, click 무관) — DELETE {n_del} / "
                f"404 stale {n_stale} / PAUSE {n_pause} / 실패 {n_fail} / 점수≤{threshold}"
            ),
            duration_ms=int((_t.monotonic() - t0) * 1000),
        )
    except Exception:
        pass
    logger.warning(
        f"[domain-cleanup] uid={user_id} cid={customer_id} basis={basis} "
        f"thr={threshold} → del={n_del} stale={n_stale} pause={n_pause} fail={n_fail}"
    )
    return {
        "customer_id": customer_id, "deleted": n_del, "stale_purged": n_stale,
        "paused": n_pause, "failed": n_fail, "basis": basis, "threshold": threshold,
        "below_threshold_total": len(targets), "total_registered": len(scored),
    }


@router.post("/keyword-pool/cron/domain-cleanup")
async def keyword_pool_cron_domain_cleanup(
    background_tasks: BackgroundTasks,
    authorization: Optional[str] = Header(None),
    threshold: int = Query(30, ge=0, le=95),
    max_delete: int = Query(500, ge=1, le=5000),
    user_id: Optional[int] = Query(None),
    customer_id: Optional[str] = Query(None),
):
    """Bearer cron — relevance_keywords 점수 ≤ threshold 등록 KW 자동 DELETE (click 무관).

    매 1시간 실행되어 100k 풀의 무관 잔재를 점진 청소. 빈 자리는 collect/register cron 이
    새 도메인 KW 로 채움 → 100k 가 100% 도메인 KW 로 수렴 (사용자 의도).
    auto_cleanup_enabled=1 광고주만 처리.
    """
    _verify_cron_token(authorization)
    from database.naver_ad_db import list_auto_cleanup_enabled_accounts, get_ad_account_auto_cleanup

    targets: List[Tuple[int, int, int]] = []
    if user_id and customer_id:
        s = get_ad_account_auto_cleanup(user_id, str(customer_id))
        thr = threshold if threshold else int(s.get("threshold") or 30)
        targets = [(user_id, int(customer_id), int(thr))]
    else:
        rows = list_auto_cleanup_enabled_accounts() or []
        for r in rows:
            uid = int(r.get("user_id"))
            cid = int(r.get("customer_id"))
            thr = threshold if threshold else int(r.get("auto_cleanup_threshold") or 30)
            targets.append((uid, cid, thr))

    if not targets:
        return {"success": True, "queued": 0, "message": "자동 cleanup ON 광고주 없음"}

    async def _run_all():
        for uid, cid, thr in targets:
            try:
                res = await _run_domain_cleanup_for_account(uid, cid, thr, max_delete=max_delete)
                logger.info(f"[domain-cleanup/cron] uid={uid} cid={cid} thr={thr} → {res}")
            except Exception as e:
                logger.error(f"[domain-cleanup/cron] uid={uid} cid={cid} 실패: {type(e).__name__}: {e}", exc_info=True)

    background_tasks.add_task(_run_all)
    return {
        "success": True, "queued": len(targets),
        "max_delete_per_account": max_delete,
        "threshold": threshold,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


@router.post("/keyword-pool/cron/seed-amplify-burst")
async def keyword_pool_cron_seed_amplify_burst(
    background_tasks: BackgroundTasks,
    authorization: Optional[str] = Header(None),
    user_id: int = Query(...),
    customer_id: int = Query(...),
    n_calls: int = Query(10, ge=1, le=30),
    target_per_call: int = Query(500, ge=100, le=500),
):
    """시드 amplify 폭발 — n_calls 번 GPT 호출 병렬 → user_seed 풀 대량 확장.

    각 호출: user_seed 100개 random sample → amplify_seeds (target ~500) → fresh seed.
    n_calls=10 → ~5,000 신규 시드 생성 시도, dedup + keywordstool 검증 → user_seed 합류.
    mt=0 시드도 합류 (drift 감수, 시드 풀 확장 우선) — 다음 collect 라운드 atom 다양성 ↑.

    사용 예 (Bearer cron):
      POST /api/naver-ad/keyword-pool/cron/seed-amplify-burst?user_id=1&customer_id=1858907&n_calls=10
      Authorization: Bearer <CRON_TOKEN>
    """
    _verify_cron_token(authorization)

    async def _run():
        from services.ai_seed_suggester import amplify_seeds as _amp
        from services.naver_ad_service import NaverAdApiClient
        from database.naver_ad_db import get_ad_account_by_customer
        from config import settings
        import time as _t
        import random as _r

        t0 = _t.monotonic()
        if not settings.OPENAI_API_KEY:
            logger.warning("[seed-amplify-burst] OPENAI_API_KEY 미설정 — abort")
            return

        account = get_ad_account_by_customer(user_id, str(customer_id))
        if not account or not account.get("is_connected"):
            logger.warning(f"[seed-amplify-burst] uid={user_id} cid={customer_id} 미연결 — abort")
            return

        pool = get_keyword_pool_db()
        user_seeds = pool.list_user_seeds(customer_id) or []
        if not user_seeds:
            logger.warning(f"[seed-amplify-burst] cid={customer_id} user_seed 0 — abort")
            return

        SAMPLE = 100
        _sem = asyncio.Semaphore(4)

        async def _one_amp(idx: int) -> List[str]:
            async with _sem:
                sample = (
                    _r.sample(user_seeds, SAMPLE)
                    if len(user_seeds) > SAMPLE else list(user_seeds)
                )
                try:
                    r = await _amp(sample, target_count=target_per_call)
                except Exception as e:
                    logger.warning(f"[seed-amplify-burst] call {idx} 예외: {e}")
                    return []
                if not r.get("success"):
                    logger.warning(f"[seed-amplify-burst] call {idx} 실패: {r.get('message')}")
                    return []
                return [s for s in (r.get("seeds") or []) if isinstance(s, str) and s.strip()]

        am_t0 = _t.monotonic()
        results = await asyncio.gather(*[_one_amp(i) for i in range(n_calls)])
        am_ms = int((_t.monotonic() - am_t0) * 1000)

        # 누적 fresh seeds — 원본 + 풀 dedup
        user_seed_set = set(user_seeds)
        pool_set = pool.list_pool_keyword_set(customer_id)
        seen: Set[str] = set()
        fresh_seeds: List[str] = []
        for batch in results:
            for s in batch:
                k = s.strip()
                if not k or k in seen or k in user_seed_set or k in pool_set:
                    continue
                seen.add(k)
                fresh_seeds.append(k)

        # 도메인 게이트 — saved_relevance 있는 계정만. amplify burst cartesian 폭발이
        # drift 증폭기 사고의 주범. cold start (relevance 없음) 만 통과시킴.
        from database.naver_ad_db import get_ad_account_relevance_keywords as _get_rel
        saved_relevance = _get_rel(user_id, str(customer_id)) or []
        burst_domain_filtered = 0
        if saved_relevance and len([s for s in saved_relevance if s and len(s) >= 2]) >= 3:
            before = len(fresh_seeds)
            fresh_seeds = [s for s in fresh_seeds if _compute_relevance_score(s, saved_relevance) >= 30]
            burst_domain_filtered = before - len(fresh_seeds)

        logger.warning(
            f"[seed-amplify-burst] cid={customer_id} amplify {n_calls}회 ({am_ms}ms) "
            f"→ 누적 raw {sum(len(b) for b in results)} → fresh {len(fresh_seeds)}"
            + (f" (도메인필터 컷 {burst_domain_filtered})" if burst_domain_filtered else "")
        )

        if not fresh_seeds:
            return

        # keywordstool 검증 — chunks 50 cap, sleep 0.3 (429 회피)
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]

        vol_t0 = _t.monotonic()
        vol_map: Dict[str, dict] = {}
        CHUNK = 5
        CHUNKS_CAP = 200  # 1,000 seed 검증 (burst 모드)
        chunks = [fresh_seeds[i:i + CHUNK] for i in range(0, len(fresh_seeds), CHUNK)][:CHUNKS_CAP]
        for chunk in chunks:
            try:
                r = await client.get_keywords_volume_batch(chunk)
                vol_map.update(r)
            except Exception as e:
                logger.debug(f"[seed-amplify-burst] volume batch 실패: {e}")
            await asyncio.sleep(0.3)
        vol_ms = int((_t.monotonic() - vol_t0) * 1000)

        # mt≥1 만 user_seed 합류 — mt=0 zerovol 시드는 등록 락 사고로 제거 (2026-05-12).
        # 과거: mt=0 도 user_seed 로 INSERT → claim_pending 필터 + INSERT OR IGNORE
        # 합세로 11k+ 등록 불가 행이 풀 점거 → 등록 throughput 영구 정지.
        items_with_vol: List[Dict] = []
        zerovol_count = 0
        for s in fresh_seeds:
            v = vol_map.get(s) or vol_map.get(s.replace(" ", ""))
            mt = int((v or {}).get("monthly_total") or 0)
            if v and mt >= 1:
                items_with_vol.append({
                    "keyword": s, "monthly_total": mt,
                    "monthly_pc": int(v.get("monthly_pc") or 0),
                    "monthly_mobile": int(v.get("monthly_mobile") or 0),
                    "comp_idx": v.get("comp_idx"),
                    "source": "user_seed", "seed": s,
                })
            else:
                zerovol_count += 1

        # 합류 cap — 한 burst 에 user_seed 최대 5000 추가
        BURST_CAP = 5000
        merged = items_with_vol[:BURST_CAP]
        items_zerovol: List[Dict] = []  # 호환성 (로그 변수)
        promoted = 0
        if merged:
            try:
                promoted = pool.add_candidates(user_id, customer_id, merged)
            except Exception as e:
                logger.warning(f"[seed-amplify-burst] add 실패: {e}")

        duration_ms = int((_t.monotonic() - t0) * 1000)
        logger.warning(
            f"[seed-amplify-burst] cid={customer_id} 완료 ({duration_ms}ms) — "
            f"amplify {n_calls}회 raw {sum(len(b) for b in results)} → "
            f"fresh {len(fresh_seeds)} → mt≥1 {len(items_with_vol)} (zerovol 컷 {zerovol_count}) "
            f"→ user_seed +{promoted} (GPT {am_ms}ms, vol {vol_ms}ms)"
        )
        pool.record_run(
            user_id, customer_id, "seed_amplify_burst",
            "success" if promoted > 0 else "no_match",
            added=promoted, seeds_count=len(user_seeds),
            error_message=(
                f"burst {n_calls}회 → fresh {len(fresh_seeds)} → "
                f"mt≥1 {len(items_with_vol)} (zerovol 컷 {zerovol_count}) → +{promoted}"
            )[:300],
            duration_ms=duration_ms,
        )

    background_tasks.add_task(_run)
    return {
        "success": True,
        "queued": True,
        "user_id": user_id,
        "customer_id": customer_id,
        "n_calls": n_calls,
        "target_per_call": target_per_call,
        "estimated_duration_seconds": n_calls * 5 + 60,  # GPT 5s × n_calls + keywordstool
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


@router.post("/keyword-pool/cron/cleanup-zerovol-seeds")
async def keyword_pool_cron_cleanup_zerovol_seeds(
    authorization: Optional[str] = Header(None),
    user_id: Optional[int] = Query(None),
    customer_id: Optional[int] = Query(None),
):
    """mt=0 user_seed 잔재 일괄 정리 — 11k+ 등록 락 해제용 (2026-05-12).

    구버전 seed_amplify_burst 가 mt=0 시드를 user_seed 로 INSERT → claim_pending
    의 mt≥1 필터에 걸려 영구 등록 불가. 삭제 후 같은 KW 가 keywordstool 에서
    mt>0 으로 재발견되면 add_candidates UPSERT 가 정상 합류시킴.

    단건: ?user_id=X&customer_id=Y / 전체: 파라미터 없이 호출.
    """
    _verify_cron_token(authorization)
    from database.naver_ad_db import list_connected_ad_accounts

    pool = get_keyword_pool_db()
    targets: List[Tuple[int, int]] = []
    if user_id and customer_id:
        targets = [(int(user_id), int(customer_id))]
    else:
        accts = list_connected_ad_accounts() or []
        for a in accts:
            uid = a.get("user_id")
            cid = a.get("customer_id")
            if uid and cid:
                targets.append((int(uid), int(cid)))

    results = []
    for uid, cid in targets:
        try:
            r = pool.cleanup_zerovol_user_seeds(cid)
            results.append({"user_id": uid, "customer_id": cid, **r})
            by_src = r.get("by_source") or {}
            src_summary = ", ".join(f"{k or '<null>'}={v}" for k, v in list(by_src.items())[:6])
            logger.warning(
                f"[cleanup-zerovol-seeds] uid={uid} cid={cid} → "
                f"삭제 {r['deleted']} (pending {r['before_pending']}→{r['after_pending']}) "
                f"source 분포: {src_summary}"
            )
        except Exception as e:
            logger.error(
                f"[cleanup-zerovol-seeds] uid={uid} cid={cid} 실패: {e}", exc_info=True,
            )
            results.append({"user_id": uid, "customer_id": cid, "error": str(e)[:200]})

    return {
        "success": True,
        "results": results,
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


@router.post("/keyword-pool/cron/auto-cleanup")
async def keyword_pool_cron_auto_cleanup(
    background_tasks: BackgroundTasks,
    authorization: Optional[str] = Header(None),
    threshold_override: Optional[int] = Query(None),
    user_id: Optional[int] = Query(None),
    customer_id: Optional[str] = Query(None),
):
    """Bearer 토큰 cron — auto_cleanup_enabled=1 인 모든 광고주에 대해 점수≤threshold 자동 삭제.
    - 단건 트리거: ?user_id=X&customer_id=Y (디버깅/수동 실행)
    - threshold_override: cron 호출 시 광고주 설정 무시하고 강제 임계값 (디버깅)
    """
    _verify_cron_token(authorization)
    from database.naver_ad_db import list_auto_cleanup_enabled_accounts, get_ad_account_auto_cleanup

    targets: List[Tuple[int, int, int]] = []  # (uid, cid, threshold)
    if user_id and customer_id:
        s = get_ad_account_auto_cleanup(user_id, str(customer_id))
        thr = threshold_override if threshold_override is not None else s["threshold"]
        targets = [(user_id, int(customer_id), int(thr))]
    else:
        rows = list_auto_cleanup_enabled_accounts() or []
        for r in rows:
            uid = int(r.get("user_id"))
            cid = int(r.get("customer_id"))
            thr = threshold_override if threshold_override is not None else int(r.get("auto_cleanup_threshold") or 30)
            targets.append((uid, cid, thr))

    if not targets:
        return {"success": True, "queued": 0, "message": "자동 cleanup ON 광고주 없음"}

    async def _run_all():
        for uid, cid, thr in targets:
            try:
                res = await _run_auto_cleanup_for_account(uid, cid, thr)
                logger.info(f"[auto-cleanup] uid={uid} cid={cid} thr={thr} → {res}")
            except Exception as e:
                logger.error(f"[auto-cleanup] uid={uid} cid={cid} 실패: {type(e).__name__}: {e}", exc_info=True)

    background_tasks.add_task(_run_all)
    return {
        "success": True,
        "queued": len(targets),
        "targets": [{"user_id": u, "customer_id": c, "threshold": t} for u, c, t in targets],
        "timestamp": datetime.now().isoformat(timespec="seconds"),
    }


# ============ 등록 KW 전체 점수 audit + 일괄 정리 ============
# auto-cleanup cron 은 click ≥ 1 KW 만 처리 (Naver stats API 호출 비용 절약).
# 이 API 는 click 무관 — registered_keywords 테이블 직접 조회 + keyword text 만으로
# user_seed 점수 매김 (stats API 호출 없이 95k KW 1초 안에 점수화).
# cascade drift 로 옛날에 등록된 무관 KW (click=0) 일괄 정리용.

class CleanupByScoreRequest(BaseModel):
    threshold: int = 30
    max_delete: int = 1000
    dry_run: bool = False
    # 사용자가 호출 시 임시로 다른 기준 키워드 쓰고 싶을 때 (저장된 광고주 설정 무시).
    # 비어있으면 ad_accounts.relevance_keywords → 비어있으면 user_seed 순으로 폴백.
    relevance_keywords_override: Optional[List[str]] = None


@router.post("/keyword-pool/registered/cleanup-by-score")
async def keyword_pool_registered_cleanup_by_score(
    request: CleanupByScoreRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """등록 KW 전체 audit — user_seed 점수 ≤ threshold 인 KW 일괄 DELETE (click 무관).
    dry_run=true: 점수 분포 + 삭제 대상 미리보기. dry_run=false: 백그라운드 실행 (수십분 소요).
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import record_auto_cleanup_run
    import sqlite3 as _sqlite3
    import time as _t
    _t0 = _t.monotonic()

    threshold = max(0, min(95, int(request.threshold)))
    # 2026-05-12: cap 5000 → 50000 상향. 한의원/한방 광고주 차 KW drift 50k+ 누적 사고에서
    # 5000 cap 이면 10+ 회 수동 호출 필요. 50000 = 한 번 호출로 ~150분 백그라운드 정리.
    max_delete = max(0, min(50000, int(request.max_delete)))
    dry_run = bool(request.dry_run)

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()

    # 점수 기준 키워드 우선순위: request.override → 광고주 저장값 → user_seed.
    # 사용자 의도 ("내가 원하는 키워드 기준으로 연관성 잡아야지") 반영.
    from database.naver_ad_db import get_ad_account_relevance_keywords
    if request.relevance_keywords_override:
        score_basis = [s.strip() for s in request.relevance_keywords_override
                       if s and len(s.strip()) >= 2]
        basis_source = "override"
    else:
        saved = get_ad_account_relevance_keywords(user_id, str(cid))
        if saved:
            score_basis = saved
            basis_source = "saved"
        else:
            score_basis = [s for s in (pool.list_user_seeds(cid) or []) if s and len(s) >= 2]
            basis_source = "user_seed_fallback"

    if not score_basis:
        raise HTTPException(
            status_code=400,
            detail=(
                "점수 기준 키워드 없음 — '연관성 기준 키워드' 입력 또는 user_seed 추가 후 재시도. "
                "예: 피부질환,피부,피부과,아토피,여드름"
            ),
        )
    user_seeds = score_basis  # 이하 코드와 호환 (변수명 유지)

    with _sqlite3.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL",
            (cid,),
        ).fetchall()

    # 점수 매김 — atoms 1회 precompute (95k KW 호출 시 atoms 95k 번 재빌드 → GIL 폭주 방지).
    # PERF: 95k × 30 atoms = 2.85M ops. asyncio.to_thread 로 워커 thread 분리해
    # 이벤트 루프 보호 (fly.io health check timeout 차단).
    def _score_all() -> Tuple[List[Tuple[str, str, int]], Dict[int, int]]:
        atoms_3plus: set = set()
        atoms_2: set = set()
        for s in user_seeds:
            if not s or len(s) < 2:
                continue
            if len(s) >= 4:
                atoms_3plus.add(s)
            for n in (2, 3):
                for i in range(len(s) - n + 1):
                    a = s[i:i + n]
                    (atoms_2 if len(a) == 2 else atoms_3plus).add(a)
        _scored: List[Tuple[str, str, int]] = []
        _dist: Dict[int, int] = {}
        for kw_text, kid in rows:
            if not kw_text:
                _scored.append((kid, kw_text or "", 0))
                _dist[0] = _dist.get(0, 0) + 1
                continue
            sc = 0
            full = False
            for s in user_seeds:
                if not s or len(s) < 2:
                    continue
                if s in kw_text:
                    sc = 100; full = True; break
                if kw_text in s:
                    sc = 95; full = True; break
            if not full:
                n_3 = sum(1 for a in atoms_3plus if a in kw_text)
                n_2 = sum(1 for a in atoms_2 if a in kw_text)
                sc = min(95, min(80, n_3 * 20) + min(30, n_2 * 5))
            _scored.append((kid, kw_text, sc))
            bucket = (sc // 10) * 10
            _dist[bucket] = _dist.get(bucket, 0) + 1
        return _scored, _dist

    _t_db = _t.monotonic() - _t0
    scored, score_dist = await asyncio.to_thread(_score_all)
    _t_score = _t.monotonic() - _t0 - _t_db
    # 앵커 모드 — required_tokens 있으면 (앵커없음 OR negative) 가 삭제기준(점수 무시). 진짜 대출 보존.
    _req_tok = []; _neg_tok = []
    try:
        from database.naver_ad_db import get_domain_profile as _gdp_cbs
        _pf_cbs = _gdp_cbs(user_id, str(cid)) or {}
        _req_tok = [t for t in _pf_cbs.get("required_tokens", []) if t and len(t) >= 2]
        _neg_tok = [n for n in _pf_cbs.get("negative_keywords", []) if n and len(n) >= 2]
    except Exception:
        pass
    if _req_tok:
        targets = [
            (kid, kw, s) for kid, kw, s in scored
            if kw and (not any(rt in kw for rt in _req_tok)
                       or (_neg_tok and any(nt in kw for nt in _neg_tok)))
        ]
    else:
        # 점수<threshold (Option B: boundary 보존) 또는 negative 포함(substring 오매칭 off-domain 차단)
        targets = [
            (kid, kw, s) for kid, kw, s in scored
            if s < threshold or (_neg_tok and kw and any(nt in kw for nt in _neg_tok))
        ]
    targets.sort(key=lambda x: x[2])  # 무관한 것부터
    targets_capped = targets[:max_delete]
    logger.warning(
        f"[cleanup-by-score] uid={user_id} cid={cid} dry_run={dry_run} "
        f"db_query={_t_db:.2f}s score={_t_score:.2f}s "
        f"total={len(scored)} below={len(targets)} threshold={threshold} "
        f"basis={basis_source} basis_count={len(user_seeds)} "
        f"basis_sample={user_seeds[:5]}"
    )

    if dry_run:
        # 화면 표시용 — targets 전체 (max_delete 적용 전) 중 max 1000 개. keyword_id 포함해서
        # frontend 에서 체크박스 선택 후 /clicked-keywords/bulk-delete 로 삭제 가능.
        DISPLAY_LIMIT = 1000
        return {
            "success": True,
            "dry_run": True,
            "customer_id": cid,
            "threshold": threshold,
            "total_registered": len(scored),
            "score_distribution": dict(sorted(score_dist.items())),
            "targets_below_threshold": len(targets),
            "will_delete_now": len(targets_capped),
            "max_delete": max_delete,
            "displayed": min(len(targets), DISPLAY_LIMIT),
            "targets": [
                {"keyword_id": kid, "keyword": kw, "score": s}
                for kid, kw, s in targets[:DISPLAY_LIMIT]
            ],
        }

    if not targets_capped:
        return {
            "success": True, "dry_run": False,
            "customer_id": cid, "threshold": threshold,
            "queued_targets": 0, "message": f"임계값 {threshold} 이하 등록 KW 없음",
        }

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    # 동시 실행 방어 — 같은 광고주에서 이미 bulk cleanup 진행 중이면 409.
    # 연타/멀티탭에서 50k 작업이 N배 쌓이면 event loop CPU + Naver rate limit 사고.
    if cid in _BULK_CLEANUP_RUNNING:
        raise HTTPException(
            status_code=409,
            detail=(
                f"이 광고주 (cid={cid}) 의 일괄 삭제 작업이 이미 진행 중입니다. "
                f"완료까지 기다려주세요 (예상 {round(len(targets_capped) * 0.18 / 60, 1)}분)."
            ),
        )

    async def _run():
        _BULK_CLEANUP_RUNNING.add(cid)
        try:
            n_del, n_pause, n_fail = 0, 0, 0
            affected: List[str] = []
            for kid, kw_text, _s in targets_capped:
                try:
                    await client.delete_keyword(kid)
                    with _sqlite3.connect(reg.db_path) as c:
                        c.execute(
                            "DELETE FROM registered_keywords "
                            "WHERE account_customer_id=? AND ncc_keyword_id=?",
                            (cid, kid),
                        )
                    n_del += 1
                    affected.append(kw_text)
                except Exception:
                    try:
                        await client.pause_keyword(kid)
                        n_pause += 1
                        affected.append(kw_text)
                    except Exception:
                        n_fail += 1
                await asyncio.sleep(0.15)
            if affected:
                pool.mark_rejected_by_naver(
                    cid,
                    [{"keyword": kw, "reason": f"수동 점수 정리(≤{threshold})"} for kw in affected],
                )
            try:
                pool.record_run(
                    user_id, cid, "inspect",
                    "success" if n_del > 0 else "no_new",
                    registered=0, failed=n_fail, skipped=n_del,
                    seeds_count=len(targets_capped),
                    error_message=(
                        f"수동 점수 정리 (점수≤{threshold}) — "
                        f"DELETE {n_del} / PAUSE {n_pause} / 실패 {n_fail}"
                    ),
                )
            except Exception:
                pass
            record_auto_cleanup_run(user_id, str(cid), n_del + n_pause)
            logger.warning(
                f"[manual-cleanup] uid={user_id} cid={cid} thr={threshold} "
                f"→ del={n_del} pause={n_pause} fail={n_fail}"
            )
        finally:
            _BULK_CLEANUP_RUNNING.discard(cid)

    background_tasks.add_task(_run)
    return {
        "success": True,
        "dry_run": False,
        "customer_id": cid,
        "threshold": threshold,
        "queued_targets": len(targets_capped),
        "below_threshold_total": len(targets),
        "estimated_minutes": round(len(targets_capped) * 0.18 / 60, 1),
        "message": f"백그라운드 실행 시작 — {len(targets_capped)}개 KW 삭제 진행 (예상 {round(len(targets_capped) * 0.18 / 60, 1)}분)",
    }


class ReactivateFailedRequest(BaseModel):
    threshold: int = Field(50, ge=0, le=95, description="이 점수 이상(온도메인)만 pending 재활성화")
    min_volume: int = Field(10, ge=0, le=100000, description="월 검색량 최소 (실볼륨만)")
    max_reactivate: int = Field(50000, ge=0, le=200000, description="이번 호출 최대 재활성화 수")
    dry_run: bool = Field(True)
    relevance_keywords_override: Optional[List[str]] = None
    include_statuses: Optional[List[str]] = Field(
        None, description="재활성화 대상 status (기본 ['failed']; deleted 는 과거 off-domain 퍼지분이라 권장 제외)"
    )
    phantom_registered: bool = Field(
        False,
        description="True 면 status='registered'인데 라이브 추적(registered_keywords.ncc_keyword_id)이 없는 phantom 행만 스캔 — reconcile 버그로 등록 처리됐으나 실제 네이버엔 없는 것 재등록. include_statuses 무시.",
    )


@router.post("/keyword-pool/registered/reactivate-failed")
async def keyword_pool_reactivate_failed(
    request: ReactivateFailedRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """잠든 failed(옵션 deleted/rejected) 키워드 중 온도메인(relevance≥threshold)을 pending 으로 되살림.

    10만 채우기 supply 보충 — 이미 발굴+검색량 검증된 키워드 재활용 (신규 발굴 0, dedup 무관).
    드리프트 방지: relevance 게이트 통과분만. deleted 는 기본 제외 (과거 off-domain 퍼지 부활 차단).
    dry_run=true: 점수 분포 + 대상 미리보기. false: status='pending' UPDATE → register cron(30s) 소진.
    """
    import sqlite3 as _sqlite3
    import time as _t
    _t0 = _t.monotonic()

    threshold = max(0, min(95, int(request.threshold)))
    min_volume = max(0, int(request.min_volume))
    max_reactivate = max(0, min(200000, int(request.max_reactivate)))
    dry_run = bool(request.dry_run)
    # domain_skipped: register 하드게이트가 off-domain 으로 뺀 것. 게이트 완화(키네스 강제등록) 후
    # threshold=0 으로 되살리면 재게이팅됨(완화된 gate 가 재판정). include_statuses 로 opt-in.
    allowed = {"failed", "deleted", "rejected_by_naver", "domain_skipped"}
    statuses = [s for s in (request.include_statuses or ["failed"]) if s in allowed] or ["failed"]

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    pool = get_keyword_pool_db()

    from database.naver_ad_db import get_ad_account_relevance_keywords
    if request.relevance_keywords_override:
        score_basis = [s.strip() for s in request.relevance_keywords_override if s and len(s.strip()) >= 2]
        basis_source = "override"
    else:
        saved = get_ad_account_relevance_keywords(user_id, str(cid))
        if saved:
            score_basis = saved; basis_source = "saved"
        else:
            score_basis = [s for s in (pool.list_user_seeds(cid) or []) if s and len(s) >= 2]
            basis_source = "user_seed_fallback"
    if not score_basis:
        raise HTTPException(status_code=400, detail="점수 기준 키워드 없음 — relevance_keywords 저장 또는 override 필요")
    user_seeds = score_basis

    def _fetch_rows():
        """행 조회 — (rows, statuses_label). 270k 행도 SQLite 라 빠름."""
        if request.phantom_registered:
            # phantom = pool.status='registered' 인데 registered_keywords(별도 DB)에 ncc_id 없음.
            reg = get_registered_keywords_db()
            with _sqlite3.connect(reg.db_path, timeout=30.0) as rc:
                live_set = {
                    k for (k,) in rc.execute(
                        "SELECT keyword FROM registered_keywords "
                        "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL",
                        (cid,),
                    ).fetchall()
                }
            with _sqlite3.connect(pool.db_path, timeout=30.0) as conn:
                all_reg = conn.execute(
                    """SELECT id, keyword, COALESCE(monthly_total,0)
                       FROM naverad_keyword_pool
                       WHERE account_customer_id=? AND status='registered'
                         AND COALESCE(monthly_total,0) >= ?""",
                    (cid, min_volume),
                ).fetchall()
            return [(rid, kw, mt) for rid, kw, mt in all_reg if kw not in live_set], ["registered(phantom)"]
        placeholders = ",".join("?" * len(statuses))
        with _sqlite3.connect(pool.db_path, timeout=30.0) as conn:
            rows = conn.execute(
                f"""SELECT id, keyword, COALESCE(monthly_total,0)
                    FROM naverad_keyword_pool
                    WHERE account_customer_id=?
                      AND status IN ({placeholders})
                      AND COALESCE(monthly_total,0) >= ?""",
                (cid, *statuses, min_volume),
            ).fetchall()
        return rows, list(statuses)

    # cleanup-by-score 와 동일한 atom 점수. rows 를 인자로 받아 to_thread 로 event loop 보호.
    def _score_all(rows):
        atoms_3plus: set = set()
        atoms_2: set = set()
        for s in user_seeds:
            if not s or len(s) < 2:
                continue
            if len(s) >= 4:
                atoms_3plus.add(s)
            for n in (2, 3):
                for i in range(len(s) - n + 1):
                    a = s[i:i + n]
                    (atoms_2 if len(a) == 2 else atoms_3plus).add(a)
        keep: List[Tuple[int, str, int, int]] = []
        dist: Dict[int, int] = {}
        for rid, kw, mt in rows:
            if not kw:
                dist[0] = dist.get(0, 0) + 1
                continue
            sc = 0
            full = False
            for s in user_seeds:
                if not s or len(s) < 2:
                    continue
                if s in kw:
                    sc = 100; full = True; break
                if kw in s:
                    sc = 95; full = True; break
            if not full:
                n_3 = sum(1 for a in atoms_3plus if a in kw)
                n_2 = sum(1 for a in atoms_2 if a in kw)
                sc = min(95, min(80, n_3 * 20) + min(30, n_2 * 5))
            bucket = (sc // 10) * 10
            dist[bucket] = dist.get(bucket, 0) + 1
            if sc >= threshold:
                keep.append((rid, kw, mt, sc))
        return keep, dist

    def _apply_update(capped):
        ids = [rid for rid, _, _, _ in capped]
        n = 0
        with _sqlite3.connect(pool.db_path, timeout=30.0) as conn:
            for i in range(0, len(ids), 900):
                chunk = ids[i:i + 900]
                ph = ",".join("?" * len(chunk))
                cur = conn.execute(
                    f"""UPDATE naverad_keyword_pool
                        SET status='pending', error_message=NULL, registered_at=NULL
                        WHERE account_customer_id=? AND id IN ({ph})""",
                    (cid, *chunk),
                )
                n += cur.rowcount or 0
            conn.commit()
        return n

    # ── dry_run: 인라인 (작은 계정 미리보기). 소잠 등 270k 는 60s 초과 가능 → 실사용은 dry_run=false ──
    if dry_run:
        rows, statuses_label = _fetch_rows()
        keep, dist = await asyncio.to_thread(_score_all, rows)
        keep.sort(key=lambda x: (-x[3], -x[2]))
        capped = keep[:max_reactivate]
        logger.warning(
            f"[reactivate-failed] dry_run uid={user_id} cid={cid} statuses={statuses_label} "
            f"scanned={len(rows)} on_domain={len(keep)} thr={threshold} elapsed={_t.monotonic()-_t0:.2f}s"
        )
        return {
            "success": True, "dry_run": True, "customer_id": cid,
            "basis_source": basis_source, "basis_count": len(user_seeds),
            "scanned_statuses": statuses_label, "min_volume": min_volume, "threshold": threshold,
            "total_scanned": len(rows),
            "score_distribution": dict(sorted(dist.items())),
            "on_domain_reactivatable": len(keep),
            "will_reactivate_now": len(capped),
            "samples": [{"keyword": kw, "score": sc, "mt": mt} for _, kw, mt, sc in capped[:30]],
        }

    # ── 실제 재활성화: 스캔+스코어+UPDATE 전부 백그라운드 (fly 60s 초과 방지) → 즉시 응답 ──
    async def _run_reactivate():
        try:
            rows, statuses_label = await asyncio.to_thread(_fetch_rows)
            logger.warning(f"[reactivate-failed] 시작 — statuses={statuses_label} scanned={len(rows)} thr={threshold} minvol={min_volume} basis={basis_source}({len(user_seeds)})")
            keep, _dist = await asyncio.to_thread(_score_all, rows)
            keep.sort(key=lambda x: (-x[3], -x[2]))
            capped = keep[:max_reactivate]
            logger.warning(f"[reactivate-failed] 온도메인 {len(keep)} → 재활성화 대상 {len(capped)} (cap {max_reactivate})")
            if not capped:
                logger.warning("[reactivate-failed] 대상 0 — 종료")
                return
            n = await asyncio.to_thread(_apply_update, capped)
            logger.warning(f"[reactivate-failed] 완료 — {n}개 → pending (register cron 30s 소진). elapsed={_t.monotonic()-_t0:.1f}s")
        except Exception as e:
            import traceback as _tb
            logger.error(f"[reactivate-failed] 태스크 예외 — {type(e).__name__}: {str(e)[:200]}\n{_tb.format_exc()[:1000]}")

    background_tasks.add_task(_run_reactivate)
    return {
        "success": True, "started": True, "dry_run": False, "customer_id": cid,
        "scanned_statuses": statuses, "threshold": threshold, "min_volume": min_volume,
        "max_reactivate": max_reactivate, "basis_source": basis_source,
        "message": "백그라운드 재활성화 시작 — failed/rejected 온도메인 스캔+pending 전환. "
                   "결과는 fly logs 의 [reactivate-failed] 라인. pending 은 register cron(30s)이 등록.",
    }


# ============ 긴급 drift 일괄 정리 — registered + user_seed + pending 한 번에 ============
# cleanup-by-score 는 registered 만. 100k drift 사고 시 user_seed 도 오염돼 있어 그것도
# 같이 갈아야 다음 amplify 가 또 차 KW 안 만듦. 이 endpoint 는 3가지 정리를 한 번에:
#  1) registered (registered_keywords + naverad_keyword_pool status='registered') 점수≤thr Naver DELETE
#  2) user_seed (source='user_seed' AND status NOT IN registered/failed) 점수≤thr DB DELETE
#  3) pending (status='pending') 점수≤thr DB DELETE (등록 전 차단)

@router.post("/keyword-pool/admin/purge-drift")
async def keyword_pool_admin_purge_drift(
    background_tasks: BackgroundTasks,
    customer_id: int = Query(..., description="대상 광고주 customer_id"),
    threshold: int = Query(30, ge=0, le=95),
    max_delete_registered: int = Query(50000, ge=0, le=100000),
    dry_run: bool = Query(False),
    relevance_keywords: Optional[str] = Query(
        None,
        description="콤마구분 도메인 KW (예: 아토피,습진,건선,한의원,한방). 비우면 ad_accounts.relevance_keywords 사용",
    ),
):
    """긴급 drift 정리 — registered + user_seed + pending 한 번에 (인증 없음, customer_id 명시).

    cleanup-by-score 는 registered 만 정리 → user_seed 가 오염된 채로 남으면 다음 amplify 가
    또 drift 키워드 생성 → 정리 의미 없음. 이 endpoint 는 3 stage 동시 처리:

    Stage 1 (즉시): user_seed + pending DB cleanup (네이버 호출 없음, 1초)
    Stage 2 (background): registered Naver DELETE (KW × 0.18s × 부하)

    relevance_keywords 우선순위: query param > ad_accounts.relevance_keywords > user_seed.
    user_seed 폴백은 위험 (이미 오염) — query param 또는 saved 권장.
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import (
        list_connected_ad_accounts,
        get_ad_account_relevance_keywords,
        record_auto_cleanup_run,
    )
    import sqlite3 as _sqlite3
    import time as _t

    t0 = _t.monotonic()
    accts = list_connected_ad_accounts() or []
    matched = next((a for a in accts if int(a.get("customer_id") or 0) == int(customer_id)), None)
    if not matched:
        raise HTTPException(status_code=404, detail=f"customer_id {customer_id} 미연결")
    uid = int(matched.get("user_id") or 0)

    # 도메인 기준 빌드
    if relevance_keywords:
        score_basis = [s.strip() for s in relevance_keywords.replace("\n", ",").split(",") if s.strip() and len(s.strip()) >= 2]
        basis_source = "query"
    else:
        saved = get_ad_account_relevance_keywords(uid, str(customer_id)) or []
        if saved:
            score_basis = saved
            basis_source = "saved"
        else:
            raise HTTPException(
                status_code=400,
                detail="relevance_keywords 없음 (query 또는 saved). 예: ?relevance_keywords=아토피,습진,한의원,한방",
            )
    if len(score_basis) < 3:
        raise HTTPException(status_code=400, detail=f"기준 KW 부족 ({len(score_basis)}/3). 최소 3개 권장.")

    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()

    # ===== Stage 1: user_seed + pending DB cleanup (즉시) =====
    user_seed_deleted = 0
    pending_deleted = 0
    sample_user_seed: List[str] = []
    sample_pending: List[str] = []

    with _sqlite3.connect(pool.db_path) as conn:
        conn.row_factory = _sqlite3.Row

        # user_seed 정리
        user_seed_rows = conn.execute(
            """SELECT keyword FROM naverad_keyword_pool
               WHERE account_customer_id=? AND source='user_seed'
                 AND status NOT IN ('registered', 'failed')""",
            (customer_id,),
        ).fetchall()
        user_seed_to_delete: List[str] = []
        for r in user_seed_rows:
            kw = r["keyword"]
            sc = _compute_relevance_score(kw, score_basis)
            if sc < threshold:  # Option B: boundary 보존
                user_seed_to_delete.append(kw)

        # pending 정리
        pending_rows = conn.execute(
            """SELECT keyword FROM naverad_keyword_pool
               WHERE account_customer_id=? AND status='pending'""",
            (customer_id,),
        ).fetchall()
        pending_to_delete: List[str] = []
        for r in pending_rows:
            kw = r["keyword"]
            sc = _compute_relevance_score(kw, score_basis)
            if sc < threshold:  # Option B: boundary 보존
                pending_to_delete.append(kw)

        sample_user_seed = user_seed_to_delete[:10]
        sample_pending = pending_to_delete[:10]

        if not dry_run:
            # user_seed: source='user_seed' AND status NOT registered/failed → DELETE
            CHUNK = 500
            for i in range(0, len(user_seed_to_delete), CHUNK):
                chunk = user_seed_to_delete[i:i + CHUNK]
                placeholders = ",".join("?" * len(chunk))
                cur = conn.execute(
                    f"""DELETE FROM naverad_keyword_pool
                        WHERE account_customer_id=? AND source='user_seed'
                          AND status NOT IN ('registered', 'failed')
                          AND keyword IN ({placeholders})""",
                    (customer_id, *chunk),
                )
                user_seed_deleted += cur.rowcount

            # pending: status='pending' → DELETE
            for i in range(0, len(pending_to_delete), CHUNK):
                chunk = pending_to_delete[i:i + CHUNK]
                placeholders = ",".join("?" * len(chunk))
                cur = conn.execute(
                    f"""DELETE FROM naverad_keyword_pool
                        WHERE account_customer_id=? AND status='pending'
                          AND keyword IN ({placeholders})""",
                    (customer_id, *chunk),
                )
                pending_deleted += cur.rowcount

    # ===== Stage 2: registered Naver DELETE (background) =====
    with _sqlite3.connect(reg.db_path) as conn:
        conn.row_factory = _sqlite3.Row
        reg_rows = conn.execute(
            "SELECT keyword, ncc_keyword_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND removed_at IS NULL",
            (customer_id,),
        ).fetchall()

    def _score_reg() -> List[Tuple[str, str, int]]:
        atoms_3plus: set = set()
        atoms_2: set = set()
        for s in score_basis:
            if not s or len(s) < 2:
                continue
            if len(s) >= 4:
                atoms_3plus.add(s)
            for n in (2, 3):
                for i in range(len(s) - n + 1):
                    a = s[i:i + n]
                    (atoms_2 if len(a) == 2 else atoms_3plus).add(a)
        out: List[Tuple[str, str, int]] = []
        for r in reg_rows:
            kw_text = r["keyword"]
            kid = r["ncc_keyword_id"]
            if not kw_text:
                out.append((kid, "", 0))
                continue
            sc = 0
            full = False
            for s in score_basis:
                if not s or len(s) < 2:
                    continue
                if s in kw_text:
                    sc = 100; full = True; break
                if kw_text in s:
                    sc = 95; full = True; break
            if not full:
                n_3 = sum(1 for a in atoms_3plus if a in kw_text)
                n_2 = sum(1 for a in atoms_2 if a in kw_text)
                sc = min(95, min(80, n_3 * 20) + min(30, n_2 * 5))
            out.append((kid, kw_text, sc))
        return out

    reg_scored = await asyncio.to_thread(_score_reg)
    reg_targets = [(kid, kw, s) for kid, kw, s in reg_scored if s < threshold]  # Option B: boundary 보존
    reg_targets.sort(key=lambda x: x[2])
    reg_capped = reg_targets[:max_delete_registered]

    if dry_run:
        return {
            "success": True,
            "dry_run": True,
            "customer_id": customer_id,
            "threshold": threshold,
            "basis_source": basis_source,
            "basis_count": len(score_basis),
            "basis_sample": score_basis[:8],
            "user_seed_total": len(user_seed_rows),
            "user_seed_to_delete": len(user_seed_to_delete),
            "user_seed_samples": sample_user_seed,
            "pending_total": len(pending_rows),
            "pending_to_delete": len(pending_to_delete),
            "pending_samples": sample_pending,
            "registered_total": len(reg_scored),
            "registered_below_threshold": len(reg_targets),
            "registered_will_delete_now": len(reg_capped),
            "registered_samples": [{"keyword": kw, "score": s} for _kid, kw, s in reg_capped[:10]],
            "estimated_minutes": round(len(reg_capped) * 0.18 / 60, 1),
        }

    # Stage 2 background — registered Naver DELETE
    account = matched
    client = NaverAdApiClient()
    client.customer_id = account.get("customer_id")
    client.api_key = account.get("api_key")
    client.secret_key = account.get("secret_key")

    async def _run_reg_purge():
        n_del, n_pause, n_fail = 0, 0, 0
        affected: List[str] = []
        for kid, kw_text, _s in reg_capped:
            try:
                await client.delete_keyword(kid)
                with _sqlite3.connect(reg.db_path) as c:
                    c.execute(
                        "DELETE FROM registered_keywords "
                        "WHERE account_customer_id=? AND ncc_keyword_id=?",
                        (customer_id, kid),
                    )
                n_del += 1
                affected.append(kw_text)
            except Exception:
                try:
                    await client.pause_keyword(kid)
                    n_pause += 1
                    affected.append(kw_text)
                except Exception:
                    n_fail += 1
            await asyncio.sleep(0.15)
        if affected:
            pool.mark_rejected_by_naver(
                customer_id,
                [{"keyword": kw, "reason": f"purge-drift(score≤{threshold})"} for kw in affected],
            )
        try:
            record_auto_cleanup_run(uid, str(customer_id), n_del + n_pause)
        except Exception:
            pass
        logger.warning(
            f"[purge-drift] uid={uid} cid={customer_id} thr={threshold} "
            f"basis={basis_source}({len(score_basis)}) → "
            f"user_seed_del={user_seed_deleted} pending_del={pending_deleted} "
            f"reg_del={n_del} reg_pause={n_pause} reg_fail={n_fail}"
        )

    background_tasks.add_task(_run_reg_purge)
    return {
        "success": True,
        "dry_run": False,
        "customer_id": customer_id,
        "threshold": threshold,
        "basis_source": basis_source,
        "basis_count": len(score_basis),
        "user_seed_deleted": user_seed_deleted,
        "pending_deleted": pending_deleted,
        "registered_queued": len(reg_capped),
        "registered_below_threshold_total": len(reg_targets),
        "estimated_minutes": round(len(reg_capped) * 0.18 / 60, 1),
        "message": (
            f"즉시 정리: user_seed -{user_seed_deleted}, pending -{pending_deleted}. "
            f"백그라운드 정리: registered {len(reg_capped)}개 Naver DELETE 진행 "
            f"(예상 {round(len(reg_capped) * 0.18 / 60, 1)}분)."
        ),
        "stage1_duration_ms": int((_t.monotonic() - t0) * 1000),
    }


# ============ 비도메인 시드 일괄 정리 ============
# 과거 POOL_DOMAIN_TOKENS bridge 누수로 의료 광고주(소잠한의원)에 "렌탈/임대/요가/피자
# /펜션/점포/파우치" 같은 무관 시드들이 박혀있음. cold_start-only fix 이후 신규는 안
# 들어오지만, 이미 등록된 KW (시드별 1500~2200개) 가 남아있어 광고비/도메인 평판 손실.
# 시드 단위로 그 lineage 의 모든 KW 일괄 DELETE.

class CleanupNonDomainSeedsRequest(BaseModel):
    domain_keywords: Optional[List[str]] = Field(
        None,
        description="도메인 정의 — 이 키워드와 atom 매칭 안 되는 시드는 모두 비도메인. "
                    "미입력 시 광고주 저장 relevance_keywords → user_seed 폴백.",
    )
    dry_run: bool = Field(True, description="True 면 삭제 대상 미리보기만, False 면 실제 삭제 시작.")
    max_delete: int = Field(5000, ge=0, le=20000, description="이번 실행 KW 삭제 상한 (네이버 API rate)")


@router.post("/keyword-pool/seeds/cleanup-non-domain")
async def cleanup_non_domain_seeds(
    request: CleanupNonDomainSeedsRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """비도메인 시드 lineage 일괄 정리 — 과거 POOL bridge 누수 잔재 제거.

    1. domain_keywords 로 도메인 토큰셋 빌드
    2. naverad_keyword_pool 의 distinct seed (status=registered) 조회 + KW 수
    3. 도메인 atom 안 맞는 seed = 비도메인 → 그 lineage 모든 KW DELETE 대상
    4. dry_run=True: 시드/KW 카운트 미리보기. dry_run=False: 백그라운드 삭제.
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import get_ad_account_relevance_keywords, record_auto_cleanup_run
    import sqlite3 as _sqlite3
    import time as _t

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    pool = get_keyword_pool_db()
    reg = get_registered_keywords_db()

    # 1) 도메인 키워드 결정
    if request.domain_keywords:
        domain_kws = [s.strip() for s in request.domain_keywords if s and len(s.strip()) >= 2]
        basis_source = "input"
    else:
        saved = get_ad_account_relevance_keywords(user_id, str(cid))
        if saved:
            domain_kws = saved
            basis_source = "saved_relevance"
        else:
            domain_kws = [s for s in (pool.list_user_seeds(cid) or []) if s and len(s) >= 2]
            basis_source = "user_seed_fallback"
    if not domain_kws:
        raise HTTPException(
            status_code=400,
            detail="도메인 키워드 없음 — domain_keywords 입력 또는 광고주 relevance_keywords 저장 필요.",
        )
    domain_tokens = _build_domain_token_set(domain_kws) | _build_seed_atoms(domain_kws)

    def _matches_domain(seed: str) -> bool:
        if not seed or len(seed) < 2:
            return False
        s = seed.replace(" ", "")
        return any(t in s for t in domain_tokens)

    # 2) 시드별 등록 KW 수 집계 (status=registered)
    with _sqlite3.connect(pool.db_path) as conn:
        seed_rows = conn.execute(
            """SELECT coalesce(seed,''), COUNT(*) AS n
               FROM naverad_keyword_pool
               WHERE account_customer_id=? AND status='registered'
               GROUP BY coalesce(seed,'')
               ORDER BY n DESC""",
            (cid,),
        ).fetchall()

    domain_seeds: List[Tuple[str, int]] = []
    non_domain_seeds: List[Tuple[str, int]] = []
    for s, n in seed_rows:
        if not s:
            non_domain_seeds.append((s, n))
            continue
        (domain_seeds if _matches_domain(s) else non_domain_seeds).append((s, n))

    # 3) 비도메인 시드의 KW + ncc_keyword_id 조회 (JOIN)
    if not non_domain_seeds:
        return {
            "success": True, "dry_run": request.dry_run, "customer_id": cid,
            "basis_source": basis_source, "domain_keywords_count": len(domain_kws),
            "domain_seeds": len(domain_seeds), "non_domain_seeds": 0,
            "total_targets": 0,
            "message": "비도메인 시드 없음 — 모든 등록 시드가 도메인 매칭됨",
        }

    non_domain_seed_set = {s for s, _ in non_domain_seeds}
    placeholders = ",".join("?" * len(non_domain_seed_set))
    with _sqlite3.connect(pool.db_path) as conn:
        conn.row_factory = _sqlite3.Row
        kw_rows = conn.execute(
            f"""SELECT p.keyword, p.seed, r.ncc_keyword_id
                FROM naverad_keyword_pool p
                LEFT JOIN registered_keywords r
                  ON r.account_customer_id = p.account_customer_id
                 AND r.keyword = p.keyword
                WHERE p.account_customer_id = ?
                  AND p.status = 'registered'
                  AND coalesce(p.seed,'') IN ({placeholders})
                  AND r.ncc_keyword_id IS NOT NULL""",
            (cid, *non_domain_seed_set),
        ).fetchall()
    targets = [(r["ncc_keyword_id"], r["keyword"], r["seed"]) for r in kw_rows]
    targets_capped = targets[:max(0, min(20000, int(request.max_delete)))]

    if request.dry_run:
        # Top 20 비도메인 시드 + 대상 KW sample
        return {
            "success": True, "dry_run": True, "customer_id": cid,
            "basis_source": basis_source,
            "domain_keywords_count": len(domain_kws),
            "domain_keywords_sample": domain_kws[:10],
            "domain_tokens_count": len(domain_tokens),
            "domain_seeds": len(domain_seeds),
            "non_domain_seeds": len(non_domain_seeds),
            "non_domain_top": [
                {"seed": s, "registered_count": n} for s, n in non_domain_seeds[:30]
            ],
            "total_targets": len(targets),
            "will_delete_now": len(targets_capped),
            "max_delete": request.max_delete,
            "estimated_minutes": round(len(targets_capped) * 0.18 / 60, 1),
            "samples": [
                {"keyword": kw, "seed": sd}
                for _, kw, sd in targets[:20]
            ],
        }

    if not targets_capped:
        return {
            "success": True, "dry_run": False, "customer_id": cid,
            "queued_targets": 0, "message": "삭제 대상 KW 없음",
        }

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    async def _run():
        n_del, n_pause, n_fail = 0, 0, 0
        affected: List[str] = []
        for kid, kw_text, _seed in targets_capped:
            try:
                await client.delete_keyword(kid)
                with _sqlite3.connect(reg.db_path) as c:
                    c.execute(
                        "DELETE FROM registered_keywords "
                        "WHERE account_customer_id=? AND ncc_keyword_id=?",
                        (cid, kid),
                    )
                with _sqlite3.connect(pool.db_path) as c:
                    c.execute(
                        "UPDATE naverad_keyword_pool SET status='deleted' "
                        "WHERE account_customer_id=? AND keyword=?",
                        (cid, kw_text),
                    )
                n_del += 1
                affected.append(kw_text)
            except Exception:
                try:
                    await client.pause_keyword(kid)
                    n_pause += 1
                    affected.append(kw_text)
                except Exception:
                    n_fail += 1
            await asyncio.sleep(0.15)
        try:
            pool.record_run(
                user_id, cid, "inspect",
                "success" if n_del > 0 else "no_new",
                registered=0, failed=n_fail, skipped=n_del,
                seeds_count=len(non_domain_seed_set),
                error_message=(
                    f"비도메인 시드 정리 ({len(non_domain_seed_set)} 시드) — "
                    f"DELETE {n_del} / PAUSE {n_pause} / 실패 {n_fail}"
                ),
            )
        except Exception:
            pass
        record_auto_cleanup_run(user_id, str(cid), n_del + n_pause)
        logger.warning(
            f"[cleanup-non-domain] uid={user_id} cid={cid} "
            f"non_domain_seeds={len(non_domain_seed_set)} → del={n_del} pause={n_pause} fail={n_fail}"
        )

    background_tasks.add_task(_run)
    return {
        "success": True, "dry_run": False, "customer_id": cid,
        "queued_targets": len(targets_capped),
        "non_domain_seeds": len(non_domain_seed_set),
        "estimated_minutes": round(len(targets_capped) * 0.18 / 60, 1),
        "message": (
            f"백그라운드 실행 시작 — 비도메인 시드 {len(non_domain_seed_set)}개의 "
            f"KW {len(targets_capped)}개 삭제 진행 (예상 {round(len(targets_capped) * 0.18 / 60, 1)}분)"
        ),
    }


# ============ 네이버 ↔ DB 한도 sync ============
# 사용자가 네이버 광고 콘솔에서 직접 캠페인/광고그룹 삭제 → 우리 DB row 는 stale.
# 한도 사용량 표시가 잘못됨 (네이버 active != DB count). 이 endpoint 가 cross-check
# 후 사라진 캠페인의 DB row 삭제.

@router.post("/keyword-pool/admin/reconcile-naver")
async def keyword_pool_reconcile_naver(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """네이버 광고 콘솔에서 직접 삭제한 캠페인 — 우리 DB 정리. 한도 사용량 정확화."""
    from services.naver_ad_service import NaverAdApiClient
    import sqlite3 as _sqlite3
    from asyncio import sleep as _sleep

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    # 1) 네이버 active 캠페인 list
    try:
        campaigns = await client.get_campaigns()
        live_campaign_ids = set(
            c.get("nccCampaignId") for c in (campaigns or [])
            if c.get("nccCampaignId")
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"네이버 캠페인 조회 실패: {type(e).__name__}: {str(e)[:200]}",
        )

    reg = get_registered_keywords_db()
    # 2) 우리 DB 의 distinct campaign_id
    with _sqlite3.connect(reg.db_path) as conn:
        rows = conn.execute(
            "SELECT DISTINCT campaign_id FROM registered_keywords "
            "WHERE account_customer_id=? AND campaign_id IS NOT NULL",
            (cid,),
        ).fetchall()
    db_campaign_ids = set(r[0] for r in rows if r[0])

    # SAFETY: 네이버 API 가 빈 응답 (rate limit / 일시 장애) 일 때 DB 통째 wipe 차단.
    # 과거 사고: 50k 일괄 삭제 트래픽 중 reconcile 클릭 → live=[] → db_campaigns 전체가
    # "사라진 캠페인" 으로 분류 → registered_keywords 테이블 전멸 → 한도 0 표시 사고.
    if not live_campaign_ids and db_campaign_ids:
        raise HTTPException(
            status_code=503,
            detail=(
                "네이버 캠페인 list 가 비었음 — API 일시 장애 가능성. "
                f"DB 에는 {len(db_campaign_ids)}개 캠페인 등록됨. "
                "이 상태에서 sync 진행 시 DB 통째 삭제 위험. 잠시 후 재시도."
            ),
        )

    # 3) DB 에 있지만 네이버에 없는 캠페인 → 그 캠페인의 모든 row 삭제
    deleted_campaigns = db_campaign_ids - live_campaign_ids
    n_rows_deleted = 0
    with _sqlite3.connect(reg.db_path) as conn:
        for cid_to_delete in deleted_campaigns:
            cur = conn.execute(
                "DELETE FROM registered_keywords WHERE account_customer_id=? AND campaign_id=?",
                (cid, cid_to_delete),
            )
            n_rows_deleted += cur.rowcount or 0
        conn.commit()

    # 4) 광고그룹 단위 cross-check — 병렬 + sem=8 + 타임아웃 보호.
    # 이전: 100 캠페인 × sequential API 호출 = 60초+ 로 frontend timeout. 병렬화로
    # 단축 (8개 동시 → ~10~15초). 추가로 전체 작업 30초 cap — 초과 시 캠페인 단위 sync 만으로 끝.
    n_orphan_groups = 0
    try:
        live_ad_group_ids: set = set()
        sem_ag = asyncio.Semaphore(8)
        async def _fetch_ag(live_cid: str) -> List[str]:
            async with sem_ag:
                try:
                    if hasattr(client, "get_ad_groups"):
                        ags = await client.get_ad_groups(campaign_id=live_cid)
                        return [ag.get("nccAdgroupId") for ag in (ags or []) if ag.get("nccAdgroupId")]
                except Exception:
                    return []
                return []
        # 30초 안에 끝나는 만큼만 처리 — 더 긴 광고주는 캠페인 sync 로 이미 대부분 정리됨
        try:
            results = await asyncio.wait_for(
                asyncio.gather(*[_fetch_ag(c) for c in list(live_campaign_ids)]),
                timeout=30.0,
            )
            for batch in results:
                live_ad_group_ids.update(batch)
        except asyncio.TimeoutError:
            logger.warning(f"[reconcile] 광고그룹 cross-check 30s 초과 — 캠페인 sync 만 적용")
            live_ad_group_ids = set()  # 부분 결과로 잘못 삭제하지 않도록 비움

        if live_ad_group_ids:
            with _sqlite3.connect(reg.db_path) as conn:
                rows2 = conn.execute(
                    "SELECT DISTINCT ad_group_id FROM registered_keywords "
                    "WHERE account_customer_id=? AND ad_group_id IS NOT NULL",
                    (cid,),
                ).fetchall()
                db_ad_group_ids = set(r[0] for r in rows2 if r[0])
                orphan_ag = db_ad_group_ids - live_ad_group_ids
                for agid in orphan_ag:
                    cur = conn.execute(
                        "DELETE FROM registered_keywords WHERE account_customer_id=? AND ad_group_id=?",
                        (cid, agid),
                    )
                    n_orphan_groups += cur.rowcount or 0
                conn.commit()
    except Exception as e:
        logger.warning(f"[reconcile] 광고그룹 cross-check 실패: {e}")

    # 5) 한도 재계산
    new_active = int((reg.stats(cid) or {}).get("active") or 0)
    logger.warning(
        f"[reconcile] uid={user_id} cid={cid} "
        f"live_campaigns={len(live_campaign_ids)} db_campaigns={len(db_campaign_ids)} "
        f"deleted_campaigns={len(deleted_campaigns)} kw_rows_deleted={n_rows_deleted} "
        f"orphan_ag_kws_deleted={n_orphan_groups} new_active={new_active}"
    )
    return {
        "success": True,
        "live_campaigns": len(live_campaign_ids),
        "db_campaigns": len(db_campaign_ids),
        "deleted_campaigns": len(deleted_campaigns),
        "deleted_kw_rows": n_rows_deleted + n_orphan_groups,
        "new_active": new_active,
    }


@router.post("/keyword-pool/admin/wipe-customer-db")
async def keyword_pool_wipe_customer_db(
    customer_id: Optional[str] = None,
    confirm: str = Query(..., description="WIPE 입력 시에만 실행 (안전장치)"),
    user_id: int = Depends(get_user_id_with_fallback),
):
    """광고주의 registered_keywords + naverad_keyword_pool row 일괄 wipe.

    네이버 광고 콘솔에서 사용자가 KW 들을 직접 일괄 삭제한 경우, 우리 DB sync 용도.
    pool 전체 wipe 라서 새 explode 가 깨끗하게 시작 가능 (이전 'registered'/'failed'
    dedup 안 됨). 캠페인/광고그룹 row 는 건드리지 않음 — Naver 에 존재하면 재사용됨.

    안전: confirm="WIPE" 필수. customer_id 명시 필수.
    """
    if confirm != "WIPE":
        raise HTTPException(status_code=400, detail="confirm=WIPE 명시 필요 (안전장치)")
    if not customer_id:
        raise HTTPException(status_code=400, detail="customer_id 필수")
    import sqlite3 as _sqlite3
    account = _resolve_account(user_id, customer_id)
    if not account:
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    reg = get_registered_keywords_db()
    pool = get_keyword_pool_db()

    n_reg = 0
    n_pool = 0
    with _sqlite3.connect(reg.db_path) as conn:
        cur = conn.execute(
            "DELETE FROM registered_keywords WHERE account_customer_id=?",
            (cid,),
        )
        n_reg = cur.rowcount or 0
        conn.commit()
    with _sqlite3.connect(pool.db_path) as conn:
        cur = conn.execute(
            "DELETE FROM naverad_keyword_pool WHERE account_customer_id=?",
            (cid,),
        )
        n_pool = cur.rowcount or 0
        conn.commit()

    logger.warning(
        f"[admin/wipe-customer-db] uid={user_id} cid={cid} "
        f"registered_rows={n_reg} pool_rows={n_pool}"
    )
    return {
        "success": True,
        "customer_id": cid,
        "registered_rows_deleted": n_reg,
        "pool_rows_deleted": n_pool,
    }


@router.post("/keyword-pool/admin/rebuild-from-naver")
async def keyword_pool_rebuild_from_naver(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """네이버에 실제 등록된 KW 를 전부 pull → registered_keywords 테이블 재구성.

    복구용. reconcile 버그 (2026-05-19) 로 DB row 전멸 사고 후 한도 사용량 0 표시
    되는 계정을 실제 네이버 상태로 동기화. UPSERT 라서 여러 번 실행해도 안전.
    """
    import sqlite3 as _sqlite3
    from services.naver_ad_service import NaverAdApiClient

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    # 1) campaigns
    try:
        campaigns = await client.get_campaigns()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"네이버 캠페인 조회 실패: {type(e).__name__}: {str(e)[:200]}")
    live_campaigns = [c for c in (campaigns or []) if c.get("nccCampaignId")]
    if not live_campaigns:
        raise HTTPException(status_code=503, detail="네이버 캠페인 list 가 비었음 — API 일시 장애 가능성")

    # 2) ad_groups (병렬 sem=8)
    sem = asyncio.Semaphore(8)

    async def _fetch_groups(camp_id: str):
        async with sem:
            try:
                ags = await client.get_ad_groups(campaign_id=camp_id) or []
                return camp_id, [ag.get("nccAdgroupId") for ag in ags if ag.get("nccAdgroupId")]
            except Exception as e:
                logger.warning(f"[rebuild] get_ad_groups({camp_id}) 실패: {e}")
                return camp_id, []

    ag_results = await asyncio.gather(*[_fetch_groups(c["nccCampaignId"]) for c in live_campaigns])
    ag_to_camp: Dict[str, str] = {}
    for camp_id, ag_ids in ag_results:
        for ag_id in ag_ids:
            ag_to_camp[ag_id] = camp_id

    if not ag_to_camp:
        raise HTTPException(status_code=503, detail=f"네이버 광고그룹 0개 — 캠페인 {len(live_campaigns)}개 있는데 그룹 못 가져옴")

    # 3) keywords per ad_group (병렬 sem=8)
    async def _fetch_kws(ag_id: str):
        async with sem:
            try:
                return ag_id, (await client.get_keywords(ad_group_id=ag_id) or [])
            except Exception as e:
                logger.warning(f"[rebuild] get_keywords({ag_id}) 실패: {e}")
                return ag_id, []

    kw_results = await asyncio.gather(*[_fetch_kws(ag_id) for ag_id in ag_to_camp.keys()])

    rows: List[Dict] = []
    for ag_id, kws in kw_results:
        camp_id = ag_to_camp.get(ag_id)
        for kw in kws:
            text = (kw.get("keyword") or "").strip()
            if not text:
                continue
            rows.append({
                "keyword": text,
                "ad_group_id": ag_id,
                "campaign_id": camp_id,
                "bid_amt": kw.get("bidAmt"),
                "ncc_keyword_id": kw.get("nccKeywordId"),
            })

    if not rows:
        raise HTTPException(
            status_code=503,
            detail=f"네이버에서 KW 0개 발견 — 캠페인 {len(live_campaigns)}개 / 그룹 {len(ag_to_camp)}개 있는데 KW 없음. 진짜 빈 상태이거나 API 부분 장애.",
        )

    # 4) UPSERT — 기존 row 의 removed_at 도 클리어 (네이버에 실제 있으면 live).
    reg = get_registered_keywords_db()
    with reg._conn() as conn:
        cur = conn.cursor()
        for r in rows:
            try:
                cur.execute(
                    """INSERT INTO registered_keywords
                       (user_id, account_customer_id, keyword, ad_group_id,
                        campaign_id, bid_amt, ncc_keyword_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?)
                       ON CONFLICT(account_customer_id, keyword) DO UPDATE SET
                         removed_at = NULL,
                         ad_group_id = excluded.ad_group_id,
                         campaign_id = excluded.campaign_id,
                         bid_amt = excluded.bid_amt,
                         ncc_keyword_id = excluded.ncc_keyword_id""",
                    (user_id, cid, r["keyword"], r["ad_group_id"], r["campaign_id"],
                     r["bid_amt"], r["ncc_keyword_id"]),
                )
            except _sqlite3.Error as e:
                logger.warning(f"[rebuild] upsert 실패 {r['keyword']}: {e}")

    new_active = int((reg.stats(cid) or {}).get("active") or 0)
    logger.warning(
        f"[rebuild] uid={user_id} cid={cid} campaigns={len(live_campaigns)} "
        f"ad_groups={len(ag_to_camp)} pulled={len(rows)} new_active={new_active}"
    )
    return {
        "success": True,
        "campaigns": len(live_campaigns),
        "ad_groups": len(ag_to_camp),
        "pulled": len(rows),
        "new_active": new_active,
    }


# 큰 계정용 백그라운드 rebuild 진행상태 (cid -> 상태dict). 워커 1개 가정.
_REBUILD_BG_STATUS: Dict[int, dict] = {}


@router.post("/keyword-pool/admin/rebuild-from-naver-bg")
async def keyword_pool_rebuild_from_naver_bg(
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """rebuild-from-naver 의 **백그라운드** 버전. 큰 계정(메디론 등)은 동기 호출이
    오래 걸려 클라 disconnect → Starlette 가 핸들러 취소 → UPSERT 안 됨. 이 버전은
    즉시 반환하고 백그라운드로 campaigns→adgroups→keywords pull 후 UPSERT(removed_at 클리어).
    진행상태: GET .../rebuild-from-naver-bg/status?customer_id= . UPSERT 라 여러 번 안전."""
    import sqlite3 as _sqlite3
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    _REBUILD_BG_STATUS[cid] = {"state": "running", "campaigns": 0, "ad_groups": 0,
                              "pulled": 0, "new_active": None, "error": None}

    async def _do():
        st = _REBUILD_BG_STATUS[cid]
        try:
            client = NaverAdApiClient()
            client.customer_id = account["customer_id"]
            client.api_key = account["api_key"]
            client.secret_key = account["secret_key"]
            campaigns = await client.get_campaigns()
            live_campaigns = [c for c in (campaigns or []) if c.get("nccCampaignId")]
            st["campaigns"] = len(live_campaigns)
            if not live_campaigns:
                st["state"] = "error"; st["error"] = "campaigns empty"; return
            sem = asyncio.Semaphore(8)

            async def _fg(camp_id: str):
                async with sem:
                    try:
                        ags = await client.get_ad_groups(campaign_id=camp_id) or []
                        return [ag.get("nccAdgroupId") for ag in ags if ag.get("nccAdgroupId")]
                    except Exception:
                        return []

            ag_lists = await asyncio.gather(*[_fg(c["nccCampaignId"]) for c in live_campaigns])
            ag_to_camp: Dict[str, str] = {}
            for c, ags in zip(live_campaigns, ag_lists):
                for ag in ags:
                    ag_to_camp[ag] = c["nccCampaignId"]
            st["ad_groups"] = len(ag_to_camp)
            if not ag_to_camp:
                st["state"] = "error"; st["error"] = "ad_groups empty"; return

            async def _fk(ag_id: str):
                async with sem:
                    try:
                        return ag_id, (await client.get_keywords(ad_group_id=ag_id) or [])
                    except Exception:
                        return ag_id, []

            kw_results = await asyncio.gather(*[_fk(a) for a in ag_to_camp.keys()])
            rows = []
            for ag_id, kws in kw_results:
                camp_id = ag_to_camp.get(ag_id)
                for kw in kws:
                    text = (kw.get("keyword") or "").strip()
                    if not text:
                        continue
                    rows.append((text, ag_id, camp_id, kw.get("bidAmt"), kw.get("nccKeywordId")))
            st["pulled"] = len(rows)
            if not rows:
                st["state"] = "error"; st["error"] = "0 keywords pulled"; return
            reg = get_registered_keywords_db()
            with reg._conn() as conn:
                cur = conn.cursor()
                for text, ag_id, camp_id, bid, nid in rows:
                    try:
                        cur.execute(
                            """INSERT INTO registered_keywords
                               (user_id, account_customer_id, keyword, ad_group_id,
                                campaign_id, bid_amt, ncc_keyword_id)
                               VALUES (?, ?, ?, ?, ?, ?, ?)
                               ON CONFLICT(account_customer_id, keyword) DO UPDATE SET
                                 removed_at = NULL,
                                 ad_group_id = excluded.ad_group_id,
                                 campaign_id = excluded.campaign_id,
                                 bid_amt = excluded.bid_amt,
                                 ncc_keyword_id = excluded.ncc_keyword_id""",
                            (user_id, cid, text, ag_id, camp_id, bid, nid),
                        )
                    except _sqlite3.Error:
                        pass
            st["new_active"] = int((reg.stats(cid) or {}).get("active") or 0)
            st["state"] = "done"
            logger.warning(
                f"[rebuild-bg] cid={cid} campaigns={st['campaigns']} "
                f"ad_groups={st['ad_groups']} pulled={st['pulled']} active={st['new_active']}"
            )
        except Exception as e:
            st["state"] = "error"; st["error"] = f"{type(e).__name__}: {str(e)[:200]}"
            logger.error(f"[rebuild-bg] cid={cid} 실패: {e}", exc_info=True)

    background_tasks.add_task(_do)
    return {"success": True, "started": True, "customer_id": cid,
            "message": "백그라운드 rebuild 시작 — status 엔드포인트로 진행 확인"}


@router.get("/keyword-pool/admin/rebuild-from-naver-bg/status")
async def keyword_pool_rebuild_from_naver_bg_status(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """백그라운드 rebuild 진행상태 조회."""
    account = _resolve_account(user_id, customer_id)
    if not account:
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    return {"success": True, "customer_id": cid,
            "status": _REBUILD_BG_STATUS.get(cid, {"state": "none"})}


@router.delete("/keyword-pool/keywords/{keyword}")
async def keyword_pool_delete_keyword(
    keyword: str,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """단일 키워드를 풀에서 삭제 (이미 네이버 등록된 건 영향 없음)."""
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        customer_id = int(account.get("customer_id"))
        pool = get_keyword_pool_db()
        n = pool.delete_keywords(customer_id, [keyword])
        return {"success": True, "deleted": n, "keyword": keyword}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/keywords DELETE 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.delete("/keyword-pool/seeds/{seed}")
async def keyword_pool_delete_seed(
    seed: str,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """시드와 그 시드로 발굴된 자식 키워드를 풀에서 모두 삭제."""
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        customer_id = int(account.get("customer_id"))
        pool = get_keyword_pool_db()
        n = pool.delete_seed_with_children(customer_id, seed)
        return {"success": True, "deleted": n, "seed": seed}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/seeds DELETE 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.post("/keyword-pool/seeds")
async def keyword_pool_add_seeds(
    request: PoolSeedsRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """초기 시드 추가 — 자동 수집의 첫 input."""
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        customer_id = int(account.get("customer_id"))
        pool = get_keyword_pool_db()
        items = [
            {"keyword": s.strip(), "seed": s.strip(), "source": "user_seed", "monthly_total": 0}
            for s in request.seeds if s and s.strip()
        ]
        added = pool.add_candidates(user_id, customer_id, items)
        return {"success": True, "added": added, "total_input": len(items)}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/seeds 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


class AiSeedExpandRequest(BaseModel):
    base_seeds: List[str] = Field(..., min_items=1, description="원본 시드 — 사용자 도메인 의도. 풀이 오염되어도 이 입력만 사용.")
    cycles: int = Field(1, ge=1, le=3, description="확장 사이클. 1=직접 확장, 2+=직전 결과를 다시 시드로.")
    keywords_per_cycle: int = Field(80, ge=10, le=200, description="LLM 한 번에 생성 후보 수")
    min_volume: int = Field(5, ge=0, le=10000, description="검증 통과 최소 월 검색량")


@router.post("/keyword-pool/seeds/ai-expand")
async def keyword_pool_ai_expand_seeds(
    request: AiSeedExpandRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """LLM 으로 사용자 시드 도메인 일관성 있게 확장 → keywordstool 검증 → user_seed INSERT.

    풀이 한약재/식물 등으로 오염되어 collect 가 cross-domain drift 하는 사례를 우회.
    base_seeds 만 도메인 기준으로 사용 — DB 풀의 잡음 시드는 무시.

    1. base_seeds → GPT-4o-mini → keywords_per_cycle 개 후보
    2. base_seeds 에서 derive 한 도메인 토큰으로 1차 필터 (LLM drift 차단)
    3. keywordstool 5개씩 배치 → 검색량 ≥ min_volume 만 통과
    4. user_seed 로 INSERT (source='ai_seed_expansion')
    5. cycles 회 반복 (직전 결과를 다음 base 로)
    """
    import json as _json
    from config import settings as _settings
    if not getattr(_settings, "OPENAI_API_KEY", ""):
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY 미설정")

    base_seeds = [s.strip() for s in request.base_seeds if s and s.strip()]
    if not base_seeds:
        raise HTTPException(status_code=400, detail="base_seeds 비어있음")

    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    customer_id = int(account.get("customer_id"))

    # base_seeds 만으로 도메인 토큰셋 빌드 (DB 풀의 한약재 오염 시드 무시)
    domain_tokens = _build_domain_token_set(base_seeds) | _build_seed_atoms(base_seeds)

    def _matches_user_domain(kw: str) -> bool:
        k = (kw or "").replace(" ", "")
        if len(k) < 2:
            return False
        return any(t in k for t in domain_tokens)

    from services.naver_ad_service import NaverAdApiClient
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    pool = get_keyword_pool_db()
    cycle_results: List[Dict] = []
    current_seeds = list(base_seeds)

    for cycle_idx in range(request.cycles):
        # 1) LLM 호출 — 도메인 일관성 강제 프롬프트
        prompt_seeds = ", ".join(current_seeds[:60])  # 토큰 절약
        prompt = (
            f"다음 한국어 키워드들과 동일 도메인의 검색 가능성 있는 한국어 키워드를 정확히 "
            f"{request.keywords_per_cycle}개 생성해줘.\n\n"
            f"입력 키워드:\n{prompt_seeds}\n\n"
            f"규칙:\n"
            f"- 입력 키워드와 동일한 분야/도메인 안에서만 확장 (예: 의료면 의료, 부동산이면 부동산)\n"
            f"- 다른 도메인의 단어 절대 금지 (예: 의료 시드면 한약재/식물/대출/렌탈 등 절대 포함 금지)\n"
            f"- 띄어쓰기 가능 (네이버 검색 사용자 자연스러운 형태)\n"
            f"- 한 줄당 1개, 일련번호/설명 없이 키워드만\n"
            f"- 결과는 JSON array (예: [\"키워드1\", \"키워드2\", ...])"
        )
        try:
            async with httpx.AsyncClient(timeout=60.0) as oai:
                resp = await oai.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={
                        "Authorization": f"Bearer {_settings.OPENAI_API_KEY}",
                        "Content-Type": "application/json",
                    },
                    json={
                        "model": "gpt-4o-mini",
                        "messages": [
                            {"role": "system", "content": "한국어 검색광고 키워드 전문가. 도메인 일관성을 절대 위반하지 마. JSON array 만 반환."},
                            {"role": "user", "content": prompt},
                        ],
                        "temperature": 0.7,
                        "max_tokens": 3000,
                    },
                )
                resp.raise_for_status()
                content = resp.json()["choices"][0]["message"]["content"]
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"LLM 호출 실패: {type(e).__name__}: {str(e)[:200]}")

        # JSON 파싱 — code block 제거
        import re as _re
        cb = _re.search(r"```(?:json)?\s*(.*?)\s*```", content, _re.DOTALL)
        if cb:
            content = cb.group(1)
        try:
            generated = _json.loads(content.strip())
            if isinstance(generated, dict):
                generated = list(generated.values()) if generated else []
            if not isinstance(generated, list):
                raise ValueError("not a list")
        except Exception as e:
            logger.warning(f"[ai-expand] JSON 파싱 실패 cycle={cycle_idx}: {e} content[:200]={content[:200]!r}")
            generated = []

        # 후보 정규화
        candidates: List[str] = []
        seen = set()
        for k in generated:
            if not isinstance(k, str):
                continue
            kw = k.strip()
            if not kw or len(kw) < 2 or kw in seen:
                continue
            seen.add(kw)
            candidates.append(kw)

        # 2) 도메인 1차 필터 (LLM drift 컷)
        domain_pass = [k for k in candidates if _matches_user_domain(k)]
        domain_fail = len(candidates) - len(domain_pass)

        # 3) keywordstool 배치 검증 (5개씩) — 네이버 응답의 keywordList 에서
        # 입력 hint 와 normalized 매칭. 매칭 안 되면 (= 검색량 데이터 없음) skip.
        validated: List[Dict] = []
        validated_seen: Set[str] = set()

        def _norm(s: str) -> str:
            return (s or "").replace(" ", "").upper()

        for i in range(0, len(domain_pass), 5):
            chunk = domain_pass[i:i + 5]
            try:
                vol_map = await client.get_keywords_volume_batch(chunk)
            except Exception as e:
                logger.warning(f"[ai-expand] volume batch 실패 {chunk}: {e}")
                continue
            # vol_map key 정규화 — Naver 응답 키는 공백 제거된 형태로 옴
            norm_vol_map: Dict[str, Dict] = {_norm(k): v for k, v in vol_map.items()}
            for kw in chunk:
                vinfo = norm_vol_map.get(_norm(kw))
                if not vinfo:
                    continue
                mt = int(vinfo.get("monthly_total") or 0)
                if mt < request.min_volume:
                    continue
                if kw in validated_seen:
                    continue
                validated_seen.add(kw)
                validated.append({
                    "keyword": kw,
                    "monthly_total": mt,
                    "monthly_pc": int(vinfo.get("monthly_pc") or 0),
                    "monthly_mobile": int(vinfo.get("monthly_mobile") or 0),
                    "comp_idx": vinfo.get("comp_idx"),
                    "seed": f"ai_cycle{cycle_idx+1}",
                })
            await asyncio.sleep(0.3)

        # 4) user_seed 로 INSERT — 다음 cycle 의 시드로 활용
        # add_candidates 는 source=user_seed 가 아니어도 INSERT 함. 시드 효과 위해
        # source 를 명시적으로 'user_seed' 로 바꿔 INSERT (확장 시드도 collect 가 사용).
        seed_items = [
            {**v, "source": "user_seed", "monthly_total": v["monthly_total"]}
            for v in validated
        ]
        added = pool.add_candidates(user_id, customer_id, seed_items)

        cycle_results.append({
            "cycle": cycle_idx + 1,
            "llm_generated": len(candidates),
            "domain_filter_pass": len(domain_pass),
            "domain_filter_fail": domain_fail,
            "volume_validated": len(validated),
            "inserted_as_seed": added,
            "samples": [v["keyword"] for v in validated[:8]],
        })
        logger.warning(
            f"[ai-expand] user={user_id} cid={customer_id} cycle {cycle_idx+1}/{request.cycles}: "
            f"LLM {len(candidates)} → domain {len(domain_pass)} → vol≥{request.min_volume} {len(validated)} → INSERT {added}"
        )

        # 다음 cycle 의 base 는 이번 통과 키워드들 (없으면 종료)
        if not validated:
            break
        current_seeds = [v["keyword"] for v in validated][:80]

    total_added = sum(r["inserted_as_seed"] for r in cycle_results)
    return {
        "success": True,
        "total_added_seeds": total_added,
        "cycles": cycle_results,
        "base_seeds_count": len(base_seeds),
        "domain_tokens_count": len(domain_tokens),
    }


class BidBulkUpdateRequest(BaseModel):
    bid: int = Field(..., ge=70, le=100000, description="새 입찰가 (네이버 최소 70원)")
    scope: str = Field("pool", description="'pool' = auto_ 프리픽스 캠페인만, 'all' = 전체 캠페인")
    only_if_bid: Optional[int] = Field(None, description="설정 시 현재 입찰가가 이 값인 키워드만 변경 (예: 70 → 70원짜리만)")


@router.post("/keyword-pool/bid/bulk-update")
async def keyword_pool_bid_bulk_update(
    request: BidBulkUpdateRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """광고주의 default 입찰가를 DB 에 저장 + 광고그룹 default bid + **모든 키워드 bidAmt** 일괄 변경.

    - scope='pool': 풀 자동 등록 캠페인 (이름 'auto_*') 의 광고그룹만 변경
    - scope='all': 그 광고주의 모든 활성 캠페인 광고그룹 변경
    - 키워드별 bidAmt 도 같이 업데이트 — 자동 등록은 useGroupBidAmt=False 라 그룹 default 만
      바꿔서는 키워드별 표시가 안 바뀜. 광고관리자에 즉시 반영되도록 키워드 PUT 도 수행.
    """
    from services.naver_ad_service import NaverAdApiClient
    from database.naver_ad_db import update_ad_account_default_bid
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        cid = int(account.get("customer_id"))
        # 네이버 입찰가는 10원 단위만 유효 (아니면 code 3904 'Invalid bid amount'). 10단위 반올림 + 최소 70.
        new_bid = max(70, round(int(request.bid) / 10) * 10)

        # 1. DB 저장 — 앞으로 cron 이 이 값 사용. (only_if_bid 조건부 변경 시엔 default 유지)
        if request.only_if_bid is None:
            update_ad_account_default_bid(user_id, str(cid), new_bid)

        # 2. 백그라운드로 네이버 일괄 변경 (45k 키워드는 HTTP 타임아웃 초과 → bg 완주)
        async def _run():
            # 2. 네이버 API 광고그룹 일괄 변경
            client = NaverAdApiClient()
            client.customer_id = account["customer_id"]
            client.api_key = account["api_key"]
            client.secret_key = account["secret_key"]

            campaigns = await client.get_campaigns() or []
            if request.scope == "pool":
                campaigns = [c for c in campaigns if (c.get("name") or "").startswith("auto_")]
            else:
                # 'all' = 파워링크(WEB_SITE) 키워드 캠페인만 — 파워컨텐츠/플레이스/브랜드검색 제외
                campaigns = [c for c in campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]

            ad_group_ids: List[Tuple[str, str]] = []  # (campaign_name, ad_group_id)
            for c in campaigns:
                try:
                    groups = await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or []
                    for g in groups:
                        gid = g.get("nccAdgroupId")
                        if gid:
                            ad_group_ids.append((c.get("name") or "", gid))
                except Exception as e:
                    logger.warning(f"[bid/bulk] 광고그룹 list 실패 cid={c.get('nccCampaignId')}: {e}")
                await asyncio.sleep(0.15)

            # 광고그룹 default bid 변경 — only_if_bid 조건부 변경 시엔 그룹 default 유지(키워드만 변경)
            ag_success = 0
            ag_failed: List[Dict] = []
            if request.only_if_bid is None:
                for cname, gid in ad_group_ids:
                    try:
                        await client.update_ad_group_bid(gid, new_bid)
                        ag_success += 1
                    except Exception as e:
                        ag_failed.append({"ad_group_id": gid, "campaign": cname, "error": f"{type(e).__name__}: {str(e)[:120]}"})
                    await asyncio.sleep(0.15)

            # 키워드 bidAmt **bulk 변경** — PUT /ncc/keywords?fields=bidAmt 배열(최대 100개/콜).
            # 이전: 키워드 1개씩 PUT(10만회) → 느림+429. 이제 100개씩 묶어 ~100배 적은 호출.
            kw_total = 0
            kw_success = 0
            kw_skipped = 0
            kw_failed: List[Dict] = []
            progress_logged_at = 0
            from services.naver_ad_service import _naver_api_breaker, NaverApiCircuitOpenError

            async def _flush(batch: List[Dict]):
                nonlocal kw_success
                if not batch:
                    return
                try:
                    await client.update_keywords_bid_bulk(batch)
                    kw_success += len(batch)
                except NaverApiCircuitOpenError:
                    kw_failed.append({"error": "circuit_open", "count": len(batch)})
                except Exception as e:
                    kw_failed.append({"error": f"{type(e).__name__}: {str(e)[:120]}", "count": len(batch)})

            logger.warning(f"[bid/bulk] 시작(bulk) — scope={request.scope} ad_groups={len(ad_group_ids)} new_bid={new_bid}원 only_if={request.only_if_bid}")
            for idx, (cname, gid) in enumerate(ad_group_ids):
                try:
                    kws = await client.get_keywords(ad_group_id=gid) or []
                except Exception as e:
                    logger.warning(f"[bid/bulk] keywords list 실패 ag={gid}: {e}")
                    continue
                kw_total += len(kws)
                # 변경 대상만 추림: only_if_bid 일치(설정 시) + 이미 목표값 아님
                items = []
                for k in kws:
                    cur = k.get("bidAmt")
                    if request.only_if_bid is not None and cur != request.only_if_bid:
                        kw_skipped += 1; continue
                    if cur == new_bid and k.get("useGroupBidAmt") is False:
                        kw_skipped += 1; continue
                    items.append({
                        "nccKeywordId": k.get("nccKeywordId"),
                        "nccAdgroupId": gid,
                        "bidAmt": new_bid,
                        "useGroupBidAmt": False,
                    })
                # 100개씩 bulk PUT
                for i in range(0, len(items), 100):
                    await _flush(items[i:i+100])
                    await asyncio.sleep(0.1)
                if kw_success - progress_logged_at >= 2000:
                    progress_logged_at = kw_success
                    logger.warning(
                        f"[bid/bulk] 진행(bulk) — 광고그룹 {idx+1}/{len(ad_group_ids)} · "
                        f"변경 {kw_success} / skip {kw_skipped} / 스캔 {kw_total}"
                    )
                await asyncio.sleep(0.03)
            logger.warning(
                f"[bid/bulk] 완료(bulk) — 광고그룹 {len(ad_group_ids)} · "
                f"변경 {kw_success} / skip {kw_skipped} / 스캔 {kw_total} ({len(kw_failed)} 배치실패)"
            )

            return {
                "success": True,
                "customer_id": str(cid),
                "new_bid": new_bid,
                "scope": request.scope,
                "campaigns_scanned": len(campaigns),
                "ad_groups_total": len(ad_group_ids),
                "ad_groups_updated": ag_success,
                "ad_groups_failed": len(ag_failed),
                "keywords_total": kw_total,
                "keywords_updated": kw_success,
                "keywords_failed": len(kw_failed),
                "failed_samples": (ag_failed[:5] + kw_failed[:5])[:10],
            }
        background_tasks.add_task(_run)
        return {"success": True, "started": True, "scope": request.scope, "new_bid": new_bid,
                "message": f"백그라운드 일괄 변경 시작 — scope={request.scope} 모든 키워드 {new_bid}원 적용 (수십분 소요, 로그에서 진행 확인)"}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/bid/bulk-update 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.post("/keyword-pool/bid/debug-one")
async def keyword_pool_bid_debug_one(
    customer_id: Optional[str] = None,
    bid: int = 70,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """디버그 — 첫 번째 키워드 1개만 업데이트하고 Naver before/after + 응답 raw 그대로 돌려줌.

    bulk-update 가 silent 하게 실패하는 원인 추적용. 로그/응답 body 다 보여줌.
    """
    from services.naver_ad_service import NaverAdApiClient
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]

        # 키워드가 있는 첫 그룹을 찾을 때까지 WEB_SITE 캠페인을 스캔 (빈 그룹 skip).
        # 이전엔 auto_[0]/groups[0] 만 봐서 그 그룹이 비면 no_keyword 로 검증 불가했음.
        campaigns = await client.get_campaigns() or []
        cand = [c for c in campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]
        if not cand:
            cand = campaigns
        first_camp = None; gid = None; first_kw = None; scanned_groups = 0
        for c in cand:
            try:
                groups = await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or []
            except Exception:
                continue
            for g in groups:
                scanned_groups += 1
                _gid = g.get("nccAdgroupId")
                if not _gid:
                    continue
                try:
                    kws = await client.get_keywords(ad_group_id=_gid) or []
                except Exception:
                    continue
                if kws:
                    first_camp = c; gid = _gid; first_kw = kws[0]; break
                if scanned_groups >= 60:  # 안전 cap — 너무 많이 스캔 방지
                    break
            if first_kw:
                break
            if scanned_groups >= 60:
                break
        if not first_kw:
            return {"success": False, "step": "no_keyword", "scanned_groups": scanned_groups}
        kid = first_kw.get("nccKeywordId")

        # before
        before = {
            "nccKeywordId": kid,
            "keyword": first_kw.get("keyword"),
            "bidAmt": first_kw.get("bidAmt"),
            "useGroupBidAmt": first_kw.get("useGroupBidAmt"),
        }

        # PUT — 응답 그대로
        try:
            put_response = await client.update_keyword_bid(kid, max(70, int(bid)), ad_group_id=first_kw.get("nccAdgroupId"))
        except Exception as e:
            import traceback
            return {
                "success": False,
                "step": "put_failed",
                "before": before,
                "error": f"{type(e).__name__}: {str(e)[:500]}",
                "trace": traceback.format_exc()[:1000],
            }

        # after — GET 다시
        after_kw = await client.get_keyword(kid) if hasattr(client, 'get_keyword') else None
        if after_kw is None:
            kws_after = await client.get_keywords(ad_group_id=gid) or []
            after_kw = next((k for k in kws_after if k.get("nccKeywordId") == kid), {})
        after = {
            "bidAmt": after_kw.get("bidAmt"),
            "useGroupBidAmt": after_kw.get("useGroupBidAmt"),
        }
        return {
            "success": True,
            "campaign": first_camp.get("name"),
            "ad_group_id": gid,
            "keyword_id": kid,
            "before": before,
            "put_response": put_response,
            "after": after,
            "changed": before["bidAmt"] != after["bidAmt"],
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/bid/debug-one 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.post("/keyword-pool/bid/test-min-exposure")
async def keyword_pool_bid_test_min_exposure(
    customer_id: Optional[str] = None,
    keywords: str = "두드러기치료,아토피한의원,건선치료,소잠한의원",
    user_id: int = Depends(get_user_id_with_fallback),
):
    """디버그 — PC 최소노출입찰가 API 응답 형식 확인용. keywords=콤마구분."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
    kws = [k.strip() for k in keywords.split(",") if k.strip()]
    try:
        raw = await client.try_min_exposure_formats(kws)
        return {"success": True, "keywords": kws, "raw": raw}
    except Exception as e:
        import traceback
        return {"success": False, "error": f"{type(e).__name__}: {str(e)[:400]}", "trace": traceback.format_exc()[:600]}


class CoreMinExposureRequest(BaseModel):
    core_tokens: Optional[List[str]] = Field(None, description="핵심 토큰(이 중 하나라도 포함된 키워드만 대상). 미지정 시 기본 소잠 핵심셋")
    keywords: Optional[List[str]] = Field(None, description="명시적 키워드 리스트. 설정 시 토큰 무시하고 '정확히 이 키워드들'만 입찰 대상(정밀 적용)")
    whole_pool: bool = Field(False, description="true: 토큰 필터 무시하고 등록된 전체 키워드를 PC 최소노출가로")
    domain_tokens: Optional[List[str]] = Field(None, description="도메인(질환) 토큰 — 키워드가 이 중 하나는 반드시 포함(AND 조건)")
    intent_tokens: Optional[List[str]] = Field(None, description="의향 토큰 — 키워드가 이 중 하나는 반드시 포함(병원오기직전 bottom-funnel)")
    exclude_tokens: Optional[List[str]] = Field(None, description="제외 토큰 — 이 중 하나라도 포함하면 제외(원인/증상 등 정보검색)")
    device: str = Field("PC", description="PC 또는 MOBILE — 최소노출입찰가 산출 기준 디바이스")
    target_position: int = Field(10, description="목표 노출 순위 — 10=최저노출(맨아래), 5=기본화면 보임권, 3=상위. 해당 순위 입찰가로 설정")
    scope: str = Field("all", description="'all' WEB_SITE 전체")
    dry_run: bool = Field(True, description="true: 대상 개수+샘플 미리보기, false: 실제 적용")
    max_keywords: int = Field(150000, description="안전 상한")
    bid_cap: Optional[int] = Field(None, description="입찰가 상한(원). 설정 시 estimate 가 이보다 크면 cap 적용(고CPC 폭마 방지). 미지정 시 상한 없음")


@router.post("/keyword-pool/bid/core-min-exposure")
async def keyword_pool_bid_core_min_exposure(
    request: CoreMinExposureRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """핵심 키워드(핵심 질환+브랜드 토큰 포함)를 PC 최소노출입찰가로 설정.
    average-position-bid 를 순위 10→5→3 캐스케이드로 조회해 키워드별 '노출되는 최저 입찰가' 산출 후
    bulk PUT. dry_run=true 면 대상 개수+샘플만."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    core = request.core_tokens or [
        "소잠", "두드러기", "아토피", "건선", "지루성피부염", "지루피부염", "한포진", "습진",
        "피부질환", "피부염", "백반증", "대상포진", "주사비", "자반증", "피부가려움", "한방피부",
    ]

    async def _collect_core(client):
        # DB(registered_keywords)에서 대상 키워드 추출. 두 모드:
        #  (A) domain_tokens+intent_tokens 지정 시: (질환 AND 의향) AND NOT 제외 — '병원오기직전' 정밀 필터
        #  (B) 그 외: core 토큰 OR (기존 동작)
        import sqlite3 as _sq
        from database.registered_keywords_db import get_registered_keywords_db
        reg = get_registered_keywords_db()
        base = ("SELECT keyword, ad_group_id, ncc_keyword_id, bid_amt FROM registered_keywords "
                "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL "
                "AND removed_at IS NULL")
        params = [cid]
        if request.keywords:
            # 명시적 키워드 리스트 모드 — 전체 actionable 행을 가져와 파이썬에서 정확 매칭(SQL IN 한계 회피)
            kset = {k.strip() for k in request.keywords if k and k.strip()}
            with _sq.connect(reg.db_path) as conn:
                allrows = conn.execute(base, params).fetchall()
            rows = [r for r in allrows if (r[0] or "").strip() in kset]
            return [{"id": r[2], "gid": r[1], "text": r[0], "bid": r[3]} for r in rows]
        dom = request.domain_tokens or []
        itn = request.intent_tokens or []
        exc = request.exclude_tokens or []
        if request.whole_pool:
            # 전체 키워드 — positive 필터 없음, 제외 토큰(junk/타지역)만 NOT 적용
            if exc:
                base += " AND NOT (" + " OR ".join(["keyword LIKE ?"] * len(exc)) + ")"
                params += [f"%{t}%" for t in exc]
        elif dom and itn:
            base += " AND (" + " OR ".join(["keyword LIKE ?"] * len(dom)) + ")"
            params += [f"%{t}%" for t in dom]
            base += " AND (" + " OR ".join(["keyword LIKE ?"] * len(itn)) + ")"
            params += [f"%{t}%" for t in itn]
            if exc:
                base += " AND NOT (" + " OR ".join(["keyword LIKE ?"] * len(exc)) + ")"
                params += [f"%{t}%" for t in exc]
        else:
            base += " AND (" + " OR ".join(["keyword LIKE ?"] * len(core)) + ")"
            params += [f"%{t}%" for t in core]
        base += " LIMIT ?"
        params += [request.max_keywords]
        with _sq.connect(reg.db_path) as conn:
            rows = conn.execute(base, params).fetchall()
        return [{"id": r[2], "gid": r[1], "text": r[0], "bid": r[3]} for r in rows]

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]

    _dev = (request.device or "PC").upper()
    if _dev not in ("PC", "MOBILE"):
        _dev = "PC"

    if request.dry_run:
        core_kws = await _collect_core(client)
        import random as _rnd
        _samp = core_kws if len(core_kws) <= 25 else _rnd.sample(core_kws, 25)
        return {
            "success": True, "dry_run": True, "customer_id": cid, "device": _dev,
            "mode": "domain+intent" if (request.domain_tokens and request.intent_tokens) else "core_or",
            "matched": len(core_kws),
            "samples": [{"keyword": k["text"], "current_bid": k["bid"]} for k in _samp],
        }

    # 목표 순위 우선 캐스케이드 — target_position 의 입찰가를 우선 사용, 없으면 인접 순위로 폴백.
    _tgt = request.target_position if request.target_position in (1, 2, 3, 4, 5, 6, 7, 8, 9, 10) else 10
    _cascade = []
    for _p in [_tgt, 5, 3, 10, 1, 2, 7]:
        if _p not in _cascade:
            _cascade.append(_p)

    async def _run():
        core_kws = await _collect_core(client)
        logger.warning(f"[core-min-bid] 대상 {len(core_kws)}개 — {_dev} 순위{_tgt} 입찰가 산출 시작")
        # target_position 우선 캐스케이드로 텍스트별 입찰가 맵
        texts = list({k["text"] for k in core_kws if k.get("text")})
        bidmap = {}
        # estimate API 는 배치 100 이면 WriteTimeout/Server disconnected 빈발 → 15로 축소.
        EST_BATCH = 15
        for pos in _cascade:
            remaining = [t for t in texts if t not in bidmap]
            if not remaining:
                break
            n_fail = 0
            for i in range(0, len(remaining), EST_BATCH):
                batch = remaining[i:i+EST_BATCH]
                try:
                    r = await client.get_avg_position_bids(batch, pos, device=_dev)
                    for e in (r.get("estimate") or []):
                        kt, bd = e.get("keyword"), e.get("bid")
                        # 네이버 응답 키워드와 DB 텍스트의 공백 차이로 조인 실패 → 스킵되던 버그.
                        # 양쪽 다 strip 정규화해서 매칭.
                        kt = (kt or "").strip()
                        if kt and bd and kt not in bidmap:
                            bidmap[kt] = bd
                except Exception as ex:
                    n_fail += 1
                    if n_fail <= 3:
                        logger.warning(f"[core-min-bid] pos={pos} 추정 실패: {str(ex)[:80]}")
                await asyncio.sleep(0.2)
            logger.warning(f"[core-min-bid] pos={pos} 산출 진행 — bidmap {len(bidmap)}/{len(texts)} (배치실패 {n_fail})")
        logger.warning(f"[core-min-bid] 입찰가 산출 완료 — {len(bidmap)}/{len(texts)} 텍스트")
        # 키워드별 bidAmt 설정 (10원 단위, 최소 70). 광고그룹별로 묶어 bulk PUT.
        from collections import defaultdict
        by_gid = defaultdict(list)
        _cap = int(request.bid_cap) if request.bid_cap else None
        for k in core_kws:
            bd = bidmap.get((k.get("text") or "").strip())
            if not bd:
                continue
            nb = max(70, round(int(bd) / 10) * 10)
            if _cap:
                nb = min(nb, _cap)
            by_gid[k["gid"]].append({"nccKeywordId": k["id"], "nccAdgroupId": k["gid"], "bidAmt": nb, "useGroupBidAmt": False})
        done = 0
        for gid, items in by_gid.items():
            for i in range(0, len(items), 100):
                try:
                    await client.update_keywords_bid_bulk(items[i:i+100])
                    done += len(items[i:i+100])
                except Exception as ex:
                    logger.warning(f"[core-min-bid] bulk PUT 실패 gid={gid}: {str(ex)[:80]}")
                await asyncio.sleep(0.1)
        logger.warning(f"[core-min-bid] 완료 — {done}개 키워드 PC 최소노출입찰가 설정")
    background_tasks.add_task(_run)
    return {"success": True, "started": True, "message": f"핵심 키워드 PC 최소노출입찰가 설정 백그라운드 시작 (핵심토큰 {len(core)}종, 로그 확인)"}


@router.post("/keyword-pool/bid/force-by-name")
async def keyword_pool_bid_force_by_name(
    names: List[str],
    bid: int = 70,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """이름으로 키워드 찾아 강제 입찰가 변경 + 검증.

    bulk-update 가 silent-ignore 되는지, 또는 그 키워드가 scope 밖 캠페인에 있는지
    추적용. 모든 캠페인 (auto_ 외 포함) 스캔해서 매칭되는 키워드 다 업데이트.
    """
    from services.naver_ad_service import NaverAdApiClient
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        target_set = {n.strip() for n in names if n and n.strip()}
        if not target_set:
            raise HTTPException(status_code=400, detail="names 비어있음")
        new_bid = max(70, int(bid))

        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]

        campaigns = await client.get_campaigns() or []
        results: List[Dict] = []
        not_found = set(target_set)

        for c in campaigns:
            cname = c.get("name") or ""
            cid_str = c.get("nccCampaignId")
            try:
                groups = await client.get_ad_groups(campaign_id=cid_str) or []
            except Exception:
                continue
            for g in groups:
                gid = g.get("nccAdgroupId")
                try:
                    kws = await client.get_keywords(ad_group_id=gid) or []
                except Exception:
                    continue
                for k in kws:
                    kname = (k.get("keyword") or "").strip()
                    if kname not in target_set:
                        continue
                    not_found.discard(kname)
                    kid = k.get("nccKeywordId")
                    before = {"bidAmt": k.get("bidAmt"), "useGroupBidAmt": k.get("useGroupBidAmt")}
                    try:
                        await client.update_keyword_bid(kid, new_bid)
                        # 검증 — 재조회
                        try:
                            after_kw = await client.get_keyword(kid)
                        except Exception:
                            after_kw = {}
                        after = {"bidAmt": after_kw.get("bidAmt"), "useGroupBidAmt": after_kw.get("useGroupBidAmt")}
                        results.append({
                            "keyword": kname,
                            "keyword_id": kid,
                            "campaign": cname,
                            "ad_group_id": gid,
                            "before": before,
                            "after": after,
                            "changed": before["bidAmt"] != after["bidAmt"],
                        })
                    except Exception as e:
                        results.append({
                            "keyword": kname,
                            "keyword_id": kid,
                            "campaign": cname,
                            "ad_group_id": gid,
                            "before": before,
                            "error": f"{type(e).__name__}: {str(e)[:200]}",
                        })
                await asyncio.sleep(0.05)
            await asyncio.sleep(0.05)

        return {
            "success": True,
            "new_bid": new_bid,
            "matched": len(results),
            "not_found": sorted(not_found),
            "results": results,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/bid/force-by-name 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


@router.post("/keyword-pool/bid/inspect-by-name")
async def keyword_pool_bid_inspect_by_name(
    names: List[str],
    customer_id: Optional[str] = None,
    like: bool = False,
    max_groups: int = 80,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """이름으로 키워드 찾아 **네이버 실시간 입찰가를 읽기만** 함(변경 없음, 검증용).

    core-min-exposure / bulk-update 가 네이버에 적용한 입찰가를 확인할 때 사용.
    DB(registered_keywords) 의 bid_amt 는 core-min-exposure 가 갱신 안 하므로 이 엔드포인트로 직접 조회.
    - like=false: names 와 정확히 일치하는 키워드
    - like=true: names(토큰) 중 하나라도 포함하는 키워드(부분일치) — core-min 과 동일 매칭
    """
    from services.naver_ad_service import NaverAdApiClient
    try:
        account = _resolve_account(user_id, customer_id)
        if not account or not account.get("is_connected"):
            raise HTTPException(status_code=400, detail="광고 계정 미연결")
        target_set = {n.strip() for n in names if n and n.strip()}
        if not target_set:
            raise HTTPException(status_code=400, detail="names 비어있음")

        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]
        client.api_key = account["api_key"]
        client.secret_key = account["secret_key"]
        cid = int(account.get("customer_id"))

        # DB(registered_keywords)에서 대상 키워드의 ad_group_id 를 찾아 그 그룹만 네이버 조회(전체 스캔 회피).
        import sqlite3 as _sq
        from database.registered_keywords_db import get_registered_keywords_db
        reg = get_registered_keywords_db()
        names_list = sorted(target_set)
        base = ("SELECT keyword, ad_group_id, bid_amt FROM registered_keywords "
                "WHERE account_customer_id=? AND removed_at IS NULL AND ad_group_id IS NOT NULL ")
        if like:
            base += "AND (" + " OR ".join(["keyword LIKE ?"] * len(names_list)) + ")"
            params = [cid] + [f"%{n}%" for n in names_list]
        else:
            base += "AND keyword IN (" + ",".join(["?"] * len(names_list)) + ")"
            params = [cid] + names_list
        with _sq.connect(reg.db_path) as conn:
            rows = conn.execute(base, params).fetchall()
        # DB 에 매칭된 키워드 텍스트 집합 + 그 그룹들. 네이버 응답과의 공백차 매칭 위해 strip 키로 정규화.
        db_match = {(r[0] or "").strip(): {"gid": r[1], "db_bid": r[2], "raw": r[0]} for r in rows}
        gids = sorted({v["gid"] for v in db_match.values()})[:max_groups]
        match_texts = set(db_match.keys())

        found: List[Dict] = []
        groups_scanned = 0
        groups_failed = 0
        naver_seen: List[str] = []  # 진단: 스캔한 그룹들이 네이버에서 실제 반환한 키워드 텍스트
        for gid in gids:
            groups_scanned += 1
            try:
                kws = await client.get_keywords(ad_group_id=gid) or []
            except Exception:
                groups_failed += 1
                continue
            for k in kws:
                kname = (k.get("keyword") or "").strip()
                if len(naver_seen) < 40:
                    naver_seen.append(kname)
                if kname not in match_texts:
                    continue
                found.append({
                    "keyword": kname,
                    "ad_group_id": gid,
                    "naver_bid": k.get("bidAmt"),
                    "db_bid": db_match.get(kname, {}).get("db_bid"),
                    "useGroupBidAmt": k.get("useGroupBidAmt"),
                })
            await asyncio.sleep(0.05)
        found_texts = {f["keyword"] for f in found}
        not_found = sorted(match_texts - found_texts) if not like else []
        raised = [f for f in found if (f.get("naver_bid") or 0) > 70]
        return {
            "success": True,
            "matched": len(found),
            "groups_scanned": groups_scanned,
            "groups_failed": groups_failed,
            "db_matched": len(db_match),
            "db_sample": [db_match[k]["raw"] for k in list(db_match)[:20]],
            "naver_seen_sample": naver_seen,
            "raised_above_70": len(raised),
            "still_70_or_below": len(found) - len(raised),
            "not_found": sorted(not_found),
            "keywords": found,
        }
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"keyword-pool/bid/inspect-by-name 실패: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {str(e)[:300]}")


# ============ P4: 소재 템플릿 CRUD ============
from database.ad_templates_db import get_ad_templates_db


class AdTemplateCreate(BaseModel):
    headline_pc: str
    description_pc: str
    display_url: str
    final_url_pc: str
    headline_mobile: Optional[str] = None
    description_mobile: Optional[str] = None
    final_url_mobile: Optional[str] = None
    is_active: bool = True


class AdExtensionCreate(BaseModel):
    kind: str  # PHONE_NUMBER / DESCRIPTION_EXTENSION / SUBLINK ...
    payload: Dict[str, Any]


class ImageExtBackfillRequest(BaseModel):
    image_path: Optional[str] = None  # 직접 지정 (없으면 기존 POWER_LINK_IMAGE 에서 자동 탐색)
    scope: str = "pool"               # pool=auto_ 캠페인만, all=전체
    mode: str = "test_one"            # test_one=1개 그룹 테스트(raw 반환) / backfill=전체 백그라운드
    disable_others: bool = False
    discover_campaign: Optional[str] = None  # 지정 시 해당 이름 포함 캠페인의 전체 그룹을 스캔(이미지 위치 타겟)


@router.get("/keyword-pool/image-library")
async def keyword_pool_image_library(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """DB에서 ad_group_id 샘플 꺼내 → 확장소재 조회 → POWER_LINK_IMAGE imagePath 반환."""
    from services.naver_ad_service import NaverAdApiClient
    from database.keyword_pool_db import get_keyword_pool_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list): return x
        if isinstance(x, dict): return x.get("data") or x.get("list") or []
        return []

    try:
        db = get_keyword_pool_db()
        # DB에서 ad_group_id 샘플 (최대 30개, 다양한 그룹)
        with db._conn() as conn:
            rows = conn.execute(
                "SELECT DISTINCT ad_group_id FROM registered_keywords "
                "WHERE account_customer_id=? AND removed_at IS NULL AND ad_group_id IS NOT NULL LIMIT 30",
                (str(account["customer_id"]),)
            ).fetchall()
        gids = [r[0] for r in rows if r[0]]

        for gid in gids:
            try:
                exts = _as_list(await client.get_ad_extensions(owner_id=gid, owner_type="ADGROUP") or [])
            except Exception:
                continue
            for ex in exts:
                if isinstance(ex, dict) and ex.get("type") == "POWER_LINK_IMAGE":
                    ad = ex.get("adExtension") or {}
                    ip = ad.get("imagePath") if isinstance(ad, dict) else None
                    if ip:
                        return {"success": True, "imagePath": ip, "adgroup_id": gid, "scanned": gids.index(gid)+1}
        return {"success": False, "scanned": len(gids), "hint": "POWER_LINK_IMAGE 없음 — 이미지 미부착 상태"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@router.post("/keyword-pool/image-upload-and-backfill")
async def keyword_pool_image_upload_and_backfill(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """이미지 업로드 → imagePath 획득 → 전 캠페인 파워링크 이미지 백필.

    POST multipart/form-data, file 필드에 이미지(jpg/png).
    1) Naver /ncc/uploads 로 이미지 업로드 → imagePath
    2) 전 광고그룹에 POWER_LINK_IMAGE 확장소재 backfill (백그라운드)
    """
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list): return x
        if isinstance(x, dict): return x.get("data") or x.get("list") or []
        return []

    # 1) 이미지 업로드
    img_bytes = await file.read()
    try:
        up = await client.upload_image(img_bytes, filename=file.filename or "image.jpg")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Naver 이미지 업로드 실패: {e}")

    image_path = (up.get("imagePath") or up.get("image_path") or
                  (up.get("data") or {}).get("imagePath") if isinstance(up, dict) else None)
    if not image_path:
        raise HTTPException(status_code=502, detail=f"imagePath 없음: {up}")

    # 2) 전 캠페인 backfill 백그라운드
    content = {"adExtension": {"imagePath": image_path}}

    async def _run():
        import asyncio, logging
        logger = logging.getLogger("image-backfill")
        try:
            camps = _as_list(await client.get_campaigns() or [])
            done = 0; skip = 0; fail = 0
            for c in camps:
                try:
                    groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                except Exception:
                    continue
                for g in groups:
                    gid = g.get("nccAdgroupId")
                    if not gid:
                        continue
                    try:
                        exts = _as_list(await client.get_ad_extensions(owner_id=gid, owner_type="ADGROUP") or [])
                        if any(isinstance(e, dict) and e.get("type") == "POWER_LINK_IMAGE" for e in exts):
                            skip += 1; continue
                        await client.create_ad_extension(owner_id=gid, kind="POWER_LINK_IMAGE",
                                                         content=content, owner_type="ADGROUP")
                        done += 1
                    except Exception as ex:
                        fail += 1
                        if fail % 50 == 0:
                            logger.warning(f"[img-backfill] fail={fail} err={ex}")
                    await asyncio.sleep(0.05)
            logger.info(f"[img-backfill] 완료 done={done} skip={skip} fail={fail}")
        except Exception as ex:
            logger.error(f"[img-backfill] 오류: {ex}")

    background_tasks.add_task(_run)
    return {
        "success": True,
        "imagePath": image_path,
        "upload_response": up,
        "message": "이미지 업로드 완료 + 전 캠페인 파워링크 이미지 백필 백그라운드 시작"
    }


@router.post("/keyword-pool/extension/image-backfill")
async def keyword_pool_image_ext_backfill(
    request: ImageExtBackfillRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """POWER_LINK_IMAGE(파워링크 이미지) 확장소재를 광고그룹에 일괄 등록.

    - image_path 미지정 → 기존 확장소재에서 POWER_LINK_IMAGE 의 imagePath 자동 탐색(ownerType=ADGROUP)
    - mode=test_one → 첫 그룹 1개에만 생성 + Naver raw 응답 반환 (본문 포맷 검증용)
    - mode=backfill → 전체 그룹 백그라운드 생성 (이미 있으면 skip)
    """
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    # 전부 백그라운드로 — 102 캠페인 동기 스캔이 fly 502 유발하므로 즉시 응답 후 _run 에서 처리.
    # 멱등(이미 POWER_LINK_IMAGE 있으면 skip) → 중간에 끊겨도 재실행하면 이어짐.
    img_path_param = request.image_path
    req_scope = request.scope
    req_mode = request.mode

    # ★ imagePath 인터리브 탐색 — 캠페인→그룹 순회하며 그룹마다 즉시 확장소재 검사, 첫 POWER_LINK_IMAGE 에서 중단.
    #   (owner_id=None 은 네이버 404. 기존 버그는 3,757 그룹 gid 를 다 모은 뒤 검사해 느렸음 → 이제 즉시 검사·조기중단.)
    async def _campaigns_retry():
        # get_campaigns 간헐적 빈 응답/429 → 최대 5회 재시도(creative-backfill 패턴).
        for _att in range(5):
            try:
                cs = _as_list(await client.get_campaigns() or [])
                if cs:
                    return cs
            except Exception:
                pass
            await asyncio.sleep(2.5)
        return []

    image_path = img_path_param
    discovered_from = None
    types_seen: Set[str] = set()
    _raw_ext_dump: List[dict] = []
    # POWER_LINK_IMAGE 생성 필수 부가필드 — discover 스캔에서 기존 확장소재로부터 캡처.
    img_w = img_h = None
    pc_channel_id = mobile_channel_id = None
    if not image_path:
        try:
            all_campaigns = await _campaigns_retry()
            dc = (request.discover_campaign or "").strip()
            if dc:
                cand = [c for c in all_campaigns if dc in (c.get("name") or "")]
                per_cap = 12     # 지정 캠페인도 첫 12그룹만(그룹 수백개 캠페인 타임아웃 방지, 이미지는 grp_0001)
            else:
                cand = all_campaigns
                per_cap = 2      # 미지정이면 캠페인당 첫 2그룹만
            scanned = 0
            img_variants: dict = {}   # imagePath -> {w,h,pc,mo,from} (치수별 후보 수집)
            MIN_DIM = 640             # 네이버 API 파워링크 이미지 생성 최소 규격(640×640)
            _scan_cap = 150 if not dc else 900   # 전수 스캔은 fly 60s 프록시 내로 제한
            qualified = None          # >=640 이미지(생성 가능) 발견 시 저장
            for c in cand:
                try:
                    groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                except Exception:
                    continue
                for g in groups[:per_cap]:
                    gid = g.get("nccAdgroupId")
                    if not gid:
                        continue
                    try:
                        exts = _as_list(await client.get_ad_extensions(owner_id=gid, owner_type="ADGROUP") or [])
                    except Exception:
                        continue
                    scanned += 1
                    for e in exts:
                        if isinstance(e, dict):
                            t = e.get("type")
                            if t:
                                types_seen.add(str(t))
                            if t == "POWER_LINK_IMAGE":
                                ad = e.get("adExtension") or {}
                                ip = ad.get("imagePath") if isinstance(ad, dict) else None
                                if ip and ip not in img_variants:
                                    w = ad.get("imageWidth"); h = ad.get("imageHeight")
                                    img_variants[ip] = {"w": w, "h": h, "from": gid,
                                                        "pc": e.get("pcChannelId"),
                                                        "mo": e.get("mobileChannelId")}
                                    if not _raw_ext_dump:
                                        _raw_ext_dump.append(e)
                                    if (isinstance(w, int) and isinstance(h, int)
                                            and w >= MIN_DIM and h >= MIN_DIM and qualified is None):
                                        qualified = (ip, img_variants[ip])
                    if qualified or scanned >= _scan_cap:
                        break
                if qualified or scanned >= _scan_cap:
                    break
            # 선택: 640+ 우선, 없으면 첫 발견분(생성 시 640 미달로 거부되지만 진단용)
            chosen_ip, chosen = (qualified if qualified else
                                 (next(iter(img_variants.items())) if img_variants else (None, None)))
            if chosen:
                image_path = chosen_ip
                discovered_from = chosen["from"]
                img_w, img_h = chosen["w"], chosen["h"]
                pc_channel_id, mobile_channel_id = chosen["pc"], chosen["mo"]
        except Exception as ex:
            return {"success": False, "step": "discover_failed", "error": f"{type(ex).__name__}: {str(ex)[:250]}"}

    if req_mode == "discover":
        return {"success": bool(image_path), "mode": "discover", "image_path": image_path,
                "discovered_from": discovered_from, "types_seen": sorted(types_seen),
                "scanned": locals().get("scanned", 0),
                "chosen_dims": {"w": img_w, "h": img_h},
                "qualified_640": bool(locals().get("qualified")),
                "variants": [{"w": v["w"], "h": v["h"], "path": p[:48]}
                             for p, v in list(locals().get("img_variants", {}).items())[:20]],
                "raw_ext": _raw_ext_dump[:1]}

    if not image_path:
        return {"success": False, "step": "no_image_path",
                "hint": "계정에 POWER_LINK_IMAGE 확장소재가 없음 — image_path 직접 지정 필요"}

    # POWER_LINK_IMAGE 생성 body — imagePath 만으론 code 1010. imageWidth/Height + 비즈채널 필수.
    _ad_ext = {"imagePath": image_path}
    if img_w is not None:
        _ad_ext["imageWidth"] = img_w
    if img_h is not None:
        _ad_ext["imageHeight"] = img_h
    content = {"adExtension": _ad_ext}
    if pc_channel_id:
        content["pcChannelId"] = pc_channel_id
    if mobile_channel_id:
        content["mobileChannelId"] = mobile_channel_id

    # test_one: 첫 캠페인 첫 그룹 1개에만 동기 생성 + 네이버 응답 반환(검증용, 빠름).
    if req_mode == "test_one":
        all_campaigns = await _campaigns_retry()
        target_camps = ([c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]
                        if req_scope == "pool" else all_campaigns)
        gid = None
        for c in target_camps:
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
            except Exception:
                continue
            for g in groups:
                if g.get("nccAdgroupId"):
                    gid = g["nccAdgroupId"]; break
            if gid:
                break
        if not gid:
            return {"success": False, "step": "no_ad_groups"}
        try:
            res = await client.create_ad_extension(owner_id=gid, kind="POWER_LINK_IMAGE", content=content, owner_type="ADGROUP")
            return {"success": True, "mode": "test_one", "ad_group_id": gid,
                    "image_path": image_path, "discovered_from": discovered_from, "naver_response": res}
        except Exception as e:
            return {"success": False, "mode": "test_one", "ad_group_id": gid,
                    "image_path": image_path, "error": f"{type(e).__name__}: {str(e)[:400]}"}

    async def _run():
        all_campaigns = await _campaigns_retry()
        target_camps = ([c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]
                        if req_scope == "pool" else all_campaigns)
        ad_group_ids: List[str] = []
        for c in target_camps:
            try:
                for g in _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or []):
                    if g.get("nccAdgroupId"):
                        ad_group_ids.append(g["nccAdgroupId"])
            except Exception:
                pass
            await asyncio.sleep(0.05)
        logger.warning(f"[img-backfill] 대상 그룹 {len(ad_group_ids)} (imagePath={str(image_path)[:60]} from {discovered_from})")
        created = skipped = failed = 0
        for gid in ad_group_ids:
            try:
                exts = _as_list(await client.get_ad_extensions(owner_id=gid, owner_type="ADGROUP") or [])
                if any(isinstance(e, dict) and e.get("type") == "POWER_LINK_IMAGE" for e in exts):
                    skipped += 1
                else:
                    await client.create_ad_extension(owner_id=gid, kind="POWER_LINK_IMAGE", content=content, owner_type="ADGROUP")
                    created += 1
            except Exception as e:
                failed += 1
                logger.warning(f"[img-backfill] ag={gid} 실패: {type(e).__name__}: {str(e)[:120]}")
            done = created + skipped + failed
            if done % 200 == 0:
                logger.warning(f"[img-backfill] 진행 {done}/{len(ad_group_ids)} — 생성{created} skip{skipped} 실패{failed}")
            await asyncio.sleep(0.12)
        logger.warning(f"[img-backfill] 완료 — 생성 {created} / skip {skipped} / 실패 {failed} / 총 {len(ad_group_ids)}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "scope": req_scope, "image_path": image_path,
            "discovered_from": discovered_from,
            "message": "백그라운드 시작 — POWER_LINK_IMAGE 부착(이미 있으면 skip). imagePath 1콜 탐색 완료."}


class CampaignBudgetBulkRequest(BaseModel):
    daily_budget: int = Field(..., ge=70, le=100000000, description="일 예산(원)")
    scope: str = Field("all", description="'all' 전체 캠페인, 'pool' auto_ 캠페인만")
    name_prefix: Optional[str] = Field(None, description="캠페인명 접두어 필터(예 '의료대출') — 지정 시 그 접두어로 시작하는 캠페인만 대상")
    name_contains: Optional[str] = Field(None, description="캠페인명 포함 필터 — 지정 시 그 문자열을 포함하는 캠페인만 대상")
    campaign_ids: Optional[List[str]] = Field(None, description="캠페인 ID 직접 지정 — 이름 규칙이 제각각인 소수 캠페인만 정밀 변경할 때. 지정 시 다른 필터보다 우선")
    dry_run: bool = Field(False)


@router.post("/keyword-pool/campaign/budget-bulk")
async def keyword_pool_campaign_budget_bulk(
    request: CampaignBudgetBulkRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """모든(또는 auto_) 캠페인의 일예산을 일괄 변경. dry_run=true 면 현재 예산 미리보기."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    campaigns = _as_list(await client.get_campaigns() or [])
    # ID 직접 지정 — 이름 규칙이 제각각인 소수 캠페인만 올릴 때(입찰 상향으로 예산이
    # 부족해진 캠페인 등). name_prefix 로 '소잠_' 를 잡으면 125개가 통째로 걸려
    # 예산 안전망이 사라진다.
    if request.campaign_ids:
        want = {c for c in request.campaign_ids if c}
        campaigns = [c for c in campaigns if c.get("nccCampaignId") in want]
    elif request.scope == "pool":
        campaigns = [c for c in campaigns if (c.get("name") or "").startswith("auto_")]
    else:
        # 'all' = 파워링크(WEB_SITE) 키워드 캠페인만 — 파워컨텐츠/플레이스/브랜드검색 제외
        campaigns = [c for c in campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]
    # 이름 필터 (의료대출_* 등 특정 캠페인만 선별 변경)
    if request.name_prefix:
        campaigns = [c for c in campaigns if (c.get("name") or "").startswith(request.name_prefix)]
    if request.name_contains:
        campaigns = [c for c in campaigns if request.name_contains in (c.get("name") or "")]
    if not campaigns:
        return {"success": False, "step": "no_campaigns"}

    new_budget = int(request.daily_budget)
    if request.dry_run:
        return {
            "success": True, "dry_run": True, "scope": request.scope,
            "target_daily_budget": new_budget, "campaigns_total": len(campaigns),
            "campaigns": [
                {"name": c.get("name"), "id": c.get("nccCampaignId"),
                 "current_budget": c.get("dailyBudget"),
                 "useDailyBudget": c.get("useDailyBudget")}
                for c in campaigns
            ][:100],
        }

    async def _run():
        ok = 0
        failed: List[Dict] = []
        for c in campaigns:
            cid_camp = c.get("nccCampaignId")
            if not cid_camp:
                continue
            try:
                await client.update_campaign_budget(cid_camp, new_budget, base=c)
                ok += 1
            except Exception as e:
                failed.append({"campaign": c.get("name"), "id": cid_camp,
                               "error": f"{type(e).__name__}: {str(e)[:120]}"})
            await asyncio.sleep(0.15)
        logger.warning(f"[budget-bulk] 완료 — {ok}/{len(campaigns)} 캠페인 예산={new_budget}원 ({len(failed)} 실패)")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "scope": request.scope,
            "daily_budget": new_budget, "campaigns_total": len(campaigns),
            "message": f"백그라운드 시작 — {len(campaigns)}개 캠페인 일예산 {new_budget}원 적용"}


class CampaignConfigItem(BaseModel):
    campaign_id: str = Field(..., description="nccCampaignId")
    new_name: Optional[str] = Field(None, description="새 캠페인명(30자 한도). None 이면 이름 유지")
    daily_budget: Optional[int] = Field(None, description="일예산(원). None 이면 예산 유지")
    user_lock: Optional[bool] = Field(None, description="False=ON(활성), True=OFF(중지). None 이면 상태 유지")


class CampaignBulkConfigureRequest(BaseModel):
    items: List[CampaignConfigItem] = Field(..., description="캠페인별 개별 설정 리스트")
    dry_run: bool = Field(True)


@router.post("/keyword-pool/campaign/bulk-configure")
async def keyword_pool_campaign_bulk_configure(
    request: CampaignBulkConfigureRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """캠페인별로 이름/일예산/활성상태(userLock)를 개별 설정 — full-body PUT 한 번에.
    지점 캠페인처럼 캠페인마다 다른 이름·예산이 필요할 때. 이름변경은 fields없는 전체 PUT 필수(원본 body echo).
    user_lock=False 로 중지 캠페인 활성. dry_run 으로 계획 미리보기."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    items = request.items or []
    if not items:
        return {"success": False, "step": "no_items"}
    for it in items:
        if it.new_name and len(it.new_name) > 30:
            raise HTTPException(status_code=400, detail=f"캠페인명 30자 초과: {it.new_name}")

    if request.dry_run:
        # 현재 캠페인 목록으로 현행값 대조(미리보기)
        try:
            campaigns = await client.get_campaigns() or []
        except Exception:
            campaigns = []
        by_id = {c.get("nccCampaignId"): c for c in campaigns if isinstance(c, dict)}
        preview = []
        for it in items:
            base = by_id.get(it.campaign_id) or {}
            preview.append({
                "campaign_id": it.campaign_id,
                "old_name": base.get("name"),
                "new_name": it.new_name,
                "old_budget": base.get("dailyBudget"),
                "new_budget": it.daily_budget,
                "old_userLock": base.get("userLock"),
                "new_userLock": it.user_lock,
                "found": bool(base),
            })
        return {"success": True, "dry_run": True, "customer_id": int(account["customer_id"]),
                "count": len(items), "preview": preview}

    async def _run():
        ok = 0
        failed: List[Dict] = []
        for it in items:
            try:
                base = await client.get_campaign(it.campaign_id)
                body = dict(base) if isinstance(base, dict) else {}
                body["nccCampaignId"] = it.campaign_id
                if it.new_name:
                    body["name"] = it.new_name
                if it.daily_budget is not None:
                    body["dailyBudget"] = int(it.daily_budget)
                    body["useDailyBudget"] = True
                if it.user_lock is not None:
                    body["userLock"] = bool(it.user_lock)
                # 이름변경 포함 = fields 없는 전체 PUT(full replace)
                await client._request("PUT", f"/ncc/campaigns/{it.campaign_id}", body)
                ok += 1
            except Exception as e:
                failed.append({"campaign_id": it.campaign_id, "name": it.new_name,
                               "error": f"{type(e).__name__}: {str(e)[:150]}"})
            await asyncio.sleep(0.15)
        logger.warning(f"[campaign-bulk-configure] 완료 — {ok}/{len(items)} 적용 ({len(failed)} 실패)")
        if failed:
            logger.warning(f"[campaign-bulk-configure] 실패 상세: {failed[:10]}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "count": len(items),
            "message": f"백그라운드 시작 — {len(items)}개 캠페인 개별설정(이름/예산/활성) 적용 (로그 [campaign-bulk-configure])"}


class CreativeBackfillRequest(BaseModel):
    scope: str = Field("all", description="'all' 전체 캠페인, 'pool' auto_ 캠페인만")
    template_id: Optional[int] = Field(None, description="특정 템플릿 id 강제(없으면 첫 활성)")
    source_ad_id: Optional[str] = Field(None, description="특정 소재(nccAdId)를 복사 소스로 사용 — 그 소재 내용(헤드라인/URL/심의필)을 소재 없는 그룹에 복제")
    mode: str = Field("backfill", description="test_one | backfill")
    medical_no: Optional[str] = Field(None, description="의료광고 심의필 번호 (예: 한42606). 의료 광고주 필수")


@router.post("/keyword-pool/ads/backfill-creative")
async def keyword_pool_ads_backfill_creative(
    request: CreativeBackfillRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """텍스트 소재(T&D)를 소재 없는 모든 광고그룹에 일괄 등록. 이미 소재 있으면 skip."""
    from services.naver_ad_service import NaverAdApiClient
    from database.ad_templates_db import get_ad_templates_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    from database.naver_ad_db import get_domain_profile as _gdp_med
    _src_medno = ""
    tpl = None
    if request.source_ad_id:
        # ── 특정 소재 복사 모드: nad-...528271364 같은 최신 심의필 소재 내용을 그대로 복제 ──
        try:
            src = await client.get_ad_by_id(request.source_ad_id) or {}
        except Exception as e:
            return {"success": False, "step": "source_ad_fetch_failed",
                    "error": f"{type(e).__name__}: {str(e)[:300]}"}
        adobj = src.get("ad") if isinstance(src, dict) else None
        if not isinstance(adobj, dict):
            return {"success": False, "step": "source_ad_no_content", "raw": str(src)[:300]}
        # MEDICAL_AD(병의원 업종 특화) — type 명시 + ad.basic 통째 복제 모드.
        # 텍스트만 복제하는 TEXT_45 와 달리 이미지·파비콘·심의필·사이트명·태그가 한 묶음.
        _src_type = (src.get("type") or "").strip().upper()
        _is_medical_ad = (_src_type == "MEDICAL_AD" and isinstance(adobj.get("basic"), dict))
        # 일부 소재 타입(브랜드형/RSA)은 'basic' 래퍼 안에 헤드라인/URL이 들어옴 → flatten.
        if isinstance(adobj.get("basic"), dict):
            _basic = adobj["basic"]
            for _k, _v in _basic.items():
                if _k not in adobj or not adobj.get(_k):
                    adobj[_k] = _v
        pc = adobj.get("pc") if isinstance(adobj.get("pc"), dict) else {}
        mob = adobj.get("mobile") if isinstance(adobj.get("mobile"), dict) else {}
        def _url(o, *keys):
            for k in keys:
                v = o.get(k)
                if isinstance(v, dict):
                    v = v.get("url") or v.get("final")
                if isinstance(v, str) and v.strip():
                    return v.strip()
            return ""
        head = (adobj.get("headline") or pc.get("headline") or "").strip()
        desc = (adobj.get("description") or pc.get("description") or "").strip()
        # 반응형(RSA) 소재 — 제목/설명이 배열(headlines/descriptions)에 들어있음. 단일 필드 비면 배열 첫 항목 사용.
        def _first_text(v):
            if isinstance(v, list) and v:
                el = v[0]
                if isinstance(el, str):
                    return el.strip()
                if isinstance(el, dict):
                    return (el.get("headline") or el.get("description") or el.get("text")
                            or el.get("value") or el.get("title") or "").strip()
            return ""
        if not head:
            for _k in ("headlines", "titles", "pcHeadlines", "headlineList"):
                head = _first_text(adobj.get(_k)) or _first_text(pc.get(_k))
                if head:
                    break
        if not desc:
            for _k in ("descriptions", "pcDescriptions", "descriptionList"):
                desc = _first_text(adobj.get(_k)) or _first_text(pc.get(_k))
                if desc:
                    break
        disp = (adobj.get("displayUrl") or adobj.get("pcDisplayUrl") or "").strip()
        final_pc = _url(adobj, "finalUrl", "pcFinalUrl") or _url(pc, "final", "finalUrl")
        final_mo = _url(adobj, "finalMobileUrl", "mobileFinalUrl") or _url(mob, "final", "finalUrl")
        _src_medno = (adobj.get("medicalNo") or src.get("medicalNo")
                      or (pc.get("medicalNo") if isinstance(pc, dict) else "") or "").strip()
        # MEDICAL_AD — basic 구조 안에서 검증(headline/final이 basic.pc.final 등에 있음).
        if _is_medical_ad:
            _b = adobj.get("basic") or {}
            _bpc = _b.get("pc") if isinstance(_b.get("pc"), dict) else {}
            _med_head = (_b.get("headline") or "").strip()
            _med_final = (_bpc.get("final") or _bpc.get("display") or "").strip()
            if not _med_head or not _med_final:
                return {"success": False, "step": "source_medical_ad_incomplete",
                        "headline": _med_head, "final": _med_final,
                        "raw_basic": str(_b)[:1800]}
            tpl = {
                "id": request.source_ad_id,
                "type": "MEDICAL_AD",
                "medical_basic": _b,  # 원본 basic 통째 복제
                "headline_pc": _med_head, "description_pc": (_b.get("description") or "").strip(),
                "final_url_pc": _med_final,
            }
        elif not head or not final_pc:
            return {"success": False, "step": "source_ad_incomplete",
                    "headline": head, "final_url_pc": final_pc, "medical_no": _src_medno,
                    "raw": str(adobj)[:1800]}  # 반응형 구조 진단용 — 더 길게
        else:
            tpl = {
                "id": request.source_ad_id,
                "headline_pc": head, "description_pc": desc,
                "display_url": disp or (final_pc.replace("https://", "").replace("http://", "").split("/")[0]),
                "final_url_pc": final_pc,
                "headline_mobile": (mob.get("headline") or head) if mob else head,
                "description_mobile": (mob.get("description") or desc) if mob else desc,
                "final_url_mobile": final_mo or final_pc,
            }
    else:
        tpl_db = get_ad_templates_db()
        all_tpls = tpl_db.list_templates(user_id, cid) or []
        if request.template_id is not None:
            tpl = next((t for t in all_tpls if int(t.get("id")) == int(request.template_id)), None)
        else:
            tpl = next((t for t in all_tpls if t.get("is_active")), None) or (all_tpls[0] if all_tpls else None)
        if not tpl:
            return {"success": False, "step": "no_template", "hint": "ad_templates 비어있음"}

    # 의료광고 심의필 — 요청값 > 소스 소재 medicalNo > 광고주 프로파일 순.
    _med_no = (request.medical_no or "").strip() or _src_medno or (_gdp_med(user_id, str(cid)) or {}).get("medical_no", "")

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    async def _create_one(gid: str):
        # MEDICAL_AD(병의원 업종 특화) — basic 구조 통째 복제. 텍스트+이미지+심의필+사이트명+태그 1소재.
        if tpl.get("type") == "MEDICAL_AD" and isinstance(tpl.get("medical_basic"), dict):
            _basic = dict(tpl["medical_basic"])  # 얕은 복사 — 광고주별 medicalNo 오버라이드 지원
            if _med_no:
                _basic["medicalNo"] = _med_no
            return await client.create_medical_ad(ad_group_id=gid, basic=_basic)
        return await client.create_ad(
            ad_group_id=gid,
            headline_pc=tpl["headline_pc"], description_pc=tpl["description_pc"],
            display_url=tpl["display_url"], final_url_pc=tpl["final_url_pc"],
            headline_mobile=tpl.get("headline_mobile"),
            description_mobile=tpl.get("description_mobile"),
            final_url_mobile=tpl.get("final_url_mobile"),
            medical_no=(_med_no or None),  # 의료광고 심의필 번호 (요청 or 프로파일)
        )

    def _target_campaigns(all_campaigns):
        if request.scope == "all":
            # 파워링크(WEB_SITE) 키워드 캠페인만 — 파워컨텐츠/플레이스/브랜드검색 제외
            return [c for c in all_campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]
        return [c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]

    # ── test_one: 첫 캠페인의 첫 그룹에만 — 전체 순회 없이 빠르게 검증 ──
    if request.mode == "test_one":
        all_campaigns = _as_list(await client.get_campaigns() or [])
        gid = None
        for c in _target_campaigns(all_campaigns):
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
            except Exception:
                continue
            for g in groups:
                if g.get("nccAdgroupId"):
                    gid = g["nccAdgroupId"]
                    break
            if gid:
                break
        if not gid:
            return {"success": False, "step": "no_ad_groups"}
        try:
            res = await _create_one(gid)
            return {"success": True, "mode": "test_one", "ad_group_id": gid,
                    "template_id": tpl.get("id"), "medical_no": (_med_no or None),
                    "naver_response": res}
        except Exception as e:
            return {"success": False, "mode": "test_one", "ad_group_id": gid,
                    "error": f"{type(e).__name__}: {str(e)[:400]}"}

    # ── backfill: 캠페인 순회 + 소재 생성 전부 백그라운드 실행 ──
    # 소잠 등 캠페인 多(69) 광고주는 동기 순회가 fly 60s 프록시를 초과해 500/타임아웃.
    # 그룹 수집(get_campaigns + 캠페인별 get_ad_groups)을 _run() 안으로 이전 → 즉시 응답.
    async def _run():
        # 시작 즉시 로그 — 태스크 기동 확인용(cron event-loop 점유로 굶거나 죽었는지 진단).
        logger.warning(f"[creative-backfill] 시작 — scope={request.scope} tpl={tpl.get('id')} medno={_med_no or '-'}")
        try:
            # get_campaigns — cron 점유/429 로 실패 가능 → 재시도(이전엔 무방비라 태스크 silent death).
            all_campaigns = []
            for _att in range(5):
                try:
                    all_campaigns = _as_list(await client.get_campaigns() or [])
                    if all_campaigns:
                        break
                except Exception as _e:
                    logger.warning(f"[creative-backfill] get_campaigns 재시도 {_att+1}/5: {type(_e).__name__}: {str(_e)[:80]}")
                    await asyncio.sleep(3.0)
            target_camps = _target_campaigns(all_campaigns)
            logger.warning(f"[creative-backfill] 캠페인 {len(all_campaigns)} → 대상(파워링크) {len(target_camps)}")
            if not target_camps:
                logger.warning("[creative-backfill] 대상 캠페인 0 — 종료")
                return
            ad_group_ids: List[str] = []
            _camp_fail = 0
            for c in target_camps:
                # 캠페인별 그룹 조회 — 429/타임아웃 시 최대 3회 재시도(불완전 롤아웃 방지).
                ok = False
                for _attempt in range(3):
                    try:
                        groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                        for g in groups:
                            gid = g.get("nccAdgroupId")
                            if gid:
                                ad_group_ids.append(gid)
                        ok = True
                        break
                    except Exception:
                        await asyncio.sleep(1.5)
                if not ok:
                    _camp_fail += 1
                await asyncio.sleep(0.08)
            logger.warning(
                f"[creative-backfill] 그룹 수집 — 캠페인 {len(target_camps)}(실패 {_camp_fail}) → 그룹 {len(ad_group_ids)}"
            )
            if not ad_group_ids:
                logger.warning("[creative-backfill] 대상 그룹 0 — 종료")
                return
            # get_ads 확인 생략하고 바로 생성 — 동일 내용 소재가 이미 있으면 네이버가 3822
            # ("same content exists") 로 거부 → '기존보유'로 처리. 더 빠르고 완전한 롤아웃.
            created = existed = failed = 0
            for i, gid in enumerate(ad_group_ids):
                try:
                    await _create_one(gid)
                    created += 1
                except Exception as e:
                    msg = str(e)
                    if "3822" in msg or "same content" in msg.lower():
                        existed += 1  # 동일 승인 소재 이미 부착됨 → 정상
                    else:
                        failed += 1
                        logger.warning(f"[creative-backfill] ag={gid} 실패: {type(e).__name__}: {msg[:120]}")
                # 진행 로그 — 500개마다(긴 롤아웃 관측).
                if (i + 1) % 500 == 0:
                    logger.warning(f"[creative-backfill] 진행 {i+1}/{len(ad_group_ids)} — 생성 {created} 기존보유 {existed} 실패 {failed}")
                await asyncio.sleep(0.12)
            logger.warning(
                f"[creative-backfill] 완료 — 생성 {created} / 기존보유 {existed} / 실패 {failed} / "
                f"총 {len(ad_group_ids)} (캠페인조회실패 {_camp_fail})"
            )
        except Exception as e:
            import traceback as _tb
            logger.error(f"[creative-backfill] 태스크 예외 — {type(e).__name__}: {str(e)[:200]}\n{_tb.format_exc()[:1000]}")

    background_tasks.add_task(_run)
    return {"success": True, "mode": "backfill", "started": True, "template_id": tpl.get("id"),
            "medical_no": (_med_no or None),
            "message": "백그라운드 시작 — 캠페인 순회 + 소재 부착 진행(즉시 응답). "
                       "중복은 기존보유 처리. 결과는 fly logs 의 [creative-backfill] 라인에서 확인."}


class RsaBackfillRequest(BaseModel):
    source_ad_id: str = Field(..., description="복제 소스 RSA_AD 소재(nccAdId) — 이 소재의 assets(헤드라인/설명)+URL을 소재 없는 그룹에 복제")
    scope: str = Field("all", description="'all' 전체 파워링크(WEB_SITE) 캠페인 / 'pool' auto_ 캠페인만")
    mode: str = Field("discover", description="discover(추출만) | test_one(첫 그룹 1개 생성+raw) | backfill(전체 백그라운드)")


@router.post("/keyword-pool/ads/backfill-rsa")
async def keyword_pool_ads_backfill_rsa(
    request: RsaBackfillRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """반응형 소재(RSA_AD)를 소스 소재 통째 복제로 소재 없는 모든 광고그룹에 부착.

    - create_ad(TEXT_45)는 헤드라인 15/설명 45자로 잘라 RSA 재현 불가 → 소스 assets 배열을 그대로 복제.
    - mode=discover: 소스에서 뽑은 assets/URL 요약만 반환(생성 없음).
    - mode=test_one: 첫 (소재 없는) 그룹 1개에 생성 + 네이버 raw 응답 반환(payload 계약 검증).
    - mode=backfill: WEB_SITE 캠페인 전체 순회, 이미 소재 있는 그룹은 skip, 나머지에 RSA 생성(백그라운드).
    """
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    # ── 1) 소스 RSA 소재 조회 + assets/URL 추출 ──
    try:
        src = await client.get_ad_by_id(request.source_ad_id) or {}
    except Exception as e:
        return {"success": False, "step": "source_fetch_failed", "error": f"{type(e).__name__}: {str(e)[:300]}"}
    src_type = (src.get("type") or "").strip().upper()
    if src_type != "RSA_AD":
        return {"success": False, "step": "source_not_rsa", "type": src_type,
                "hint": "이 엔드포인트는 RSA_AD 전용. TEXT_45 는 /backfill-creative 사용."}
    src_assets = src.get("assets") or (src.get("ad") or {}).get("assets") or []
    adobj = src.get("ad") if isinstance(src.get("ad"), dict) else {}
    pc_src = adobj.get("pc") if isinstance(adobj.get("pc"), dict) else {}
    mo_src = adobj.get("mobile") if isinstance(adobj.get("mobile"), dict) else {}

    def _clean_url_block(o):
        # 생성 시 필요한 URL 필드만 — final/display (punyCode 등 메타 제외)
        out = {}
        if isinstance(o, dict):
            if o.get("final"):
                out["final"] = o["final"]
            if o.get("display"):
                out["display"] = o["display"]
        return out

    pc = _clean_url_block(pc_src)
    mobile = _clean_url_block(mo_src) or pc

    # assets 배열 → 생성용 최소 필드로 재구성(nccAssetId/링크ID/시각 메타 제거, pin 보존)
    create_assets: List[dict] = []
    for a in src_assets:
        if not isinstance(a, dict):
            continue
        ad_data = a.get("assetData") or {}
        text_v = ad_data.get("text") if isinstance(ad_data, dict) else None
        link_type = a.get("linkType")
        if not text_v or link_type not in ("HEADLINE", "DESCRIPTION"):
            continue
        item = {"assetType": a.get("assetType") or "TEXT",
                "assetData": {"text": text_v},
                "linkType": link_type}
        if a.get("pin"):
            item["pin"] = str(a["pin"])
        create_assets.append(item)

    n_head = sum(1 for a in create_assets if a["linkType"] == "HEADLINE")
    n_desc = sum(1 for a in create_assets if a["linkType"] == "DESCRIPTION")
    if not (create_assets and n_head and n_desc and pc.get("final")):
        return {"success": False, "step": "source_incomplete",
                "headlines": n_head, "descriptions": n_desc, "pc": pc}

    summary = {"headlines": n_head, "descriptions": n_desc, "pc": pc, "mobile": mobile,
               "pinned": [a for a in create_assets if a.get("pin")]}

    if request.mode == "discover":
        return {"success": True, "mode": "discover", "source_ad_id": request.source_ad_id,
                "assets_total": len(create_assets), **summary,
                "assets_preview": [a["assetData"]["text"][:24] for a in create_assets[:6]]}

    async def _create_one(gid: str):
        return await client.create_rsa_ad(gid, assets=create_assets, pc=pc, mobile=mobile)

    def _target_campaigns(all_campaigns):
        if request.scope == "all":
            return [c for c in all_campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]
        return [c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]

    async def _campaigns_retry():
        for _att in range(5):
            try:
                cs = _as_list(await client.get_campaigns() or [])
                if cs:
                    return cs
            except Exception:
                pass
            await asyncio.sleep(2.5)
        return []

    # ── test_one: 소재 없는 첫 그룹 1개에만 생성 + raw 반환(계약 검증) ──
    if request.mode == "test_one":
        all_campaigns = await _campaigns_retry()
        gid = None
        for c in _target_campaigns(all_campaigns):
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
            except Exception:
                continue
            for g in groups:
                cand = g.get("nccAdgroupId")
                if not cand:
                    continue
                try:
                    existing = _as_list(await client.get_ads(ad_group_id=cand) or [])
                except Exception:
                    existing = []
                if not existing:            # 소재 없는 그룹만 대상(중복 방지)
                    gid = cand
                    break
            if gid:
                break
        if not gid:
            return {"success": False, "step": "no_empty_ad_group",
                    "hint": "모든 그룹에 이미 소재 있음 or 그룹 없음"}
        try:
            res = await _create_one(gid)
            return {"success": True, "mode": "test_one", "ad_group_id": gid,
                    "assets_total": len(create_assets), "naver_response": res}
        except Exception as e:
            return {"success": False, "mode": "test_one", "ad_group_id": gid,
                    "assets_total": len(create_assets),
                    "error": f"{type(e).__name__}: {str(e)[:500]}"}

    # ── backfill: 전체 백그라운드 ──
    async def _run():
        logger.warning(f"[rsa-backfill] 시작 — scope={request.scope} src={request.source_ad_id} "
                       f"assets={len(create_assets)}(H{n_head}/D{n_desc})")
        all_campaigns = await _campaigns_retry()
        target_camps = _target_campaigns(all_campaigns)
        logger.warning(f"[rsa-backfill] 캠페인 {len(all_campaigns)} → 대상(파워링크) {len(target_camps)}")
        if not target_camps:
            logger.warning("[rsa-backfill] 대상 캠페인 0 — 종료")
            return
        ad_group_ids: List[str] = []
        for c in target_camps:
            for _attempt in range(3):
                try:
                    groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                    for g in groups:
                        if g.get("nccAdgroupId"):
                            ad_group_ids.append(g["nccAdgroupId"])
                    break
                except Exception:
                    await asyncio.sleep(1.5)
            await asyncio.sleep(0.08)
        logger.warning(f"[rsa-backfill] 대상 그룹 {len(ad_group_ids)}")
        created = existed = failed = 0
        for i, gid in enumerate(ad_group_ids):
            try:
                # 이미 소재 있으면 skip(멱등 — 중간에 끊겨도 재실행하면 이어짐)
                existing = _as_list(await client.get_ads(ad_group_id=gid) or [])
                if existing:
                    existed += 1
                else:
                    await _create_one(gid)
                    created += 1
            except Exception as e:
                msg = str(e)
                if "3822" in msg or "same content" in msg.lower():
                    existed += 1
                else:
                    failed += 1
                    logger.warning(f"[rsa-backfill] ag={gid} 실패: {type(e).__name__}: {msg[:120]}")
            if (i + 1) % 300 == 0:
                logger.warning(f"[rsa-backfill] 진행 {i+1}/{len(ad_group_ids)} — 생성{created} 기존{existed} 실패{failed}")
            await asyncio.sleep(0.12)
        logger.warning(f"[rsa-backfill] 완료 — 생성 {created} / 기존 {existed} / 실패 {failed} / 총 {len(ad_group_ids)}")

    background_tasks.add_task(_run)
    return {"success": True, "mode": "backfill", "started": True,
            "source_ad_id": request.source_ad_id, "assets_total": len(create_assets),
            "message": "백그라운드 시작 — RSA 소재 전 그룹 부착(이미 소재 있으면 skip). 로그 [rsa-backfill]."}


class CreativePurgeRequest(BaseModel):
    customer_id: Optional[str] = None
    # 삭제 조건 — 모두 AND 매칭. type 필수, 나머지는 옵션.
    type: str = Field(..., description="삭제 대상 광고 type (예: TEXT_45)")
    headline_equals: Optional[str] = Field(None, description="ad.headline(또는 ad.pc.headline) 완전 일치")
    headline_contains: Optional[str] = Field(None, description="ad.headline 부분 일치")
    dry_run: bool = Field(True, description="True: 매칭만 카운트하고 실제 삭제 안함")
    scope: str = Field("all", description="all=파워링크 전체 / auto=auto_ prefix")


@router.post("/keyword-pool/ads/purge")
async def keyword_pool_ads_purge(
    request: CreativePurgeRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """type+headline 매칭 광고소재 일괄 삭제 (백그라운드).

    예) MEDICAL_AD로 재백필한 뒤 잘못 만든 TEXT_45 사본 정리:
       {"type":"TEXT_45","headline_equals":"강남,피부질환진료,소잠한의원","dry_run":false}
    dry_run=true 면 매칭 카운트만 로그로 출력."""
    from services.naver_ad_service import NaverAdApiClient
    cid = request.customer_id or customer_id
    account = _resolve_account(user_id, cid)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    def _target_campaigns(all_campaigns):
        if request.scope == "all":
            return [c for c in all_campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]
        return [c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]

    target_type = (request.type or "").strip().upper()
    he = (request.headline_equals or "").strip()
    hc = (request.headline_contains or "").strip()

    def _ad_headline(adobj):
        if not isinstance(adobj, dict):
            return ""
        h = adobj.get("headline") or ""
        if not h and isinstance(adobj.get("pc"), dict):
            h = adobj["pc"].get("headline") or ""
        if not h and isinstance(adobj.get("basic"), dict):
            h = adobj["basic"].get("headline") or ""
        return (h or "").strip()

    def _matches(ad_row):
        t = (ad_row.get("type") or "").strip().upper()
        if t != target_type:
            return False
        head = _ad_headline(ad_row.get("ad") or {})
        if he and head != he:
            return False
        if hc and hc not in head:
            return False
        return True

    async def _run():
        logger.warning(
            f"[creative-purge] 시작 — cid={account['customer_id']} type={target_type} "
            f"eq={he!r} contains={hc!r} dry_run={request.dry_run}"
        )
        try:
            camps = []
            for _att in range(5):
                try:
                    camps = _as_list(await client.get_campaigns() or [])
                    if camps:
                        break
                except Exception as e:
                    logger.warning(f"[creative-purge] get_campaigns 재시도 {_att+1}/5: {str(e)[:80]}")
                    await asyncio.sleep(3.0)
            targets = _target_campaigns(camps)
            logger.warning(f"[creative-purge] 캠페인 {len(camps)} → 대상 {len(targets)}")
            gids: List[str] = []
            for c in targets:
                for _att in range(3):
                    try:
                        gs = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                        for g in gs:
                            if g.get("nccAdgroupId"):
                                gids.append(g["nccAdgroupId"])
                        break
                    except Exception:
                        await asyncio.sleep(1.5)
                await asyncio.sleep(0.08)
            logger.warning(f"[creative-purge] 그룹 {len(gids)}")

            matched = deleted = failed = skipped = 0
            for i, gid in enumerate(gids):
                try:
                    ads = _as_list(await client.get_ads(ad_group_id=gid) or [])
                except Exception:
                    await asyncio.sleep(0.2)
                    continue
                victims = [a for a in ads if isinstance(a, dict) and _matches(a)]
                # 안전장치: 그룹에 광고가 1개뿐인데 그게 victim 이면 삭제 보류(빈 그룹 방지).
                if victims and len(ads) <= len(victims):
                    skipped += len(victims)
                    if i % 100 == 0:
                        logger.warning(f"[creative-purge] {i+1}/{len(gids)} skip(only_ad) gid={gid}")
                    await asyncio.sleep(0.08)
                    continue
                for v in victims:
                    matched += 1
                    if request.dry_run:
                        continue
                    aid = v.get("nccAdId")
                    if not aid:
                        continue
                    try:
                        await client.delete_ad(aid)
                        deleted += 1
                    except Exception as e:
                        failed += 1
                        if failed <= 5:
                            logger.warning(f"[creative-purge] delete fail {aid}: {str(e)[:120]}")
                    await asyncio.sleep(0.08)
                if (i + 1) % 200 == 0:
                    logger.warning(
                        f"[creative-purge] {i+1}/{len(gids)} matched={matched} deleted={deleted} failed={failed} skipped={skipped}"
                    )
                await asyncio.sleep(0.03)
            logger.warning(
                f"[creative-purge] 완료 — matched={matched} deleted={deleted} failed={failed} skipped(only_ad)={skipped} dry_run={request.dry_run}"
            )
        except Exception as e:
            logger.exception(f"[creative-purge] 예외: {e}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "dry_run": request.dry_run,
            "type": target_type, "headline_equals": he, "headline_contains": hc,
            "message": "백그라운드 시작. fly logs 의 [creative-purge] 라인에서 확인."}


@router.get("/keyword-pool/diagnostics/find-ad-by-headline")
async def keyword_pool_find_ad_by_headline(
    customer_id: Optional[str] = None,
    headline: str = Query(..., description="찾을 헤드라인(부분일치)"),
    max_matches: int = Query(3, ge=1, le=50),
    scope: str = Query("all", description="all=WEB_SITE 전체 / auto=auto_"),
    user_id: int = Depends(get_user_id_with_fallback),
):
    """헤드라인(부분일치)으로 소재의 캠페인/그룹/nccAdId + 그룹 POWER_LINK_IMAGE imagePath 조회(읽기).
    소재 1개가 어느 캠페인/그룹에 있는지 특정 + 복제 소스(nccAdId)·이미지(imagePath) 확보용."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    def _ad_headline(adobj):
        if not isinstance(adobj, dict):
            return ""
        h = adobj.get("headline") or ""
        if not h and isinstance(adobj.get("pc"), dict):
            h = adobj["pc"].get("headline") or ""
        if not h and isinstance(adobj.get("basic"), dict):
            h = adobj["basic"].get("headline") or ""
        return (h or "").strip()

    def _ad_final(adobj):
        if not isinstance(adobj, dict):
            return ""
        for k in ("finalUrl", "displayUrl"):
            v = adobj.get(k)
            if isinstance(v, str) and v.strip():
                return v.strip()
        pc = adobj.get("pc") if isinstance(adobj.get("pc"), dict) else {}
        return (pc.get("final") or pc.get("finalUrl") or "") if isinstance(pc, dict) else ""

    target = (headline or "").strip()
    camps = _as_list(await client.get_campaigns() or [])
    if scope == "auto":
        camps = [c for c in camps if (c.get("name") or "").startswith("auto_")]
    else:
        camps = [c for c in camps if (c.get("campaignTp") or "") == "WEB_SITE"]

    matches: List[Dict] = []
    groups_scanned = 0
    for c in camps:
        try:
            groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
        except Exception:
            continue
        for g in groups:
            gid = g.get("nccAdgroupId")
            if not gid:
                continue
            groups_scanned += 1
            try:
                ads = _as_list(await client.get_ads(ad_group_id=gid) or [])
            except Exception:
                continue
            for a in ads:
                if not isinstance(a, dict):
                    continue
                head = _ad_headline(a.get("ad") or {})
                if target and target in head:
                    image_path = None
                    try:
                        exts = _as_list(await client.get_ad_extensions(owner_id=gid, owner_type="ADGROUP") or [])
                        for e in exts:
                            if isinstance(e, dict) and e.get("type") == "POWER_LINK_IMAGE":
                                ext = e.get("adExtension") or {}
                                image_path = ext.get("imagePath") if isinstance(ext, dict) else None
                                if image_path:
                                    break
                    except Exception:
                        pass
                    matches.append({
                        "campaign_id": c.get("nccCampaignId"),
                        "campaign_name": c.get("name"),
                        "campaign_tp": c.get("campaignTp"),
                        "ad_group_id": gid,
                        "ncc_ad_id": a.get("nccAdId"),
                        "type": a.get("type"),
                        "headline": head,
                        "final_url": _ad_final(a.get("ad") or {}),
                        "inspect_status": a.get("inspectStatus") or a.get("status"),
                        "image_path": image_path,
                    })
                    if len(matches) >= max_matches:
                        return {"success": True, "customer_id": int(account["customer_id"]),
                                "headline_query": target, "matches": matches,
                                "groups_scanned": groups_scanned, "early_exit": True}
            await asyncio.sleep(0.02)
    return {"success": True, "customer_id": int(account["customer_id"]),
            "headline_query": target, "matches": matches,
            "groups_scanned": groups_scanned, "early_exit": False}


class AdOffByHeadlineRequest(BaseModel):
    customer_id: Optional[str] = None
    type: str = Field(..., description="대상 소재 type (예: TEXT_45)")
    headline_equals: Optional[str] = Field(None, description="ad.headline 완전 일치")
    headline_contains: Optional[str] = Field(None, description="ad.headline 부분 일치")
    lock: bool = Field(True, description="True=off(userLock ON) / False=재개(userLock OFF)")
    dry_run: bool = Field(True, description="True: 매칭만 카운트, 실제 토글 안함")
    scope: str = Field("all", description="all=WEB_SITE / auto=auto_")
    test_ad_id: Optional[str] = Field(None, description="지정 시 이 소재 1개만 토글 후 raw 반환(메커니즘 검증)")


@router.post("/keyword-pool/ads/off-by-headline")
async def keyword_pool_ads_off_by_headline(
    request: AdOffByHeadlineRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """type+headline 매칭 소재를 userLock 으로 off(또는 재개). 삭제 아님 — 가역(백그라운드)."""
    from services.naver_ad_service import NaverAdApiClient
    cid = request.customer_id or customer_id
    account = _resolve_account(user_id, cid)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    async def _toggle(aid: str):
        # Naver: PUT /ncc/ads/{adId}?fields=userLock  body {nccAdId, userLock}
        return await client._request(
            "PUT", f"/ncc/ads/{aid}?fields=userLock",
            {"nccAdId": aid, "userLock": bool(request.lock)},
        )

    # ── 단일 소재 검증 모드 — userLock PUT 계약 확인용 ──
    if request.test_ad_id:
        try:
            res = await _toggle(request.test_ad_id)
            return {"success": True, "mode": "test_one", "ad_id": request.test_ad_id,
                    "lock": request.lock, "naver_response": res}
        except Exception as e:
            return {"success": False, "mode": "test_one", "ad_id": request.test_ad_id,
                    "error": f"{type(e).__name__}: {str(e)[:400]}"}

    target_type = (request.type or "").strip().upper()
    he = (request.headline_equals or "").strip()
    hc = (request.headline_contains or "").strip()

    def _ad_headline(adobj):
        if not isinstance(adobj, dict):
            return ""
        h = adobj.get("headline") or ""
        if not h and isinstance(adobj.get("pc"), dict):
            h = adobj["pc"].get("headline") or ""
        if not h and isinstance(adobj.get("basic"), dict):
            h = adobj["basic"].get("headline") or ""
        return (h or "").strip()

    def _matches(ad_row):
        t = (ad_row.get("type") or "").strip().upper()
        if t != target_type:
            return False
        head = _ad_headline(ad_row.get("ad") or {})
        if he and head != he:
            return False
        if hc and hc not in head:
            return False
        return True

    def _target_campaigns(all_campaigns):
        if request.scope == "auto":
            return [c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]
        return [c for c in all_campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]

    async def _run():
        logger.warning(f"[creative-off] 시작 cid={account['customer_id']} type={target_type} "
                       f"eq={he!r} contains={hc!r} lock={request.lock} dry_run={request.dry_run}")
        try:
            camps = []
            for _att in range(5):
                try:
                    camps = _as_list(await client.get_campaigns() or [])
                    if camps:
                        break
                except Exception as e:
                    logger.warning(f"[creative-off] get_campaigns 재시도 {_att+1}/5: {str(e)[:80]}")
                    await asyncio.sleep(3.0)
            targets = _target_campaigns(camps)
            logger.warning(f"[creative-off] 캠페인 {len(camps)} → 대상 {len(targets)}")
            gids: List[str] = []
            for c in targets:
                for _att in range(3):
                    try:
                        gs = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                        for g in gs:
                            if g.get("nccAdgroupId"):
                                gids.append(g["nccAdgroupId"])
                        break
                    except Exception:
                        await asyncio.sleep(1.5)
                await asyncio.sleep(0.08)
            logger.warning(f"[creative-off] 그룹 {len(gids)}")
            matched = toggled = failed = 0
            for i, gid in enumerate(gids):
                try:
                    ads = _as_list(await client.get_ads(ad_group_id=gid) or [])
                except Exception:
                    await asyncio.sleep(0.2)
                    continue
                victims = [a for a in ads if isinstance(a, dict) and _matches(a)]
                for v in victims:
                    matched += 1
                    if request.dry_run:
                        continue
                    aid = v.get("nccAdId")
                    if not aid:
                        continue
                    try:
                        await _toggle(aid)
                        toggled += 1
                    except Exception as e:
                        failed += 1
                        if failed <= 5:
                            logger.warning(f"[creative-off] toggle fail {aid}: {str(e)[:120]}")
                    await asyncio.sleep(0.08)
                if (i + 1) % 200 == 0:
                    logger.warning(f"[creative-off] {i+1}/{len(gids)} matched={matched} toggled={toggled} failed={failed}")
                await asyncio.sleep(0.03)
            logger.warning(f"[creative-off] 완료 matched={matched} toggled={toggled} failed={failed} dry_run={request.dry_run}")
        except Exception as e:
            logger.exception(f"[creative-off] 예외: {e}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "dry_run": request.dry_run, "lock": request.lock,
            "type": target_type, "headline_equals": he, "headline_contains": hc,
            "message": "백그라운드 시작. fly logs 의 [creative-off] 라인에서 확인."}


class ExplicitBackfillRequest(BaseModel):
    customer_id: Optional[str] = None
    headline: str = Field(..., description="헤드라인 (PC=모바일 동일 적용)")
    description: str = Field(..., description="설명 (PC=모바일 동일 적용)")
    final_url: str = Field(..., description="연결 URL")
    display_url: Optional[str] = Field(None, description="표시 URL (없으면 final_url 도메인)")
    scope: str = Field("all", description="all=WEB_SITE / auto=auto_")
    mode: str = Field("backfill", description="test_one | backfill")


@router.post("/keyword-pool/ads/backfill-explicit")
async def keyword_pool_ads_backfill_explicit(
    request: ExplicitBackfillRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """명시적 텍스트 소재(TEXT_45)를 전 그룹에 생성 — PC·모바일 헤드라인/설명 동일 통일.
    불량 소재(source pc/mobile 불일치) 복제 회피용. 이미 동일 소재 있으면 3822=기존보유."""
    from services.naver_ad_service import NaverAdApiClient
    cid = request.customer_id or customer_id
    account = _resolve_account(user_id, cid)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    H = request.headline.strip()
    D = request.description.strip()
    U = request.final_url.strip()
    disp = (request.display_url or "").strip() or U.replace("https://", "").replace("http://", "").split("/")[0]

    async def _create_one(gid: str):
        return await client.create_ad(
            ad_group_id=gid,
            headline_pc=H, description_pc=D, display_url=disp, final_url_pc=U,
            headline_mobile=H, description_mobile=D, final_url_mobile=U,
        )

    def _target_campaigns(all_campaigns):
        if request.scope == "auto":
            return [c for c in all_campaigns if (c.get("name") or "").startswith("auto_")]
        return [c for c in all_campaigns if (c.get("campaignTp") or "") == "WEB_SITE"]

    if request.mode == "test_one":
        all_campaigns = _as_list(await client.get_campaigns() or [])
        gid = None
        for c in _target_campaigns(all_campaigns):
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
            except Exception:
                continue
            for g in groups:
                if g.get("nccAdgroupId"):
                    gid = g["nccAdgroupId"]; break
            if gid:
                break
        if not gid:
            return {"success": False, "step": "no_ad_groups"}
        try:
            res = await _create_one(gid)
            return {"success": True, "mode": "test_one", "ad_group_id": gid, "naver_response": res}
        except Exception as e:
            return {"success": False, "mode": "test_one", "ad_group_id": gid,
                    "error": f"{type(e).__name__}: {str(e)[:400]}"}

    async def _run():
        logger.warning(f"[backfill-explicit] 시작 cid={account['customer_id']} head={H!r} scope={request.scope}")
        try:
            all_campaigns = []
            for _att in range(5):
                try:
                    all_campaigns = _as_list(await client.get_campaigns() or [])
                    if all_campaigns:
                        break
                except Exception as _e:
                    logger.warning(f"[backfill-explicit] get_campaigns 재시도 {_att+1}/5: {str(_e)[:80]}")
                    await asyncio.sleep(3.0)
            target_camps = _target_campaigns(all_campaigns)
            logger.warning(f"[backfill-explicit] 캠페인 {len(all_campaigns)} → 대상 {len(target_camps)}")
            if not target_camps:
                logger.warning("[backfill-explicit] 대상 캠페인 0 — 종료")
                return
            ad_group_ids: List[str] = []
            _camp_fail = 0
            for c in target_camps:
                ok = False
                for _attempt in range(3):
                    try:
                        groups = _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or [])
                        for g in groups:
                            gid = g.get("nccAdgroupId")
                            if gid:
                                ad_group_ids.append(gid)
                        ok = True
                        break
                    except Exception:
                        await asyncio.sleep(1.5)
                if not ok:
                    _camp_fail += 1
                await asyncio.sleep(0.08)
            logger.warning(f"[backfill-explicit] 그룹 수집 — 캠페인 {len(target_camps)}(실패 {_camp_fail}) → 그룹 {len(ad_group_ids)}")
            if not ad_group_ids:
                logger.warning("[backfill-explicit] 대상 그룹 0 — 종료")
                return
            created = existed = failed = 0
            for i, gid in enumerate(ad_group_ids):
                try:
                    await _create_one(gid)
                    created += 1
                except Exception as e:
                    msg = str(e)
                    if "3822" in msg or "same content" in msg.lower():
                        existed += 1
                    else:
                        failed += 1
                        if failed <= 10:
                            logger.warning(f"[backfill-explicit] ag={gid} 실패: {type(e).__name__}: {msg[:120]}")
                if (i + 1) % 500 == 0:
                    logger.warning(f"[backfill-explicit] 진행 {i+1}/{len(ad_group_ids)} — 생성 {created} 기존보유 {existed} 실패 {failed}")
                await asyncio.sleep(0.12)
            logger.warning(f"[backfill-explicit] 완료 — 생성 {created} / 기존보유 {existed} / 실패 {failed} / 총 {len(ad_group_ids)} (캠페인조회실패 {_camp_fail})")
        except Exception as e:
            import traceback as _tb
            logger.error(f"[backfill-explicit] 태스크 예외 — {type(e).__name__}: {str(e)[:200]}\n{_tb.format_exc()[:1000]}")

    background_tasks.add_task(_run)
    return {"success": True, "mode": "backfill", "started": True, "headline": H,
            "message": "백그라운드 시작 — 명시적 소재 전 그룹 부착. fly logs 의 [backfill-explicit] 라인에서 확인."}


class CampaignRenameKoreanRequest(BaseModel):
    customer_id: Optional[str] = None
    prefix: str = Field("리베리", description="캠페인명 접두어")
    top_n: int = Field(3, ge=1, le=6, description="이름에 넣을 대표 키워드 수")
    sample_groups: int = Field(2, ge=1, le=10, description="캠페인당 샘플링할 광고그룹 수(키워드 대표성 vs API 콜)")
    scope: str = Field("all", description="all=WEB_SITE 전체 / auto=auto_ 이름만 / unnamed=prefix 로 시작 안하는 것만")
    skip_already: bool = Field(True, description="이미 prefix 로 시작하는(=한글화 완료) 캠페인은 건너뜀")
    max_len: int = Field(30, description="캠페인명 최대 길이 가드(네이버 한도 30자)")
    max_campaigns: int = Field(2000, description="안전 상한")
    body_mode: str = Field("fullreplace", description="fullreplace=fields없는 전체 PUT(이름변경 정답) / fields_name=예전 잘못된 경로(실패)")
    test_campaign_id: Optional[str] = Field(None, description="지정 시 이 캠페인 1개만 즉시 개명 후 raw 반환(계약 검증, 샘플링 생략)")
    name_override: Optional[str] = Field(None, description="test 모드에서 강제할 새 이름")
    dry_run: bool = Field(True, description="true: rename 계획 미리보기(동기), false: 실제 rename(백그라운드)")


@router.post("/keyword-pool/campaigns/rename-korean")
async def keyword_pool_campaigns_rename_korean(
    request: CampaignRenameKoreanRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """계정의 캠페인을 **안에 든 대표 키워드로 한글 개명** — 'auto_001_178658_0' 같은 무의미 이름을
    '리베리_여드름_모공_흉터_#178658' 처럼 바꿔 어떤 키워드 버킷인지 한눈에.
    캠페인당 광고그룹 sample_groups 개를 샘플링해 head 키워드(짧을수록 대표) top_n 을 뽑음.
    cid tail 로 유니크 보장(네이버 캠페인명 중복 불가). 이미 개명된 건 skip → 재실행 안전.
    dry_run=true: rename 계획만 동기 반환(승인용). false: 백그라운드 rename(로그 [campaign-rename])."""
    from services.naver_ad_service import NaverAdApiClient
    cid_arg = request.customer_id or customer_id
    account = _resolve_account(user_id, cid_arg)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    prefix = (request.prefix or "리베리").strip()
    top_n = int(request.top_n)
    max_len = max(20, int(request.max_len))

    def _rep_keywords(texts):
        # head term 우선: 글자수 짧고 토큰 적은 것 = 대표. 서로 포함관계인 건 하나만.
        uniq = sorted({t.strip() for t in texts if t and t.strip()},
                      key=lambda s: (len(s.replace(" ", "")), len(s.split())))
        picks: List[str] = []
        for t in uniq:
            tn = t.replace(" ", "")
            if any(tn in p.replace(" ", "") or p.replace(" ", "") in tn for p in picks):
                continue
            picks.append(t)
            if len(picks) >= top_n:
                break
        return picks, len(uniq)

    def _uniq_tag(camp_id: str) -> str:
        digits = re.findall(r"\d{4,}", camp_id or "")
        if digits:
            return max(digits, key=len)[-6:]
        return (camp_id or "")[-6:]

    def _build_name(reps, camp_id):
        # 네이버 한도 30자. 태그는 언더스코어+숫자만(#/[] 금지, 유니크·추적용).
        # core 는 예산 안에서 키워드를 통째로 그리디 패킹(중간 잘림 방지).
        tag = f"_{_uniq_tag(camp_id)}"
        avail = max_len - len(prefix) - 1 - len(tag)  # '리베리' + '_' + core + tag
        picks: List[str] = []
        used = 0
        for r in reps:
            add = len(r) + (1 if picks else 0)  # 키워드 사이 언더스코어
            if used + add <= avail:
                picks.append(r)
                used += add
        core = "_".join(picks)
        if not core:  # 첫 키워드가 avail 보다 길면 하나만 잘라서라도 표기
            core = (reps[0][:max(1, avail)] if reps else "빈버킷")
        return f"{prefix}_{core}{tag}"

    async def _do_rename(camp_id: str, new_name: str, mode: str):
        # 네이버: fields 부분수정 화이트리스트=userLock/budget/period (name 불가).
        # 이름 변경은 fields 없는 전체 PUT(full replace)로만 가능 → 원본 body echo + name 교체.
        base = await client.get_campaign(camp_id)
        body = dict(base) if isinstance(base, dict) else {}
        body["nccCampaignId"] = camp_id
        body["name"] = new_name
        if mode == "fields_name":  # (실패 확인용) 예전 잘못된 경로
            return await client.update_campaign(camp_id, body, fields="name")
        # fullreplace (기본): fields 쿼리 없이 전체 PUT
        return await client._request("PUT", f"/ncc/campaigns/{camp_id}", body)

    # ── 단일 테스트 모드: 샘플링(80s) 생략, 한 캠페인만 즉시 개명 후 raw 반환 ──
    if request.test_campaign_id:
        tid = request.test_campaign_id
        nm = (request.name_override or f"{prefix}_테스트개명_{_uniq_tag(tid)}").strip()
        try:
            res = await _do_rename(tid, nm, request.body_mode)
            return {"success": True, "mode": "test_one", "campaign_id": tid, "new": nm,
                    "body_mode": request.body_mode, "naver_response": res}
        except Exception as e:
            return {"success": False, "mode": "test_one", "campaign_id": tid, "new": nm,
                    "body_mode": request.body_mode, "error": f"{type(e).__name__}: {str(e)[:600]}"}

    def _target_campaigns(all_campaigns):
        out = []
        for c in all_campaigns:
            name = c.get("name") or ""
            if request.scope == "auto":
                if not name.startswith("auto_"):
                    continue
            elif request.scope == "unnamed":
                if name.startswith(prefix):
                    continue
            else:  # all
                if (c.get("campaignTp") or "") != "WEB_SITE":
                    continue
            if request.skip_already and name.startswith(prefix):
                continue
            out.append(c)
        return out

    sem = asyncio.Semaphore(6)

    want = int(request.sample_groups)
    scan_cap = max(want + 4, want * 4)  # 빈 그룹 건너뛰고 키워드 있는 그룹을 찾을 때까지 (상한)

    async def _sample(camp):
        cid0 = camp.get("nccCampaignId")
        async with sem:
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=cid0) or [])
            except Exception as e:
                return camp, [], 0, f"adgroups_err:{type(e).__name__}"
            texts: List[str] = []
            filled = 0
            for g in groups[:scan_cap]:
                gid = g.get("nccAdgroupId")
                if not gid:
                    continue
                try:
                    kws = [k.get("keyword") for k in _as_list(await client.get_keywords(ad_group_id=gid) or [])
                           if k.get("keyword")]
                except Exception:
                    kws = []
                if kws:
                    texts.extend(kws)
                    filled += 1
                    if filled >= want:  # 키워드 있는 그룹 want 개 확보하면 중단
                        break
            return camp, texts, len(groups), None

    all_campaigns = _as_list(await client.get_campaigns() or [])
    targets = _target_campaigns(all_campaigns)[: int(request.max_campaigns)]
    if not targets:
        return {"success": True, "customer_id": int(account["customer_id"]),
                "campaigns_total": len(all_campaigns), "targets": 0,
                "message": "대상 캠페인 0 (이미 개명됐거나 scope 미해당)."}

    sampled = await asyncio.gather(*[_sample(c) for c in targets])

    rows = []
    seen = set()
    for camp, texts, ngroups, err in sampled:
        cid0 = camp.get("nccCampaignId")
        old = camp.get("name") or ""
        reps, uniq_kw = _rep_keywords(texts)
        new = _build_name(reps, cid0)
        if new in seen:  # 극히 드문 tag 충돌 방어 (30자 유지: 뒤 2자를 cid 끝 2자로 치환)
            new = (new[:max_len - 2] + cid0[-2:]) if len(new) >= max_len - 2 else f"{new}{cid0[-2:]}"
        seen.add(new)
        rows.append({
            "campaign_id": cid0, "old": old, "new": new,
            "ad_groups": ngroups, "sampled_uniq_kw": uniq_kw,
            "reps": reps, "unchanged": (old == new), "err": err or "",
        })
    rows.sort(key=lambda r: r["old"])
    changed = [r for r in rows if not r["unchanged"] and not r["err"]]

    if request.dry_run:
        return {
            "success": True, "dry_run": True, "customer_id": int(account["customer_id"]),
            "campaigns_total": len(all_campaigns), "targets": len(targets),
            "will_rename": len(changed), "unchanged": sum(1 for r in rows if r["unchanged"]),
            "errors": sum(1 for r in rows if r["err"]),
            "preview": [{"old": r["old"], "new": r["new"], "ad_groups": r["ad_groups"],
                         "reps": r["reps"], "err": r["err"]} for r in rows[:80]],
            "note": f"미리보기 최대 80개 표시(총 {len(rows)}). 확인 후 dry_run=false 로 실제 개명.",
        }

    async def _run():
        logger.warning(f"[campaign-rename] 시작 cid={account['customer_id']} 대상 {len(changed)} "
                       f"scope={request.scope} prefix={prefix!r} body_mode={request.body_mode}")
        ok = fail = 0
        for r in changed:
            cid0, new = r["campaign_id"], r["new"]
            try:
                await _do_rename(cid0, new, request.body_mode)
                ok += 1
                if ok % 20 == 0:
                    logger.warning(f"[campaign-rename] 진행 {ok}/{len(changed)}")
            except Exception as e:
                fail += 1
                if fail <= 10:
                    logger.warning(f"[campaign-rename] 실패 {r['old']}→{new}: {type(e).__name__} {str(e)[:400]}")
            await asyncio.sleep(0.2)
        logger.warning(f"[campaign-rename] 완료 — 개명 {ok} / 실패 {fail} / 대상 {len(changed)}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "dry_run": False,
            "customer_id": int(account["customer_id"]),
            "targets": len(targets), "will_rename": len(changed),
            "message": "백그라운드 개명 시작. fly logs 의 [campaign-rename] 라인에서 확인."}


# ══════════════════════════════════════════════════════════════════════════
#  지역 기반 캠페인 개명 + 지역(REGIONAL_TARGET) 근사 타겟팅
#  ─ 파워링크는 반경(km) 타겟팅이 없음 → 시/도 코드 집합으로 근사.
#    서울지점=서울(9), 경기/인천지점=수도권, 지방지점=해당광역시+인접도.
# ══════════════════════════════════════════════════════════════════════════

# 시/도 코드 (네이버 검색광고 REGIONAL_TARGET location.KR — 공식 FAQ 검증)
_SIDO_CODE = {
    "서울": "9", "부산": "8", "대구": "6", "인천": "11", "광주": "5", "대전": "7",
    "울산": "10", "세종": "17", "경기": "2", "강원": "1", "충북": "16", "충남": "15",
    "전북": "13", "전남": "12", "경북": "4", "경남": "3", "제주": "14",
}
# 시/도 코드 → 권역 표시명 (개명 접두어)
_SIDO_KWON = {
    "9": "서울", "11": "인천", "2": "경기",
    "8": "영남", "6": "영남", "10": "영남", "3": "영남", "4": "영남",
    "7": "충청", "15": "충청", "16": "충청", "17": "충청",
    "5": "호남", "12": "호남", "13": "호남",
    "1": "강원", "14": "제주",
}

# 지점 토큰 → (권역표시, 지점라벨, 근사반경 시도코드집합)
#   서울권=서울시(가장 좁음 ~7km) / 수도권=경기+인천+서울(~15km) / 지방=광역+인접도(~30km)
_KINESS_BRANCH = {
    "강남": ("서울", "강남점", ["9"]),
    "마포": ("서울", "마포점", ["9"]),
    "목동": ("서울", "목동점", ["9"]),
    "반포": ("서울", "반포점", ["9"]),
    "성북": ("서울", "성북점", ["9"]),
    "잠실": ("서울", "잠실점", ["9"]),
    "부천": ("경기", "부천점", ["2", "11", "9"]),
    "분당": ("경기", "분당점", ["2", "11", "9"]),
    "송도": ("인천", "송도점", ["11", "2", "9"]),
    "수원": ("경기", "수원점", ["2", "11", "9"]),
    "수지": ("경기", "수지점", ["2", "11", "9"]),
    "용인수지": ("경기", "수지점", ["2", "11", "9"]),
    "일산": ("경기", "일산점", ["2", "11", "9"]),
    "평촌": ("경기", "평촌점", ["2", "11", "9"]),
    "평택": ("경기", "평택점", ["2", "15"]),      # 경기 남단 → 충남 인접
    "대구": ("영남", "대구점", ["6", "4"]),        # 대구+경북
    "부산": ("영남", "부산점", ["8", "3", "10"]),  # 부산+경남+울산
    "창원": ("영남", "창원점", ["3", "8"]),        # 경남+부산
}

# 도시/구/군 토큰 → 시도코드 (풀·롱테일 캠페인 대표지역 추정용)
_REGION_TOKEN = {
    # 서울
    "강남": "9", "서초": "9", "반포": "9", "방배": "9", "마포": "9", "목동": "9",
    "양천": "9", "잠실": "9", "송파": "9", "성북": "9", "노원": "9", "강북": "9",
    "관악": "9", "성동": "9", "은평": "9", "구로": "9", "동작": "9", "영등포": "9",
    # 경기·인천
    "부천": "2", "분당": "2", "성남": "2", "판교": "2", "수원": "2", "영통": "2",
    "용인": "2", "수지": "2", "동백": "2", "고양": "2", "일산": "2", "파주": "2",
    "운정": "2", "안양": "2", "평촌": "2", "평택": "2", "광명": "2", "이천": "2",
    "안산": "2", "화성": "2", "동탄": "2", "광교": "2", "과천": "2", "김포": "2",
    "남양주": "2", "의정부": "2", "인천": "11", "송도": "11", "청라": "11", "부평": "11",
    # 영남
    "대구": "6", "구미": "4", "경산": "4", "포항": "4", "경주": "4", "안동": "4",
    "부산": "8", "서면": "8", "해운대": "8", "울산": "10", "창원": "3", "김해": "3",
    "양산": "3", "거제": "3", "진주": "3", "율하": "3",
    # 충청
    "대전": "7", "천안": "15", "아산": "15", "논산": "15", "청주": "16", "충주": "16",
    "세종": "17",
    # 호남·제주
    "광주": "5", "전주": "13", "익산": "13", "군산": "13", "목포": "12", "여수": "12",
    "순천": "12", "강진": "12", "제주": "14",
    # 강원
    "원주": "1", "춘천": "1", "강릉": "1", "속초": "1", "동해": "1",
}

# 풀 지역 캠페인명(키네스풀_07_지역_서울 등) 접미 → 권역
_POOL_REGION_SUFFIX = {
    "서울": "서울", "경기인천": "경기", "경기": "경기", "인천": "인천",
    "영남": "영남", "충청": "충청", "호남제주": "호남", "호남": "호남", "강원": "강원",
}


def _kiness_region_uniq_tag(camp_id: str) -> str:
    digits = re.findall(r"\d{4,}", camp_id or "")
    if digits:
        return max(digits, key=len)[-6:]
    return (camp_id or "")[-6:]


def _classify_campaign_region(name: str, sample_keywords: List[str]):
    """캠페인 → (권역표시, 개명label, 지역타겟 시도코드리스트, method).
    1) 지점 캠페인(이름에 지점토큰) → _KINESS_BRANCH 확정 매핑
    2) 풀 지역 캠페인(키네스풀_..지역_XX) → 접미 권역
    3) 그 외(카테고리/롱테일) → 샘플 키워드 지역토큰 최빈 시도 → 권역, 없으면 전국
    """
    nm = name or ""
    nmz = nm.replace(" ", "")
    # 자동생성 롱테일(끝이 _숫자6자리)·풀 카테고리는 지점 매칭에서 제외
    #  (이름 안에 우연히 든 지역 substring 이 지점으로 오인식되는 것 방지)
    is_longtail = bool(re.search(r"_\d{4,}$", nmz))
    is_pool_cat = nmz.startswith("키네스풀") and "지역_" not in nm
    # 1) 지점 캠페인 — 실제 지점 형태만. 긴 토큰 우선(용인수지 > 수지)
    if not is_longtail and not is_pool_cat:
        for tok in sorted(_KINESS_BRANCH.keys(), key=len, reverse=True):
            if tok in nmz:
                kwon, label, codes = _KINESS_BRANCH[tok]
                region = label[:-1] if label.endswith("점") else label
                # 성장판/사춘기 클린 시리즈는 지점 시리즈와 구분되게 label 다르게
                if "성장판" in nmz or "사춘기" in nmz:
                    label = f"{region}_성장"
                return kwon, label, codes, "branch"
    # 2) 풀 지역 캠페인
    if "지역_" in nm:
        suf = nm.split("지역_")[-1].strip().replace(" ", "")
        for key, kwon in _POOL_REGION_SUFFIX.items():
            if suf.startswith(key):
                codes = _kwon_target_codes(kwon)
                return kwon, "통합풀", codes, "pool_region"
    # 2.5) 풀 카테고리 캠페인(02핵심·03성장지연·13건기식 등) → 전국 고정
    if is_pool_cat:
        cat = re.sub(r"^키네스풀_", "", nm).strip()
        return "전국", cat or "통합", [], "pool_cat"
    # 3) 샘플 키워드 지역토큰 최빈
    tally: Dict[str, int] = {}
    for kw in sample_keywords or []:
        kwz = (kw or "").replace(" ", "")
        for tok in sorted(_REGION_TOKEN.keys(), key=len, reverse=True):
            if tok in kwz:
                tally[_REGION_TOKEN[tok]] = tally.get(_REGION_TOKEN[tok], 0) + 1
                break
    if tally:
        top_sido = max(tally.items(), key=lambda x: x[1])[0]
        kwon = _SIDO_KWON.get(top_sido, "전국")
        # 최빈이 과반이 아니면 혼재로 간주 → 전국
        total = sum(tally.values())
        if tally[top_sido] * 2 < total:
            return "전국", "혼합", [], "mixed"
        return kwon, "롱테일", _kwon_target_codes(kwon), "keyword"
    return "전국", "", [], "none"


def _kwon_target_codes(kwon: str) -> List[str]:
    """권역 표시명 → 근사반경 시도코드 집합 (풀/롱테일 캠페인용 기본값)."""
    return {
        "서울": ["9"],
        "경기": ["2", "11", "9"],
        "인천": ["11", "2", "9"],
        "영남": ["8", "6", "3", "10", "4"],
        "충청": ["7", "15", "16", "17"],
        "호남": ["5", "12", "13"],
        "강원": ["1"],
        "제주": ["14"],
        "전국": [],
    }.get(kwon, [])


class RenameByRegionRequest(BaseModel):
    customer_id: Optional[str] = None
    sample_groups: int = Field(1, ge=1, le=5, description="캠페인당 샘플링 광고그룹 수(대표지역 추정용)")
    max_campaigns: int = Field(2000)
    max_len: int = Field(30, description="네이버 캠페인명 한도")
    skip_if_prefixed: bool = Field(True, description="이미 권역_ 로 시작하면 skip(재실행 안전)")
    dry_run: bool = Field(True, description="true=계획 미리보기 / false=실제 개명(백그라운드)")


@router.post("/keyword-pool/campaigns/rename-by-region")
async def keyword_pool_campaigns_rename_by_region(
    request: RenameByRegionRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """캠페인을 **권역_지점** 형식으로 개명 (서울_강남점 / 경기_분당점 / 영남_대구점).
    지점 캠페인은 확정 매핑, 풀/롱테일은 대표 지역 추정. register 라우팅은 campaign_id
    기반이라 개명해도 자동등록 안 깨짐. dry_run=true 로 계획 확인 후 false 실행."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, request.customer_id or customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    _KWON_SET = {"서울", "경기", "인천", "영남", "충청", "호남", "강원", "제주", "전국"}

    def _already(nm: str) -> bool:
        head = (nm or "").split("_", 1)[0]
        return head in _KWON_SET

    max_len = max(20, int(request.max_len))
    sem = asyncio.Semaphore(6)

    async def _sample(camp):
        cid0 = camp.get("nccCampaignId")
        async with sem:
            texts: List[str] = []
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=cid0) or [])
            except Exception:
                return camp, texts, 0
            filled = 0
            for g in groups[: max(request.sample_groups + 4, 8)]:
                gid = g.get("nccAdgroupId")
                if not gid:
                    continue
                try:
                    kws = [k.get("keyword") for k in _as_list(await client.get_keywords(ad_group_id=gid) or [])
                           if k.get("keyword")]
                except Exception:
                    kws = []
                if kws:
                    texts.extend(kws)
                    filled += 1
                    if filled >= request.sample_groups:
                        break
            return camp, texts, len(groups)

    all_campaigns = [c for c in _as_list(await client.get_campaigns() or [])
                     if (c.get("campaignTp") or "") == "WEB_SITE"][: int(request.max_campaigns)]
    targets = [c for c in all_campaigns
               if not (request.skip_if_prefixed and _already(c.get("name") or ""))]
    if not targets:
        return {"success": True, "customer_id": int(account["customer_id"]),
                "campaigns_total": len(all_campaigns), "targets": 0,
                "message": "대상 0 (이미 권역_ 로 개명됐거나 WEB_SITE 없음)."}

    sampled = await asyncio.gather(*[_sample(c) for c in targets])

    rows = []
    seen = set()
    for camp, texts, ngroups in sampled:
        cid0 = camp.get("nccCampaignId")
        old = camp.get("name") or ""
        kwon, label, codes, method = _classify_campaign_region(old, texts)
        tag = f"_{_kiness_region_uniq_tag(cid0)}"
        # core: label 있으면 label, 없으면 기존이름 축약
        core = label or re.sub(r"^(키네스풀?_|[AB]\d\s*|C\d\s*|키네스_)", "", old).strip()
        core = core.replace(" ", "")[: max(1, max_len - len(kwon) - 1)]
        base = (f"{kwon}_{core}" if core else kwon)[:max_len]
        # 깔끔한 이름 우선 — 충돌(중복 지점/혼합)일 때만 캠페인ID 6자리 tag 부착
        new = base
        if new in seen:
            new = f"{base}{tag}"[:max_len]
        if new in seen:
            new = (new[:max_len - 2] + cid0[-2:])
        seen.add(new)
        rows.append({"campaign_id": cid0, "old": old, "new": new, "kwon": kwon,
                     "codes": codes, "method": method, "ad_groups": ngroups,
                     "unchanged": (old == new)})
    rows.sort(key=lambda r: (r["kwon"], r["old"]))
    changed = [r for r in rows if not r["unchanged"]]

    async def _do_rename(camp_id: str, new_name: str):
        base = await client.get_campaign(camp_id)
        body = dict(base) if isinstance(base, dict) else {}
        body["nccCampaignId"] = camp_id
        body["name"] = new_name
        return await client._request("PUT", f"/ncc/campaigns/{camp_id}", body)

    if request.dry_run:
        from collections import Counter
        by_kwon = Counter(r["kwon"] for r in rows)
        return {
            "success": True, "dry_run": True, "customer_id": int(account["customer_id"]),
            "campaigns_total": len(all_campaigns), "targets": len(targets),
            "will_rename": len(changed), "by_kwon": dict(by_kwon),
            "preview": [{"old": r["old"], "new": r["new"], "권역": r["kwon"],
                         "지역코드": r["codes"], "method": r["method"]} for r in rows[:120]],
            "note": f"미리보기 최대 120(총 {len(rows)}). method: branch=확정지점 / pool_region=풀지역 "
                    f"/ keyword=키워드추정 / mixed=다지역혼합(전국) / none=지역미상(전국). "
                    f"확인 후 dry_run=false 로 실제 개명.",
        }

    async def _run():
        logger.warning(f"[rename-region] 시작 cid={account['customer_id']} 대상 {len(changed)}")
        ok = fail = 0
        for r in changed:
            try:
                await _do_rename(r["campaign_id"], r["new"])
                ok += 1
            except Exception as e:
                fail += 1
                if fail <= 10:
                    logger.warning(f"[rename-region] 실패 {r['old']}→{r['new']}: {type(e).__name__} {str(e)[:300]}")
            await asyncio.sleep(0.2)
        logger.warning(f"[rename-region] 완료 — 개명 {ok} / 실패 {fail} / 대상 {len(changed)}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "dry_run": False,
            "customer_id": int(account["customer_id"]), "will_rename": len(changed),
            "message": "백그라운드 개명 시작. fly logs 의 [rename-region] 라인에서 확인."}


@router.get("/keyword-pool/business-channels")
async def keyword_pool_list_business_channels(
    customer_id: Optional[str] = None,
    channel_tp: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """READ-ONLY. 계정 비즈채널 목록 — 지점 플레이스 채널의 **주소**를 뽑는 용도.
    반경(km) 지역타게팅은 중심 주소가 필요한데, 지점 주소가 여기 들어있다.
    channel_tp 로 필터(예: PLACE / SITE / PHONE)."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]
    try:
        chans = await client.list_business_channels() or []
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"비즈채널 조회 실패: {str(e)[:200]}")
    if isinstance(chans, dict):
        chans = chans.get("data") or chans.get("list") or []
    out = []
    for ch in chans:
        if not isinstance(ch, dict):
            continue
        tp = ch.get("channelTp") or ch.get("businessChannelTp")
        if channel_tp and tp != channel_tp:
            continue
        out.append({k: v for k, v in ch.items() if k not in ("customerId",)})
    from collections import Counter
    return {"success": True, "customer_id": int(account["customer_id"]),
            "total": len(out),
            "by_tp": dict(Counter((c.get("channelTp") or c.get("businessChannelTp")) for c in out)),
            "channels": out}


class InspectRegionalTargetRequest(BaseModel):
    customer_id: Optional[str] = None
    campaign_name_contains: Optional[str] = Field(None, description="특정 캠페인명 부분일치만 조회(예: '강남')")
    groups_per_campaign: int = Field(2, description="캠페인당 조회할 광고그룹 수(샘플)")
    max_campaigns: int = Field(80)
    raw: bool = Field(False, description="true=REGIONAL_TARGET 원본 JSON 그대로(코드체계 역공학용)")


@router.post("/keyword-pool/adgroups/inspect-regional-target")
async def keyword_pool_inspect_regional_target(
    request: InspectRegionalTargetRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """READ-ONLY. 광고그룹에 **실제로 걸려있는** REGIONAL_TARGET 을 조회.

    set-regional-target 의 dry_run 은 '계획'만 보여줘서 실제 적용 여부를 알 수 없다.
    본 엔드포인트는 GET /ncc/adgroups/{id}/targets 를 실호출해 현재 상태를 확인한다.
    raw=true 면 원본 JSON 을 그대로 반환 — 읍/면/동 코드나 주소기준 반경(km) 타게팅이
    어떤 스키마로 저장되는지 역공학하는 용도(UI 로 한 그룹 설정 후 읽어보면 됨).
    """
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, request.customer_id or customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    camps = [c for c in _as_list(await client.get_campaigns() or [])
             if (c.get("campaignTp") or "") == "WEB_SITE"]
    if request.campaign_name_contains:
        needle = request.campaign_name_contains.strip()
        camps = [c for c in camps if needle in (c.get("name") or "")]
    camps = camps[: int(request.max_campaigns)]

    sem = asyncio.Semaphore(5)

    async def _one(camp):
        async with sem:
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=camp.get("nccCampaignId")) or [])
            except Exception as e:
                return {"campaign": camp.get("name"), "error": f"adgroups: {str(e)[:120]}"}
            out = []
            for g in groups[: int(request.groups_per_campaign)]:
                gid = g.get("nccAdgroupId")
                if not gid:
                    continue
                try:
                    tmap = await client.get_ad_group_targets(gid) or {}
                except Exception as e:
                    out.append({"adgroup": g.get("name"), "id": gid, "error": str(e)[:120]})
                    continue
                if isinstance(tmap, list):
                    tmap = {t.get("targetTp"): t for t in tmap if isinstance(t, dict)}
                reg = tmap.get("REGIONAL_TARGET") if isinstance(tmap, dict) else None
                loc = ((reg or {}).get("target") or {}).get("location") or {}
                kr = loc.get("KR")
                item = {"adgroup": g.get("name"), "id": gid,
                        "has_regional_target": bool(reg),
                        "kr_codes": kr,
                        "kr_count": len(kr) if isinstance(kr, list) else None,
                        "targetTps": sorted(tmap.keys()) if isinstance(tmap, dict) else None}
                if request.raw:
                    item["raw_regional"] = reg
                out.append(item)
            return {"campaign": camp.get("name"), "campaign_id": camp.get("nccCampaignId"),
                    "groups_checked": len(out), "groups": out}

    results = await asyncio.gather(*[_one(c) for c in camps])
    targeted = sum(1 for r in results for g in (r.get("groups") or []) if g.get("has_regional_target"))
    total_g = sum(len(r.get("groups") or []) for r in results)
    return {"success": True, "customer_id": int(account["customer_id"]),
            "campaigns_checked": len(results),
            "adgroups_checked": total_g,
            "adgroups_with_regional_target": targeted,
            "adgroups_without": total_g - targeted,
            "results": results}


class NaverRawRequest(BaseModel):
    customer_id: Optional[str] = None
    method: str = Field("GET", description="GET|POST|PUT|DELETE")
    path: str = Field(..., description="네이버 API 경로. /ncc/ 로 시작해야 함 (예: /ncc/targets)")
    # 네이버는 키워드 생성(POST /ncc/keywords)·입찰 bulk(PUT /ncc/keywords) 를 **배열 body** 로만 받는다.
    # dict 전용이면 400 "Cannot deserialize ArrayList from Object value" → list 도 허용.
    body: Optional[Union[dict, list]] = Field(None, description="요청 body(JSON). 객체 또는 배열")


@router.post("/keyword-pool/debug/naver-raw")
async def keyword_pool_debug_naver_raw(
    request: NaverRawRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """진단 전용 — 네이버 SearchAd API 를 직접 호출하고 원본 응답/에러를 반환.

    지역타게팅(REGIONAL_TARGET) 생성 경로를 찾기 위해 추가. 광고그룹 PUT 은
    기존 타겟 수정만 되고 신규 생성은 200 OK 로 조용히 무시되기 때문에
    /ncc/targets 등 다른 경로를 재배포 없이 탐색해야 한다.

    경로는 /ncc/ 하위로 제한(계정 자격증명은 서버가 보유, 호출자가 지정 불가).
    """
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, request.customer_id or customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    path = (request.path or "").strip()
    if not path.startswith("/ncc/"):
        raise HTTPException(status_code=400, detail="path 는 /ncc/ 로 시작해야 함")
    method = (request.method or "GET").upper()
    if method not in ("GET", "POST", "PUT", "DELETE"):
        raise HTTPException(status_code=400, detail="method 불가")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]
    try:
        resp = await client._request(method, path, request.body)
        return {"success": True, "method": method, "path": path, "response": resp}
    except Exception as e:
        return {"success": False, "method": method, "path": path,
                "error": f"{type(e).__name__}: {str(e)[:1000]}"}


class DebugRegionalTargetRequest(BaseModel):
    customer_id: Optional[str] = None
    adgroup_id: str = Field(..., description="대상 광고그룹 nccAdgroupId")
    kr_codes: List[str] = Field(default_factory=list, description="시/도·시군구·읍면동 코드 배열. 빈배열=해제")
    fields: str = Field("targetLocation", description="PUT fields 파라미터. 예: 'targetLocation' / 'targetLocation,targetMedia,targetTime'")
    include_all_targets: bool = Field(False, description="true=기존 모든 타겟(MEDIA/PC_MOBILE 등)을 targets 배열에 함께 전송")
    echo_adgroup: bool = Field(False, description="true=광고그룹 원본 body 를 통째로 echo 하고 targets 만 교체(full replace)")
    raw_target: Optional[dict] = Field(None, description="지정 시 target 객체를 이 값으로 그대로 사용(반경 스키마 실험용)")


@router.post("/keyword-pool/adgroups/debug-regional-target")
async def keyword_pool_debug_regional_target(
    request: DebugRegionalTargetRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """진단 전용 — 광고그룹 1개에 REGIONAL_TARGET 을 **동기적으로** 설정하고
    네이버 원본 응답/에러를 그대로 반환. set-regional-target 이 백그라운드라
    실패 원인이 로그에 묻히는 문제 때문에 추가.

    fields / include_all_targets / echo_adgroup / raw_target 조합을 바꿔가며
    어떤 방식이 실제로 먹히는지 실험할 수 있다(재배포 없이).
    """
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, request.customer_id or customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]
    gid = request.adgroup_id

    out: Dict[str, Any] = {"adgroup_id": gid, "fields": request.fields}

    # 1) 기존 타겟 조회
    try:
        tmap = await client.get_ad_group_targets(gid) or {}
    except Exception as e:
        return {"success": False, "step": "get_targets", "error": f"{type(e).__name__}: {str(e)[:400]}"}
    if isinstance(tmap, list):
        tmap = {t.get("targetTp"): t for t in tmap if isinstance(t, dict)}
    out["before_targetTps"] = sorted(tmap.keys()) if isinstance(tmap, dict) else None
    out["before_regional"] = tmap.get("REGIONAL_TARGET")

    # 2) regional 객체 구성
    codes = [str(c) for c in (request.kr_codes or []) if str(c).strip()]
    regional = dict(tmap.get("REGIONAL_TARGET") or {})
    regional["targetTp"] = "REGIONAL_TARGET"
    regional.setdefault("ownerId", gid)
    regional["target"] = request.raw_target if request.raw_target is not None \
        else {"location": {"KR": codes, "OTHERS": []}}

    targets = [regional]
    if request.include_all_targets and isinstance(tmap, dict):
        for tp, obj in tmap.items():
            if tp != "REGIONAL_TARGET" and isinstance(obj, dict):
                targets.append(obj)

    if request.echo_adgroup:
        try:
            base = await client.get_ad_group(gid) if hasattr(client, "get_ad_group") else None
        except Exception:
            base = None
        body = dict(base) if isinstance(base, dict) else {"nccAdgroupId": gid}
        body["nccAdgroupId"] = gid
        body["targets"] = targets
    else:
        body = {"nccAdgroupId": gid, "targets": targets}

    out["request_body"] = body

    # 3) PUT 실행 — 에러를 그대로 반환
    try:
        resp = await client._request("PUT", f"/ncc/adgroups/{gid}?fields={request.fields}", body)
        out["put_response"] = resp
        out["put_ok"] = True
    except Exception as e:
        out["put_ok"] = False
        out["put_error"] = f"{type(e).__name__}: {str(e)[:800]}"

    # 4) 사후 재조회로 실제 반영 확인
    try:
        tmap2 = await client.get_ad_group_targets(gid) or {}
        if isinstance(tmap2, list):
            tmap2 = {t.get("targetTp"): t for t in tmap2 if isinstance(t, dict)}
        out["after_targetTps"] = sorted(tmap2.keys()) if isinstance(tmap2, dict) else None
        out["after_regional"] = tmap2.get("REGIONAL_TARGET")
        out["applied"] = bool(tmap2.get("REGIONAL_TARGET"))
    except Exception as e:
        out["after_error"] = str(e)[:300]
    return {"success": True, **out}


class SetRegionalTargetRequest(BaseModel):
    customer_id: Optional[str] = None
    only_branch: bool = Field(True, description="지점 캠페인(확정 지역)만 타겟 설정(권장). false=풀/롱테일도 추정지역으로")
    campaign_name_contains: Optional[str] = Field(None, description="특정 캠페인명 부분일치만 대상(예: '서울')")
    max_campaigns: int = Field(2000)
    max_groups_per_campaign: int = Field(2000)
    dry_run: bool = Field(True, description="true=대상/코드 미리보기 / false=실제 REGIONAL_TARGET 설정(백그라운드)")


@router.post("/keyword-pool/adgroups/set-regional-target")
async def keyword_pool_set_regional_target(
    request: SetRegionalTargetRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """캠페인의 모든 광고그룹에 **지역(REGIONAL_TARGET) 근사 타겟팅** 설정.
    파워링크는 반경(km) 불가 → 시/도 코드 집합으로 근사(서울지점=서울, 경기=수도권, 지방=광역+인접).
    only_branch=true(기본): 지역이 확정된 지점 캠페인만. dry_run=true 로 대상·코드 확인 후 실행."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, request.customer_id or customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    sem = asyncio.Semaphore(6)

    async def _plan(camp):
        cid0 = camp.get("nccCampaignId")
        name = camp.get("name") or ""
        async with sem:
            texts: List[str] = []
            try:
                groups = _as_list(await client.get_ad_groups(campaign_id=cid0) or [])
            except Exception:
                return {"campaign_id": cid0, "name": name, "error": "adgroups_fetch"}
            # 대표지역 추정용 소량 샘플
            for g in groups[:2]:
                gid = g.get("nccAdgroupId")
                if not gid:
                    continue
                try:
                    kws = [k.get("keyword") for k in _as_list(await client.get_keywords(ad_group_id=gid) or [])
                           if k.get("keyword")]
                    texts.extend(kws)
                except Exception:
                    pass
        kwon, label, codes, method = _classify_campaign_region(name, texts)
        gids = [g.get("nccAdgroupId") for g in groups if g.get("nccAdgroupId")][: request.max_groups_per_campaign]
        return {"campaign_id": cid0, "name": name, "kwon": kwon, "codes": codes,
                "method": method, "adgroup_ids": gids, "adgroups": len(gids)}

    all_campaigns = [c for c in _as_list(await client.get_campaigns() or [])
                     if (c.get("campaignTp") or "") == "WEB_SITE"]
    if request.campaign_name_contains:
        needle = request.campaign_name_contains.strip()
        all_campaigns = [c for c in all_campaigns if needle in (c.get("name") or "")]
    all_campaigns = all_campaigns[: int(request.max_campaigns)]

    plans = await asyncio.gather(*[_plan(c) for c in all_campaigns])
    # 적용 대상: 코드가 있고(전국=제한없음 제외), only_branch면 branch method만
    def _eligible(p):
        if p.get("error") or not p.get("codes"):
            return False
        if request.only_branch and p.get("method") != "branch":
            return False
        return True

    apply_plans = [p for p in plans if _eligible(p)]
    total_groups = sum(p["adgroups"] for p in apply_plans)

    if request.dry_run:
        from collections import Counter
        by_kwon = Counter(p["kwon"] for p in apply_plans)
        return {
            "success": True, "dry_run": True, "customer_id": int(account["customer_id"]),
            "campaigns_total": len(all_campaigns), "will_apply_campaigns": len(apply_plans),
            "will_apply_adgroups": total_groups, "by_kwon": dict(by_kwon),
            "preview": [{"campaign": p["name"], "권역": p["kwon"], "지역코드": p["codes"],
                         "광고그룹": p["adgroups"], "method": p["method"]}
                        for p in sorted(apply_plans, key=lambda x: x["kwon"])[:120]],
            "skipped": [{"campaign": p["name"], "method": p.get("method"), "reason":
                         ("no_codes/전국" if not p.get("codes") else "not_branch")}
                        for p in plans if not _eligible(p) and not p.get("error")][:40],
            "note": "코드 배열=적용할 시/도(9=서울,2=경기,11=인천,8=부산,6=대구,3=경남,10=울산,4=경북,15=충남 등). "
                    "only_branch=false 로 풀/롱테일 추정지역까지 확장 가능. 확인 후 dry_run=false 실행.",
        }

    async def _run():
        logger.warning(f"[regional-target] 시작 cid={account['customer_id']} "
                       f"캠페인 {len(apply_plans)} / 광고그룹 {total_groups}")
        ok = fail = 0
        gsem = asyncio.Semaphore(5)

        async def _one(gid, codes):
            nonlocal ok, fail
            async with gsem:
                try:
                    await client.set_ad_group_regional_target(gid, codes)
                    ok += 1
                except Exception as e:
                    fail += 1
                    if fail <= 15:
                        logger.warning(f"[regional-target] 실패 gid={gid}: {type(e).__name__} {str(e)[:250]}")
                await asyncio.sleep(0.1)

        for p in apply_plans:
            await asyncio.gather(*[_one(gid, p["codes"]) for gid in p["adgroup_ids"]])
            logger.warning(f"[regional-target] {p['name']} → {p['codes']} (누적 ok={ok} fail={fail})")
        logger.warning(f"[regional-target] 완료 — 성공 {ok} / 실패 {fail} / 광고그룹 {total_groups}")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "dry_run": False,
            "customer_id": int(account["customer_id"]),
            "will_apply_campaigns": len(apply_plans), "will_apply_adgroups": total_groups,
            "message": "백그라운드 지역타겟 설정 시작. fly logs 의 [regional-target] 라인에서 확인."}


class PromoteCoreRequest(BaseModel):
    """중요도 상위(pos1-3) 핵심 키워드를 풀에서 빼내 **고예산 전용 캠페인**으로 이동 + 의료심의 소재 부착.
    100원에 묶여 노출 못하던 강남/역삼 피부과추천류를 실제로 밀어주기 위한 '핵심 전용 분리'."""
    campaign_name: str = Field("소잠_핵심_강남피부", description="생성할 전용 캠페인명(30자 한도)")
    daily_budget: int = Field(30000, ge=100, le=10000000, description="전용 캠페인 일예산(원)")
    score_min: int = Field(48, description="이 점수 이상만 핵심으로 승격. 48=pos3, 60=pos2, 75=pos1")
    init_bid: int = Field(990, ge=70, le=100000, description="핵심 키워드 초기 입찰가(이후 bulk-rank-bid로 순위 최적화)")
    keywords_per_group: int = Field(300, ge=1, le=1000, description="광고그룹당 키워드 수")
    max_keywords: int = Field(20000)
    delete_pool_copies: bool = Field(True, description="승격 후 기존 100원 풀 복사본을 네이버에서 삭제(중복 제거). 실패해도 무해(코어가 경매 우선)")
    defund_campaign_id: Optional[str] = Field(None, description="코어 예산 재원 마련용으로 감액할 기존 캠페인 ID(예: 생 파워링크). 총 예산 상한 유지")
    defund_to_budget: int = Field(100, ge=70, description="defund 캠페인을 이 일예산으로 낮춤(사실상 정지)")
    # 의료심의 소재(기존 승인본 복제 — 심의필 한42606)
    headline: str = Field("두드러기치료, 소잠한의원", description="소재 헤드라인(15자). 심의필과 세트인 승인본 유지 권장")
    description: str = Field("만성두드러기, 콜린성두드러기, 만성피부질환, 해독, 체질개선, 면역강화", description="소재 설명(45자)")
    display_url: str = Field("https://sojam.co.kr")
    final_url: str = Field("https://sojam.co.kr")
    medical_no: str = Field("한42606", description="의료광고 심의필 번호")
    dry_run: bool = Field(True)


@router.post("/keyword-pool/campaigns/promote-core")
async def keyword_pool_promote_core(
    request: PromoteCoreRequest,
    background_tasks: BackgroundTasks,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """중요도 상위(score>=score_min) 핵심 키워드를 풀에서 전용 고예산 캠페인으로 이동 + 의료심의 소재 부착.
    스코어링은 bulk-rank-bid 와 동일(지역/의도/질환/브랜드). dry_run: 승격 대상 분포/샘플/구조 미리보기."""
    import sqlite3 as _sq
    from services.naver_ad_service import NaverAdApiClient
    from database.registered_keywords_db import get_registered_keywords_db
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = int(account.get("customer_id"))

    # ── 스코어링 (bulk-rank-bid 동일 기준) ──
    GEO_TOP = ["강남", "역삼", "논현", "신사", "청담", "압구정", "선릉", "대치", "학동", "신논현", "강남구", "양재", "도곡"]
    GEO_ADJ = ["서초", "방배", "잠원", "개포", "수서", "일원", "세곡", "우면"]
    GEO_OTHER = ["위례", "동탄", "분당", "판교", "수지", "기흥", "일산", "평촌", "산본", "범계", "부천", "안산", "광명", "시흥",
        "인천", "부평", "송도", "부산", "해운대", "서면", "대구", "수성", "범어", "만촌", "광주", "대전", "울산", "세종", "수원", "성남",
        "용인", "고양", "천안", "아산", "청주", "전주", "익산", "군산", "포항", "경주", "구미", "창원", "김해", "양산", "진주",
        "목포", "여수", "순천", "제주", "춘천", "원주", "강릉", "노원", "송파", "마포", "은평", "구로", "관악", "동작", "성북"]
    DIS = ["아토피", "건선", "여드름", "두드러기", "습진", "지루성", "탈모", "무좀", "대상포진", "사마귀", "백반증", "기미", "모낭염",
        "한포진", "주사비", "다한증", "켈로이드", "피부", "한방", "피부염", "피부질환", "가려움", "곤지름", "헤르페스", "티눈",
        "뾰루지", "구내염", "색소침착", "흉터", "모공", "땀띠", "주근깨", "비듬"]
    BRAND = ["소잠"]
    I_BOOK = ["예약", "상담", "문의", "예약문의", "전화상담", "예약하기", "당일"]
    I_COST = ["비용", "가격", "얼마"]
    I_CHOICE = ["추천", "후기", "잘하는곳", "명의", "유명"]
    I_CLINIC = ["한의원", "한방병원", "병원", "의원", "피부과", "클리닉"]
    I_TREAT = ["치료", "한약", "약침", "한방치료", "봉독", "완치", "낫는법"]
    I_INFO = ["증상", "원인", "사진", "이미지", "뜻", "종류", "에좋은", "음식", "민간요법", "전염", "옮나", "초기증상"]

    def _score(kw: str) -> int:
        t = (kw or "").replace(" ", "")
        s = 0
        if any(b in t for b in BRAND): s += 50
        if any(g in t for g in GEO_TOP): s += 40
        elif any(g in t for g in GEO_ADJ): s += 25
        elif any(g in t for g in GEO_OTHER): s -= 25
        if any(i in t for i in I_BOOK): s += 30
        if any(i in t for i in I_COST): s += 25
        if any(i in t for i in I_CHOICE): s += 20
        if any(i in t for i in I_CLINIC): s += 15
        elif any(i in t for i in I_TREAT): s += 8
        if any(i in t for i in I_INFO): s -= 25
        if any(d in t for d in DIS): s += 10
        return s

    reg = get_registered_keywords_db()
    with _sq.connect(reg.db_path, timeout=30.0) as conn:
        rows = conn.execute(
            "SELECT keyword, ncc_keyword_id, ad_group_id FROM registered_keywords "
            "WHERE account_customer_id=? AND ncc_keyword_id IS NOT NULL AND ad_group_id IS NOT NULL AND removed_at IS NULL",
            (cid,),
        ).fetchall()
    # 핵심 선별 + keyword 중복 제거(UNIQUE(account,keyword) — 최고점 1건 유지)
    best: Dict[str, Tuple[int, str, str]] = {}  # keyword -> (score, old_ncc_id, old_gid)
    for kw, nid, gid in rows:
        sc = _score(kw)
        if sc < request.score_min:
            continue
        if kw not in best or sc > best[kw][0]:
            best[kw] = (sc, nid, gid)
    core = [(kw, v[1], v[2], v[0]) for kw, v in best.items()][: request.max_keywords]  # (kw, old_ncc, old_gid, score)

    from collections import Counter as _Counter
    band = _Counter()
    for _, _, _, sc in core:
        b = "pos1(>=75)" if sc >= 75 else "pos2(>=60)" if sc >= 60 else "pos3(>=48)" if sc >= 48 else "sub"
        band[b] += 1
    n_groups = (len(core) + request.keywords_per_group - 1) // max(1, request.keywords_per_group)

    if request.dry_run:
        return {
            "success": True, "dry_run": True, "customer_id": cid,
            "core_selected": len(core), "score_min": request.score_min,
            "band_distribution": dict(band),
            "plan": {
                "new_campaign": request.campaign_name, "daily_budget": request.daily_budget,
                "ad_groups": n_groups, "keywords_per_group": request.keywords_per_group,
                "init_bid": request.init_bid,
                "creative": {"type": "TEXT_45", "headline": request.headline,
                             "medical_no": request.medical_no, "final_url": request.final_url},
                "delete_pool_copies": request.delete_pool_copies,
                "defund_campaign_id": request.defund_campaign_id,
                "defund_to_budget": request.defund_to_budget,
            },
            "samples": sorted([c[0] for c in core[:60]]),
            "note": "dry_run=false 로 실제 승격(캠페인+그룹+키워드+의료소재 생성, DB 이동, 풀 복사본 삭제).",
        }

    if not core:
        return {"success": False, "step": "no_core_keywords", "score_min": request.score_min}

    async def _run():
        client = NaverAdApiClient()
        client.customer_id = account["customer_id"]; client.api_key = account["api_key"]; client.secret_key = account["secret_key"]
        # 1) 비즈채널(WEB_SITE) 확보
        ch_id = None
        try:
            channels = await client.list_business_channels() or []
            ws = [c for c in channels if c.get("channelTp") == "WEB_SITE"] or channels
            if ws:
                ch = ws[0]
                ch_id = (ch.get("nccBusinessChannelId") or ch.get("businessChannelId")
                         or ch.get("nccChannelId") or ch.get("id"))
        except Exception as e:
            logger.error(f"[promote-core] 비즈채널 조회 실패: {e}")
        if not ch_id:
            logger.error("[promote-core] 비즈채널 없음 — 중단"); return
        # 2) 전용 캠페인 생성 (이름 중복 시 suffix)
        cname = request.campaign_name
        new_cid = None
        for attempt in range(3):
            try:
                camp = await client.create_campaign(name=cname, daily_budget=request.daily_budget, campaign_tp="WEB_SITE")
                new_cid = camp.get("nccCampaignId")
                if new_cid:
                    break
            except Exception as e:
                if "already in use" in str(e) or "3506" in str(e):
                    cname = f"{request.campaign_name}_{attempt+2}"; await asyncio.sleep(0.3); continue
                logger.error(f"[promote-core] 캠페인 생성 실패: {e}"); return
        if not new_cid:
            logger.error("[promote-core] 캠페인 ID 없음 — 중단"); return
        logger.warning(f"[promote-core] 캠페인 생성 ✓ {cname} ({new_cid}) 예산={request.daily_budget} 핵심={len(core)}")
        # 3) 그룹별: 광고그룹 → 키워드(100/콜) → 의료소재
        moved: List[Tuple[str, str, str]] = []  # (keyword, new_ncc_id, gid)
        old_ncc_by_kw = {c[0]: c[1] for c in core}
        grp_idx = 0
        for i in range(0, len(core), request.keywords_per_group):
            chunk = core[i:i + request.keywords_per_group]
            grp_idx += 1
            gname = f"{request.campaign_name}_grp_{grp_idx:03d}"
            try:
                ag = await client.create_ad_group(campaign_id=new_cid, name=gname, bid_amt=request.init_bid,
                                                   business_channel_id=ch_id)
                gid = ag.get("nccAdgroupId")
                if not gid:
                    logger.warning(f"[promote-core] 그룹 {gname} ID 없음 — skip"); continue
            except Exception as e:
                logger.warning(f"[promote-core] 그룹 생성 실패 {gname}: {str(e)[:120]}"); continue
            # 키워드 100개씩
            for j in range(0, len(chunk), 100):
                sub = chunk[j:j + 100]
                body = [{"nccAdgroupId": gid, "keyword": c[0], "bidAmt": request.init_bid, "useGroupBidAmt": False}
                        for c in sub]
                try:
                    res = await client.create_keywords(body, ad_group_id=gid) or []
                    for k in res:
                        kt, knid = (k.get("keyword") or "").strip(), k.get("nccKeywordId")
                        if kt and knid:
                            moved.append((kt, knid, gid))
                except Exception as e:
                    logger.warning(f"[promote-core] 키워드 생성 실패 grp{grp_idx} sub{j}: {str(e)[:120]}")
                await asyncio.sleep(0.15)
            # 의료심의 소재 1개
            try:
                await client.create_ad(
                    ad_group_id=gid, headline_pc=request.headline, description_pc=request.description,
                    display_url=request.display_url, final_url_pc=request.final_url,
                    medical_no=request.medical_no,
                )
            except Exception as e:
                logger.warning(f"[promote-core] 소재 생성 실패 grp{grp_idx}: {str(e)[:120]}")
            await asyncio.sleep(0.2)
        logger.warning(f"[promote-core] 이동 생성 완료 — {len(moved)}개 (그룹 {grp_idx})")
        # 4) DB 이동 반영 (keyword 기준 UPDATE → 코어 캠페인 지시)
        try:
            with _sq.connect(reg.db_path, timeout=30.0) as conn:
                for kt, knid, gid in moved:
                    conn.execute(
                        "UPDATE registered_keywords SET campaign_id=?, ad_group_id=?, ncc_keyword_id=?, bid_amt=? "
                        "WHERE account_customer_id=? AND keyword=?",
                        (new_cid, gid, knid, request.init_bid, cid, kt),
                    )
                conn.commit()
        except Exception as e:
            logger.warning(f"[promote-core] DB 업데이트 실패: {str(e)[:120]}")
        # 5) 풀 복사본 삭제 (성공 이동분의 옛 ncc_id — 실패해도 무해)
        if request.delete_pool_copies:
            old_ids = [old_ncc_by_kw[kt] for kt, _, _ in moved if kt in old_ncc_by_kw and old_ncc_by_kw[kt]]
            done = 0; fail = 0
            for i in range(0, len(old_ids), 100):
                try:
                    await client.delete_keywords_bulk(old_ids[i:i + 100]); done += len(old_ids[i:i + 100])
                except Exception:
                    fail += len(old_ids[i:i + 100])
                await asyncio.sleep(0.1)
            logger.warning(f"[promote-core] 풀 복사본 삭제 — {done} (실패 {fail})")
        # 6) 코어 예산 재원 마련 — 지정 캠페인 감액(총 상한 유지)
        if request.defund_campaign_id:
            try:
                base = await client.get_campaign(request.defund_campaign_id)
                await client.update_campaign_budget(
                    request.defund_campaign_id, request.defund_to_budget,
                    base=base if isinstance(base, dict) else None)
                logger.warning(f"[promote-core] defund {request.defund_campaign_id} → {request.defund_to_budget}원")
            except Exception as e:
                logger.warning(f"[promote-core] defund 실패: {str(e)[:120]}")
        logger.warning(f"[promote-core] 전체 완료 — 캠페인 {cname}({new_cid}) 핵심 {len(moved)}개 이동")

    background_tasks.add_task(_run)
    return {"success": True, "started": True, "customer_id": cid,
            "core_selected": len(core), "band_distribution": dict(band),
            "new_campaign": request.campaign_name, "daily_budget": request.daily_budget,
            "ad_groups_planned": n_groups,
            "message": "핵심 전용 캠페인 승격 백그라운드 시작 (로그 [promote-core])"}


@router.get("/keyword-pool/diagnostics/ad-by-id")
async def keyword_pool_diag_ad_by_id(
    ad_id: str = Query(..., description="조회할 광고소재 nccAdId"),
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """특정 광고소재(nccAdId) raw 조회 — type/payload 구조 확인용 (진단 전용)."""
    from services.naver_ad_service import NaverAdApiClient
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]
    try:
        src = await client.get_ad_by_id(ad_id)
        return {"success": True, "ad_id": ad_id, "raw": src}
    except Exception as e:
        return {"success": False, "error": f"{type(e).__name__}: {str(e)[:300]}"}


@router.get("/keyword-pool/diagnostics/ad-creatives")
async def keyword_pool_diag_ad_creatives(
    customer_id: Optional[str] = None,
    sample: int = 30,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """광고그룹 샘플의 소재(헤드라인+심의상태) 조회 — 소재 백필/심의 진행 확인용 (읽기)."""
    from services.naver_ad_service import NaverAdApiClient
    import random as _rnd
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    def _as_list(x):
        if isinstance(x, list):
            return x
        if isinstance(x, dict):
            return x.get("data") or x.get("list") or []
        return []

    camps = _as_list(await client.get_campaigns() or [])
    gids: List[str] = []
    for c in camps:
        try:
            for g in _as_list(await client.get_ad_groups(campaign_id=c.get("nccCampaignId")) or []):
                if g.get("nccAdgroupId"):
                    gids.append(g["nccAdgroupId"])
        except Exception:
            pass
        await asyncio.sleep(0.05)
    if not gids:
        return {"success": False, "step": "no_ad_groups"}
    sampled = _rnd.sample(gids, min(int(sample), len(gids)))
    from collections import Counter as _C
    head_ct, status_ct, url_ct, no_ad = _C(), _C(), _C(), 0
    samples = []
    raw_with_simui = None  # 한42606 등 심의필 들어간 raw ad 객체 (필드 위치 파악용)
    for gid in sampled:
        try:
            ads = _as_list(await client.get_ads(ad_group_id=gid) or [])
        except Exception:
            continue
        if not ads:
            no_ad += 1
            continue
        if raw_with_simui is None:
            import json as _json
            for a in ads:
                _txt = _json.dumps(a, ensure_ascii=False)
                if "42606" in _txt or "심의" in _txt:
                    raw_with_simui = a
                    break
        for a in ads:
            adobj = a.get("ad") if isinstance(a, dict) else {}
            adobj = adobj if isinstance(adobj, dict) else {}
            h = (adobj.get("headline") or (adobj.get("pc") or {}).get("headline") or "")[:30]
            url = (adobj.get("finalUrl") or adobj.get("displayUrl")
                   or (adobj.get("pc") or {}).get("final") or "")
            host = "sojam.co.kr" if "sojam.co.kr" in url else ("blog.naver" if "blog.naver" in url else (url[:24] or "?"))
            st = a.get("inspectStatus") or a.get("status") or "?"
            head_ct[h] += 1
            status_ct[st] += 1
            url_ct[host] += 1
            if len(samples) < 15:
                samples.append({"headline": h, "status": st, "url_host": host})
        await asyncio.sleep(0.05)
    return {
        "success": True, "customer_id": int(account.get("customer_id")),
        "ad_groups_total": len(gids), "sampled": len(sampled), "groups_without_ad": no_ad,
        "headline_distribution": dict(head_ct.most_common(10)),
        "status_distribution": dict(status_ct),
        "url_host_distribution": dict(url_ct),
        "samples": samples,
        "raw_ad_with_simui": raw_with_simui,
    }


# ============ 전자동 광맥 발굴 — Domain Profile API (Stage 2) ============

class DomainProfileGenerateRequest(BaseModel):
    description: str = Field(..., description="사업 설명 한 줄 (예: 의료인 대상 대출 — 병원/약사/한의사대출)")
    target_count: int = Field(100000, ge=1000, le=100000)


class DomainProfileSaveRequest(BaseModel):
    description: Optional[str] = None
    atom_library: Optional[Dict[str, Any]] = None
    relevance_keywords: Optional[List[str]] = None
    negative_keywords: Optional[List[str]] = None
    enabled: Optional[bool] = None
    min_score: Optional[int] = None
    target_count: Optional[int] = None
    daily_budget: Optional[int] = None
    default_bid: Optional[int] = None
    ad_template_id: Optional[int] = None
    category_split: Optional[bool] = None
    nonmedical_budget: Optional[int] = None
    required_tokens: Optional[List[str]] = None
    medical_no: Optional[str] = None


@router.get("/keyword-pool/domain-profile")
async def keyword_pool_get_domain_profile(
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """저장된 도메인 프로파일 조회 (자동화 설정 화면용)."""
    from database.naver_ad_db import get_domain_profile
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    return {"success": True, "profile": get_domain_profile(user_id, str(account.get("customer_id")))}


@router.post("/keyword-pool/domain-profile/generate")
async def keyword_pool_generate_domain_profile(
    request: DomainProfileGenerateRequest,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """사업 설명 → LLM 이 atom_library/relevance/negative/예시시드 생성 (검수용, 저장 X)."""
    from services.ai_seed_suggester import generate_domain_profile
    return await generate_domain_profile(request.description, request.target_count)


@router.post("/keyword-pool/domain-profile/save")
async def keyword_pool_save_domain_profile(
    request: DomainProfileSaveRequest,
    customer_id: Optional[str] = None,
    user_id: int = Depends(get_user_id_with_fallback),
):
    """검수한 도메인 프로파일 저장 + 자동화 ON/OFF. None 필드는 건너뜀."""
    from database.naver_ad_db import update_domain_profile, get_domain_profile
    account = _resolve_account(user_id, customer_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    cid = str(account.get("customer_id"))
    fields = {}
    for k in ("description", "atom_library", "relevance_keywords", "negative_keywords",
              "enabled", "min_score", "target_count", "daily_budget", "default_bid", "ad_template_id",
              "category_split", "nonmedical_budget", "required_tokens", "medical_no"):
        v = getattr(request, k, None)
        if v is not None:
            fields[k] = v
    ok = update_domain_profile(user_id, cid, **fields)
    return {"success": ok, "profile": get_domain_profile(user_id, cid)}


@router.get("/ad-templates")
async def list_ad_templates(user_id: int = Depends(get_user_id_with_fallback)):
    account = get_ad_account(user_id)
    if not account:
        return {"success": False, "templates": [], "extensions": []}
    customer_id = int(account.get("customer_id"))
    db = get_ad_templates_db()
    return {
        "success": True,
        "customer_id": customer_id,
        "templates": db.list_templates(user_id, customer_id),
        "extensions": db.list_extensions(user_id, customer_id, active_only=False),
    }


@router.post("/ad-templates")
async def create_ad_template(
    request: AdTemplateCreate,
    user_id: int = Depends(get_user_id_with_fallback),
):
    account = get_ad_account(user_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    customer_id = int(account.get("customer_id"))
    db = get_ad_templates_db()
    tpl_id = db.create_template(
        user_id, customer_id,
        headline_pc=request.headline_pc,
        description_pc=request.description_pc,
        display_url=request.display_url,
        final_url_pc=request.final_url_pc,
        headline_mobile=request.headline_mobile,
        description_mobile=request.description_mobile,
        final_url_mobile=request.final_url_mobile,
        is_active=request.is_active,
    )
    return {"success": True, "id": tpl_id}


@router.patch("/ad-templates/{tpl_id}/active")
async def toggle_ad_template(
    tpl_id: int,
    is_active: bool = Query(...),
    user_id: int = Depends(get_user_id_with_fallback),
):
    db = get_ad_templates_db()
    db.update_active(tpl_id, is_active)
    return {"success": True}


@router.delete("/ad-templates/{tpl_id}")
async def delete_ad_template(
    tpl_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
):
    db = get_ad_templates_db()
    db.delete_template(tpl_id, user_id)
    return {"success": True}


@router.post("/ad-templates/extensions")
async def create_ad_extension_template(
    request: AdExtensionCreate,
    user_id: int = Depends(get_user_id_with_fallback),
):
    account = get_ad_account(user_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    customer_id = int(account.get("customer_id"))
    db = get_ad_templates_db()
    ext_id = db.create_extension(user_id, customer_id, request.kind, request.payload)
    return {"success": True, "id": ext_id}


@router.delete("/ad-templates/extensions/{ext_id}")
async def delete_ad_extension_template(
    ext_id: int,
    user_id: int = Depends(get_user_id_with_fallback),
):
    db = get_ad_templates_db()
    db.delete_extension(ext_id, user_id)
    return {"success": True}


# 확장소재 응답에서 payload로 보존할 키 (ownerId/Type, ID, 시각 메타 제외)
_EXT_META_KEYS = {
    "nccAdExtensionId", "ownerId", "ownerType", "customerId",
    "type", "status", "statusReason", "regTm", "editTm", "delFlag",
    "userLock", "inspectStatus", "label", "name",
}


def _extract_ext_payload(item: Dict[str, Any]) -> Dict[str, Any]:
    """네이버 확장소재 응답에서 payload(= create 시 보낼 본문)만 추출."""
    out: Dict[str, Any] = {}
    for k, v in (item or {}).items():
        if k in _EXT_META_KEYS:
            continue
        if v is None:
            continue
        out[k] = v
    return out


@router.post("/ad-templates/import")
async def import_ad_templates_from_naver(
    user_id: int = Depends(get_user_id_with_fallback),
):
    """네이버에 이미 등록된 광고 소재(T&D) + 확장소재를 끌어와 템플릿으로 저장.

    - 광고그룹 전체 순회 → 각 그룹의 ads, adextensions GET
    - 동일 콘텐츠는 중복 저장 안 함 (헤드라인+설명+URL 4-tuple / kind+payload 일치)
    - 비활성 소재(userLock 등)는 그대로 가져오되 is_active=1로 저장 (사용자가 화면에서 토글)
    """
    from services.naver_ad_service import NaverAdApiClient

    account = get_ad_account(user_id)
    if not account or not account.get("is_connected"):
        raise HTTPException(status_code=400, detail="광고 계정 미연결")
    customer_id = int(account.get("customer_id"))

    client = NaverAdApiClient()
    client.customer_id = account["customer_id"]
    client.api_key = account["api_key"]
    client.secret_key = account["secret_key"]

    db = get_ad_templates_db()

    # 1) 광고그룹 전체 조회
    try:
        ad_groups = await client.get_ad_groups()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"광고그룹 조회 실패: {e}")
    if not isinstance(ad_groups, list):
        ad_groups = []

    tpl_imported = 0
    tpl_skipped = 0
    ext_imported = 0
    ext_skipped = 0
    ads_total_seen = 0
    ads_missing_field = 0
    exts_total_seen = 0
    errors: List[str] = []
    sample_ad: Optional[Dict[str, Any]] = None
    sample_ext: Optional[Dict[str, Any]] = None
    sample_field_check: Optional[Dict[str, Any]] = None

    def _unwrap_list(resp: Any) -> List[Any]:
        """네이버 API 응답이 raw list 또는 {list:[...]}/{data:[...]}/{items:[...]} 등 wrap된 경우 풀어냄."""
        if isinstance(resp, list):
            return resp
        if isinstance(resp, dict):
            for k in ("list", "data", "items", "ads", "extensions", "results", "content"):
                v = resp.get(k)
                if isinstance(v, list):
                    return v
            # dict인데 wrap key 없으면 단일 객체로 보고 [resp] 반환
            if resp.get("nccAdId") or resp.get("nccAdExtensionId"):
                return [resp]
        return []

    def _first_str(*candidates) -> str:
        """여러 후보 중 비어있지 않은 첫 문자열 반환. 리스트면 첫 원소 사용."""
        for c in candidates:
            if c is None:
                continue
            if isinstance(c, list):
                for item in c:
                    if isinstance(item, str) and item.strip():
                        return item.strip()
                    if isinstance(item, dict):
                        # RSP_AD: [{text: "..."}, ...] 패턴
                        s = item.get("text") or item.get("value") or item.get("headline") or item.get("description")
                        if isinstance(s, str) and s.strip():
                            return s.strip()
            elif isinstance(c, str) and c.strip():
                return c.strip()
            elif isinstance(c, dict):
                s = c.get("text") or c.get("value")
                if isinstance(s, str) and s.strip():
                    return s.strip()
        return ""

    seen_ext_ids: set = set()

    # ── 확장소재: 다중 fallback 전략 ──
    # 1) 계정 전체 (no params) — 가장 깔끔한 케이스
    # 2) 비어있으면 → 캠페인 ID 모두 돌면서 ?ownerId={cp_id}
    # 광고그룹 단위는 네이버가 404 반환하므로 시도 안 함
    all_exts: List[Any] = []
    try:
        all_ext_resp = await client.get_ad_extensions(owner_id=None)
        all_exts = _unwrap_list(all_ext_resp)
    except Exception as e:
        errors.append(f"exts-all: {str(e)[:120]}")

    # 폴백: 계정 전체 0건이면 캠페인 ID 별로 시도
    if not all_exts:
        try:
            cps = await client.get_campaigns()
            cps_list = _unwrap_list(cps)
        except Exception as e:
            errors.append(f"campaigns: {str(e)[:120]}")
            cps_list = []
        for cp in cps_list:
            cid = cp.get("nccCampaignId") if isinstance(cp, dict) else None
            if not cid:
                continue
            try:
                cp_ext_resp = await client.get_ad_extensions(owner_id=cid)
                cp_list = _unwrap_list(cp_ext_resp)
                all_exts.extend(cp_list)
            except Exception as e:
                errors.append(f"exts-cp({cid}): {str(e)[:120]}")
            await asyncio.sleep(0.15)

    for ex in all_exts:
        ext_id = (ex or {}).get("nccAdExtensionId") if isinstance(ex, dict) else None
        if ext_id and ext_id in seen_ext_ids:
            continue
        if ext_id:
            seen_ext_ids.add(ext_id)
        exts_total_seen += 1
        if sample_ext is None and isinstance(ex, dict):
            sample_ext = ex
        try:
            kind = (ex or {}).get("type") or (ex or {}).get("kind") or ""
            if not kind:
                continue
            payload = _extract_ext_payload(ex)
            if not payload:
                payload = {
                    k: v for k, v in (ex or {}).items()
                    if k not in (
                        "ownerId", "nccAdExtensionId", "createdDate", "editedDate",
                        "regTime", "editTime", "customerId",
                    )
                }
            if not payload:
                continue
            res = db.get_or_create_extension(user_id, customer_id, kind, payload)
            if res.get("created"):
                ext_imported += 1
            else:
                ext_skipped += 1
        except Exception as e:
            errors.append(f"ext-parse: {str(e)[:120]}")

    # 광고그룹 응답 wrap 처리
    ad_groups = _unwrap_list(ad_groups)

    for ag in ad_groups:
        ag_id = ag.get("nccAdgroupId") if isinstance(ag, dict) else None
        if not ag_id:
            continue

        # 2) 소재 조회
        try:
            ads_resp = await client.get_ads(ag_id)
        except Exception as e:
            errors.append(f"ads({ag_id}): {str(e)[:120]}")
            ads_resp = []
        ads = _unwrap_list(ads_resp)

        for a in ads:
            ads_total_seen += 1
            if sample_ad is None and isinstance(a, dict):
                sample_ad = a
            try:
                ad_type = (a or {}).get("type") or ""
                raw_ad = (a or {}).get("ad")
                if isinstance(raw_ad, str):
                    try:
                        ad = _json_lib.loads(raw_ad)
                    except Exception:
                        ad = {}
                elif isinstance(raw_ad, dict):
                    ad = raw_ad
                else:
                    ad = a if isinstance(a, dict) else {}

                pc = ad.get("pc") if isinstance(ad.get("pc"), dict) else {}
                mo = ad.get("mobile") if isinstance(ad.get("mobile"), dict) else {}

                # ─── RSA_AD: assets 배열에서 HEADLINE/DESCRIPTION/URL 분리 ───
                # 새 시스템(RSA_AD)은 assets 배열 사용. 각 asset에 linkType과 assetData.text.
                headlines: List[str] = []
                descriptions: List[str] = []
                if ad_type == "RSA_AD" or (a or {}).get("assets"):
                    assets = (a or {}).get("assets") or ad.get("assets") or []
                    if isinstance(assets, list):
                        for asset in assets:
                            if not isinstance(asset, dict):
                                continue
                            link_type = asset.get("linkType") or ""
                            asset_data = asset.get("assetData") or {}
                            text_v = asset_data.get("text") if isinstance(asset_data, dict) else None
                            if not text_v or not isinstance(text_v, str):
                                continue
                            text_v = text_v.strip()
                            if not text_v:
                                continue
                            if link_type == "HEADLINE":
                                headlines.append(text_v)
                            elif link_type == "DESCRIPTION":
                                descriptions.append(text_v)

                # ─── URLs (RSA_AD 기준: pc.display, pc.final / 레거시 폴백 포함) ───
                display_url = _first_str(
                    pc.get("display"), pc.get("displayUrl"), pc.get("display_url"),
                    ad.get("displayUrl"), ad.get("display_url"),
                    (a or {}).get("displayUrl"),
                )
                final_url_pc = _first_str(
                    pc.get("final"), pc.get("finalUrl"), pc.get("landingUrl"),
                    ad.get("finalUrl"), ad.get("landingUrl"), ad.get("finalPcUrl"),
                    (a or {}).get("finalUrl"),
                )
                final_url_mobile = _first_str(
                    mo.get("final"), mo.get("finalUrl"), mo.get("landingUrl"),
                    ad.get("finalMobileUrl"),
                ) or final_url_pc

                # 레거시 TEXT_45: 단일 headline/description 폴백
                if not headlines:
                    legacy_h = _first_str(
                        pc.get("headline"), pc.get("title"),
                        ad.get("headline"), ad.get("title"),
                        ad.get("headlines"), pc.get("headlines"),
                    )
                    if legacy_h:
                        headlines = [legacy_h]
                if not descriptions:
                    legacy_d = _first_str(
                        pc.get("description"), pc.get("desc"),
                        ad.get("description"), ad.get("desc"),
                        ad.get("descriptions"), pc.get("descriptions"),
                    )
                    if legacy_d:
                        descriptions = [legacy_d]

                # display_url이 비면 final_url_pc 도메인으로 폴백
                if not display_url and final_url_pc:
                    try:
                        from urllib.parse import urlparse
                        u = urlparse(final_url_pc)
                        if u.netloc:
                            display_url = f"{u.scheme or 'https'}://{u.netloc}"
                    except Exception:
                        pass
                if not display_url:
                    display_url = final_url_pc

                if sample_field_check is None:
                    sample_field_check = {
                        "ad_type": ad_type,
                        "headlines_count": len(headlines),
                        "descriptions_count": len(descriptions),
                        "headlines_sample": headlines[:3],
                        "descriptions_sample": [d[:30] for d in descriptions[:2]],
                        "display_url": display_url,
                        "final_url_pc": final_url_pc,
                        "ad_top_keys": list((a or {}).keys()) if isinstance(a, dict) else [],
                        "ad_inner_keys": list(ad.keys()) if isinstance(ad, dict) else [],
                        "pc_keys": list(pc.keys()) if isinstance(pc, dict) else [],
                    }

                if not (headlines and descriptions and final_url_pc):
                    ads_missing_field += 1
                    continue

                # RSA_AD: 헤드라인 × 설명 페어를 N=min(len(h), len(d)) 만큼 생성
                # (cross-product는 너무 많아짐 — 인덱스 매칭이 자연스러움)
                # headlines가 더 많으면 description을 라운드로빈
                pair_n = max(len(headlines), len(descriptions))
                pair_n = min(pair_n, 10)  # 광고당 최대 10개 템플릿
                for i in range(pair_n):
                    h = headlines[i % len(headlines)]
                    d = descriptions[i % len(descriptions)]
                    res = db.get_or_create_template(
                        user_id, customer_id,
                        headline_pc=h[:15],
                        description_pc=d[:45],
                        display_url=display_url,
                        final_url_pc=final_url_pc,
                        headline_mobile=h[:15],
                        description_mobile=d[:45],
                        final_url_mobile=final_url_mobile or final_url_pc,
                    )
                    if res.get("created"):
                        tpl_imported += 1
                    else:
                        tpl_skipped += 1
            except Exception as e:
                errors.append(f"ad-parse: {str(e)[:120]}")

        await asyncio.sleep(0.2)

    return {
        "success": True,
        "ad_groups_scanned": len(ad_groups),
        "templates_imported": tpl_imported,
        "templates_skipped_duplicate": tpl_skipped,
        "extensions_imported": ext_imported,
        "extensions_skipped_duplicate": ext_skipped,
        "ads_total_seen": ads_total_seen,
        "ads_missing_field": ads_missing_field,
        "exts_total_seen": exts_total_seen,
        "sample_ad_raw": sample_ad,
        "sample_ext_raw": sample_ext,
        "sample_field_check": sample_field_check,
        "errors": errors[:20],
    }
